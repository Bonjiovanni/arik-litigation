"""
fw_walk_grp4.py
---------------
Walk_Coverage and Walk_History sheet management for fw_walk.py.

Provides:
    ensure_walk_coverage_sheet   — create/return Walk_Coverage sheet
    get_next_coverage_id         — generate next WC##### ID
    update_walk_coverage         — append one directory walk record
    ensure_walk_history_sheet    — create/return Walk_History sheet
    log_walk_run                 — append one full-run summary record

Header style: bold white on dark blue (#2F5496), centered, frozen row 1,
auto-filter enabled (matches all other sheets in this workbook).

Python 3.11, Windows 10. Requires openpyxl.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Shared header style
# ---------------------------------------------------------------------------

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# Walk_Coverage sheet
# ---------------------------------------------------------------------------

_WC_SHEET_NAME = "Walk_Coverage"

_WC_COLUMNS = [
    ("CoverageID",      12),
    ("RunID",           20),
    ("DirPath",         70),
    ("ProcessingLevel", 18),
    ("FilesInserted",   14),
    ("FilesUpdated",    14),
    ("FilesSkipped",    14),
    ("Errors",          10),
    ("WalkedAt",        20),
]

_WC_COL_INDEX: dict[str, int] = {
    name: idx + 1 for idx, (name, _) in enumerate(_WC_COLUMNS)
}


# ---------------------------------------------------------------------------
# FUNCTION 1: ensure_walk_coverage_sheet
# ---------------------------------------------------------------------------

def ensure_walk_coverage_sheet(wb: Workbook) -> Worksheet:
    """Create and style the Walk_Coverage sheet if absent; return it.

    Idempotent — returns the existing sheet unchanged if already present.

    Parameters
    ----------
    wb : Workbook

    Returns
    -------
    Worksheet
    """
    if _WC_SHEET_NAME in wb.sheetnames:
        return wb[_WC_SHEET_NAME]

    ws = wb.create_sheet(title=_WC_SHEET_NAME)

    for col_idx, (col_name, col_width) in enumerate(_WC_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_WC_COLUMNS))}1"

    return ws


# ---------------------------------------------------------------------------
# FUNCTION 2: get_next_coverage_id
# ---------------------------------------------------------------------------

def get_next_coverage_id(ws: Worksheet) -> str:
    """Return the next available CoverageID (e.g. "WC00042").

    Scans column A for values matching the pattern WC##### and returns
    one above the current maximum. Returns "WC00001" if the sheet has
    no data rows or no parseable IDs.

    Parameters
    ----------
    ws : Worksheet  — Walk_Coverage sheet

    Returns
    -------
    str
    """
    max_num = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = str(row[0] or "").strip().upper()
        if val.startswith("WC") and val[2:].isdigit():
            num = int(val[2:])
            if num > max_num:
                max_num = num
    return f"WC{max_num + 1:05d}"


# ---------------------------------------------------------------------------
# FUNCTION 3: update_walk_coverage
# ---------------------------------------------------------------------------

def update_walk_coverage(
    ws: Worksheet,
    run_id: str,
    dir_path: str,
    processing_level: str,
    files_inserted: int,
    files_updated: int,
    files_skipped: int,
    errors: int,
) -> int:
    """Append a new row to Walk_Coverage for one directory walk.

    Sets WalkedAt to datetime.now().isoformat() at call time.

    Parameters
    ----------
    ws : Worksheet
        Walk_Coverage sheet (use ensure_walk_coverage_sheet first).
    run_id : str
    dir_path : str
        Forward-slash normalized directory path.
    processing_level : str
        Processing level reached (e.g. "file_listed").
    files_inserted : int
    files_updated : int
    files_skipped : int
    errors : int

    Returns
    -------
    int
        1-based row number written.
    """
    next_row = max(ws.max_row + 1, 2)
    coverage_id = get_next_coverage_id(ws)

    row_data = {
        "CoverageID":      coverage_id,
        "RunID":           run_id,
        "DirPath":         dir_path.replace("\\", "/"),
        "ProcessingLevel": processing_level,
        "FilesInserted":   files_inserted,
        "FilesUpdated":    files_updated,
        "FilesSkipped":    files_skipped,
        "Errors":          errors,
        "WalkedAt":        datetime.now().isoformat(),
    }

    for col_name, col_idx in _WC_COL_INDEX.items():
        ws.cell(row=next_row, column=col_idx, value=row_data[col_name])

    return next_row


# ---------------------------------------------------------------------------
# Walk_History sheet
# ---------------------------------------------------------------------------

_WH_SHEET_NAME = "Walk_History"

_WH_COLUMNS = [
    ("RunID",             20),
    ("StartedAt",         20),
    ("CompletedAt",       20),
    ("ScanDirs",          80),
    ("ProcessingLevel",   18),
    ("TotalInserted",     14),
    ("TotalUpdated",      14),
    ("TotalSkippedFiles", 16),
    ("TotalSkippedDirs",  15),
    ("TotalErrors",       12),
    ("OverlapAction",     14),
    ("Notes",             40),
]

_WH_COL_INDEX: dict[str, int] = {
    name: idx + 1 for idx, (name, _) in enumerate(_WH_COLUMNS)
}


# ---------------------------------------------------------------------------
# FUNCTION 4: ensure_walk_history_sheet
# ---------------------------------------------------------------------------

def ensure_walk_history_sheet(wb: Workbook) -> Worksheet:
    """Create and style the Walk_History sheet if absent; return it.

    Idempotent — returns the existing sheet unchanged if already present.

    Parameters
    ----------
    wb : Workbook

    Returns
    -------
    Worksheet
    """
    if _WH_SHEET_NAME in wb.sheetnames:
        return wb[_WH_SHEET_NAME]

    ws = wb.create_sheet(title=_WH_SHEET_NAME)

    for col_idx, (col_name, col_width) in enumerate(_WH_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_WH_COLUMNS))}1"

    return ws


# ---------------------------------------------------------------------------
# FUNCTION 5: log_walk_run
# ---------------------------------------------------------------------------

def log_walk_run(
    wb: Workbook,
    run_id: str,
    started_at: str,
    scan_dirs: "list[str]",
    processing_level: str,
    stats: dict,
    overlap_action: str,
    notes: str = "",
) -> int:
    """Append one run-summary row to Walk_History.

    Ensures the Walk_History sheet exists before writing.
    Sets CompletedAt to datetime.now().isoformat() at call time.

    Parameters
    ----------
    wb : Workbook
    run_id : str
    started_at : str
        ISO datetime string recorded when the run began.
    scan_dirs : list[str]
        Directories that were walked; joined with " | " for the ScanDirs cell.
    processing_level : str
    stats : dict
        Must contain: inserted, updated, skipped_files, skipped_dirs, errors.
    overlap_action : str
        "skip", "rewalk", or "abort".
    notes : str, optional

    Returns
    -------
    int
        1-based row number written.
    """
    ws = ensure_walk_history_sheet(wb)
    next_row = max(ws.max_row + 1, 2)

    row_data = {
        "RunID":             run_id,
        "StartedAt":         started_at,
        "CompletedAt":       datetime.now().isoformat(),
        "ScanDirs":          " | ".join(d.replace("\\", "/") for d in scan_dirs),
        "ProcessingLevel":   processing_level,
        "TotalInserted":     stats.get("inserted", 0),
        "TotalUpdated":      stats.get("updated", 0),
        "TotalSkippedFiles": stats.get("skipped_files", 0),
        "TotalSkippedDirs":  stats.get("skipped_dirs", 0),
        "TotalErrors":       stats.get("errors", 0),
        "OverlapAction":     overlap_action,
        "Notes":             notes,
    }

    for col_name, col_idx in _WH_COL_INDEX.items():
        ws.cell(row=next_row, column=col_idx, value=row_data[col_name])

    return next_row
