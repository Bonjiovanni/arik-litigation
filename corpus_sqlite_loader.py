"""
corpus_sqlite_loader.py — Loads email + attachment data into litigation_corpus.db.

Reads directly from Excel via openpyxl (read-only mode).
File paths and sheet name are CLI args — not hardcoded.

Usage:
    python corpus_sqlite_loader.py \
        --emails "path/to/V11.xlsx" \
        --sheet "Merge1" \
        --attachments "path/to/manifest.xlsx" \
        [--db "path/to/litigation_corpus.db"]

Rebuild behavior:
    - Email domain tables are dropped and recreated on every run
    - Entity tables and other domain tables are NEVER touched
    - _column_map rows for emails/attachments are replaced
"""

import argparse
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from corpus_sqlite_schema import (
    DEFAULT_DB_PATH,
    build_attachments_column_mapping,
    build_email_column_mapping,
    connect,
    create_email_tables,
    create_shared_tables,
)


def read_excel_headers(path: str, sheet_name: str | None = None) -> list[str]:
    """Read header row from an Excel file."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    return headers


def read_excel_data(path: str, sheet_name: str | None = None):
    """Yield data rows from an Excel file as lists of values."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(min_row=2)  # skip header
    for row in rows:
        yield [cell.value for cell in row]
    wb.close()


def coerce_value(value, data_type: str):
    """Coerce a cell value to the declared type, preserving NULLs."""
    if value is None:
        return None
    if data_type == "INTEGER":
        if isinstance(value, (int, float)):
            return int(value)
        # String that looks numeric
        s = str(value).strip()
        if not s:
            return None
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None
    # TEXT — convert to string, but keep None as None
    if isinstance(value, str):
        return value if value.strip() else None
    return str(value)


def load_emails(conn: sqlite3.Connection, emails_path: str, sheet_name: str,
                email_mapping: list[tuple[str, str, str]]) -> int:
    """Load emails from Excel into emails_master. Returns row count."""
    snake_names = [snake for _, snake, _ in email_mapping]
    types = [dtype for _, _, dtype in email_mapping]
    placeholders = ", ".join(["?"] * len(snake_names))
    col_names = ", ".join(snake_names)

    insert_sql = f"INSERT INTO emails_master ({col_names}) VALUES ({placeholders})"

    count = 0
    for row_values in read_excel_data(emails_path, sheet_name):
        # Pad or truncate to match column count
        while len(row_values) < len(snake_names):
            row_values.append(None)
        row_values = row_values[:len(snake_names)]

        # Coerce types
        coerced = [coerce_value(v, t) for v, t in zip(row_values, types)]
        conn.execute(insert_sql, coerced)
        count += 1

    return count


def load_attachments(conn: sqlite3.Connection, attachments_path: str,
                     attach_mapping: list[tuple[str, str, str]]) -> tuple[int, int]:
    """Load attachments from Excel into attachments table.

    Returns (total_loaded, orphan_count).
    All rows get extraction_status='pending' and extracted_text=NULL.
    Orphan attachments (message_id not in emails_master) are loaded with
    extraction_status='pending_orphan' so they can be identified.
    """
    manifest_snake = [snake for _, snake, _ in attach_mapping]
    manifest_types = [dtype for _, _, dtype in attach_mapping]

    # Build set of valid email message_ids for orphan detection
    valid_ids = set(
        r[0] for r in conn.execute(
            "SELECT rfc_822_message_id FROM emails_master"
        ).fetchall()
    )

    all_cols = manifest_snake + ["extraction_status"]
    placeholders = ", ".join(["?"] * len(all_cols))
    col_names = ", ".join(all_cols)

    insert_sql = f"INSERT INTO attachments ({col_names}) VALUES ({placeholders})"

    count = 0
    orphan_count = 0

    # Temporarily disable FK for loading — we handle orphans ourselves
    # PRAGMA foreign_keys can only be changed outside a transaction
    conn.commit()
    conn.execute("PRAGMA foreign_keys = OFF")

    for row_values in read_excel_data(attachments_path):
        while len(row_values) < len(manifest_snake):
            row_values.append(None)
        row_values = row_values[:len(manifest_snake)]

        coerced = [coerce_value(v, t) for v, t in zip(row_values, manifest_types)]

        # Check if this attachment's message_id exists in emails_master
        msg_id_idx = manifest_snake.index("message_id")
        msg_id = coerced[msg_id_idx]

        if msg_id and msg_id not in valid_ids:
            coerced.append("pending_orphan")
            orphan_count += 1
        else:
            coerced.append("pending")

        conn.execute(insert_sql, coerced)
        count += 1

    # Re-enable FK enforcement
    conn.commit()
    conn.execute("PRAGMA foreign_keys = ON")

    return count, orphan_count


def populate_email_fts(conn: sqlite3.Connection, email_mapping: list[tuple[str, str, str]]) -> int:
    """Populate _fts_emails from emails_master data.

    Finds the snake_case column names for subject and body_clean and
    reads them from the table.
    """
    # Find the actual column name for body_clean (might be body_clean_1, body_clean_2, etc.)
    body_col = None
    subject_col = None
    for _, snake, _ in email_mapping:
        if snake == "subject":
            subject_col = snake
        if snake.startswith("body_clean"):
            body_col = snake

    if not subject_col:
        print("WARNING: No 'subject' column found — FTS will be incomplete")
        subject_col = "subject"
    if not body_col:
        print("WARNING: No 'body_clean' column found — FTS will be incomplete")
        body_col = "subject"  # fallback

    conn.execute(f"""
        INSERT INTO _fts_emails (subject, body_clean)
        SELECT {subject_col}, {body_col}
        FROM emails_master
    """)
    count = conn.execute("SELECT COUNT(*) FROM _fts_emails").fetchone()[0]
    return count


def populate_attachments_fts(conn: sqlite3.Connection) -> int:
    """Populate _fts_attachments from attachments data."""
    conn.execute("""
        INSERT INTO _fts_attachments (original_filename, extracted_text)
        SELECT original_filename, extracted_text
        FROM attachments
    """)
    count = conn.execute("SELECT COUNT(*) FROM _fts_attachments").fetchone()[0]
    return count


def save_column_map(conn: sqlite3.Connection,
                    email_mapping: list[tuple[str, str, str]],
                    attach_mapping: list[tuple[str, str, str]]) -> None:
    """Write column mappings to _column_map table."""
    # Clear existing email/attachment mappings (domain rebuild)
    conn.execute("DELETE FROM _column_map WHERE source_table IN ('emails_master', 'attachments')")

    for original, snake, dtype in email_mapping:
        conn.execute(
            "INSERT INTO _column_map (source_table, original_name, snake_case_name, data_type) "
            "VALUES (?, ?, ?, ?)",
            ("emails_master", original, snake, dtype)
        )

    for original, snake, dtype in attach_mapping:
        conn.execute(
            "INSERT INTO _column_map (source_table, original_name, snake_case_name, data_type) "
            "VALUES (?, ?, ?, ?)",
            ("attachments", original, snake, dtype)
        )


def save_schema_notes(conn: sqlite3.Connection) -> None:
    """Write schema documentation notes."""
    note_text = (
        "emails_master contains attachment columns imported as-is from V11 "
        "(attachments_count, attachment_names, email_attachments, attachments). "
        "These are REFERENCE ONLY -- stale summary blobs, not the source of truth. "
        "For attachment queries, always use the attachments table joined on message_id."
    )
    conn.execute(
        "INSERT OR REPLACE INTO _schema_notes (note_key, note_text, created_date) "
        "VALUES (?, ?, ?)",
        ("attachment_overlap_warning", note_text, datetime.now().isoformat())
    )


def print_summary(conn: sqlite3.Connection) -> None:
    """Print a summary of what was loaded."""
    email_count = conn.execute("SELECT COUNT(*) FROM emails_master").fetchone()[0]
    attach_count = conn.execute("SELECT COUNT(*) FROM attachments").fetchone()[0]
    fts_email_count = conn.execute("SELECT COUNT(*) FROM _fts_emails").fetchone()[0]
    fts_attach_count = conn.execute("SELECT COUNT(*) FROM _fts_attachments").fetchone()[0]
    col_map_count = conn.execute("SELECT COUNT(*) FROM _column_map").fetchone()[0]

    print("\n" + "=" * 60)
    print("LITIGATION CORPUS — LOAD SUMMARY")
    print("=" * 60)
    print(f"  emails_master:    {email_count:,} rows")
    print(f"  attachments:      {attach_count:,} rows")
    print(f"  _fts_emails:      {fts_email_count:,} entries")
    print(f"  _fts_attachments: {fts_attach_count:,} entries")
    print(f"  _column_map:      {col_map_count:,} mappings")

    # Extraction status breakdown
    print("\n  Extraction status by file type:")
    rows = conn.execute("""
        SELECT file_type, extraction_status, COUNT(*)
        FROM attachments
        GROUP BY file_type, extraction_status
        ORDER BY COUNT(*) DESC
    """).fetchall()
    for ftype, status, cnt in rows:
        print(f"    {ftype or '(none)':<12} {status:<12} {cnt:,}")

    # Entity tables (should be empty on first load)
    entity_count = conn.execute("SELECT COUNT(*) FROM entity_master").fetchone()[0]
    candidate_count = conn.execute("SELECT COUNT(*) FROM entity_candidates").fetchone()[0]
    print(f"\n  entity_master:    {entity_count:,} (shared, not touched by email load)")
    print(f"  entity_candidates: {candidate_count:,}")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(
        description="Load email + attachment data into litigation_corpus.db"
    )
    parser.add_argument("--emails", required=True,
                        help="Path to emails Excel workbook (e.g. V11.xlsx)")
    parser.add_argument("--sheet", required=True,
                        help="Sheet name in emails workbook (e.g. Merge1)")
    parser.add_argument("--attachments", required=True,
                        help="Path to attachments manifest Excel workbook")
    parser.add_argument("--db", default=DEFAULT_DB_PATH,
                        help=f"Path to database file (default: {DEFAULT_DB_PATH})")
    args = parser.parse_args()

    # Validate inputs
    if not Path(args.emails).exists():
        print(f"ERROR: Emails file not found: {args.emails}")
        sys.exit(1)
    if not Path(args.attachments).exists():
        print(f"ERROR: Attachments file not found: {args.attachments}")
        sys.exit(1)

    # Ensure DB directory exists
    Path(args.db).parent.mkdir(parents=True, exist_ok=True)

    print(f"Database: {args.db}")
    print(f"Emails:   {args.emails} [sheet: {args.sheet}]")
    print(f"Attachments: {args.attachments}")
    print()

    # Step 1: Read headers and build column mappings
    print("Reading headers...")
    email_headers = read_excel_headers(args.emails, args.sheet)
    attach_headers = read_excel_headers(args.attachments)

    email_mapping = build_email_column_mapping(email_headers)
    attach_mapping = build_attachments_column_mapping(attach_headers)

    print(f"  Email columns: {len(email_mapping)}")
    print(f"  Attachment columns: {len(attach_mapping)}")

    # Step 2: Create/rebuild schema
    print("\nCreating schema...")
    conn = connect(args.db)

    try:
        create_shared_tables(conn)
        print("  Shared tables: OK (preserved)")

        create_email_tables(conn, email_mapping, rebuild=True)
        print("  Email tables: rebuilt")

        # Step 3: Load emails
        print("\nLoading emails...")
        email_count = load_emails(conn, args.emails, args.sheet, email_mapping)
        print(f"  Loaded {email_count:,} emails")

        # Step 4: Load attachments
        print("Loading attachments...")
        attach_count, orphan_count = load_attachments(conn, args.attachments, attach_mapping)
        print(f"  Loaded {attach_count:,} attachments")
        if orphan_count:
            print(f"  WARNING: {orphan_count:,} orphan attachments (message_id not in emails_master)")
            print(f"    These are flagged as 'pending_orphan' - real files, parent email not in V11")

        # Step 5: Populate FTS
        print("Building FTS indexes...")
        fts_email = populate_email_fts(conn, email_mapping)
        fts_attach = populate_attachments_fts(conn)
        print(f"  _fts_emails: {fts_email:,} entries")
        print(f"  _fts_attachments: {fts_attach:,} entries")

        # Step 6: Save metadata
        print("Saving column mappings and schema notes...")
        save_column_map(conn, email_mapping, attach_mapping)
        save_schema_notes(conn)

        conn.commit()
        print("\nAll data committed.")

        # Step 7: Summary
        print_summary(conn)

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
