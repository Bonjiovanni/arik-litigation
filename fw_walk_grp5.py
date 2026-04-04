"""
fw_walk_grp5.py
---------------
main() orchestration for the fw_walk file inventory agent.

Coordinates all other fw_walk_grp modules:
    grp1 — file utilities
    grp2 — Master_File_Inventory + FileFamily_Config sheet management
    grp3 — overlap detection + walk_files orchestration
    grp4 — Walk_Coverage + Walk_History logging
    fw_dirmap_grp1 — folder picker + path validation

Writes to: C:\\Users\\arika\\OneDrive\\Litigation\\filewalker_master.xlsx
Python 3.11, Windows 10.
"""

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

import fw_walk_grp1  # noqa: F401  (imported for completeness; grp3 calls it)
import fw_walk_grp2
import fw_walk_grp3
import fw_walk_grp4
import fw_dirmap_grp1

WORKBOOK_PATH = r"C:\Users\arika\OneDrive\Litigation\filewalker_master.xlsx"


def main() -> None:
    """Run a complete fw_walk file inventory pass.

    Steps:
      1.  Open or create the master workbook.
      2.  Ensure all required sheets exist.
      3.  Generate a unique run ID.
      4.  Choose processing level (interactive).
      5.  Pick scan directories via native folder picker.
      6.  Validate and normalize chosen paths.
      7.  Overlap detection against Dir_Processing_Status.
      8.  Walk files and accumulate stats.
      9.  Log Walk_Coverage (per dir) and Walk_History (one run row).
      10. Save workbook.
      11. Print final summary.
    """
    try:
        # ------------------------------------------------------------------
        # 1. Open or create workbook
        # ------------------------------------------------------------------
        wb_path = Path(WORKBOOK_PATH)
        if wb_path.exists():
            print(f"[fw_walk] Loading existing workbook: {wb_path}")
            wb = load_workbook(wb_path)
        else:
            print(f"[fw_walk] Creating new workbook: {wb_path}")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ------------------------------------------------------------------
        # 2. Ensure all required sheets exist
        # ------------------------------------------------------------------
        fw_walk_grp2.ensure_master_file_inventory_sheet(wb)
        fw_walk_grp2.ensure_file_family_config_sheet(wb)
        fw_walk_grp4.ensure_walk_coverage_sheet(wb)
        fw_walk_grp4.ensure_walk_history_sheet(wb)

        # ------------------------------------------------------------------
        # 3. Generate run ID
        # ------------------------------------------------------------------
        run_id: str = datetime.now().strftime("WALK_%Y%m%d_%H%M%S")
        print(f"[fw_walk] Run ID: {run_id}")

        # ------------------------------------------------------------------
        # 4. Choose processing level
        # ------------------------------------------------------------------
        levels = fw_walk_grp3.PROCESSING_LEVEL_ORDER
        print("\n[fw_walk] Select processing level:")
        for i, level in enumerate(levels, start=1):
            suffix = "  ← default" if level == "file_listed" else ""
            print(f"  {i}. {level}{suffix}")

        raw = input("\nEnter number or name [default: file_listed]: ").strip()

        if not raw:
            processing_level = "file_listed"
        elif raw.isdigit():
            idx = int(raw) - 1
            processing_level = levels[idx] if 0 <= idx < len(levels) else "file_listed"
            if not (0 <= idx < len(levels)):
                print(f"[fw_walk] Invalid selection — defaulting to 'file_listed'.")
        elif raw in levels:
            processing_level = raw
        else:
            print(f"[fw_walk] Unknown level '{raw}' — defaulting to 'file_listed'.")
            processing_level = "file_listed"

        print(f"[fw_walk] Processing level: {processing_level}")

        # ------------------------------------------------------------------
        # 5 & 6. Pick and validate scan directories
        # ------------------------------------------------------------------
        print("\n[fw_walk] Opening folder picker — select directories to scan.")
        chosen_dirs: list[str] = fw_dirmap_grp1.pick_scan_dirs()

        if not chosen_dirs:
            print("[fw_walk] No directories selected. Exiting.")
            sys.exit(0)

        valid_dirs: list[str] = []
        for raw_dir in chosen_dirs:
            try:
                norm = fw_dirmap_grp1.validate_and_normalize_path(raw_dir)
                valid_dirs.append(norm)
            except ValueError as exc:
                print(f"[fw_walk] WARNING: skipping invalid path '{raw_dir}': {exc}")

        if not valid_dirs:
            print("[fw_walk] No valid directories after validation. Exiting.")
            sys.exit(1)

        print(f"[fw_walk] Directories to scan ({len(valid_dirs)}):")
        for d in valid_dirs:
            print(f"  {d}")

        # ------------------------------------------------------------------
        # 7. Overlap detection
        # ------------------------------------------------------------------
        overlap_action: str = "proceed"

        if "Dir_Processing_Status" in wb.sheetnames:
            ws_status = wb["Dir_Processing_Status"]
            covered = fw_walk_grp3.get_covered_paths(ws_status, processing_level)
            overlaps = fw_walk_grp3.check_overlap(valid_dirs, covered)

            if overlaps:
                overlap_action = fw_walk_grp3.prompt_overlap_decision(overlaps)

                if overlap_action == "abort":
                    print("[fw_walk] Aborted by user.")
                    sys.exit(0)

                if overlap_action == "skip":
                    overlap_paths = {o["path"] for o in overlaps}
                    valid_dirs = [d for d in valid_dirs if d not in overlap_paths]
                    print(f"[fw_walk] Skipped {len(overlaps)} overlapping dir(s). "
                          f"Remaining: {len(valid_dirs)}")
                    if not valid_dirs:
                        print("[fw_walk] No directories left after skipping. Exiting.")
                        sys.exit(0)
            else:
                print("[fw_walk] No overlaps detected.")
        else:
            print("[fw_walk] Dir_Processing_Status not found — skipping overlap check.")

        # ------------------------------------------------------------------
        # 8. Walk files
        # ------------------------------------------------------------------
        started_at_str: str = datetime.now().isoformat()
        print(f"\n[fw_walk] Starting walk at {started_at_str} ...")

        stats: dict = fw_walk_grp3.walk_files(
            scan_dirs=valid_dirs,
            run_id=run_id,
            processing_level=processing_level,
            wb=wb,
            overlap_action=overlap_action,
        )

        inserted      = stats.get("inserted", 0)
        updated       = stats.get("updated", 0)
        skipped_files = stats.get("skipped_files", 0)
        skipped_dirs  = stats.get("skipped_dirs", 0)
        errors        = stats.get("errors", 0)

        # ------------------------------------------------------------------
        # 9a. Log Walk_Coverage (one row per dir; totals on first dir)
        # ------------------------------------------------------------------
        ws_cov = wb["Walk_Coverage"]
        for i, dir_path in enumerate(valid_dirs):
            if i == 0:
                fi, fu, fs, fe = inserted, updated, skipped_files, errors
            else:
                fi = fu = fs = fe = 0
            fw_walk_grp4.update_walk_coverage(
                ws_cov, run_id, dir_path, processing_level, fi, fu, fs, fe
            )

        # ------------------------------------------------------------------
        # 9b. Log Walk_History (one row for the whole run)
        # ------------------------------------------------------------------
        fw_walk_grp4.log_walk_run(
            wb=wb,
            run_id=run_id,
            started_at=started_at_str,
            scan_dirs=valid_dirs,
            processing_level=processing_level,
            stats=stats,
            overlap_action=overlap_action,
        )

        # ------------------------------------------------------------------
        # 10. Save
        # ------------------------------------------------------------------
        wb.save(WORKBOOK_PATH)

        # ------------------------------------------------------------------
        # 11. Summary
        # ------------------------------------------------------------------
        print(f"\n[fw_walk] Run complete: {run_id}")
        print(f"  Inserted:      {inserted:,}")
        print(f"  Updated:       {updated:,}")
        print(f"  Skipped files: {skipped_files:,}")
        print(f"  Skipped dirs:  {skipped_dirs:,}")
        print(f"  Errors:        {errors:,}")
        print(f"Saved: {WORKBOOK_PATH}")

    except KeyboardInterrupt:
        print("\n[fw_walk] Interrupted.")
        sys.exit(0)


if __name__ == "__main__":
    main()
