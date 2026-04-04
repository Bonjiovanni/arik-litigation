"""
fw_classify.py
--------------
First-pass content signal extraction for the filewalker evidence pipeline.

Reads Master_File_Inventory rows with ProcessingStatus = "file_listed",
extracts text signals, and writes results back to the workbook:
    TextSample, MoneyDetected, DateDetected, KeywordHits,
    LikelyTextBearing, NeedsOCR,
    DocType, DocSubtype, DocTypeConfidence,
    ProcessingStatus → "classified"

Also manages:
    Keywords_Config  — editable keyword list loaded at runtime
    Classify_History — one-row-per-run audit log

Merged from: fw_classify_grp1 through fw_classify_grp5.

Writes to: C:\\Users\\arika\\OneDrive\\Litigation\\filewalker_master.xlsx
Python 3.11, Windows 10. Requires openpyxl, pymupdf (optional).
"""

import re
import csv
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from fw_walk_grp2 import _COL_INDEX

try:
    import fitz  # pymupdf
    _FITZ_AVAILABLE = True
except ImportError:
    _FITZ_AVAILABLE = False


WORKBOOK_PATH = r"C:\Users\arika\OneDrive\Litigation\filewalker_master.xlsx"

# Shared header style (used by Keywords_Config and Classify_History sheets)
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ===========================================================================
# grp1 — Text extraction + signal detection
# ===========================================================================

def extract_text_from_pdf(filepath: str, max_chars: int = 5000) -> tuple[str, bool]:
    """Extract text from a PDF file using pymupdf (fitz).

    Returns (text_sample, is_complete).
    is_complete is True if the full PDF text fits within max_chars.
    Returns ("", False) if fitz is unavailable or extraction fails.
    """
    if not _FITZ_AVAILABLE:
        return ("", False)

    try:
        doc = fitz.open(filepath)
        parts = []
        total = 0
        for page in doc:
            page_text = page.get_text("text")
            if not page_text:
                continue
            remaining = max_chars - total
            if remaining <= 0:
                break
            chunk = page_text[:remaining]
            parts.append(chunk)
            total += len(chunk)
            if total >= max_chars:
                break
        doc.close()
        text = "".join(parts)
        is_complete = total < max_chars
        return (text, is_complete)
    except Exception:
        return ("", False)


def extract_text_from_text_file(filepath: str, max_chars: int = 5000) -> str:
    """Extract text from a plain-text file (txt, log, md, json, xml, etc.).

    Uses UTF-8 with errors='replace'. Returns up to max_chars characters.
    Returns "" on any read error.
    """
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            return f.read(max_chars)
    except Exception:
        return ""


def extract_text_from_spreadsheet(filepath: str, max_cells: int = 200) -> str:
    """Extract a text sample from a spreadsheet (.xlsx/.xlsm or .csv).

    Returns a space-separated string of up to max_cells non-empty cell values.
    Returns "" on any error or unsupported format.
    """
    path = Path(filepath)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xlsm", ".xls"):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            parts = []
            count = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None and str(cell).strip():
                            parts.append(str(cell).strip())
                            count += 1
                            if count >= max_cells:
                                break
                    if count >= max_cells:
                        break
                if count >= max_cells:
                    break
            wb.close()
            return " ".join(parts)
        except Exception:
            return ""

    if suffix == ".csv":
        try:
            parts = []
            count = 0
            with open(filepath, "r", encoding="utf-8", errors="replace", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    for field in row:
                        if field.strip():
                            parts.append(field.strip())
                            count += 1
                            if count >= max_cells:
                                break
                    if count >= max_cells:
                        break
            return " ".join(parts)
        except Exception:
            return ""

    return ""


def get_text_sample(filepath: str, file_family: str,
                    max_chars: int = 500) -> tuple[str, bool]:
    """Dispatcher: extract a short text sample based on file_family.

    Returns (text_sample, extraction_succeeded).
    extraction_succeeded is False for unsupported families or extraction failures.
    """
    family = (file_family or "").lower()

    if family == "pdf":
        text, _ = extract_text_from_pdf(filepath, max_chars=max_chars)
        return (text, bool(text))

    if family == "text_file":
        text = extract_text_from_text_file(filepath, max_chars=max_chars)
        return (text, bool(text))

    if family == "spreadsheet":
        text = extract_text_from_spreadsheet(filepath, max_cells=max(50, max_chars // 10))
        return (text[:max_chars], bool(text))

    if family == "word_doc":
        try:
            import docx as _docx
            doc = _docx.Document(filepath)
            text = " ".join(p.text for p in doc.paragraphs if p.text.strip())
            return (text[:max_chars], bool(text))
        except Exception:
            return ("", False)

    # image, audio, video, email_file, archive — no text extraction
    return ("", False)


# Signal detection patterns
_MONEY_PATTERNS = [
    re.compile(r'\$\s*[\d,]+(?:\.\d{1,2})?'),
    re.compile(r'\b(?:USD|dollars?|cents?)\b', re.IGNORECASE),
]

_DATE_PATTERNS = [
    re.compile(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b'),
    re.compile(r'\b\d{4}-\d{2}-\d{2}\b'),
    re.compile(
        r'\b(?:January|February|March|April|May|June|July|August|September|'
        r'October|November|December|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'
        r'\s+\d{1,2}(?:st|nd|rd|th)?,?\s+\d{4}\b',
        re.IGNORECASE,
    ),
]


def detect_money(text: str) -> str:
    """Return "Y" / "N" / "".

    "Y"  — money signal found.
    "N"  — text present but no money signal.
    ""   — no text.
    """
    if not text or not text.strip():
        return ""
    for pattern in _MONEY_PATTERNS:
        if pattern.search(text):
            return "Y"
    return "N"


def detect_dates(text: str) -> str:
    """Return "Y" / "N" / "".

    "Y"  — date pattern found.
    "N"  — text present but no date pattern.
    ""   — no text.
    """
    if not text or not text.strip():
        return ""
    for pattern in _DATE_PATTERNS:
        if pattern.search(text):
            return "Y"
    return "N"


def match_keywords(text: str, keywords: list[str]) -> str:
    """Return semicolon-joined keywords found in text (case-insensitive).

    Returns "" if no text, no keywords, or no matches.
    Preserves original keyword casing in output.
    """
    if not text or not text.strip() or not keywords:
        return ""
    text_lower = text.lower()
    hits = [kw for kw in keywords if kw.lower() in text_lower]
    return ";".join(hits)


# ===========================================================================
# grp2 — Master_File_Inventory Excel I/O
# ===========================================================================

# Columns safe to write during classification (all others ignored)
_PROTECTED_COLS = {"ManualReviewStatus", "KeepForCase", "PossibleExhibit"}


def get_classifiable_rows(ws) -> list[tuple[int, dict]]:
    """Return (row_num, record_dict) for rows needing classification.

    Criteria:
      - ProcessingStatus is None, "", or "file_listed"
      - SkipReason is None or empty

    Stops at first empty FileID (end of data).

    record_dict keys: FileID, FilePath, FileName, FileFamily,
                      LikelyTextBearing, NeedsOCR, ProcessingStatus
    """
    _FIELDS = ["FileID", "FilePath", "FileName", "FileFamily",
               "LikelyTextBearing", "NeedsOCR", "ProcessingStatus"]

    result = []
    for row_num in range(2, ws.max_row + 1):
        file_id = ws.cell(row=row_num, column=_COL_INDEX["FileID"]).value
        if file_id is None or str(file_id).strip() == "":
            break

        processing_status = ws.cell(row=row_num, column=_COL_INDEX["ProcessingStatus"]).value
        skip_reason       = ws.cell(row=row_num, column=_COL_INDEX["SkipReason"]).value

        if processing_status not in (None, "", "file_listed"):
            continue
        if skip_reason is not None and str(skip_reason).strip() != "":
            continue

        record = {field: ws.cell(row=row_num, column=_COL_INDEX[field]).value
                  for field in _FIELDS}
        result.append((row_num, record))

    return result


def write_classify_signals(ws, row_num: int, signals: dict) -> None:
    """Write signal column values to a specific row.

    Writes any key in signals dict that exists in _COL_INDEX,
    unless it is in _PROTECTED_COLS.
    """
    for key, value in signals.items():
        if key in _PROTECTED_COLS:
            continue
        if key not in _COL_INDEX:
            continue
        ws.cell(row=row_num, column=_COL_INDEX[key], value=value)


def mark_row_classified(ws, row_num: int) -> None:
    """Set ProcessingStatus = 'classified' for the given row."""
    ws.cell(row=row_num, column=_COL_INDEX["ProcessingStatus"], value="classified")


# ===========================================================================
# grp3 — Keywords_Config sheet + DocType inference
# ===========================================================================

KEYWORDS_CONFIG_SHEET = "Keywords_Config"

_DEFAULT_KEYWORDS = [
    ("trust",          "Y", "entity",    ""),
    ("trustee",        "Y", "entity",    ""),
    ("beneficiary",    "Y", "entity",    ""),
    ("distribution",   "Y", "financial", ""),
    ("fiduciary",      "Y", "entity",    ""),
    ("amendment",      "Y", "legal",     ""),
    ("restatement",    "Y", "legal",     ""),
    ("invoice",        "Y", "financial", ""),
    ("premium",        "Y", "financial", ""),
    ("claim",          "Y", "ltc",       ""),
    ("long-term care", "Y", "ltc",       ""),
    ("LTC",            "Y", "ltc",       ""),
    ("benefit",        "Y", "ltc",       ""),
    ("policy",         "Y", "ltc",       ""),
    ("coverage",       "Y", "ltc",       ""),
    ("caregiver",      "Y", "ltc",       ""),
    ("facility",       "Y", "ltc",       ""),
    ("nursing",        "Y", "ltc",       ""),
    ("home health",    "Y", "ltc",       ""),
]


def ensure_keywords_config_sheet(wb: Workbook) -> Worksheet:
    """Create and style the Keywords_Config sheet if absent; return it.

    Idempotent. Seeds default litigation keywords on first creation.
    """
    if KEYWORDS_CONFIG_SHEET in wb.sheetnames:
        return wb[KEYWORDS_CONFIG_SHEET]

    ws = wb.create_sheet(title=KEYWORDS_CONFIG_SHEET)

    headers    = ["Keyword", "Active", "Category", "Notes"]
    col_widths = [25, 8, 14, 40]
    for col_idx, (header_text, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    for row_idx, (keyword, active, category, notes) in enumerate(_DEFAULT_KEYWORDS, start=2):
        ws.cell(row=row_idx, column=1, value=keyword)
        ws.cell(row=row_idx, column=2, value=active)
        ws.cell(row=row_idx, column=3, value=category)
        ws.cell(row=row_idx, column=4, value=notes)

    return ws


def load_keywords(ws: Worksheet) -> list[str]:
    """Return list of active keyword strings from Keywords_Config.

    Active = column B == "Y" (case-insensitive).
    Skips header and rows with no keyword.
    """
    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        keyword = row[0] if len(row) > 0 else None
        active  = row[1] if len(row) > 1 else None
        if not keyword:
            continue
        if active is not None and str(active).strip().upper() == "Y":
            keywords.append(str(keyword))
    return keywords


def infer_likely_text_bearing(file_family: str, text_sample: str,
                               extraction_succeeded: bool) -> str:
    """Return "Y" / "N" / "" based on whether file likely contains extractable text."""
    fam = (file_family or "").lower()

    if fam in ("text_file", "word_doc", "spreadsheet", "presentation"):
        return "Y"
    if fam == "pdf":
        return "Y" if (extraction_succeeded and text_sample) else "N"
    if fam in ("image", "audio", "video"):
        return "N"
    if fam == "email_file":
        return "Y"
    return ""


def infer_needs_ocr(file_family: str, text_sample: str,
                    extraction_succeeded: bool) -> str:
    """Return "Y" / "N" / "" based on whether file needs OCR."""
    fam = (file_family or "").lower()

    if fam == "pdf":
        return "N" if (extraction_succeeded and text_sample) else "Y"
    if fam == "image":
        return "Y"
    if fam in ("text_file", "word_doc", "spreadsheet", "presentation", "email_file"):
        return "N"
    if fam in ("audio", "video", "archive"):
        return "N"
    return ""


def classify_doc_type(text_sample: str, keyword_hits: str,
                      file_family: str) -> dict:
    """Lightweight keyword-driven DocType inference.

    Returns dict: {DocType, DocSubtype, DocTypeConfidence}.
    First matching rule wins.
    """
    sig = (text_sample or "").lower() + " " + (keyword_hits or "").lower()

    if "invoice" in sig:
        return {"DocType": "Invoice",         "DocSubtype": "",          "DocTypeConfidence": "medium"}
    if "claim" in sig and ("long-term care" in sig or "ltc" in sig):
        return {"DocType": "LTC_Claim",       "DocSubtype": "",          "DocTypeConfidence": "medium"}
    if ("benefit" in sig or "coverage" in sig) and "claim" in sig:
        return {"DocType": "LTC_Claim",       "DocSubtype": "",          "DocTypeConfidence": "low"}
    if "amendment" in sig or "restatement" in sig:
        return {"DocType": "Trust_Document",  "DocSubtype": "Amendment", "DocTypeConfidence": "medium"}
    if "trust" in sig and ("trustee" in sig or "beneficiary" in sig or "fiduciary" in sig):
        return {"DocType": "Trust_Document",  "DocSubtype": "",          "DocTypeConfidence": "medium"}
    if "trust" in sig:
        return {"DocType": "Trust_Document",  "DocSubtype": "",          "DocTypeConfidence": "low"}
    if "premium" in sig or ("policy" in sig and "long-term care" in sig):
        return {"DocType": "Insurance_Policy","DocSubtype": "",          "DocTypeConfidence": "low"}
    return {"DocType": "",                    "DocSubtype": "",          "DocTypeConfidence": ""}


# ===========================================================================
# grp4 — Classify_History sheet
# ===========================================================================

CLASSIFY_HISTORY_SHEET = "Classify_History"

_CH_COLUMNS = [
    ("RunID",          25),
    ("StartedAt",      22),
    ("CompletedAt",    22),
    ("RowsProcessed",  14),
    ("RowsUpdated",    12),
    ("RowsSkipped",    12),
    ("Errors",         10),
    ("Notes",          40),
]

_CH_COL_MAP = {name: idx + 1 for idx, (name, _) in enumerate(_CH_COLUMNS)}


def ensure_classify_history_sheet(wb: Workbook) -> Worksheet:
    """Create and style the Classify_History sheet if absent; return it. Idempotent."""
    if CLASSIFY_HISTORY_SHEET in wb.sheetnames:
        return wb[CLASSIFY_HISTORY_SHEET]

    ws = wb.create_sheet(title=CLASSIFY_HISTORY_SHEET)

    for col_idx, (name, width) in enumerate(_CH_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    return ws


def log_classify_run(wb: Workbook, run_id: str, started_at: str,
                     stats: dict, notes: str = "") -> int:
    """Append one classify-run summary row to Classify_History.

    stats keys (default 0): rows_processed, rows_updated, rows_skipped, errors.
    Returns the row number written.
    """
    ws = ensure_classify_history_sheet(wb)

    row_num = ws.max_row + 1
    if row_num == 2 and ws.cell(row=2, column=1).value is None:
        row_num = 2

    ws.cell(row=row_num, column=_CH_COL_MAP["RunID"],         value=run_id)
    ws.cell(row=row_num, column=_CH_COL_MAP["StartedAt"],     value=started_at)
    ws.cell(row=row_num, column=_CH_COL_MAP["CompletedAt"],   value=datetime.now().isoformat())
    ws.cell(row=row_num, column=_CH_COL_MAP["RowsProcessed"], value=stats.get("rows_processed", 0))
    ws.cell(row=row_num, column=_CH_COL_MAP["RowsUpdated"],   value=stats.get("rows_updated",   0))
    ws.cell(row=row_num, column=_CH_COL_MAP["RowsSkipped"],   value=stats.get("rows_skipped",   0))
    ws.cell(row=row_num, column=_CH_COL_MAP["Errors"],        value=stats.get("errors",         0))
    ws.cell(row=row_num, column=_CH_COL_MAP["Notes"],         value=notes)

    return row_num


# ===========================================================================
# grp5 — main() orchestration
# ===========================================================================

def main() -> None:
    """Run a complete fw_classify first-pass signal extraction pass.

    Steps:
      1.  Open or create the master workbook.
      2.  Ensure Keywords_Config and Classify_History sheets exist.
      3.  Generate run ID: CLASSIFY_YYYYMMDD_HHMMSS.
      4.  Load active keywords from Keywords_Config.
      5.  Get classifiable rows from Master_File_Inventory.
      6.  For each row: extract text, detect signals, infer flags, classify doc type.
      7.  Write signals back and mark row classified.
      8.  Log Classify_History row.
      9.  Save workbook.
      10. Print final summary.
    """
    try:
        # ------------------------------------------------------------------
        # 1. Open or create workbook
        # ------------------------------------------------------------------
        wb_path = Path(WORKBOOK_PATH)
        if wb_path.exists():
            print(f"[fw_classify] Loading existing workbook: {wb_path}")
            wb = load_workbook(wb_path)
        else:
            print(f"[fw_classify] Creating new workbook: {wb_path}")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ------------------------------------------------------------------
        # 2. Ensure required sheets exist
        # ------------------------------------------------------------------
        ensure_keywords_config_sheet(wb)
        ensure_classify_history_sheet(wb)

        # ------------------------------------------------------------------
        # 3. Generate run ID
        # ------------------------------------------------------------------
        run_id: str = datetime.now().strftime("CLASSIFY_%Y%m%d_%H%M%S")
        print(f"[fw_classify] Run ID: {run_id}")

        # ------------------------------------------------------------------
        # 4. Load keywords
        # ------------------------------------------------------------------
        ws_kw    = wb[KEYWORDS_CONFIG_SHEET]
        keywords = load_keywords(ws_kw)
        print(f"[fw_classify] Loaded {len(keywords)} active keyword(s).")

        # ------------------------------------------------------------------
        # 5. Get classifiable rows
        # ------------------------------------------------------------------
        if "Master_File_Inventory" not in wb.sheetnames:
            print("[fw_classify] ERROR: Master_File_Inventory not found. Run fw_walk first.")
            sys.exit(1)

        ws_inv = wb["Master_File_Inventory"]
        rows   = get_classifiable_rows(ws_inv)
        print(f"[fw_classify] Rows to classify: {len(rows):,}")

        if not rows:
            print("[fw_classify] Nothing to classify.")
            sys.exit(0)

        # ------------------------------------------------------------------
        # 6–7. Process each row
        # ------------------------------------------------------------------
        started_at_str: str = datetime.now().isoformat()
        rows_processed = 0
        rows_updated   = 0
        rows_skipped   = 0
        errors         = 0

        for row_num, record in rows:
            filepath    = record.get("FilePath",   "") or ""
            file_family = record.get("FileFamily", "") or ""

            try:
                text_sample, extraction_ok = get_text_sample(filepath, file_family, max_chars=500)

                money   = detect_money(text_sample)
                dates   = detect_dates(text_sample)
                kw_hits = match_keywords(text_sample, keywords)

                likely_text = infer_likely_text_bearing(file_family, text_sample, extraction_ok)
                needs_ocr   = infer_needs_ocr(file_family, text_sample, extraction_ok)

                doc_info = classify_doc_type(text_sample, kw_hits, file_family)

                signals = {
                    "TextSample":        (text_sample or "")[:500],
                    "MoneyDetected":     money,
                    "DateDetected":      dates,
                    "KeywordHits":       kw_hits,
                    "LikelyTextBearing": likely_text,
                    "NeedsOCR":          needs_ocr,
                    "DocType":           doc_info.get("DocType",           ""),
                    "DocSubtype":        doc_info.get("DocSubtype",        ""),
                    "DocTypeConfidence": doc_info.get("DocTypeConfidence", ""),
                }
                write_classify_signals(ws_inv, row_num, signals)
                mark_row_classified(ws_inv, row_num)
                rows_updated += 1

            except Exception as exc:
                print(f"[fw_classify] ERROR row {row_num} ({filepath}): {exc}")
                errors += 1
            finally:
                rows_processed += 1

            if rows_processed % 100 == 0:
                print(f"[fw_classify]   {rows_processed:,}/{len(rows):,} processed...")

        # ------------------------------------------------------------------
        # 8. Log Classify_History
        # ------------------------------------------------------------------
        log_classify_run(
            wb=wb,
            run_id=run_id,
            started_at=started_at_str,
            stats={
                "rows_processed": rows_processed,
                "rows_updated":   rows_updated,
                "rows_skipped":   rows_skipped,
                "errors":         errors,
            },
        )

        # ------------------------------------------------------------------
        # 9. Save
        # ------------------------------------------------------------------
        wb.save(WORKBOOK_PATH)

        # ------------------------------------------------------------------
        # 10. Summary
        # ------------------------------------------------------------------
        print(f"\n[fw_classify] Run complete: {run_id}")
        print(f"  Rows processed: {rows_processed:,}")
        print(f"  Updated:        {rows_updated:,}")
        print(f"  Skipped:        {rows_skipped:,}")
        print(f"  Errors:         {errors:,}")
        print(f"Saved: {WORKBOOK_PATH}")

    except KeyboardInterrupt:
        print("\n[fw_classify] Interrupted.")
        sys.exit(0)


if __name__ == "__main__":
    main()
