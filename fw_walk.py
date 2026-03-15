"""
fw_walk.py
----------
File inventory agent for the litigation fileWalker pipeline.

Scans one or more Windows directories, inventories every file with metadata
and optional SHA-256 hash, and writes/updates rows in Master_File_Inventory
inside filewalker_master.xlsx.

Functions (grouped by concern):

  Group 1 — File utilities
    load_file_family_config, classify_file_family, should_skip_file,
    get_file_metadata, compute_sha256, peek_archive_contents

  Group 2 — Master_File_Inventory + FileFamily_Config sheet management
    ensure_master_file_inventory_sheet, get_next_file_id,
    find_existing_file_row, insert_file_record, update_file_record,
    write_or_update_file_record, ensure_file_family_config_sheet

  Group 3 — Overlap detection + walk orchestration
    PROCESSING_LEVEL_ORDER, get_covered_paths, check_overlap,
    prompt_overlap_decision, walk_files

  Group 4 — Walk_Coverage + Walk_History logging
    ensure_walk_coverage_sheet, get_next_coverage_id, update_walk_coverage,
    ensure_walk_history_sheet, log_walk_run

  Group 5 — Entry point
    main()

Writes to: C:\\Users\\arika\\OneDrive\\Litigation\\filewalker_master.xlsx
Python 3.11, Windows 10. Requires openpyxl.
"""

import hashlib
import os
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import fw_dirmap_grp1

WORKBOOK_PATH = r"C:\Users\arika\OneDrive\Litigation\filewalker_master.xlsx"

# ---------------------------------------------------------------------------
# Shared header styling (used by all sheet-creation functions)
# ---------------------------------------------------------------------------

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ===========================================================================
# GROUP 1: File utilities
# ===========================================================================

_FAMILY_MAP: dict[str, set[str]] = {
    "pdf":          {".pdf"},
    "word_doc":     {".doc", ".docx", ".odt", ".rtf", ".dot", ".dotx"},
    "spreadsheet":  {".xls", ".xlsx", ".xlsm", ".xlsb", ".csv", ".ods", ".numbers"},
    "text_file":    {".txt", ".md", ".log", ".json", ".xml", ".html", ".htm",
                     ".yaml", ".yml", ".ini", ".toml", ".cfg"},
    "presentation": {".ppt", ".pptx", ".odp", ".key"},
    "image":        {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif",
                     ".heic", ".heif", ".webp", ".svg", ".raw", ".cr2", ".nef",
                     ".arw", ".dng", ".ico"},
    "email_file":   {".eml", ".msg", ".pst", ".ost", ".mbox"},
    "archive":      {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz", ".cab"},
    "audio":        {".mp3", ".wav", ".m4a", ".flac", ".aac", ".ogg", ".wma", ".aiff"},
    "video":        {".mp4", ".mov", ".avi", ".mkv", ".wmv", ".flv", ".m4v",
                     ".mpg", ".mpeg", ".webm"},
}

_DEFAULT_SKIP_FAMILIES: set[str] = {"email_file", "archive"}

_SYSTEM_EXTENSIONS: set[str] = {
    ".tmp", ".temp", ".lock", ".lnk", ".url", ".ds_store", ".thumbs.db",
}


def load_file_family_config(ws) -> "tuple[dict[str, set[str]], set[str]]":
    """Load the file-family extension map and skip-family set from a worksheet.

    Reads the FileFamily_Config worksheet. Expected columns (row 1 = header):
        A: Family, B: Extensions (semicolon-separated), C: ShouldSkip (Y/N)

    Args:
        ws: openpyxl Worksheet for FileFamily_Config.

    Returns:
        (family_map, skip_families) where family_map = {family: {extensions}}
    """
    family_map: dict[str, set[str]] = {}
    skip_families: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        family = str(row[0]).strip().lower()
        if not family:
            continue

        raw_exts = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        ext_set: set[str] = set()
        for token in raw_exts.split(";"):
            token = token.strip().lower()
            if not token:
                continue
            if not token.startswith("."):
                token = "." + token
            ext_set.add(token)
        family_map[family] = ext_set

        should_skip_val = str(row[2]).strip().upper() if len(row) > 2 and row[2] is not None else "N"
        if should_skip_val == "Y":
            skip_families.add(family)

    return family_map, skip_families


def classify_file_family(
    filename: str,
    extension: str,
    family_map: "dict[str, set[str]] | None" = None,
) -> str:
    """Return the logical family name for a file extension.

    Args:
        filename:   Basename (reserved for future magic-byte fallback).
        extension:  Lowercased extension with leading dot (e.g. ".pdf").
        family_map: Optional map from load_file_family_config(). Uses built-in if None.

    Returns:
        Family name string, or "other" if no match.
    """
    effective_map = family_map if family_map is not None else _FAMILY_MAP
    ext = extension.lower()
    for family, extensions in effective_map.items():
        if ext in extensions:
            return family
    return "other"


def should_skip_file(
    filename: str,
    extension: str,
    file_family: str,
    skip_families: "set[str] | None" = None,
) -> "tuple[bool, str]":
    """Determine whether a file should be skipped.

    Priority order:
      1. family in skip_families
      2. filename starts with "~$"  → office_temp
      3. filename starts with "."   → hidden_file
      4. extension in system types  → system_file

    Args:
        filename:     File basename.
        extension:    Lowercased extension with dot.
        file_family:  From classify_file_family().
        skip_families: Families to skip. Defaults to {"email_file", "archive"}.

    Returns:
        (should_skip, skip_reason)
    """
    effective_skip = skip_families if skip_families is not None else _DEFAULT_SKIP_FAMILIES
    if file_family in effective_skip:
        return (True, file_family)
    if filename.startswith("~$"):
        return (True, "office_temp")
    if filename.startswith("."):
        return (True, "hidden_file")
    if extension.lower() in _SYSTEM_EXTENSIONS:
        return (True, "system_file")
    return (False, "")


def get_file_metadata(filepath: str) -> dict:
    """Collect filesystem metadata for a file path.

    Returns dict with: filename, extension, parent_folder, file_path,
    size_bytes, created_time, modified_time, accessed_time.
    Numeric fields are "" on OSError.
    """
    def _norm(p: str) -> str:
        return p.replace("\\", "/")

    filename      = os.path.basename(filepath)
    _, raw_ext    = os.path.splitext(filename)
    extension     = raw_ext.lower()
    parent_folder = _norm(os.path.dirname(os.path.abspath(filepath)))
    norm_path     = _norm(filepath)

    result: dict = {
        "filename":      filename,
        "extension":     extension,
        "parent_folder": parent_folder,
        "file_path":     norm_path,
        "size_bytes":    "",
        "created_time":  "",
        "modified_time": "",
        "accessed_time": "",
    }

    try:
        st = os.stat(filepath)
        result["size_bytes"]    = st.st_size
        result["created_time"]  = datetime.fromtimestamp(st.st_ctime).isoformat()
        result["modified_time"] = datetime.fromtimestamp(st.st_mtime).isoformat()
        try:
            result["accessed_time"] = datetime.fromtimestamp(st.st_atime).isoformat()
        except OSError:
            pass
    except OSError:
        pass

    return result


def compute_sha256(filepath: str, max_bytes: int = 524_288_000) -> "tuple[str, bool]":
    """Compute SHA-256 for a file with optional byte cap (default 500 MB).

    Returns (hash_hex, is_complete). Returns ("", False) on any read error.
    """
    _CHUNK = 65_536
    hasher      = hashlib.sha256()
    bytes_read  = 0
    is_complete = True

    try:
        with open(filepath, "rb") as fh:
            while True:
                if max_bytes > 0:
                    remaining = max_bytes - bytes_read
                    if remaining <= 0:
                        is_complete = False
                        break
                    read_size = min(_CHUNK, remaining)
                else:
                    read_size = _CHUNK
                chunk = fh.read(read_size)
                if not chunk:
                    break
                hasher.update(chunk)
                bytes_read += len(chunk)
        return (hasher.hexdigest(), is_complete)
    except (PermissionError, OSError):
        return ("", False)


def peek_archive_contents(filepath: str) -> str:
    """Return a plain-text summary of a zip archive's contents.

    Only .zip is supported natively. Other formats return an explanatory string.
    Does not extract any files.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext != ".zip":
        return f"cannot peek — unsupported format ({ext}); install rarfile/py7zr to enable"

    try:
        with zipfile.ZipFile(filepath, "r") as zf:
            names = [info.filename for info in zf.infolist() if not info.is_dir()]
    except zipfile.BadZipFile:
        return "cannot peek — corrupt or unreadable zip"
    except PermissionError:
        return "cannot peek — access denied"
    except OSError as exc:
        return f"cannot peek — OS error ({exc.strerror})"

    total = len(names)
    if total == 0:
        return "empty archive (0 files)"

    family_counts: dict[str, int] = {}
    ext_examples: dict[str, list[str]] = {}
    for name in names:
        _, inner_ext = os.path.splitext(name)
        inner_ext = inner_ext.lower()
        family = classify_file_family(os.path.basename(name), inner_ext)
        family_counts[family] = family_counts.get(family, 0) + 1
        if family not in ext_examples:
            ext_examples[family] = []
        if inner_ext and inner_ext not in ext_examples[family]:
            ext_examples[family].append(inner_ext)

    sorted_families = sorted(family_counts.items(), key=lambda kv: kv[1], reverse=True)
    parts = []
    for family, count in sorted_families:
        exts  = " ".join(sorted(ext_examples.get(family, [])))
        label = family.replace("_", " ")
        parts.append(f"{count} {label} ({exts})" if exts else f"{count} {label}")

    return f"{total} file{'s' if total != 1 else ''}: " + ", ".join(parts)


# ===========================================================================
# GROUP 2: Master_File_Inventory + FileFamily_Config sheet management
# ===========================================================================

COLUMNS = [
    # --- A. Location / identity ---
    ("FileID",               8),
    ("RunID_FirstSeen",      20),
    ("RunID_LastSeen",       20),
    ("FilePath",             70),
    ("ParentFolder",         60),
    ("FileName",             35),
    ("ScanRootPath",         50),
    ("SourceStore",          14),
    # --- B. File metadata ---
    ("SizeBytes",            12),
    ("CreatedTime",          20),
    ("ModifiedTime",         20),
    ("SHA256",               66),
    ("IsDuplicateExact",     10),
    ("DuplicateGroupID",     14),
    # --- C. Physical/source classification ---
    ("FileFamily",           14),
    ("SourceType",           20),
    ("LikelyTextBearing",    10),
    ("LikelyImage",          10),
    ("LikelySpreadsheet",    10),
    ("LikelyDocument",       10),
    ("LikelyScreenshot",     10),
    ("NeedsOCR",             10),
    ("IsContainerType",      10),
    ("SkipReason",           14),
    ("ArchiveContents",      45),
    # --- D. Existing structured joins ---
    ("MsgID",                20),
    ("AttachmentID",         20),
    ("AttachmentKind",       20),
    ("RelatedEmailSubject",  20),
    ("RelatedEmailFrom",     20),
    ("RelatedEmailDate",     20),
    ("RelatedEmailPDFPath",  20),
    ("RelatedTextThreadID",  20),
    # --- E. Cheap first-pass extraction signals ---
    ("TextSample",           20),
    ("OCRSnippet",           20),
    ("KeywordHits",          20),
    ("MoneyDetected",        20),
    ("DateDetected",         20),
    ("SignatureLike",        20),
    ("HandwritingLike",      20),
    ("LogoLetterheadLike",   20),
    ("TableLayoutLike",      20),
    # --- F. Classification / semantics ---
    ("DocType",              25),
    ("DocSubtype",           25),
    ("DocTypeConfidence",    20),
    ("CategoryTags",         30),
    ("ReviewGroup",          20),
    ("PrimaryEntity",        25),
    ("MatchedEntities",      25),
    ("NewEntityCandidates",  25),
    ("EntityMatchSource",    25),
    ("EntityConfidence",     20),
    # --- G. Triage / routing ---
    ("TriageScore",          18),
    ("TriageBand",           18),
    ("ReasonFlagged",        18),
    ("NextStep",             18),
    ("ManualReviewStatus",   18),
    ("KeepForCase",          12),
    ("PossibleExhibit",      12),
    ("ProcessingStatus",     16),
    ("Notes",                30),
]

SHEET_NAME = "Master_File_Inventory"

_COL_INDEX: dict[str, int] = {name: idx + 1 for idx, (name, _) in enumerate(COLUMNS)}

_ALWAYS_UPDATE_COLS = {
    "RunID_LastSeen", "SizeBytes", "ModifiedTime", "SHA256",
    "IsDuplicateExact", "DuplicateGroupID", "ProcessingStatus",
}
_NEVER_OVERWRITE_COLS = {"ManualReviewStatus", "KeepForCase", "PossibleExhibit"}
_CLASSIFICATION_COLS = {
    "TextSample", "OCRSnippet", "KeywordHits", "MoneyDetected", "DateDetected",
    "SignatureLike", "HandwritingLike", "LogoLetterheadLike", "TableLayoutLike",
    "DocType", "DocSubtype", "DocTypeConfidence", "CategoryTags", "ReviewGroup",
    "PrimaryEntity", "MatchedEntities", "NewEntityCandidates", "EntityMatchSource",
    "EntityConfidence", "TriageScore", "TriageBand", "ReasonFlagged", "NextStep",
}


def _col_name_to_key(col_name: str) -> str:
    return re.sub(r"(?<=[a-z])(?=[A-Z])", "_", col_name).lower()


def ensure_master_file_inventory_sheet(wb: Workbook) -> Worksheet:
    """Create and style Master_File_Inventory if absent; return it (idempotent)."""
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    ws = wb.create_sheet(title=SHEET_NAME)
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    return ws


def get_next_file_id(ws: Worksheet) -> str:
    """Return next FileID (e.g. "F00248"), or "F00001" if sheet is empty."""
    max_num = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = str(row[0] or "").strip().upper()
        if val.startswith("F") and val[1:].isdigit():
            num = int(val[1:])
            if num > max_num:
                max_num = num
    return f"F{max_num + 1:05d}"


def find_existing_file_row(ws: Worksheet, file_path: str) -> "int | None":
    """Search column D (FilePath) for a normalized case-insensitive match.

    Returns 1-based row number, or None.
    """
    col_d  = _COL_INDEX["FilePath"]
    target = file_path.replace("\\", "/").lower()
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, min_col=col_d, max_col=col_d, values_only=True), start=2
    ):
        val = row[0]
        if val is not None and str(val).replace("\\", "/").lower() == target:
            return row_idx
    return None


def insert_file_record(ws: Worksheet, record: dict, run_id: str) -> int:
    """Append a new file record row. Returns the 1-based row number written."""
    next_row = max(ws.max_row + 1, 2)
    row_data: dict[str, object] = {col_name: "" for col_name, _ in COLUMNS}

    row_data["FileID"]           = record.get("file_id", "")
    row_data["RunID_FirstSeen"]  = run_id
    row_data["RunID_LastSeen"]   = run_id
    row_data["FilePath"]         = record.get("file_path", "").replace("\\", "/")
    row_data["ParentFolder"]     = record.get("parent_folder", "").replace("\\", "/")
    row_data["FileName"]         = record.get("filename", "")
    row_data["ScanRootPath"]     = record.get("scan_root_path", "").replace("\\", "/")
    row_data["SourceStore"]      = record.get("source_store", "")
    row_data["SizeBytes"]        = record.get("size_bytes", "")
    row_data["CreatedTime"]      = record.get("created_time", "")
    row_data["ModifiedTime"]     = record.get("modified_time", "")
    row_data["SHA256"]           = record.get("sha256", "")
    row_data["IsDuplicateExact"] = record.get("is_duplicate_exact", "")
    row_data["DuplicateGroupID"] = record.get("duplicate_group_id", "")
    row_data["FileFamily"]       = record.get("file_family", "")
    row_data["SourceType"]       = record.get("source_type", "standalone_file")
    row_data["LikelyTextBearing"]= record.get("likely_text_bearing", "")
    row_data["LikelyImage"]      = record.get("likely_image", "")
    row_data["LikelySpreadsheet"]= record.get("likely_spreadsheet", "")
    row_data["LikelyDocument"]   = record.get("likely_document", "")
    row_data["LikelyScreenshot"] = record.get("likely_screenshot", "")
    row_data["NeedsOCR"]         = record.get("needs_ocr", "")
    row_data["IsContainerType"]  = record.get("is_container_type", "")
    row_data["SkipReason"]       = record.get("skip_reason", "")
    row_data["ArchiveContents"]  = record.get("archive_contents", "")
    row_data["ManualReviewStatus"] = record.get("manual_review_status", "unreviewed")
    row_data["ProcessingStatus"] = record.get("processing_status", "file_listed")

    for col_name, col_idx in _COL_INDEX.items():
        ws.cell(row=next_row, column=col_idx, value=row_data[col_name])

    return next_row


def update_file_record(ws: Worksheet, row: int, record: dict, run_id: str) -> None:
    """Update an existing row with fresh scan data.

    Always updates: RunID_LastSeen, SizeBytes, ModifiedTime, SHA256,
    IsDuplicateExact, DuplicateGroupID, ProcessingStatus.
    Never overwrites if populated: ManualReviewStatus, KeepForCase, PossibleExhibit.
    Leaves untouched: all classification/triage columns.
    """
    def _set(col_name: str, value) -> None:
        ws.cell(row=row, column=_COL_INDEX[col_name], value=value)

    def _current(col_name: str):
        return ws.cell(row=row, column=_COL_INDEX[col_name]).value

    _set("RunID_LastSeen",   run_id)
    _set("SizeBytes",        record.get("size_bytes", ""))
    _set("ModifiedTime",     record.get("modified_time", ""))
    _set("SHA256",           record.get("sha256", ""))
    _set("IsDuplicateExact", record.get("is_duplicate_exact", ""))
    _set("DuplicateGroupID", record.get("duplicate_group_id", ""))
    _set("ProcessingStatus", record.get("processing_status", "file_listed"))

    if record.get("archive_contents"):
        _set("ArchiveContents", record["archive_contents"])

    for col_name in _NEVER_OVERWRITE_COLS:
        current_val = _current(col_name)
        if current_val is None or str(current_val).strip() == "":
            key     = _col_name_to_key(col_name)
            new_val = record.get(key, "")
            if new_val:
                _set(col_name, new_val)


def write_or_update_file_record(
    wb: Workbook, record: dict, run_id: str
) -> "tuple[str, int]":
    """Insert or update a file record. Returns ("inserted"|"updated", row_number).

    Raises ValueError if record lacks a non-empty "file_path".
    """
    file_path = record.get("file_path", "")
    if not file_path:
        raise ValueError("record must contain a non-empty 'file_path' key.")

    ws = ensure_master_file_inventory_sheet(wb)
    existing_row = find_existing_file_row(ws, file_path)
    if existing_row is not None:
        update_file_record(ws, existing_row, record, run_id)
        return ("updated", existing_row)

    record = dict(record)
    record["file_id"] = get_next_file_id(ws)
    inserted_row = insert_file_record(ws, record, run_id)
    return ("inserted", inserted_row)


# --- FileFamily_Config sheet ---

_FAMILY_CONFIG_DEFAULTS = [
    ("pdf",          ".pdf",
     "N", "Y", "N", "N", "Y", ""),
    ("word_doc",     ".doc;.docx;.odt;.rtf;.dot;.dotx",
     "N", "Y", "N", "N", "Y", ""),
    ("spreadsheet",  ".xls;.xlsx;.xlsm;.xlsb;.csv;.ods;.numbers",
     "N", "Y", "N", "Y", "N", ""),
    ("text_file",    ".txt;.md;.log;.json;.xml;.html;.htm;.yaml;.yml;.ini;.toml;.cfg",
     "N", "Y", "N", "N", "N", ""),
    ("presentation", ".ppt;.pptx;.odp;.key",
     "N", "Y", "N", "N", "Y", ""),
    ("image",        ".jpg;.jpeg;.png;.gif;.bmp;.tiff;.tif;.heic;.heif;.webp;.svg;.raw;.cr2;.nef;.arw;.dng;.ico",
     "N", "N", "Y", "N", "N", ""),
    ("email_file",   ".eml;.msg;.pst;.ost;.mbox",
     "Y", "N", "N", "N", "N", "Processed by separate email pipeline"),
    ("archive",      ".zip;.rar;.7z;.tar;.gz;.bz2;.xz;.cab",
     "Y", "N", "N", "N", "N", "Content description captured in ArchiveContents column"),
    ("audio",        ".mp3;.wav;.m4a;.flac;.aac;.ogg;.wma;.aiff",
     "N", "N", "N", "N", "N", ""),
    ("video",        ".mp4;.mov;.avi;.mkv;.wmv;.flv;.m4v;.mpg;.mpeg;.webm",
     "N", "N", "N", "N", "N", ""),
]

_FC_SHEET_NAME = "FileFamily_Config"

_FC_COLUMNS = [
    ("Family",            18),
    ("Extensions",        80),
    ("ShouldSkip",        11),
    ("LikelyTextBearing", 16),
    ("LikelyImage",       13),
    ("LikelySpreadsheet", 16),
    ("LikelyDocument",    15),
    ("Notes",             50),
]


def ensure_file_family_config_sheet(wb: Workbook) -> Worksheet:
    """Create and populate FileFamily_Config if absent; return it (idempotent)."""
    if _FC_SHEET_NAME in wb.sheetnames:
        return wb[_FC_SHEET_NAME]

    ws = wb.create_sheet(title=_FC_SHEET_NAME)
    for col_idx, (col_name, col_width) in enumerate(_FC_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_FC_COLUMNS))}1"

    for row_idx, row_data in enumerate(_FAMILY_CONFIG_DEFAULTS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    return ws


# ===========================================================================
# GROUP 3: Overlap detection + walk orchestration
# ===========================================================================

PROCESSING_LEVEL_ORDER: list[str] = [
    "file_listed",
    "hashed",
    "classified",
    "text_extracted",
    "ocr_complete",
    "entity_matched",
    "triaged",
]


def _level_index(level: str) -> int:
    try:
        return PROCESSING_LEVEL_ORDER.index(level)
    except ValueError:
        return -1


def _normalize_dir_path(p: str) -> str:
    """Normalize a directory path: forward slashes, lowercase, no trailing slash."""
    return p.replace("\\", "/").rstrip("/").lower()


def get_covered_paths(ws_status, target_level: str) -> "dict[str, str]":
    """Return paths already walked to at least target_level depth.

    Reads Dir_Processing_Status (col C = DirPath, col D = ProcessingLevel).
    Returns {normalized_path: highest_level} for paths at or above target.
    """
    target_idx = _level_index(target_level)
    if target_idx < 0:
        return {}

    highest: dict[str, str] = {}
    for row in ws_status.iter_rows(min_row=2, values_only=True):
        if not row or row[2] is None or row[3] is None:
            continue
        path  = _normalize_dir_path(str(row[2]))
        level = str(row[3]).strip()
        if path not in highest or _level_index(level) > _level_index(highest[path]):
            highest[path] = level

    return {p: l for p, l in highest.items() if _level_index(l) >= target_idx}


def check_overlap(
    scan_dirs: "list[str]",
    covered_paths: "dict[str, str]",
) -> "list[dict]":
    """Return overlap dicts for any scan_dirs already in covered_paths.

    Each overlap dict: {"path": str, "existing_level": str}.
    """
    overlaps = []
    for raw_dir in scan_dirs:
        norm = _normalize_dir_path(raw_dir)
        if norm in covered_paths:
            overlaps.append({"path": norm, "existing_level": covered_paths[norm]})
    return overlaps


def prompt_overlap_decision(overlaps: "list[dict]") -> str:
    """Print overlap summary, prompt user for global action.

    Returns: "skip", "rewalk", or "abort".
    """
    print("\n[fw_walk] Overlap detected — the following directories have already")
    print("been walked to the requested depth or deeper:\n")
    for o in overlaps:
        print(f"  {o['path']}  (already: {o['existing_level']})")
    print("\nOptions:")
    print("  [S] Skip overlapping directories (only walk new/unprocessed)")
    print("  [R] Re-walk all directories (overwrite existing records)")
    print("  [A] Abort\n")
    _choices = {"s": "skip", "r": "rewalk", "a": "abort"}
    while True:
        raw = input("Enter choice [S/R/A]: ").strip().lower()
        if raw in _choices:
            return _choices[raw]
        print("  Invalid — please enter S, R, or A.")


def walk_files(
    scan_dirs: "list[str]",
    run_id: str,
    processing_level: str,
    wb: Workbook,
    overlap_action: str = "skip",
) -> dict:
    """Walk directories and write every file to Master_File_Inventory.

    Returns stats dict: inserted, updated, skipped_files, skipped_dirs,
    errors, error_paths.
    """
    do_hash = _level_index(processing_level) >= _level_index("hashed")

    fc_ws = ensure_file_family_config_sheet(wb)
    family_map, skip_families = load_file_family_config(fc_ws)

    stats: dict = {
        "inserted": 0, "updated": 0,
        "skipped_files": 0, "skipped_dirs": 0,
        "errors": 0, "error_paths": [],
    }

    for raw_dir in scan_dirs:
        norm_dir = raw_dir.replace("\\", "/").rstrip("/")

        try:
            walker = os.walk(norm_dir, topdown=True, followlinks=False)
        except PermissionError as exc:
            stats["errors"] += 1
            stats["error_paths"].append(norm_dir)
            stats["skipped_dirs"] += 1
            print(f"[fw_walk] WARNING: cannot walk {norm_dir}: {exc}", file=sys.stderr)
            continue

        for dirpath, dirnames, filenames in walker:
            accessible = []
            for d in dirnames:
                subpath = os.path.join(dirpath, d)
                try:
                    os.scandir(subpath).close()
                    accessible.append(d)
                except PermissionError:
                    stats["skipped_dirs"] += 1
                    print(f"[fw_walk] WARNING: permission denied, skipping: {subpath}",
                          file=sys.stderr)
            dirnames[:] = accessible

            for filename in filenames:
                filepath = os.path.join(dirpath, filename).replace("\\", "/")
                try:
                    meta        = get_file_metadata(filepath)
                    file_family = classify_file_family(
                        meta["filename"], meta["extension"], family_map)
                    should_skip, skip_reason = should_skip_file(
                        meta["filename"], meta["extension"], file_family, skip_families)

                    archive_contents = ""
                    if should_skip and file_family == "archive":
                        archive_contents = peek_archive_contents(filepath)

                    sha256 = ""
                    if do_hash and not should_skip:
                        hash_hex, is_complete = compute_sha256(filepath)
                        if hash_hex:
                            sha256 = hash_hex if is_complete else hash_hex + "_partial"

                    record = {
                        "file_path":          meta["file_path"],
                        "parent_folder":      meta["parent_folder"],
                        "filename":           meta["filename"],
                        "scan_root_path":     norm_dir,
                        "source_store":       "",
                        "size_bytes":         meta["size_bytes"],
                        "created_time":       meta["created_time"],
                        "modified_time":      meta["modified_time"],
                        "sha256":             sha256,
                        "is_duplicate_exact": "",
                        "duplicate_group_id": "",
                        "file_family":        file_family,
                        "source_type":        "standalone_file",
                        "likely_text_bearing": "",
                        "likely_image":       "",
                        "likely_spreadsheet": "",
                        "likely_document":    "",
                        "likely_screenshot":  "",
                        "needs_ocr":          "",
                        "is_container_type":  "Y" if file_family == "archive" else "",
                        "skip_reason":        skip_reason,
                        "archive_contents":   archive_contents,
                        "processing_status":  processing_level,
                    }

                    if should_skip:
                        stats["skipped_files"] += 1
                        write_or_update_file_record(wb, record, run_id)
                    else:
                        action, _ = write_or_update_file_record(wb, record, run_id)
                        if action == "inserted":
                            stats["inserted"] += 1
                        else:
                            stats["updated"] += 1

                except (PermissionError, OSError) as exc:
                    stats["errors"] += 1
                    stats["error_paths"].append(filepath)
                    print(f"[fw_walk] WARNING: error processing {filepath}: {exc}",
                          file=sys.stderr)

    return stats


# ===========================================================================
# GROUP 4: Walk_Coverage + Walk_History logging
# ===========================================================================

_WC_SHEET_NAME = "Walk_Coverage"
_WC_COLUMNS = [
    ("CoverageID",      12),
    ("RunID",           20),
    ("DirPath",         70),
    ("ProcessingLevel", 18),
    ("FilesInserted",   14),
    ("FilesUpdated",    14),
    ("FilesSkipped",    14),
    ("Errors",          10),
    ("WalkedAt",        20),
]
_WC_COL_INDEX: dict[str, int] = {name: idx + 1 for idx, (name, _) in enumerate(_WC_COLUMNS)}


def ensure_walk_coverage_sheet(wb: Workbook) -> Worksheet:
    """Create and style Walk_Coverage if absent; return it (idempotent)."""
    if _WC_SHEET_NAME in wb.sheetnames:
        return wb[_WC_SHEET_NAME]
    ws = wb.create_sheet(title=_WC_SHEET_NAME)
    for col_idx, (col_name, col_width) in enumerate(_WC_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_WC_COLUMNS))}1"
    return ws


def get_next_coverage_id(ws: Worksheet) -> str:
    """Return next CoverageID (e.g. "WC00042"), or "WC00001" if empty."""
    max_num = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = str(row[0] or "").strip().upper()
        if val.startswith("WC") and val[2:].isdigit():
            num = int(val[2:])
            if num > max_num:
                max_num = num
    return f"WC{max_num + 1:05d}"


def update_walk_coverage(
    ws: Worksheet,
    run_id: str,
    dir_path: str,
    processing_level: str,
    files_inserted: int,
    files_updated: int,
    files_skipped: int,
    errors: int,
) -> int:
    """Append one directory walk record to Walk_Coverage. Returns row number."""
    next_row    = max(ws.max_row + 1, 2)
    coverage_id = get_next_coverage_id(ws)
    row_data = {
        "CoverageID":      coverage_id,
        "RunID":           run_id,
        "DirPath":         dir_path.replace("\\", "/"),
        "ProcessingLevel": processing_level,
        "FilesInserted":   files_inserted,
        "FilesUpdated":    files_updated,
        "FilesSkipped":    files_skipped,
        "Errors":          errors,
        "WalkedAt":        datetime.now().isoformat(),
    }
    for col_name, col_idx in _WC_COL_INDEX.items():
        ws.cell(row=next_row, column=col_idx, value=row_data[col_name])
    return next_row


_WH_SHEET_NAME = "Walk_History"
_WH_COLUMNS = [
    ("RunID",             20),
    ("StartedAt",         20),
    ("CompletedAt",       20),
    ("ScanDirs",          80),
    ("ProcessingLevel",   18),
    ("TotalInserted",     14),
    ("TotalUpdated",      14),
    ("TotalSkippedFiles", 16),
    ("TotalSkippedDirs",  15),
    ("TotalErrors",       12),
    ("OverlapAction",     14),
    ("Notes",             40),
]
_WH_COL_INDEX: dict[str, int] = {name: idx + 1 for idx, (name, _) in enumerate(_WH_COLUMNS)}


def ensure_walk_history_sheet(wb: Workbook) -> Worksheet:
    """Create and style Walk_History if absent; return it (idempotent)."""
    if _WH_SHEET_NAME in wb.sheetnames:
        return wb[_WH_SHEET_NAME]
    ws = wb.create_sheet(title=_WH_SHEET_NAME)
    for col_idx, (col_name, col_width) in enumerate(_WH_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_WH_COLUMNS))}1"
    return ws


def log_walk_run(
    wb: Workbook,
    run_id: str,
    started_at: str,
    scan_dirs: "list[str]",
    processing_level: str,
    stats: dict,
    overlap_action: str,
    notes: str = "",
) -> int:
    """Append one run-summary row to Walk_History. Returns row number."""
    ws = ensure_walk_history_sheet(wb)
    next_row = max(ws.max_row + 1, 2)
    row_data = {
        "RunID":             run_id,
        "StartedAt":         started_at,
        "CompletedAt":       datetime.now().isoformat(),
        "ScanDirs":          " | ".join(d.replace("\\", "/") for d in scan_dirs),
        "ProcessingLevel":   processing_level,
        "TotalInserted":     stats.get("inserted", 0),
        "TotalUpdated":      stats.get("updated", 0),
        "TotalSkippedFiles": stats.get("skipped_files", 0),
        "TotalSkippedDirs":  stats.get("skipped_dirs", 0),
        "TotalErrors":       stats.get("errors", 0),
        "OverlapAction":     overlap_action,
        "Notes":             notes,
    }
    for col_name, col_idx in _WH_COL_INDEX.items():
        ws.cell(row=next_row, column=col_idx, value=row_data[col_name])
    return next_row


# ===========================================================================
# GROUP 5: main() orchestration
# ===========================================================================

def main() -> None:
    """Run a complete fw_walk file inventory pass.

    Steps:
      1.  Open or create filewalker_master.xlsx.
      2.  Ensure all required sheets exist.
      3.  Generate run ID (WALK_YYYYMMDD_HHMMSS).
      4.  Select processing level (interactive).
      5.  Pick scan directories (native folder picker).
      6.  Validate and normalize chosen paths.
      7.  Overlap detection against Dir_Processing_Status.
      8.  Walk files and accumulate stats.
      9.  Log Walk_Coverage + Walk_History.
      10. Save workbook and print summary.
    """
    try:
        # 1. Open or create workbook
        wb_path = Path(WORKBOOK_PATH)
        if wb_path.exists():
            print(f"[fw_walk] Loading: {wb_path}")
            wb = load_workbook(wb_path)
        else:
            print(f"[fw_walk] Creating: {wb_path}")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # 2. Ensure sheets
        ensure_master_file_inventory_sheet(wb)
        ensure_file_family_config_sheet(wb)
        ensure_walk_coverage_sheet(wb)
        ensure_walk_history_sheet(wb)

        # 3. Run ID
        run_id: str = datetime.now().strftime("WALK_%Y%m%d_%H%M%S")
        print(f"[fw_walk] Run ID: {run_id}")

        # 4. Processing level
        levels = PROCESSING_LEVEL_ORDER
        print("\n[fw_walk] Select processing level:")
        for i, level in enumerate(levels, start=1):
            suffix = "  ← default" if level == "file_listed" else ""
            print(f"  {i}. {level}{suffix}")
        raw = input("\nEnter number or name [default: file_listed]: ").strip()
        if not raw:
            processing_level = "file_listed"
        elif raw.isdigit():
            idx = int(raw) - 1
            processing_level = levels[idx] if 0 <= idx < len(levels) else "file_listed"
        elif raw in levels:
            processing_level = raw
        else:
            print(f"[fw_walk] Unknown level '{raw}' — defaulting to 'file_listed'.")
            processing_level = "file_listed"
        print(f"[fw_walk] Processing level: {processing_level}")

        # 5 & 6. Pick and validate directories
        print("\n[fw_walk] Opening folder picker — select directories to scan.")
        chosen_dirs = fw_dirmap_grp1.pick_scan_dirs()
        if not chosen_dirs:
            print("[fw_walk] No directories selected. Exiting.")
            sys.exit(0)

        valid_dirs: list[str] = []
        for raw_dir in chosen_dirs:
            try:
                valid_dirs.append(fw_dirmap_grp1.validate_and_normalize_path(raw_dir))
            except ValueError as exc:
                print(f"[fw_walk] WARNING: skipping '{raw_dir}': {exc}")

        if not valid_dirs:
            print("[fw_walk] No valid directories. Exiting.")
            sys.exit(1)

        print(f"[fw_walk] Scanning {len(valid_dirs)} directory/ies:")
        for d in valid_dirs:
            print(f"  {d}")

        # 7. Overlap detection
        overlap_action: str = "proceed"
        if "Dir_Processing_Status" in wb.sheetnames:
            ws_status = wb["Dir_Processing_Status"]
            covered   = get_covered_paths(ws_status, processing_level)
            overlaps  = check_overlap(valid_dirs, covered)
            if overlaps:
                overlap_action = prompt_overlap_decision(overlaps)
                if overlap_action == "abort":
                    print("[fw_walk] Aborted.")
                    sys.exit(0)
                if overlap_action == "skip":
                    skip_set   = {o["path"] for o in overlaps}
                    valid_dirs = [d for d in valid_dirs if _normalize_dir_path(d) not in skip_set]
                    print(f"[fw_walk] Skipped {len(overlaps)} overlap(s). Remaining: {len(valid_dirs)}")
                    if not valid_dirs:
                        print("[fw_walk] Nothing left to walk. Exiting.")
                        sys.exit(0)
            else:
                print("[fw_walk] No overlaps detected.")
        else:
            print("[fw_walk] Dir_Processing_Status not found — skipping overlap check.")

        # 8. Walk
        started_at_str = datetime.now().isoformat()
        print(f"\n[fw_walk] Starting walk at {started_at_str} ...")
        stats = walk_files(
            scan_dirs=valid_dirs,
            run_id=run_id,
            processing_level=processing_level,
            wb=wb,
            overlap_action=overlap_action,
        )

        inserted      = stats.get("inserted", 0)
        updated       = stats.get("updated", 0)
        skipped_files = stats.get("skipped_files", 0)
        skipped_dirs  = stats.get("skipped_dirs", 0)
        errors        = stats.get("errors", 0)

        # 9a. Walk_Coverage (totals on first dir; zeroes on rest)
        ws_cov = wb["Walk_Coverage"]
        for i, dir_path in enumerate(valid_dirs):
            fi, fu, fs, fe = (inserted, updated, skipped_files, errors) if i == 0 else (0, 0, 0, 0)
            update_walk_coverage(ws_cov, run_id, dir_path, processing_level, fi, fu, fs, fe)

        # 9b. Walk_History
        log_walk_run(wb=wb, run_id=run_id, started_at=started_at_str,
                     scan_dirs=valid_dirs, processing_level=processing_level,
                     stats=stats, overlap_action=overlap_action)

        # 10. Save
        wb.save(WORKBOOK_PATH)

        # 11. Summary
        print(f"\n[fw_walk] Run complete: {run_id}")
        print(f"  Inserted:      {inserted:,}")
        print(f"  Updated:       {updated:,}")
        print(f"  Skipped files: {skipped_files:,}")
        print(f"  Skipped dirs:  {skipped_dirs:,}")
        print(f"  Errors:        {errors:,}")
        print(f"Saved: {WORKBOOK_PATH}")

    except KeyboardInterrupt:
        print("\n[fw_walk] Interrupted.")
        sys.exit(0)


if __name__ == "__main__":
    main()
