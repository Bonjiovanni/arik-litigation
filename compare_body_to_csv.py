"""Export body comparison data to CSV for Excel MCP to pick up."""
import re, csv
from html.parser import HTMLParser
from collections import Counter
import openpyxl

SRC = r"C:\Users\arika\OneDrive\Litigation\Pipeline\label_LegalEmailExtracts - Ariks Version V11.xlsx"
OUT_CSV = r"C:\Users\arika\Repo-for-Claude-android\body_comparison_data.csv"
SHEET = "Merge1"
TARGET_COLS = ["body_clean.1", "Body.HTML", "Body.SenderText", "Body.Text"]

class HTMLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []
    def handle_data(self, d):
        self.parts.append(d)
    def get_text(self):
        return "".join(self.parts)

def strip_html(s):
    if not s:
        return ""
    s = str(s)
    h = HTMLStripper()
    try:
        h.feed(s)
        text = h.get_text()
    except:
        text = re.sub(r'<[^>]+>', '', s)
    return re.sub(r'\s+', ' ', text).strip()

def main():
    print(f"Loading {SRC} ...")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[SHEET]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip() if h else "" for h in row]
        break

    col_map = {}
    for target in TARGET_COLS:
        for i, h in enumerate(headers):
            if h.lower() == target.lower():
                col_map[target] = i
                break

    id_col = None
    for i, h in enumerate(headers):
        if "message" in h.lower() and "id" in h.lower():
            id_col = i
            break

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Message ID"] + TARGET_COLS + ["Match Code", "Result"] + [f"Len:{c}" for c in TARGET_COLS])

        row_num = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            msg_id = str(row[id_col]) if id_col is not None and row[id_col] else ""
            raw_vals = {}
            stripped_vals = {}
            for col_name in TARGET_COLS:
                if col_name in col_map:
                    raw = row[col_map[col_name]]
                    raw_vals[col_name] = str(raw) if raw else ""
                    stripped_vals[col_name] = strip_html(raw)
                else:
                    raw_vals[col_name] = ""
                    stripped_vals[col_name] = ""

            texts = [stripped_vals[c] for c in TARGET_COLS]
            char_counts = [len(t) for t in texts]
            non_empty = [t for t in texts if t]

            if not non_empty:
                code = "----"
                flag = "ALL_EMPTY"
            else:
                freq = Counter(non_empty)
                majority_text = freq.most_common(1)[0][0]
                code_chars = []
                for t in texts:
                    if not t:
                        code_chars.append("E")
                    elif t == majority_text:
                        code_chars.append("Y")
                    else:
                        code_chars.append("N")
                code = "".join(code_chars)
                y = code.count("Y"); n = code.count("N"); e = code.count("E")
                if n == 0 and e == 0:
                    flag = "ALL_MATCH"
                elif y + e == 4:
                    flag = "MATCH (some empty)"
                else:
                    unique = set(non_empty)
                    flag = "NONE_MATCH" if len(unique) == len(non_empty) else "PARTIAL"

            writer.writerow([msg_id] + [raw_vals[c] for c in TARGET_COLS] + [code, flag] + char_counts)

    wb.close()
    print(f"Done: {row_num} rows written to {OUT_CSV}")

if __name__ == "__main__":
    main()
