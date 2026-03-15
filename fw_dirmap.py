"""
fw_dirmap.py
============
Directory Mapper Agent — single-file entry point.

Walks one or more root directories chosen by the user, builds a flat inventory
of every directory found, and writes the results to `filewalker_master.xlsx`
(created in the same folder as this script if it does not already exist).

Output sheets
-------------
  Dir_Inventory         — one row per directory (Option C hybrid format)
  Dir_Processing_Status — one row per (directory × file-family) combination
  Walk_History          — one row per fw_dirmap run

Usage
-----
  python fw_dirmap.py

  A native Windows folder-picker dialog opens first.  Select one or more
  folders, then answer two short prompts (recursive? max depth?).

Platform  : Windows 10 only
Python    : 3.11+
Pip deps  : openpyxl
"""

# ---------------------------------------------------------------------------
# Imports
# ---------------------------------------------------------------------------

import ctypes
import ctypes.wintypes
import json
import os
import sys
from datetime import datetime
from typing import List

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ===========================================================================
# GROUP 1 — User input & path utilities
# ===========================================================================

# ---------------------------------------------------------------------------
# FUNCTION: pick_scan_dirs
# ---------------------------------------------------------------------------

def pick_scan_dirs() -> List[str]:
    """Open a native Windows IFileOpenDialog for multi-folder selection.

    Falls back to a plain input() prompt if COM/ctypes fails.

    Returns
    -------
    list[str]
        Selected folder paths as raw strings (not yet normalized).
        Empty list if the user cancels the dialog.
    """
    FOS_ALLOWMULTISELECT = 0x00000200
    FOS_PICKFOLDERS      = 0x00000020

    CLSID_FileOpenDialog = "{DC1C5A9C-E88A-4DDE-A5A1-60F82A20AEF7}"
    IID_IFileOpenDialog  = "{D57C7288-D4AD-4768-BE02-9D969532D960}"
    IID_IShellItemArray  = "{B63EA76D-1F85-456F-A19C-48159EFA858B}"
    IID_IShellItem       = "{43826D1E-E718-42EE-BC55-A1E261C37BFE}"

    try:
        ole32 = ctypes.windll.ole32
        ole32.CoInitialize(None)

        clsid    = ctypes.create_string_buffer(16)
        iid_dlg  = ctypes.create_string_buffer(16)
        iid_arr  = ctypes.create_string_buffer(16)

        ole32.CLSIDFromString(CLSID_FileOpenDialog, clsid)
        ole32.IIDFromString(IID_IFileOpenDialog,    iid_dlg)
        ole32.IIDFromString(IID_IShellItemArray,    iid_arr)

        dialog_ptr = ctypes.c_void_p()
        hr = ole32.CoCreateInstance(clsid, None, 0x1, iid_dlg, ctypes.byref(dialog_ptr))
        if hr != 0:
            raise OSError(f"CoCreateInstance failed: HRESULT {hr:#010x}")

        vtable     = ctypes.cast(dialog_ptr, ctypes.POINTER(ctypes.c_void_p))
        vtable_ptr = ctypes.cast(vtable[0],  ctypes.POINTER(ctypes.c_void_p))

        PROTO_GetOptions = ctypes.WINFUNCTYPE(
            ctypes.HRESULT, ctypes.c_void_p, ctypes.POINTER(ctypes.c_uint32))
        PROTO_SetOptions = ctypes.WINFUNCTYPE(
            ctypes.HRESULT, ctypes.c_void_p, ctypes.c_uint32)

        current_opts = ctypes.c_uint32(0)
        PROTO_GetOptions(vtable_ptr[10])(dialog_ptr, ctypes.byref(current_opts))

        new_opts = current_opts.value | FOS_PICKFOLDERS | FOS_ALLOWMULTISELECT
        hr = PROTO_SetOptions(vtable_ptr[9])(dialog_ptr, new_opts)
        if hr != 0:
            raise OSError(f"SetOptions failed: HRESULT {hr:#010x}")

        PROTO_Show = ctypes.WINFUNCTYPE(ctypes.HRESULT, ctypes.c_void_p, ctypes.c_void_p)
        hr = PROTO_Show(vtable_ptr[3])(dialog_ptr, None)

        CANCELLED = ctypes.c_long(0x800704C7).value
        if hr == CANCELLED:
            return []
        if hr != 0:
            raise OSError(f"Show failed: HRESULT {hr:#010x}")

        PROTO_GetResults = ctypes.WINFUNCTYPE(
            ctypes.HRESULT, ctypes.c_void_p, ctypes.POINTER(ctypes.c_void_p))
        arr_ptr = ctypes.c_void_p()
        hr = PROTO_GetResults(vtable_ptr[27])(dialog_ptr, ctypes.byref(arr_ptr))
        if hr != 0:
            raise OSError(f"GetResults failed: HRESULT {hr:#010x}")

        arr_vtable = ctypes.cast(
            ctypes.cast(arr_ptr, ctypes.POINTER(ctypes.c_void_p))[0],
            ctypes.POINTER(ctypes.c_void_p),
        )

        PROTO_GetCount  = ctypes.WINFUNCTYPE(
            ctypes.HRESULT, ctypes.c_void_p, ctypes.POINTER(ctypes.c_uint32))
        PROTO_GetItemAt = ctypes.WINFUNCTYPE(
            ctypes.HRESULT, ctypes.c_void_p, ctypes.c_uint32,
            ctypes.POINTER(ctypes.c_void_p))
        PROTO_GetDisplayName = ctypes.WINFUNCTYPE(
            ctypes.HRESULT, ctypes.c_void_p, ctypes.c_int,
            ctypes.POINTER(ctypes.c_wchar_p))
        PROTO_Release = ctypes.WINFUNCTYPE(ctypes.c_ulong, ctypes.c_void_p)

        item_count = ctypes.c_uint32(0)
        PROTO_GetCount(arr_vtable[7])(arr_ptr, ctypes.byref(item_count))

        SIGDN_FILESYSPATH = ctypes.c_int(0x80058000)
        paths: List[str] = []

        for i in range(item_count.value):
            item_ptr = ctypes.c_void_p()
            hr = PROTO_GetItemAt(arr_vtable[8])(arr_ptr, i, ctypes.byref(item_ptr))
            if hr != 0:
                continue

            item_vtable = ctypes.cast(
                ctypes.cast(item_ptr, ctypes.POINTER(ctypes.c_void_p))[0],
                ctypes.POINTER(ctypes.c_void_p),
            )
            path_buf = ctypes.c_wchar_p()
            hr = PROTO_GetDisplayName(item_vtable[5])(
                item_ptr, SIGDN_FILESYSPATH.value, ctypes.byref(path_buf))
            if hr == 0 and path_buf.value:
                paths.append(path_buf.value)
                ole32.CoTaskMemFree(path_buf)
            PROTO_Release(item_vtable[2])(item_ptr)

        PROTO_Release(arr_vtable[2])(arr_ptr)
        PROTO_Release(vtable_ptr[2])(dialog_ptr)
        ole32.CoUninitialize()
        return paths

    except Exception as exc:
        print(
            f"\n[fw_dirmap] WARNING: Native folder dialog unavailable ({exc}).\n"
            "Falling back to manual path entry.\n"
            "Enter one folder path per line.  Press Enter on a blank line when done.\n"
        )
        paths: List[str] = []
        while True:
            line = input("  Folder path (blank to finish): ")
            if not line.strip():
                break
            paths.append(line)
        return paths


# ---------------------------------------------------------------------------
# FUNCTION: validate_and_normalize_path
# ---------------------------------------------------------------------------

def validate_and_normalize_path(raw: str) -> str:
    """Clean, resolve, and validate a raw path string.

    Steps: strip whitespace → strip quotes → expand ~ → abspath →
    forward slashes → verify exists and is a directory.

    Parameters
    ----------
    raw : str

    Returns
    -------
    str
        Absolute path with forward slashes.

    Raises
    ------
    ValueError
        If path does not exist or is not a directory.
    """
    cleaned = raw.strip()
    if (cleaned.startswith('"') and cleaned.endswith('"')) or \
       (cleaned.startswith("'") and cleaned.endswith("'")):
        cleaned = cleaned[1:-1]

    cleaned = os.path.expanduser(cleaned)
    cleaned = os.path.abspath(cleaned)
    cleaned = cleaned.replace("\\", "/")

    if not os.path.exists(cleaned):
        raise ValueError(f"Path does not exist on disk: '{cleaned}'")
    if not os.path.isdir(cleaned):
        raise ValueError(f"Path exists but is not a directory: '{cleaned}'")

    return cleaned


# ---------------------------------------------------------------------------
# FUNCTION: generate_run_id
# ---------------------------------------------------------------------------

def generate_run_id() -> str:
    """Return a timestamped run ID: ``DIRMAP_YYYYMMDD_HHMMSS``."""
    return datetime.now().strftime("DIRMAP_%Y%m%d_%H%M%S")


# ===========================================================================
# GROUP 2 — Filesystem walking helpers
# ===========================================================================

# ---------------------------------------------------------------------------
# FUNCTION: detect_source_store
# ---------------------------------------------------------------------------

def detect_source_store(path: str) -> str:
    """Identify the backing store for a given normalized forward-slash path.

    Parameters
    ----------
    path : str
        Normalized path string (forward slashes).

    Returns
    -------
    str
        ``"OneDrive"``, ``"GoogleDriveSync"``, or ``"Local"``.
    """
    lower = path.lower()
    if "/onedrive/" in lower or lower.endswith("/onedrive"):
        return "OneDrive"
    if "/google drive/" in lower or "/googledrive/" in lower:
        return "GoogleDriveSync"
    return "Local"


# ---------------------------------------------------------------------------
# FUNCTION: count_dir_contents
# ---------------------------------------------------------------------------

def count_dir_contents(dir_path: str) -> tuple:
    """Count files and immediate subdirectories inside *dir_path*.

    Symlinks are excluded from both counts.  Returns ``(0, 0)`` on
    ``PermissionError`` or ``OSError``.

    Parameters
    ----------
    dir_path : str

    Returns
    -------
    tuple[int, int]
        ``(file_count, subdir_count)``
    """
    file_count = 0
    subdir_count = 0
    try:
        with os.scandir(dir_path) as scanner:
            for entry in scanner:
                if entry.is_file(follow_symlinks=False):
                    file_count += 1
                elif entry.is_dir(follow_symlinks=False):
                    subdir_count += 1
    except PermissionError:
        print(
            f"[fw_dirmap WARNING] PermissionError — cannot read directory: {dir_path}",
            file=sys.stderr,
        )
        return (0, 0)
    except OSError as exc:
        print(
            f"[fw_dirmap WARNING] OSError ({exc.strerror}) — cannot read directory: {dir_path}",
            file=sys.stderr,
        )
        return (0, 0)
    return (file_count, subdir_count)


# ---------------------------------------------------------------------------
# FUNCTION: walk_directories
# ---------------------------------------------------------------------------

def walk_directories(root: str, recursive: bool, max_depth):
    """Generator yielding ``(full_path, depth)`` tuples for a directory tree.

    Traversal is top-down.  Symlinked directories are never followed.
    All yielded paths use forward slashes.

    Parameters
    ----------
    root : str
    recursive : bool
        If ``False``, only depth-0 and depth-1 entries are yielded.
    max_depth : int | None
        Maximum depth to descend; ``None`` = unlimited.

    Yields
    ------
    tuple[str, int]
    """
    root_normalised = root.replace("\\", "/").rstrip("/")
    yield (root_normalised, 0)

    try:
        walker = os.walk(root, topdown=True, followlinks=False)
    except PermissionError:
        print(
            f"[fw_dirmap WARNING] PermissionError — cannot walk: {root}",
            file=sys.stderr,
        )
        return

    for dirpath, dirnames, _filenames in walker:
        norm_dirpath = dirpath.replace("\\", "/")

        if norm_dirpath.rstrip("/") == root_normalised:
            if not recursive:
                for name in list(dirnames):
                    yield (f"{root_normalised}/{name}", 1)
                dirnames[:] = []
            continue

        relative = norm_dirpath[len(root_normalised):].lstrip("/")
        depth = relative.count("/") + 1

        yield (norm_dirpath, depth)

        if max_depth is not None and depth >= max_depth:
            dirnames[:] = []


# ===========================================================================
# GROUP 3 — Record builder
# ===========================================================================

_PROGRESS_INTERVAL = 50


# ---------------------------------------------------------------------------
# FUNCTION: build_dir_records
# ---------------------------------------------------------------------------

def build_dir_records(
    roots: list,
    recursive: bool,
    max_depth,
    run_id: str,
    # Dependency injection hooks (used by unit tests; None = use real functions)
    _validate=None,
    _detect_store=None,
    _count_contents=None,
    _walk=None,
) -> list:
    """Walk one or more root directories and return a list of DirRecord dicts.

    Parameters
    ----------
    roots : list[str]
    recursive : bool
    max_depth : int | None
    run_id : str

    Returns
    -------
    list[dict]
        DirRecord dicts in tree order (parents before children).
        Dir IDs are sequential across all roots (D001, D002, …).
    """
    if _validate is None:
        _validate = validate_and_normalize_path
    if _detect_store is None:
        _detect_store = detect_source_store
    if _count_contents is None:
        _count_contents = count_dir_contents
    if _walk is None:
        _walk = walk_directories

    all_records = []
    dir_counter  = 0

    for raw_root in roots:
        scan_root = _validate(raw_root)
        print(f"  Scanning: {scan_root}")

        for full_path, depth in _walk(scan_root, recursive, max_depth):
            dir_counter += 1
            if dir_counter % _PROGRESS_INTERVAL == 0:
                print(f"  ... {dir_counter} dirs found")

            full_path_fwd = full_path.replace("\\", "/")

            if os.path.normcase(full_path) == os.path.normcase(scan_root):
                relative_dir = "."
            else:
                relative_dir = os.path.relpath(full_path, scan_root).replace("\\", "/")

            dir_name = (
                os.path.basename(scan_root.rstrip("/\\")) if depth == 0
                else os.path.basename(full_path)
            )

            dir_id = f"D{dir_counter:03d}"
            file_count, subdir_count = _count_contents(full_path)
            source_store = _detect_store(full_path)

            all_records.append({
                "dir_id":       dir_id,
                "scan_root":    scan_root,
                "relative_dir": relative_dir,
                "dir_name":     dir_name,
                "full_path":    full_path_fwd,
                "depth":        depth,
                "file_count":   file_count,
                "subdir_count": subdir_count,
                "source_store": source_store,
                "run_id":       run_id,
            })

    print(f"  Done. Total directories mapped: {dir_counter}")
    return all_records


# ===========================================================================
# GROUP 4 — Excel I/O for Dir_Inventory sheet
# ===========================================================================

_HEADERS: List[str] = [
    "DirID",            # A
    "ScanRoot",         # B
    "RelativeDir",      # C
    "DirName",          # D
    "FullPath",         # E
    "Depth",            # F
    "FileCount",        # G
    "SubdirCount",      # H
    "ProcessingStatus", # I
    "Notes",            # J
]

_COL_WIDTHS: dict = {
    "DirID":            8,
    "ScanRoot":         35,
    "RelativeDir":      45,
    "DirName":          25,
    "FullPath":         60,
    "Depth":            7,
    "FileCount":        10,
    "SubdirCount":      11,
    "ProcessingStatus": 16,
    "Notes":            25,
}

_STATUS_FILLS: dict = {
    "not_scanned":    PatternFill(fill_type="solid", fgColor="F2F2F2"),
    "file_walked":    PatternFill(fill_type="solid", fgColor="DAEEF3"),
    "classified":     PatternFill(fill_type="solid", fgColor="E2EFDA"),
    "triaged":        PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "deep_processed": PatternFill(fill_type="solid", fgColor="D5A6BD"),
}

_EVEN_ROW_FILL = PatternFill(fill_type="solid", fgColor="F9F9F9")
_ODD_ROW_FILL  = PatternFill(fill_type="solid", fgColor="FFFFFF")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_HEADER_FILL   = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# FUNCTION: open_or_create_workbook
# ---------------------------------------------------------------------------

def open_or_create_workbook(path: str) -> Workbook:
    """Open an existing workbook or create a fresh one.

    Parameters
    ----------
    path : str

    Returns
    -------
    openpyxl.Workbook
        Caller is responsible for calling ``wb.save(path)`` when done.
    """
    path = os.path.abspath(path)

    if os.path.exists(path):
        lock_file = os.path.join(
            os.path.dirname(path), "~$" + os.path.basename(path))
        if os.path.exists(lock_file):
            print(
                f"WARNING: Lock file detected — Excel may have '{path}' open. "
                "Proceeding anyway, but save may fail."
            )
        print(f"Opened existing: {path}")
        return load_workbook(path)

    print(f"Creating new workbook: {path}")
    wb = Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    return wb


# ---------------------------------------------------------------------------
# FUNCTION: ensure_dir_inventory_sheet
# ---------------------------------------------------------------------------

def ensure_dir_inventory_sheet(wb: Workbook) -> Worksheet:
    """Return the Dir_Inventory worksheet, creating and styling it if absent."""
    sheet_name = "Dir_Inventory"
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws: Worksheet = wb.create_sheet(title=sheet_name)
    ws.append(_HEADERS)

    for col_idx in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS[header]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}1"
    return ws


# ---------------------------------------------------------------------------
# FUNCTION: write_dir_inventory_rows
# ---------------------------------------------------------------------------

def write_dir_inventory_rows(wb: Workbook, records: List[dict]) -> None:
    """Append DirRecord rows to the Dir_Inventory sheet with full formatting."""
    ws: Worksheet = ensure_dir_inventory_sheet(wb)
    STATUS_COL_IDX = _HEADERS.index("ProcessingStatus") + 1

    for record_num, record in enumerate(records, start=1):
        depth  = int(record.get("depth", 0))
        status = record.get("processing_status", "not_scanned")
        notes  = record.get("notes", "")
        indented_dir_name = ("  " * depth) + str(record.get("dir_name", ""))

        ws.append([
            record.get("dir_id",        ""),
            record.get("scan_root",     ""),
            record.get("relative_dir",  ""),
            indented_dir_name,
            record.get("full_path",     ""),
            depth,
            int(record.get("file_count",   0)),
            int(record.get("subdir_count", 0)),
            status,
            notes,
        ])

        current_row = ws.max_row
        base_fill = _EVEN_ROW_FILL if (current_row % 2 == 0) else _ODD_ROW_FILL

        for col_idx in range(1, len(_HEADERS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = (
                _STATUS_FILLS.get(status, _STATUS_FILLS["not_scanned"])
                if col_idx == STATUS_COL_IDX
                else base_fill
            )

        ws.row_dimensions[current_row].outline_level = depth

        if record_num % 100 == 0:
            print(f"  Written {record_num} rows...")

    ws.sheet_properties.outlinePr.summaryBelow = False


# ===========================================================================
# GROUP 5 — Excel I/O for Dir_Processing_Status sheet
# ===========================================================================

_DPS_SHEET_NAME = "Dir_Processing_Status"

_DPS_HEADERS = [
    "DirID",
    "FullPath",
    "FileFamily",
    "FileCount",
    "ProcessingStatus",
    "LastProcessed",
    "Notes",
]

_DPS_FILE_FAMILIES = [
    "pdf",
    "image",
    "word_doc",
    "spreadsheet",
    "presentation",
    "email_export",
    "video",
    "audio",
    "other",
]

_DPS_COL_WIDTHS = {"A": 8, "B": 60, "C": 14, "D": 10, "E": 16, "F": 16, "G": 25}

_DPS_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_DPS_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_DPS_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_DPS_FILL_EVEN    = PatternFill(fill_type="solid", fgColor="F9F9F9")
_DPS_FILL_ODD     = PatternFill(fill_type="solid", fgColor="FFFFFF")
_DPS_FILL_EMAIL   = PatternFill(fill_type="solid", fgColor="FFF2CC")
_DPS_THIN_SIDE    = Side(style="thin", color="CCCCCC")
_DPS_CELL_BORDER  = Border(
    left=_DPS_THIN_SIDE, right=_DPS_THIN_SIDE,
    top=_DPS_THIN_SIDE,  bottom=_DPS_THIN_SIDE,
)


# ---------------------------------------------------------------------------
# FUNCTION: ensure_dir_processing_status_sheet
# ---------------------------------------------------------------------------

def ensure_dir_processing_status_sheet(wb: Workbook) -> Worksheet:
    """Return the Dir_Processing_Status worksheet, creating it if absent."""
    if _DPS_SHEET_NAME in wb.sheetnames:
        return wb[_DPS_SHEET_NAME]

    ws = wb.create_sheet(title=_DPS_SHEET_NAME)
    ws.append(_DPS_HEADERS)

    for col_idx in range(1, len(_DPS_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = _DPS_HEADER_FONT
        cell.fill      = _DPS_HEADER_FILL
        cell.alignment = _DPS_HEADER_ALIGN
        cell.border    = _DPS_CELL_BORDER

    for col_letter, width in _DPS_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_DPS_HEADERS))}1"
    return ws


# ---------------------------------------------------------------------------
# FUNCTION: initialize_processing_status_rows
# ---------------------------------------------------------------------------

def initialize_processing_status_rows(wb: Workbook, records: list) -> None:
    """Append 9 rows per directory (one per file family) to Dir_Processing_Status."""
    ws = ensure_dir_processing_status_sheet(wb)

    for dir_index, record in enumerate(records, start=1):
        dir_id    = record["dir_id"]
        full_path = record["full_path"]

        for family in _DPS_FILE_FAMILIES:
            ws.append([dir_id, full_path, family, 0, "not_scanned", "", ""])

            current_row = ws.max_row
            if family == "email_export":
                row_fill = _DPS_FILL_EMAIL
            elif current_row % 2 == 0:
                row_fill = _DPS_FILL_EVEN
            else:
                row_fill = _DPS_FILL_ODD

            for col_idx in range(1, len(_DPS_HEADERS) + 1):
                ws.cell(row=current_row, column=col_idx).fill = row_fill

        if dir_index % 100 == 0:
            print(f"  [Dir_Processing_Status] {dir_index} directories initialized...")

    print(f"  [Dir_Processing_Status] Done — {len(records)} "
          f"director{'y' if len(records) == 1 else 'ies'} written.")


# ===========================================================================
# GROUP 6 — Run logging, summary, and main orchestration
# ===========================================================================

_WALK_HISTORY_COLS = [
    "RunID", "RunType", "StartTime", "EndTime",
    "ElapsedSeconds", "RootCount", "DirCount", "Config",
]
_WH_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_WH_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_WH_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# FUNCTION: log_dirmap_run
# ---------------------------------------------------------------------------

def log_dirmap_run(wb: Workbook, run_id: str, config: dict, stats: dict) -> None:
    """Append one summary row to the Walk_History sheet.

    Parameters
    ----------
    wb : Workbook
    run_id : str
    config : dict  — keys: roots (list), recursive (bool), max_depth (int|None)
    stats  : dict  — keys: start_time, end_time (datetime), dir_count, root_count (int)
    """
    sheet_name = "Walk_History"

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        for col_idx, header in enumerate(_WALK_HISTORY_COLS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = _WH_HEADER_FONT
            cell.fill      = _WH_HEADER_FILL
            cell.alignment = _WH_HEADER_ALIGN
        ws.freeze_panes = "A2"
    else:
        ws = wb[sheet_name]

    start_time: datetime = stats["start_time"]
    end_time:   datetime = stats["end_time"]
    elapsed = round((end_time - start_time).total_seconds(), 2)

    ws.append([
        run_id,
        "DIRMAP",
        start_time.isoformat(),
        end_time.isoformat(),
        elapsed,
        stats["root_count"],
        stats["dir_count"],
        json.dumps(config, default=str),
    ])


# ---------------------------------------------------------------------------
# FUNCTION: print_run_summary
# ---------------------------------------------------------------------------

def print_run_summary(records: list, run_id: str, start_time, end_time) -> None:
    """Print a clean terminal summary after a dirmap run."""
    elapsed    = (end_time - start_time).total_seconds()
    top_level  = [r for r in records if r.get("depth", -1) == 0]
    divider    = "=" * 60

    print(divider)
    print("fw_dirmap Run Summary")
    print(divider)
    print(f"Run ID      : {run_id}")
    print(f"Directories : {len(records)}")
    print(f"Roots       : {len(top_level)}")
    print(f"Elapsed     : {elapsed:.2f} seconds")
    print("-" * 60)
    print("Top-level directories:")
    for rec in top_level:
        print(f"  {rec.get('dir_id', '?'):<6}  {rec.get('full_path', '?')}")
    print(divider)


# ---------------------------------------------------------------------------
# FUNCTION: main
# ---------------------------------------------------------------------------

def main() -> None:
    """Full fw_dirmap orchestration."""
    print("=" * 60)
    print("fw_dirmap — Directory Mapper Agent")
    print("=" * 60)

    # 1. Pick directories
    roots = pick_scan_dirs()
    if not roots:
        print("No directories selected. Exiting.")
        return

    # 2. Scan options
    recursive_input = input("Scan recursively? [Y/n]: ").strip().lower()
    recursive = recursive_input not in ("n", "no")

    max_depth_input = input("Max depth (blank = unlimited): ").strip()
    if max_depth_input == "":
        max_depth = None
    else:
        try:
            max_depth = int(max_depth_input)
        except ValueError:
            print(f"Invalid depth '{max_depth_input}' — treating as unlimited.")
            max_depth = None

    # 3. Run ID and workbook path
    run_id  = generate_run_id()
    wb_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "filewalker_master.xlsx")

    # 4. Open workbook
    wb = open_or_create_workbook(wb_path)

    # 5. Walk and build records (timed)
    start_time = datetime.now()
    records    = build_dir_records(roots, recursive, max_depth, run_id)
    end_time   = datetime.now()

    if not records:
        print("Warning: no directory records were produced. Nothing to save.")
        return

    # 6. Write sheets
    write_dir_inventory_rows(wb, records)
    initialize_processing_status_rows(wb, records)

    # 7. Log run
    config = {"roots": [str(r) for r in roots], "recursive": recursive,
              "max_depth": max_depth}
    stats  = {"start_time": start_time, "end_time": end_time,
              "dir_count": len(records),
              "root_count": len([r for r in records if r.get("depth", -1) == 0])}
    log_dirmap_run(wb, run_id, config, stats)

    # 8. Save
    wb.save(wb_path)
    print(f"Saved: {wb_path}")

    # 9. Summary
    print_run_summary(records, run_id, start_time, end_time)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    main()
