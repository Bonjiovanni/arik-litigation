"""
fw_classify_grp5.py
--------------------
main() orchestration for the fw_classify first-pass content signal pipeline.

Coordinates all fw_classify_grp modules:
    grp1 — text extraction + signal detection
    grp2 — Master_File_Inventory Excel I/O
    grp3 — Keywords_Config sheet + DocType inference
    grp4 — Classify_History logging

Reads: Master_File_Inventory rows with ProcessingStatus = "file_listed"
Writes: TextSample, MoneyDetected, DateDetected, KeywordHits,
        LikelyTextBearing, NeedsOCR, DocType, DocSubtype, DocTypeConfidence,
        ProcessingStatus → "classified"

Workbook: C:\\Users\\arika\\OneDrive\\Litigation\\filewalker_master.xlsx
Python 3.11, Windows 10.
"""

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook

import fw_classify_grp1 as grp1
import fw_classify_grp2 as grp2
import fw_classify_grp3 as grp3
import fw_classify_grp4 as grp4


WORKBOOK_PATH = r"C:\Users\arika\OneDrive\Litigation\filewalker_master.xlsx"
ENTITY_INDEX_PATH = r"C:\Users\arika\OneDrive\Copy of EntityIndex_Normalized_v24_fixed.xlsm"


def main() -> None:
    """Run a complete fw_classify first-pass signal extraction.

    Steps:
      1.  Open or create the master workbook.
      2.  Ensure Keywords_Config and Classify_History sheets exist.
      3.  Generate a unique run ID.
      4.  Load active keywords from Keywords_Config.
      5.  Get classifiable rows from Master_File_Inventory.
      6.  For each row: extract text, detect signals, infer flags, classify doc type.
      7.  Write signals back and mark row classified.
      8.  Log Classify_History row.
      9.  Save workbook.
      10. Print final summary.
    """
    try:
        # ------------------------------------------------------------------
        # 1. Open or create workbook
        # ------------------------------------------------------------------
        wb_path = Path(WORKBOOK_PATH)
        if wb_path.exists():
            print(f"[fw_classify] Loading existing workbook: {wb_path}")
            wb = load_workbook(wb_path)
        else:
            print(f"[fw_classify] Creating new workbook: {wb_path}")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ------------------------------------------------------------------
        # 2. Ensure required sheets exist
        # ------------------------------------------------------------------
        grp3.ensure_keywords_config_sheet(wb)
        grp4.ensure_classify_history_sheet(wb)

        # ------------------------------------------------------------------
        # 3. Generate run ID
        # ------------------------------------------------------------------
        run_id: str = datetime.now().strftime("CLASSIFY_%Y%m%d_%H%M%S")
        print(f"[fw_classify] Run ID: {run_id}")

        # ------------------------------------------------------------------
        # 4. Load keywords
        # ------------------------------------------------------------------
        ws_kw = wb["Keywords_Config"]
        keywords = grp3.load_keywords(ws_kw)
        print(f"[fw_classify] Loaded {len(keywords)} active keyword(s).")
        alias_map = grp3.load_entity_aliases(ENTITY_INDEX_PATH)
        print(f"[fw_classify] Loaded {len(alias_map)} entity alias(es).")

        # ------------------------------------------------------------------
        # 5. Get classifiable rows
        # ------------------------------------------------------------------
        if "Master_File_Inventory" not in wb.sheetnames:
            print("[fw_classify] ERROR: Master_File_Inventory sheet not found. "
                  "Run fw_walk first.")
            sys.exit(1)

        ws_inv = wb["Master_File_Inventory"]
        rows = grp2.get_classifiable_rows(ws_inv)
        print(f"[fw_classify] Rows to classify: {len(rows):,}")

        if not rows:
            print("[fw_classify] Nothing to classify.")
            sys.exit(0)

        # ------------------------------------------------------------------
        # 6–7. Process each row
        # ------------------------------------------------------------------
        started_at_str: str = datetime.now().isoformat()
        rows_processed = 0
        rows_updated   = 0
        rows_skipped   = 0
        errors         = 0

        for row_num, record in rows:
            filepath    = record.get("FilePath", "") or ""
            file_family = record.get("FileFamily", "") or ""

            try:
                # a. Extract text sample
                text_sample, extraction_ok = grp1.get_text_sample(
                    filepath, file_family, max_chars=500
                )

                # b. Detect signals
                money        = grp1.detect_money(text_sample)
                dates        = grp1.detect_dates(text_sample)
                kw_hits      = grp1.match_keywords(text_sample, keywords)
                entity_hits  = grp3.detect_entity_hits(text_sample, alias_map)
                has_fields   = grp1.detect_form_fields(filepath) if file_family == "pdf" else ""
                page_density = grp1.check_page_density(filepath)  if file_family == "pdf" else ""

                # c. Infer text-bearing / OCR flags
                likely_text = grp3.infer_likely_text_bearing(
                    file_family, text_sample, extraction_ok
                )
                needs_ocr = page_density if page_density else grp3.infer_needs_ocr(
                    file_family, text_sample, extraction_ok
                )

                # d. Classify doc type
                doc_info = grp3.classify_doc_type(text_sample, kw_hits, file_family)

                # e. Write signals + mark classified
                signals = {
                    "TextSample":        (text_sample or "")[:500],
                    "MoneyDetected":     money,
                    "DateDetected":      dates,
                    "KeywordHits":       kw_hits,
                    "EntityHits":        entity_hits,
                    "HasFormFields":     has_fields,
                    "LikelyTextBearing": likely_text,
                    "NeedsOCR":          needs_ocr,
                    "DocType":           doc_info.get("DocType", ""),
                    "DocSubtype":        doc_info.get("DocSubtype", ""),
                    "DocTypeConfidence": doc_info.get("DocTypeConfidence", ""),
                }
                grp2.write_classify_signals(ws_inv, row_num, signals)
                grp2.mark_row_classified(ws_inv, row_num)
                rows_updated += 1

            except Exception as exc:
                print(f"[fw_classify] ERROR row {row_num} ({filepath}): {exc}")
                errors += 1
            finally:
                rows_processed += 1

            if rows_processed % 100 == 0:
                print(f"[fw_classify]   {rows_processed:,}/{len(rows):,} processed...")

        # ------------------------------------------------------------------
        # 8. Log Classify_History
        # ------------------------------------------------------------------
        stats_dict = {
            "rows_processed": rows_processed,
            "rows_updated":   rows_updated,
            "rows_skipped":   rows_skipped,
            "errors":         errors,
        }
        grp4.log_classify_run(
            wb=wb,
            run_id=run_id,
            started_at=started_at_str,
            stats=stats_dict,
        )

        # ------------------------------------------------------------------
        # 9. Save
        # ------------------------------------------------------------------
        wb.save(WORKBOOK_PATH)

        # ------------------------------------------------------------------
        # 10. Summary
        # ------------------------------------------------------------------
        print(f"\n[fw_classify] Run complete: {run_id}")
        print(f"  Rows processed: {rows_processed:,}")
        print(f"  Updated:        {rows_updated:,}")
        print(f"  Skipped:        {rows_skipped:,}")
        print(f"  Errors:         {errors:,}")
        print(f"Saved: {WORKBOOK_PATH}")

    except KeyboardInterrupt:
        print("\n[fw_classify] Interrupted.")
        sys.exit(0)


if __name__ == "__main__":
    main()
