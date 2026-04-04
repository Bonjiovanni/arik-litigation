#!/usr/bin/env python3
"""
Batch Invoice Processor
  1. Extracts fields from page 1 of every JH Invoice PDF in INVOICE_DIR
  2. Saves individual JSON per invoice  ->  <name>_extracted.json
  3. Builds Excel summary               ->  Invoice_summary.xlsx
       Layout: fields as rows (col A = field name), docs as columns (L->R oldest->newest)
  4. Merges all page-1s into one PDF    ->  Invoice_all_p1_chrono.pdf  (oldest -> newest)
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime

import fitz          # pymupdf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Import extraction logic from the visualizer
# ---------------------------------------------------------------------------
sys.path.insert(0, str(Path(__file__).parent))
from visualize_invoice_fields import extract_field_values, FIELD_ORDER

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
INVOICE_DIR = Path(
    r"C:\Users\arika\OneDrive\Litigation\08_INCOMING"
    r"\John Hancock long term care document dpwnloads on 11-9-2025"
)
OUT_DIR = INVOICE_DIR / "outputs"
OUT_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Field type definitions  ->  controls Excel cell value type + number format
# ---------------------------------------------------------------------------
FIELD_TYPES = {
    "doc_type":                  "text",
    "doc_subtype":               "text",
    "issuing_entity":            "text",
    "invoice_number":            "int",
    "doc_date":                  "date",
    "page_count":                "int",
    "policy_number":             "int",
    "claim_number":              "text",
    "provider_name":             "text",
    "total_charges":             "currency",
    "hourly_rate":               "currency",
    "service_date_from":         "date",
    "service_date_to":           "date",
    "submitted_by":              "text",
    "date_submitted":            "date",
    "provider_phone":            "int",
    "provider_email":            "text",
    "care_at_home":              "text",
    "q1a_response":              "text",
    "shared_care":               "text",
    "q2a_who_else":              "text",
    "q2b_jh_customer":           "text",
    "q2c_other_policy":          "text",
    "q2d_other_claim":           "text",
    "assignment_of_benefits":    "text",
    "proof_of_payment_type":     "text",
    "q4a_payment_desc":          "text",
    "additional_info":           "text",
    "fraud_attestation_checked":  "text",
    "fraud_attestation_text":     "text",
    "proof_payment_attested":     "text",
    "proof_payment_attest_text":  "text",
}

XL_FMT = {
    "currency": '"$"#,##0.00',
    "date":     "MM/DD/YYYY",
    "int":      "0",
    "text":     "@",
}

XL_ALIGN = {
    "currency": "right",
    "date":     "center",
    "int":      "right",
    "text":     "left",
}


def coerce(raw, ftype):
    """Convert extracted string to a Python native type for Excel."""
    if not raw:
        return ""
    if ftype == "currency":
        try:
            return float(raw.replace("$", "").replace(",", "").strip())
        except ValueError:
            return raw
    if ftype == "date":
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(raw.strip(), fmt)
            except ValueError:
                pass
        return raw
    if ftype == "int":
        try:
            return int(raw.strip())
        except ValueError:
            return raw
    return raw  # text


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def find_invoices():
    """Return list of Invoice PDFs sorted oldest->newest by date in filename.
    Only includes files whose names START with 'Invoice_' and contain a YYYY-MM-DD date."""
    DATE_RE = re.compile(r"_(\d{4}-\d{2}-\d{2})_")
    pdfs = [
        p for p in INVOICE_DIR.glob("Invoice_*.pdf")
        if DATE_RE.search(p.name)
    ]

    def date_key(p):
        m = DATE_RE.search(p.name)
        return m.group(1) if m else "0000-00-00"

    return sorted(pdfs, key=date_key)


def extract_all(pdfs):
    """Extract field values from page 1 of every PDF. Returns list of (pdf_path, values_dict)."""
    results = []
    for pdf_path in pdfs:
        print(f"  Extracting: {pdf_path.name}")
        doc = fitz.open(str(pdf_path))
        page = doc[0]
        vals = extract_field_values(page, pdf_path=pdf_path, page_count=doc.page_count, doc=doc)
        doc.close()
        results.append((pdf_path, vals))
    return results


def save_jsons(results):
    """Save one JSON per invoice into OUT_DIR."""
    for pdf_path, vals in results:
        out = OUT_DIR / (pdf_path.stem + "_extracted.json")
        data = {
            "_source":    pdf_path.name,
            "_extracted": datetime.now().isoformat(timespec="seconds"),
            "fields":     vals,
        }
        out.write_text(json.dumps(data, indent=2), encoding="utf-8")
        print(f"  JSON saved: {out.name}")


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1A3A5C")
FIELD_FILL   = PatternFill("solid", fgColor="E8EFF8")
ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")
MISSING_FILL = PatternFill("solid", fgColor="FFF3F3")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Fields"

    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Calibri", size=10)
    mono_font = Font(name="Consolas", size=9)

    def hdr_cell(ws, row, col, value, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = hdr_font
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border    = BORDER
        return c

    # Col A header
    hdr_cell(ws, 1, 1, "Field")
    ws.column_dimensions["A"].width = 28

    # Doc columns (B onward) — use invoice number + date as header
    for i, (pdf_path, vals) in enumerate(results):
        col = i + 2
        inv_num = vals.get("invoice_number", "")
        date_str = vals.get("doc_date", "")
        # Reformat date from YYYY-MM-DD to MM/DD/YYYY for readability
        try:
            date_str = datetime.strptime(date_str, "%Y-%m-%d").strftime("%m/%d/%Y")
        except (ValueError, TypeError):
            pass
        label = f"{date_str}\nI-{inv_num}" if inv_num else date_str
        hdr_cell(ws, 1, col, label, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = 22

    ws.row_dimensions[1].height = 36

    # Data rows: one per field
    for row_idx, (schema_key, display_label, box_num, _color_key) in enumerate(FIELD_ORDER):
        row = row_idx + 2

        # Col A — field name
        a = ws.cell(row=row, column=1, value=display_label)
        a.font      = Font(name="Consolas", bold=True, size=9)
        a.fill      = FIELD_FILL
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        a.border    = BORDER

        # Doc value columns
        ftype  = FIELD_TYPES.get(schema_key, "text")
        halign = XL_ALIGN[ftype]
        for i, (pdf_path, vals) in enumerate(results):
            col   = i + 2
            raw   = vals.get(schema_key, "")
            value = coerce(raw, ftype)
            c     = ws.cell(row=row, column=col, value=value)
            c.font          = mono_font
            c.number_format = XL_FMT[ftype]
            c.alignment     = Alignment(horizontal=halign, vertical="center",
                                        indent=1 if halign == "left" else 0)
            c.border        = BORDER
            c.fill          = MISSING_FILL if not raw else (
                ALT_FILL if row_idx % 2 == 1 else WHITE_FILL
            )

        ws.row_dimensions[row].height = 16

    # Freeze panes at B2
    ws.freeze_panes = "B2"

    # Source filename row (insert after header)
    ws.insert_rows(2)
    fn_font = Font(name="Calibri", italic=True, color="666666", size=8)
    ws.cell(row=2, column=1, value="(source file)").font = fn_font
    for i, (pdf_path, _) in enumerate(results):
        c = ws.cell(row=2, column=i + 2, value=pdf_path.name)
        c.font      = fn_font
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.fill      = PatternFill("solid", fgColor="F0F4FA")
        c.border    = BORDER
    ws.row_dimensions[2].height = 13

    out_path = OUT_DIR / "invoice_summary_latest.xlsx"
    tmp_path = OUT_DIR / f"invoice_summary_tmp_{datetime.now().strftime('%H%M%S')}.xlsx"
    wb.save(str(tmp_path))
    try:
        if out_path.exists():
            out_path.unlink()
        tmp_path.rename(out_path)
        print(f"  Excel saved: {out_path}")
        return out_path
    except PermissionError:
        # File is open — find an unused invoice_summary_latest (n).xlsx
        n = 1
        while True:
            alt = OUT_DIR / f"invoice_summary_latest ({n}).xlsx"
            if not alt.exists():
                tmp_path.rename(alt)
                print(f"  WARNING: invoice_summary_latest.xlsx is open — saved as:")
                print(f"    {alt}")
                return alt
            n += 1


# ---------------------------------------------------------------------------
# Merge page 1s chronologically into one PDF
# ---------------------------------------------------------------------------
def merge_page1s(pdfs):
    merged = fitz.open()
    for pdf_path in pdfs:
        src = fitz.open(str(pdf_path))
        # Include page 2 if it has extractable text (2023+ digital cover sheet spans 2 pages)
        to_page = 0
        if src.page_count > 1 and len(src[1].get_text().strip()) > 100:
            to_page = 1
        merged.insert_pdf(src, from_page=0, to_page=to_page)
        src.close()
        print(f"  Added {to_page + 1} cover page(s): {pdf_path.name}")

    out_path = OUT_DIR / "Invoice_all_p1_chrono.pdf"
    tmp_path = OUT_DIR / f"Invoice_all_p1_chrono_tmp_{datetime.now().strftime('%H%M%S')}.pdf"
    merged.save(str(tmp_path))
    merged.close()
    try:
        if out_path.exists():
            out_path.unlink()
        tmp_path.rename(out_path)
        print(f"  Merged PDF saved: {out_path}")
        return out_path
    except PermissionError:
        print(f"  WARNING: {out_path.name} is open in another app -- new version saved as:")
        print(f"    {tmp_path}")
        print(f"  Close the old file, then rename _new to Invoice_all_p1_chrono.pdf")
        return tmp_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("\n=== Batch Invoice Processor ===\n")

    pdfs = find_invoices()
    print(f"Found {len(pdfs)} Invoice PDFs (oldest to newest):")
    for p in pdfs:
        print(f"  {p.name}")
    print()

    print("Extracting fields (page 1 only)...")
    results = extract_all(pdfs)
    print()

    print("Saving individual JSONs...")
    save_jsons(results)
    print()

    print("Building Excel summary...")
    xl_path = build_excel(results)
    print()

    print("Merging page 1s chronologically...")
    pdf_path = merge_page1s(pdfs)
    print()

    print("=== Done ===")
    print(f"  Excel  : {xl_path}")
    print(f"  PDF    : {pdf_path}")

    import subprocess
    subprocess.Popen(["cmd", "/c", "start", "", str(xl_path)])
    subprocess.Popen(["cmd", "/c", "start", "", str(pdf_path)])


if __name__ == "__main__":
    main()
