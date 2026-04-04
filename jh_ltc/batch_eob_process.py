#!/usr/bin/env python3
"""
Batch EOB Processor
  1. Extracts fields from every EOB PDF in EOB_DIR (skipping /baks)
  2. Saves individual JSON per EOB  →  <name>_extracted.json
  3. Builds Excel summary            →  EOB_summary.xlsx
       Layout: fields as rows (col A = field name), docs as columns (L→R oldest→newest)
  4. Merges all EOBs into one PDF   →  EOB_all_chrono.pdf  (oldest → newest)
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime

import fitz          # pymupdf
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Import extraction logic from the visualizer
# ---------------------------------------------------------------------------
sys.path.insert(0, str(Path(__file__).parent))
from visualize_eob_fields import extract_field_values, FIELD_ORDER

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EOB_DIR = Path(
    r"C:\Users\arika\OneDrive\Litigation\08_INCOMING"
    r"\John Hancock long term care document dpwnloads on 11-9-2025"
)
OUT_DIR = EOB_DIR  # outputs saved alongside sources

# Fields to show in Excel, in display order (schema_key, display_label, box_num, color_key)
FIELDS = FIELD_ORDER   # from visualize_eob_fields

# ---------------------------------------------------------------------------
# Field type definitions  →  controls Excel cell value type + number format
# ---------------------------------------------------------------------------
# "currency" "$1,234.56"  | "date" "M/D/YYYY"  | "int" whole number  | "text" plain string
FIELD_TYPES = {
    "issuing_entity":            "text",
    "page_number":               "int",
    "page_total":                "int",
    "claim_id":                  "text",
    "group_nbr":                 "int",
    "payment_date":              "date",
    "payment_amount":            "currency",
    "payment_recipient_name":    "text",
    "payment_recipient_address": "text",
    "insured_name":              "text",
    "insured_address":           "text",
    "transaction_seq":           "text",
    "provider":                  "text",
    "service_type":              "text",
    "service_date_from":         "date",
    "service_date_to":           "date",
    "total_charge":              "currency",
    "exceeds_plan_max":          "currency",
    "benefit_paid":              "currency",
    "lifetime_used":             "currency",
    "lifetime_maximum":          "currency",
}

XL_FMT = {
    "currency": '"$"#,##0.00',
    "date":     "MM/DD/YYYY",
    "int":      "0",
    "text":     "@",
}

XL_ALIGN = {
    "currency": "right",
    "date":     "center",
    "int":      "right",
    "text":     "left",
}


def coerce(raw, ftype):
    """Convert extracted string to a Python native type for Excel."""
    if not raw:
        return ""
    if ftype == "currency":
        try:
            return float(raw.replace("$", "").replace(",", "").strip())
        except ValueError:
            return raw
    if ftype == "date":
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%-m/%-d/%Y"):
            try:
                return datetime.strptime(raw.strip(), fmt)
            except ValueError:
                pass
        return raw          # leave as string if parse fails
    if ftype == "int":
        try:
            return int(raw.strip())
        except ValueError:
            return raw
    return raw              # text


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def find_eobs():
    """Return list of EOB PDFs sorted oldest→newest by YYYY-MM-DD prefix.
    Only includes files whose names START with a YYYY-MM-DD date (skips output files)."""
    DATE_PREFIX = re.compile(r"^\d{4}-\d{2}-\d{2}")
    pdfs = [
        p for p in EOB_DIR.glob("*.pdf")
        if DATE_PREFIX.match(p.name) and "EOB" in p.name
    ]

    def date_key(p):
        m = re.match(r"(\d{4}-\d{2}-\d{2})", p.name)
        return m.group(1) if m else "0000-00-00"

    return sorted(pdfs, key=date_key)


def extract_all(pdfs):
    """Extract field values from every PDF. Returns list of (pdf_path, values_dict)."""
    results = []
    for pdf_path in pdfs:
        print(f"  Extracting: {pdf_path.name}")
        doc  = fitz.open(str(pdf_path))
        page = doc[0]
        vals = extract_field_values(page)
        doc.close()
        results.append((pdf_path, vals))
    return results


def save_jsons(results):
    """Save one JSON per EOB into OUT_DIR."""
    for pdf_path, vals in results:
        out = OUT_DIR / (pdf_path.stem + "_extracted.json")
        data = {
            "_source":    pdf_path.name,
            "_extracted": datetime.now().isoformat(timespec="seconds"),
            "fields":     vals,
        }
        out.write_text(json.dumps(data, indent=2), encoding="utf-8")
        print(f"  JSON saved: {out.name}")


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1A3A5C")
FIELD_FILL   = PatternFill("solid", fgColor="E8EFF8")
ALT_FILL     = PatternFill("solid", fgColor="F5F5F5")
MISSING_FILL = PatternFill("solid", fgColor="FFF3F3")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EOB Fields"

    n_docs = len(results)

    # ---- Row 1: column headers ----------------------------------------
    # Col A: "Field"
    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Calibri", size=10)
    mono_font = Font(name="Consolas", size=9)

    def hdr_cell(ws, row, col, value, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = hdr_font
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border    = BORDER
        return c

    # Col A header
    hdr_cell(ws, 1, 1, "Field")
    ws.column_dimensions["A"].width = 28

    # Doc columns (B onward) — use date from filename + stem as header
    for i, (pdf_path, _) in enumerate(results):
        col = i + 2
        # Extract YYYY-MM-DD from filename for header
        m = re.match(r"(\d{4}-\d{2}-\d{2})(.*)", pdf_path.stem)
        if m:
            date_str = m.group(1)
            suffix   = m.group(2).strip("-").strip()
            label    = f"{date_str}\n{suffix}" if suffix else date_str
        else:
            label = pdf_path.stem
        hdr_cell(ws, 1, col, label, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = 22

    ws.row_dimensions[1].height = 36

    # ---- Data rows: one per field -------------------------------------
    for row_idx, (schema_key, display_label, box_num, _color_key) in enumerate(FIELDS):
        row = row_idx + 2

        # Col A — field name
        a = ws.cell(row=row, column=1, value=display_label)
        a.font      = Font(name="Consolas", bold=True, size=9)
        a.fill      = FIELD_FILL
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        a.border    = BORDER

        # Doc value columns
        ftype = FIELD_TYPES.get(schema_key, "text")
        halign = XL_ALIGN[ftype]
        for i, (pdf_path, vals) in enumerate(results):
            col    = i + 2
            raw    = vals.get(schema_key, "")
            value  = coerce(raw, ftype)
            c      = ws.cell(row=row, column=col, value=value)
            c.font         = mono_font
            c.number_format = XL_FMT[ftype]
            c.alignment    = Alignment(horizontal=halign, vertical="center", indent=1 if halign == "left" else 0)
            c.border       = BORDER
            c.fill         = MISSING_FILL if not raw else (
                ALT_FILL if row_idx % 2 == 1 else WHITE_FILL
            )

        ws.row_dimensions[row].height = 16

    # ---- Freeze panes at B2 ------------------------------------------
    ws.freeze_panes = "B2"

    # ---- Source filename row at top (row 0 = actual row 1 shifted) ---
    # Add a row 2 with just the raw filename for reference
    ws.insert_rows(2)
    fn_font = Font(name="Calibri", italic=True, color="666666", size=8)
    ws.cell(row=2, column=1, value="(source file)").font = fn_font
    for i, (pdf_path, _) in enumerate(results):
        c = ws.cell(row=2, column=i + 2, value=pdf_path.name)
        c.font      = fn_font
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.fill      = PatternFill("solid", fgColor="F0F4FA")
        c.border    = BORDER
    ws.row_dimensions[2].height = 13

    out_path = OUT_DIR / "EOB_summary.xlsx"
    wb.save(str(out_path))
    print(f"  Excel saved: {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# Merge PDFs chronologically
# ---------------------------------------------------------------------------
def merge_pdfs(pdfs):
    merged = fitz.open()
    for pdf_path in pdfs:
        src = fitz.open(str(pdf_path))
        merged.insert_pdf(src)
        src.close()
        print(f"  Added to merged PDF: {pdf_path.name}")

    out_path = OUT_DIR / "EOB_all_chrono.pdf"
    tmp_path = OUT_DIR / "EOB_all_chrono_new.pdf"
    merged.save(str(tmp_path))
    merged.close()
    # Try to replace the existing file; if locked (open in reader), keep the _new copy
    try:
        if out_path.exists():
            out_path.unlink()
        tmp_path.rename(out_path)
        print(f"  Merged PDF saved: {out_path}")
        return out_path
    except PermissionError:
        print(f"  WARNING: {out_path.name} is open in another app — new version saved as:")
        print(f"    {tmp_path}")
        print(f"  Close the old file, then rename _new to EOB_all_chrono.pdf")
        return tmp_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("\n=== Batch EOB Processor ===\n")

    pdfs = find_eobs()
    print(f"Found {len(pdfs)} EOB PDFs (oldest to newest):")
    for p in pdfs:
        print(f"  {p.name}")
    print()

    print("Extracting fields...")
    results = extract_all(pdfs)
    print()

    print("Saving individual JSONs...")
    save_jsons(results)
    print()

    print("Building Excel summary...")
    xl_path = build_excel(results)
    print()

    print("Merging PDFs chronologically...")
    pdf_path = merge_pdfs(pdfs)
    print()

    print("=== Done ===")
    print(f"  Excel  : {xl_path}")
    print(f"  PDF    : {pdf_path}")

    # Open both outputs
    import subprocess
    subprocess.Popen(["cmd", "/c", "start", "", str(xl_path)])
    subprocess.Popen(["cmd", "/c", "start", "", str(pdf_path)])


if __name__ == "__main__":
    main()
