"""
fw_dirmap_grp4.py
-----------------
Directory mapper agent helper functions for the fw_dirmap project.
Handles Excel I/O for the Dir_Inventory sheet using openpyxl.

Target workbook : filewalker_master.xlsx (same dir as script, or full path)
Sheet           : Dir_Inventory
Column layout   : Option C (Hybrid)
Platform        : Windows 10 only
"""

import os
from typing import List

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Column definitions (order matters — corresponds to A-J)
# ---------------------------------------------------------------------------

_HEADERS: List[str] = [
    "DirID",            # A
    "ScanRoot",         # B
    "RelativeDir",      # C
    "DirName",          # D
    "FullPath",         # E
    "Depth",            # F
    "FileCount",        # G
    "SubdirCount",      # H
    "ProcessingStatus", # I
    "Notes",            # J
]

_COL_WIDTHS: dict = {
    "DirID":            8,
    "ScanRoot":         35,
    "RelativeDir":      45,
    "DirName":          25,
    "FullPath":         60,
    "Depth":            7,
    "FileCount":        10,
    "SubdirCount":      11,
    "ProcessingStatus": 16,
    "Notes":            25,
}

# ---------------------------------------------------------------------------
# ProcessingStatus fill colours
# ---------------------------------------------------------------------------

_STATUS_FILLS: dict = {
    "not_scanned":    PatternFill(fill_type="solid", fgColor="F2F2F2"),
    "file_walked":    PatternFill(fill_type="solid", fgColor="DAEEF3"),
    "classified":     PatternFill(fill_type="solid", fgColor="E2EFDA"),
    "triaged":        PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "deep_processed": PatternFill(fill_type="solid", fgColor="D5A6BD"),
}

_EVEN_ROW_FILL = PatternFill(fill_type="solid", fgColor="F9F9F9")
_ODD_ROW_FILL  = PatternFill(fill_type="solid", fgColor="FFFFFF")

_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# FUNCTION 1: open_or_create_workbook
# ---------------------------------------------------------------------------

def open_or_create_workbook(path: str) -> Workbook:
    """Open an existing workbook or create a fresh one.

    Parameters
    ----------
    path : str
        Full file-system path to the target .xlsx file.

    Returns
    -------
    openpyxl.Workbook
        The loaded or newly created workbook object.
        Caller is responsible for calling wb.save(path) when done.
    """
    path = os.path.abspath(path)

    if os.path.exists(path):
        # Check for Excel lock file (~$filename)
        lock_file = os.path.join(
            os.path.dirname(path),
            "~$" + os.path.basename(path),
        )
        if os.path.exists(lock_file):
            print(
                f"WARNING: Lock file detected — Excel may have '{path}' open. "
                "Proceeding anyway, but save may fail if Excel holds an exclusive lock."
            )

        print(f"Opened existing: {path}")
        return load_workbook(path)

    print(f"Creating new workbook: {path}")
    wb = Workbook()

    # Remove the default sheet openpyxl adds
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]

    return wb


# ---------------------------------------------------------------------------
# FUNCTION 2: ensure_dir_inventory_sheet
# ---------------------------------------------------------------------------

def ensure_dir_inventory_sheet(wb: Workbook) -> Worksheet:
    """Return the Dir_Inventory worksheet, creating and styling it if absent.

    Parameters
    ----------
    wb : openpyxl.Workbook

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet
    """
    sheet_name = "Dir_Inventory"

    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws: Worksheet = wb.create_sheet(title=sheet_name)

    ws.append(_HEADERS)

    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    for col_idx, header in enumerate(_HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _COL_WIDTHS[header]

    ws.freeze_panes = "A2"

    last_col_letter = get_column_letter(len(_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    return ws


# ---------------------------------------------------------------------------
# FUNCTION 3: write_dir_inventory_rows
# ---------------------------------------------------------------------------

def write_dir_inventory_rows(wb: Workbook, records: List[dict]) -> None:
    """Append DirRecord rows to the Dir_Inventory sheet with full formatting.

    Parameters
    ----------
    wb : openpyxl.Workbook
    records : list[dict]
        List of DirRecord dicts. Required keys:
            dir_id, scan_root, relative_dir, dir_name,
            full_path, depth, file_count, subdir_count
        Optional keys (with defaults):
            processing_status → "not_scanned"
            notes → ""
    """
    ws: Worksheet = ensure_dir_inventory_sheet(wb)

    STATUS_COL_IDX = _HEADERS.index("ProcessingStatus") + 1  # column 9

    for record_num, record in enumerate(records, start=1):
        depth  = int(record.get("depth", 0))
        status = record.get("processing_status", "not_scanned")
        notes  = record.get("notes", "")

        # Indent DirName visually by depth
        indented_dir_name = ("  " * depth) + str(record.get("dir_name", ""))

        row_values = [
            record.get("dir_id",        ""),
            record.get("scan_root",     ""),
            record.get("relative_dir",  ""),
            indented_dir_name,
            record.get("full_path",     ""),
            depth,
            int(record.get("file_count",   0)),
            int(record.get("subdir_count", 0)),
            status,
            notes,
        ]

        ws.append(row_values)
        current_row = ws.max_row

        # Alternating row shading
        base_fill = _EVEN_ROW_FILL if (current_row % 2 == 0) else _ODD_ROW_FILL

        for col_idx in range(1, len(_HEADERS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            if col_idx == STATUS_COL_IDX:
                cell.fill = _STATUS_FILLS.get(status, _STATUS_FILLS["not_scanned"])
            else:
                cell.fill = base_fill

        # Excel row outline grouping for expand/collapse by depth
        ws.row_dimensions[current_row].outline_level = depth

        if record_num % 100 == 0:
            print(f"  Written {record_num} rows...")

    # +/- controls appear above each group (not below)
    ws.sheet_properties.outlinePr.summaryBelow = False
