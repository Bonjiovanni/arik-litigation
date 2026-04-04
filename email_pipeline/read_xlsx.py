"""
read_xlsx.py

Dumps an Excel file to JSON on stdout.
Claude.ai can run this via Desktop Commander to read any xlsx file.

Usage:
    python read_xlsx.py <path_to_xlsx> [sheet_name]

    If sheet_name is omitted, dumps all sheets.
    Output is JSON: { "sheet_name": [ {col: val, ...}, ... ], ... }
"""

import json
import sys
from pathlib import Path
import openpyxl


def read_xlsx(path: str, sheet_name: str = None) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
    result = {}

    for name in target_sheets:
        if name not in wb.sheetnames:
            result[name] = f"ERROR: sheet '{name}' not found. Available: {wb.sheetnames}"
            continue

        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[name] = []
            continue

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue  # skip blank rows
            data.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})

        result[name] = data

    wb.close()
    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python read_xlsx.py <path_to_xlsx> [sheet_name]", file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None

    if not Path(path).exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        sys.exit(1)

    data = read_xlsx(path, sheet)
    print(json.dumps(data, indent=2, ensure_ascii=False))
