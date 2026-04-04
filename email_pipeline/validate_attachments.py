#!/usr/bin/env python3
"""
Aid4Mail Attachment Integrity Validator
Validates attachment exports against JSON metadata and normalized Excel index.
"""

import json
import hashlib
import os
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ─── CONFIGURE THESE PATHS ────────────────────────────────────────────────────
ATTACHMENTS_JSON = r"C:\Users\arika\OneDrive\Litigation\Aid4Mail Exports\Json attachments\json_attachments.json"
ATTACHMENTS_XLSX = r"C:\Users\arika\OneDrive\Litigation\Aid4Mail Exports\Json attachments\aid4mail_attachments_normalized.xlsx"
ATTACHMENT_DIR   = r"C:\Users\arika\OneDrive\Litigation\Aid4Mail Exports\Json attachments\Attachment"
EMBEDDED_DIR     = r"C:\Users\arika\OneDrive\Litigation\Aid4Mail Exports\Json attachments\Embedded"
OUTPUT_XLSX      = r"C:\Users\arika\OneDrive\Litigation\Aid4Mail Exports\Json attachments\validation_report.xlsx"

# Set True to compute SHA-256 hashes (slow for large collections)
DO_HASHES = False

# Case sensitivity: Windows = False, macOS/Linux = True
CASE_SENSITIVE = (os.name != 'nt')
# ──────────────────────────────────────────────────────────────────────────────


def normalize_name(name: str) -> str:
    return name if CASE_SENSITIVE else name.lower()


def hash_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def load_json(path: str) -> list:
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, dict):
        for key in ('emails', 'messages', 'records', 'data'):
            if key in data:
                return data[key]
        for v in data.values():
            if isinstance(v, list):
                return v
        return [data]
    elif isinstance(data, list):
        return data
    return []


def scan_folder(folder: str) -> dict:
    """Returns {normalized_basename: [Path, ...]} for all files in folder (recursive)."""
    result = defaultdict(list)
    p = Path(folder)
    if not p.exists():
        print(f"WARNING: folder does not exist: {folder}")
        return result
    for f in p.rglob('*'):
        if f.is_file():
            result[normalize_name(f.name)].append(f)
    return result


def style_header_row(ws, row=1):
    fill = PatternFill("solid", fgColor="2F5496")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def write_df_to_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(title=sheet_name)
    if df.empty:
        ws.append(["No issues found."])
        return
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    style_header_row(ws)
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


def main():
    print("=" * 60)
    print("Aid4Mail Attachment Integrity Validator")
    print("=" * 60)

    # ── Load normalized Excel index ──────────────────────────────────────────
    print("\n[1/6] Loading normalized Excel index...")
    df_idx = pd.read_excel(ATTACHMENTS_XLSX, dtype=str)
    df_idx.columns = [c.strip() for c in df_idx.columns]
    df_idx = df_idx.fillna('')

    name_col = next(
        (c for c in df_idx.columns if 'attachment name' in c.lower() or c.lower() == 'attachment name (full filename)'),
        None
    )
    msgid_col = next((c for c in df_idx.columns if 'message-id' in c.lower()), None)
    uid_col   = next((c for c in df_idx.columns if 'uid' in c.lower()), None)
    count_col = next((c for c in df_idx.columns if 'count' in c.lower()), None)

    print(f"  Rows in index: {len(df_idx)}")
    print(f"  Name column:   {name_col}")
    print(f"  Message-ID:    {msgid_col}")

    if not name_col:
        raise ValueError("Cannot find attachment name column in Excel. Check column headers.")

    indexed_basenames = set()
    indexed_by_msgid = defaultdict(list)

    for _, row in df_idx.iterrows():
        raw_name = row[name_col].strip()
        if not raw_name:
            continue
        basename = normalize_name(Path(raw_name).name)
        indexed_basenames.add(basename)
        if msgid_col:
            indexed_by_msgid[row[msgid_col].strip()].append(raw_name)

    print(f"  Unique basenames indexed: {len(indexed_basenames)}")

    # ── Load JSON ─────────────────────────────────────────────────────────────
    print("\n[2/6] Loading JSON metadata...")
    emails = load_json(ATTACHMENTS_JSON)
    print(f"  Email records in JSON: {len(emails)}")

    json_attach_count = {}
    json_attach_files = defaultdict(list)

    for rec in emails:
        msgid = ''
        if isinstance(rec.get('Header'), dict):
            msgid = rec['Header'].get('Message-ID', '')
        elif 'Header.Message-ID' in rec:
            msgid = rec['Header.Message-ID']
        msgid = str(msgid).strip()

        count = rec.get('Attachment.Count') or rec.get('Attachment', {}).get('Count') if isinstance(rec, dict) else None
        try:
            json_attach_count[msgid] = int(count) if count not in (None, '', 'nan') else None
        except (ValueError, TypeError):
            json_attach_count[msgid] = None

        files = rec.get('Email.Attachments') or rec.get('Attachments') or []
        if isinstance(files, str):
            files = [files]
        if isinstance(files, list):
            json_attach_files[msgid].extend([Path(f).name for f in files if f])

    # ── Scan disk folders ─────────────────────────────────────────────────────
    print("\n[3/6] Scanning Attachment/ folder...")
    disk_attach = scan_folder(ATTACHMENT_DIR)
    print(f"  Files on disk (unique basenames): {len(disk_attach)}")

    print("\n[4/6] Scanning Embedded/ folder...")
    disk_embed = scan_folder(EMBEDDED_DIR)
    print(f"  Embedded files (unique basenames): {len(disk_embed)}")

    # ── Validation ────────────────────────────────────────────────────────────
    print("\n[5/6] Running validations...")

    missing = []
    for _, row in df_idx.iterrows():
        raw_name = row[name_col].strip()
        if not raw_name:
            continue
        basename = normalize_name(Path(raw_name).name)
        if basename not in disk_attach:
            r = {'Attachment Name': raw_name, 'Basename': Path(raw_name).name}
            if msgid_col:
                r['Message-ID'] = row[msgid_col]
            if uid_col:
                r['UID'] = row[uid_col]
            missing.append(r)
    df_missing = pd.DataFrame(missing)

    orphans = []
    for basename, paths in disk_attach.items():
        if basename not in indexed_basenames:
            for p in paths:
                orphans.append({'File': p.name, 'Full Path': str(p)})
    df_orphans = pd.DataFrame(orphans)

    embed_matches = []
    for basename in disk_embed:
        if basename in indexed_basenames:
            for p in disk_embed[basename]:
                embed_matches.append({
                    'Embedded File': p.name,
                    'Embedded Path': str(p),
                    'Also Indexed As Attachment': 'YES'
                })
    df_embed_matches = pd.DataFrame(embed_matches)

    count_mismatches = []
    if msgid_col:
        excel_counts = df_idx.groupby(msgid_col).size().to_dict()
        all_msgids = set(list(excel_counts.keys()) + list(json_attach_count.keys()))
        for mid in all_msgids:
            excel_n = excel_counts.get(mid, 0)
            json_n  = json_attach_count.get(mid)
            if json_n is not None and excel_n != json_n:
                count_mismatches.append({
                    'Message-ID': mid,
                    'Excel Row Count': excel_n,
                    'JSON Attachment.Count': json_n,
                    'Delta': excel_n - json_n
                })
    df_count = pd.DataFrame(count_mismatches)

    dups_disk = []
    for basename, paths in disk_attach.items():
        if len(paths) > 1:
            for p in paths:
                dups_disk.append({'Basename': Path(basename).name, 'Full Path': str(p)})
    df_dups_disk = pd.DataFrame(dups_disk)

    name_to_msgids = defaultdict(set)
    if msgid_col:
        for _, row in df_idx.iterrows():
            raw_name = row[name_col].strip()
            if raw_name:
                bn = normalize_name(Path(raw_name).name)
                name_to_msgids[bn].add(row[msgid_col].strip())
    collisions = []
    for bn, msgids in name_to_msgids.items():
        if len(msgids) > 1:
            collisions.append({
                'Basename': bn,
                'Referenced By # Message-IDs': len(msgids),
                'Message-IDs (sample)': ' | '.join(list(msgids)[:5])
            })
    df_collisions = pd.DataFrame(collisions)

    summary_data = {
        'Metric': [
            'Total indexed attachments (Excel rows)',
            'Unique basenames in index',
            'Files in Attachment/ folder',
            'Files in Embedded/ folder',
            'Email records in JSON',
            'Missing files (indexed but not on disk)',
            'Orphan files (on disk but not indexed)',
            'Embedded files matching attachment names',
            'Attachment count mismatches (Excel vs JSON)',
            'Duplicate basenames on disk',
            'Name collisions in index',
        ],
        'Count': [
            len(df_idx),
            len(indexed_basenames),
            sum(len(v) for v in disk_attach.values()),
            sum(len(v) for v in disk_embed.values()),
            len(emails),
            len(df_missing),
            len(df_orphans),
            len(df_embed_matches),
            len(df_count),
            len(df_dups_disk),
            len(df_collisions),
        ]
    }
    df_summary = pd.DataFrame(summary_data)

    df_hashes = None
    if DO_HASHES:
        print("\n  Computing SHA-256 hashes (this may take a while)...")
        hash_rows = []
        for basename, paths in disk_attach.items():
            for p in paths:
                try:
                    h = hash_file(p)
                except Exception as e:
                    h = f"ERROR: {e}"
                hash_rows.append({'Filename': p.name, 'Full Path': str(p), 'SHA-256': h})
        df_hashes = pd.DataFrame(hash_rows)

    # ── Write Excel report ────────────────────────────────────────────────────
    print("\n[6/6] Writing Excel report...")
    wb = Workbook()
    wb.remove(wb.active)

    write_df_to_sheet(wb, 'Summary', df_summary)
    write_df_to_sheet(wb, 'Missing_Files', df_missing)
    write_df_to_sheet(wb, 'Orphan_Files', df_orphans)
    write_df_to_sheet(wb, 'Count_Mismatches', df_count)
    write_df_to_sheet(wb, 'Duplicates_On_Disk', df_dups_disk)
    write_df_to_sheet(wb, 'Name_Collisions_In_Index', df_collisions)
    write_df_to_sheet(wb, 'Embedded_Matches_Attachments', df_embed_matches)
    if df_hashes is not None:
        write_df_to_sheet(wb, 'Hashes', df_hashes)

    ws_sum = wb['Summary']
    green = PatternFill("solid", fgColor="E2EFDA")
    red   = PatternFill("solid", fgColor="FFDEDE")
    for row in ws_sum.iter_rows(min_row=2):
        val = row[1].value
        if isinstance(val, int):
            fill = red if val > 0 and row[0].value not in (
                'Total indexed attachments (Excel rows)',
                'Unique basenames in index',
                'Files in Attachment/ folder',
                'Files in Embedded/ folder',
                'Email records in JSON',
            ) else green
            for cell in row:
                cell.fill = fill

    wb.save(OUTPUT_XLSX)
    print(f"\n✓ Report saved to: {OUTPUT_XLSX}")

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for _, row in df_summary.iterrows():
        print(f"  {row['Metric']:<50} {row['Count']}")
    print("=" * 60)


if __name__ == '__main__':
    main()
