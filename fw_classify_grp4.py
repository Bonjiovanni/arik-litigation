"""
fw_classify_grp4.py
--------------------
Classify_History sheet management for fw_classify.

Follows the same pattern as fw_walk_grp4.py.

Functions:
    ensure_classify_history_sheet  — creates/returns the Classify_History sheet
    log_classify_run               — appends one run summary row
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


CLASSIFY_HISTORY_SHEET = "Classify_History"

_COLUMNS = [
    ("RunID",          25),
    ("StartedAt",      22),
    ("CompletedAt",    22),
    ("RowsProcessed",  14),
    ("RowsUpdated",    12),
    ("RowsSkipped",    12),
    ("Errors",         10),
    ("Notes",          40),
]

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# FUNCTION 1: ensure_classify_history_sheet
# ---------------------------------------------------------------------------

def ensure_classify_history_sheet(wb: Workbook) -> Worksheet:
    """Create and style the Classify_History sheet if absent; return it.

    Idempotent — returns existing sheet unchanged if already present.
    """
    if CLASSIFY_HISTORY_SHEET in wb.sheetnames:
        return wb[CLASSIFY_HISTORY_SHEET]

    ws = wb.create_sheet(title=CLASSIFY_HISTORY_SHEET)

    for col_idx, (name, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    return ws


# ---------------------------------------------------------------------------
# FUNCTION 2: log_classify_run
# ---------------------------------------------------------------------------

def log_classify_run(wb: Workbook, run_id: str, started_at: str,
                     stats: dict, notes: str = "") -> int:
    """Append one classify run summary to Classify_History.

    stats keys (all default to 0 if missing):
        rows_processed, rows_updated, rows_skipped, errors

    Returns the row number written.
    """
    ws = ensure_classify_history_sheet(wb)

    # Next available row
    row_num = ws.max_row + 1
    if row_num == 2 and ws.cell(row=2, column=1).value is None:
        row_num = 2  # sheet only has header

    completed_at = datetime.now().isoformat()

    col_map = {name: idx + 1 for idx, (name, _) in enumerate(_COLUMNS)}

    ws.cell(row=row_num, column=col_map["RunID"],         value=run_id)
    ws.cell(row=row_num, column=col_map["StartedAt"],     value=started_at)
    ws.cell(row=row_num, column=col_map["CompletedAt"],   value=completed_at)
    ws.cell(row=row_num, column=col_map["RowsProcessed"], value=stats.get("rows_processed", 0))
    ws.cell(row=row_num, column=col_map["RowsUpdated"],   value=stats.get("rows_updated",   0))
    ws.cell(row=row_num, column=col_map["RowsSkipped"],   value=stats.get("rows_skipped",   0))
    ws.cell(row=row_num, column=col_map["Errors"],        value=stats.get("errors",         0))
    ws.cell(row=row_num, column=col_map["Notes"],         value=notes)

    return row_num
