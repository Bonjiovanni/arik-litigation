"""
fw_triage_grp1.py
-----------------
Config-driven scoring logic for fw_triage.

Manages two config sheets in the master workbook:
    Triage_Config — maps (SignalType, SignalValue) → ScorePoints
    Triage_Bands  — maps MinScore → (Band, NextStep)

Functions:
    ensure_triage_config_sheet  — creates/returns Triage_Config with defaults
    ensure_triage_bands_sheet   — creates/returns Triage_Bands with defaults
    load_triage_config          — loads config into nested dict
    load_triage_bands           — loads bands as sorted list (highest first)
    score_record                — computes 0-100 triage score from a record dict
    get_triage_band             — maps score → (band_name, next_step)
    get_reason_flagged          — semicolon-joined list of signals that scored > 0
    get_next_step               — looks up NextStep from bands; special-cases OCR

Python 3.11, Windows 10. Requires openpyxl.
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Shared header style
# ---------------------------------------------------------------------------

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# Triage_Config defaults
# (SignalType, SignalValue, ScorePoints, Notes)
# ---------------------------------------------------------------------------

TRIAGE_CONFIG_SHEET = "Triage_Config"

_DEFAULT_TRIAGE_CONFIG = [
    # --- DocType weights ---
    ("DocType", "LTC_Claim",            30, "Core LTC claim doc"),
    ("DocType", "Denial_Letter",         30, "Denial / adverse action"),
    ("DocType", "EOB",                   28, "Explanation of Benefits"),
    ("DocType", "Appeal_Letter",         25, "Claim appeal"),
    ("DocType", "Trust_Amendment",       25, "Amendment to trust"),
    ("DocType", "Power_of_Attorney",     25, "POA document"),
    ("DocType", "Wire_Transfer",         22, "Wire / ACH transfer record"),
    ("DocType", "Care_Assessment",       22, "Level-of-care assessment"),
    ("DocType", "Care_Plan",             22, "Care plan"),
    ("DocType", "Court_Filing",          22, "Court document / filing"),
    ("DocType", "Account_Statement",     20, "Bank / financial account statement"),
    ("DocType", "Bank_Statement",        70, "Bank statement"),
    ("DocType", "Financial_Statement",   70, "Financial statement (general)"),
    ("DocType", "Investment_Statement",  20, "Brokerage / retirement statement"),
    ("DocType", "Invoice",               65, "Invoice / bill"),
    ("DocType", "Provider_Invoice",      20, "Medical provider invoice"),
    ("DocType", "Tax_Document",          65, "Tax return / 1099 / W-2"),
    ("DocType", "Utility_Bill",          60, "Utility bill"),
    ("DocType", "Trust_Document",        20, "Trust agreement / deed"),
    ("DocType", "Legal_Correspondence",  18, "Legal letters"),
    ("DocType", "Medical_Record",        18, "Medical record"),
    ("DocType", "Insurance_Policy",      15, "Insurance policy document"),
    ("DocType", "Premium_Notice",        15, "Premium / payment notice"),
    ("DocType", "Doctor_Note",           16, "Physician note"),
    ("DocType", "Lab_Result",            14, "Lab / diagnostic result"),
    ("DocType", "Check_Copy",            15, "Check image / copy"),
    ("DocType", "Promissory_Note",       25, "Promissory note / loan doc"),
    ("DocType", "Will",                  20, "Last will and testament"),
    ("DocType", "Contract",              15, "General contract / agreement"),
    ("DocType", "Receipt",               12, "Receipt / proof of payment"),
    ("DocType", "Correspondence",        10, "General correspondence"),
    ("DocType", "Unknown",                0, "No signals matched"),
    # --- DocTypeConfidence weights ---
    ("DocTypeConfidence", "high",        15, "High confidence classification"),
    ("DocTypeConfidence", "medium",      10, "Medium confidence classification"),
    ("DocTypeConfidence", "low",          5, "Low confidence classification"),
    # --- General signal weights ---
    ("Signal", "MoneyDetected",          15, "Dollar amounts / currency found"),
    ("Signal", "DateDetected",           10, "Dates found in text"),
    ("Signal", "KeywordHits",            10, "Litigation keywords matched"),
    ("Signal", "NeedsOCR",               5,  "Scanned doc flagged for OCR"),
    ("Signal", "EntityHits",            20, "Known case entity detected in text"),
]


# ---------------------------------------------------------------------------
# Triage_Bands defaults
# (Band, MinScore, NextStep)
# Valid NextStep values include: "Priority manual review", "Manual review",
# "Batch review", "Archive", "OCR then review",
# "Review: no extractor"  — use for financial DocTypes with no extraction path
# ---------------------------------------------------------------------------

TRIAGE_BANDS_SHEET = "Triage_Bands"

_DEFAULT_TRIAGE_BANDS = [
    ("High",   60, "Priority manual review"),
    ("Medium", 35, "Manual review"),
    ("Low",    10, "Batch review"),
    ("Skip",    0, "Archive"),
]


# ---------------------------------------------------------------------------
# FUNCTION 1: ensure_triage_config_sheet
# ---------------------------------------------------------------------------

def ensure_triage_config_sheet(wb: Workbook) -> Worksheet:
    """Create and style Triage_Config sheet if absent; return it.

    Columns: SignalType | SignalValue | ScorePoints | Notes
    Seeds default rows on first creation. Idempotent.
    """
    if TRIAGE_CONFIG_SHEET in wb.sheetnames:
        return wb[TRIAGE_CONFIG_SHEET]

    ws = wb.create_sheet(title=TRIAGE_CONFIG_SHEET)

    headers    = ["SignalType", "SignalValue", "ScorePoints", "Notes"]
    col_widths = [18, 30, 12, 45]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    for row_idx, (sig_type, sig_val, points, notes) in enumerate(_DEFAULT_TRIAGE_CONFIG, start=2):
        ws.cell(row=row_idx, column=1, value=sig_type)
        ws.cell(row=row_idx, column=2, value=sig_val)
        ws.cell(row=row_idx, column=3, value=points)
        ws.cell(row=row_idx, column=4, value=notes)

    return ws


# ---------------------------------------------------------------------------
# FUNCTION 2: ensure_triage_bands_sheet
# ---------------------------------------------------------------------------

def ensure_triage_bands_sheet(wb: Workbook) -> Worksheet:
    """Create and style Triage_Bands sheet if absent; return it.

    Columns: Band | MinScore | NextStep
    Seeds default bands on first creation. Idempotent.
    """
    if TRIAGE_BANDS_SHEET in wb.sheetnames:
        return wb[TRIAGE_BANDS_SHEET]

    ws = wb.create_sheet(title=TRIAGE_BANDS_SHEET)

    headers    = ["Band", "MinScore", "NextStep"]
    col_widths = [12, 10, 35]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    for row_idx, (band, min_score, next_step) in enumerate(_DEFAULT_TRIAGE_BANDS, start=2):
        ws.cell(row=row_idx, column=1, value=band)
        ws.cell(row=row_idx, column=2, value=min_score)
        ws.cell(row=row_idx, column=3, value=next_step)

    return ws


# ---------------------------------------------------------------------------
# FUNCTION 3: load_triage_config
# ---------------------------------------------------------------------------

def load_triage_config(ws: Worksheet) -> dict:
    """Load Triage_Config into a nested dict.

    Returns:
        {
          "DocType":           {"LTC_Claim": 30, "Invoice": 20, ...},
          "DocTypeConfidence": {"high": 15, "medium": 10, "low": 5},
          "Signal":            {"MoneyDetected": 15, "DateDetected": 10, ...},
        }
    Unknown SignalTypes are also stored under their own key.
    """
    config: dict[str, dict[str, int]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        sig_type  = str(row[0]).strip() if row[0] is not None else None
        sig_val   = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        points    = row[2] if len(row) > 2 else 0

        if not sig_type or not sig_val:
            continue

        try:
            points = int(points) if points is not None else 0
        except (ValueError, TypeError):
            points = 0

        config.setdefault(sig_type, {})[sig_val] = points

    return config


# ---------------------------------------------------------------------------
# FUNCTION 4: load_triage_bands
# ---------------------------------------------------------------------------

def load_triage_bands(ws: Worksheet) -> list[tuple[str, int, str]]:
    """Load Triage_Bands as a list sorted by MinScore descending.

    Returns list of (band_name, min_score, next_step).
    Highest min_score first — iterate to find the first band a score qualifies for.
    """
    bands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        band      = str(row[0]).strip() if row[0] is not None else None
        min_score = row[1] if len(row) > 1 else 0
        next_step = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        if not band:
            continue
        try:
            min_score = int(min_score) if min_score is not None else 0
        except (ValueError, TypeError):
            min_score = 0

        bands.append((band, min_score, next_step))

    # Sort highest min_score first
    bands.sort(key=lambda x: x[1], reverse=True)
    return bands


# ---------------------------------------------------------------------------
# FUNCTION 5: score_record
# ---------------------------------------------------------------------------

def score_record(record: dict, config: dict) -> int:
    """Compute 0-100 triage score for a classified record.

    Uses the loaded config dict from load_triage_config().
    Additive: sum all matching signal points.

    record keys used:
        DocType, DocTypeConfidence, MoneyDetected, DateDetected,
        KeywordHits, NeedsOCR
    """
    total = 0

    doc_type_scores = config.get("DocType", {})
    conf_scores     = config.get("DocTypeConfidence", {})
    signal_scores   = config.get("Signal", {})

    # DocType score
    doc_type = (record.get("DocType") or "").strip()
    if doc_type and doc_type in doc_type_scores:
        total += doc_type_scores[doc_type]

    # DocTypeConfidence score
    confidence = (record.get("DocTypeConfidence") or "").strip().lower()
    if confidence and confidence in conf_scores:
        total += conf_scores[confidence]

    # Signal: MoneyDetected
    if (record.get("MoneyDetected") or "").strip().upper() == "Y":
        total += signal_scores.get("MoneyDetected", 0)

    # Signal: DateDetected
    if (record.get("DateDetected") or "").strip().upper() == "Y":
        total += signal_scores.get("DateDetected", 0)

    # Signal: KeywordHits (any non-empty value counts)
    if (record.get("KeywordHits") or "").strip():
        total += signal_scores.get("KeywordHits", 0)

    # Signal: NeedsOCR
    if (record.get("NeedsOCR") or "").strip().upper() == "Y":
        total += signal_scores.get("NeedsOCR", 0)

    if (record.get("EntityHits") or "").strip():
        total += signal_scores.get("EntityHits", 0)

    return max(0, min(100, total))


# ---------------------------------------------------------------------------
# FUNCTION 6: get_triage_band
# ---------------------------------------------------------------------------

def get_triage_band(score: int,
                    bands: list[tuple[str, int, str]]) -> tuple[str, str]:
    """Return (band_name, next_step) for a given score.

    Iterates bands (sorted highest-first) and returns the first where
    score >= min_score. Falls back to ("Skip", "Archive") if no match.
    """
    for band_name, min_score, next_step in bands:
        if score >= min_score:
            return (band_name, next_step)
    return ("Skip", "Archive")


# ---------------------------------------------------------------------------
# FUNCTION 7: get_reason_flagged
# ---------------------------------------------------------------------------

def get_reason_flagged(record: dict, config: dict) -> str:
    """Return semicolon-joined list of signals that contributed score points.

    e.g. "LTC_Claim;MoneyDetected;DateDetected"
    Returns "" if no signals scored.
    """
    reasons = []
    doc_type_scores = config.get("DocType", {})
    signal_scores   = config.get("Signal", {})

    doc_type = (record.get("DocType") or "").strip()
    if doc_type and doc_type_scores.get(doc_type, 0) > 0:
        reasons.append(doc_type)

    if (record.get("MoneyDetected") or "").strip().upper() == "Y" and signal_scores.get("MoneyDetected", 0) > 0:
        reasons.append("MoneyDetected")

    if (record.get("DateDetected") or "").strip().upper() == "Y" and signal_scores.get("DateDetected", 0) > 0:
        reasons.append("DateDetected")

    if (record.get("KeywordHits") or "").strip() and signal_scores.get("KeywordHits", 0) > 0:
        reasons.append("KeywordHits")

    if (record.get("NeedsOCR") or "").strip().upper() == "Y" and signal_scores.get("NeedsOCR", 0) > 0:
        reasons.append("NeedsOCR")

    return ";".join(reasons)


# ---------------------------------------------------------------------------
# FUNCTION 8: get_next_step
# ---------------------------------------------------------------------------

def get_next_step(band: str, record: dict,
                  bands: list[tuple[str, int, str]]) -> str:
    """Return NextStep string for a band + record.

    Special case: Medium band + NeedsOCR="Y" → "OCR then review".
    Otherwise looks up from bands list.
    """
    needs_ocr = (record.get("NeedsOCR") or "").strip().upper() == "Y"

    if band == "Medium" and needs_ocr:
        return "OCR then review"

    for band_name, _, next_step in bands:
        if band_name == band:
            return next_step

    return "Archive"
