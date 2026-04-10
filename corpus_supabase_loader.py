"""
corpus_supabase_loader.py — Loads email + attachment data into Supabase.

Replaces corpus_sqlite_loader.py. Reads from the same Excel sources,
maps columns to the Supabase schema (core columns + extended JSONB),
and inserts via the Supabase Python client (PostgREST).

Usage:
    python corpus_supabase_loader.py \
        --emails "path/to/V11.xlsx" \
        --sheet "Merge1" \
        --attachments "path/to/manifest.xlsx"

Rebuild behavior:
    - Truncates emails_master, attachments, extraction_log on every run
    - Entity tables are NEVER touched
    - _column_map rows are replaced
"""

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

# Reuse column mapping logic from the SQLite schema module
from corpus_sqlite_schema import (
    build_attachments_column_mapping,
    build_email_column_mapping,
    to_snake_case,
)

# ─── .env loading ───

ENV_PATH = r"C:\Users\arika\OneDrive\Litigation\Pipeline\.env"
load_dotenv(ENV_PATH)


def get_supabase_client():
    """Create and return a Supabase client using .env credentials."""
    from supabase import create_client

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SECRET_KEY")
    if not url or not key:
        raise RuntimeError(
            f"SUPABASE_URL and SUPABASE_SECRET_KEY must be set in {ENV_PATH}"
        )
    return create_client(url, key)


# ─── Core column definitions ───
# These match the Supabase schema (user-writable columns only).
# Auto-generated columns (created_at, fts_vector, attachment_id) are excluded.

CORE_EMAIL_COLUMNS = {
    "message_id",
    "from_addr",
    "from_name",
    "to_addr",
    "to_name",
    "cc",
    "bcc",
    "subject",
    "date_sent",
    "date_received",
    "labels",
    "body_clean",
    "body_text",
    "body_sender_text",
    "strip_method",
    "drafts2",
    "is_real_reply",
    "is_auto_reply",
    "body_snippet",
    "thread_id",
    "email_in_pdf",
    "attachments_count",
    "source",
    "sha256",
}

CORE_ATTACHMENT_COLUMNS = {
    "message_id",
    "source",
    "from_addr",
    "to_addr",
    "subject",
    "date_sent",
    "attachment_ordinal",
    "original_filename",
    "saved_filename",
    "file_type",
    "saved_full_path",
    "storage_folder",
    "size_bytes",
    "sha256",
    "extraction_status",
    "extracted_text",
    "extraction_method",
    "extraction_log",
}

# ─── Snake case → Supabase column name overrides ───
# Most snake_case names from the Excel mapping match the Supabase column names.
# These are the ones that DON'T — they need explicit renaming.

SNAKE_TO_SUPABASE = {
    "rfc_822_message_id": "message_id",
    "body_clean_1": "body_clean",
    "date_time_sent": "date_sent",
    "date_time_received": "date_received",
    "from_name": "from_name",
    "to_name": "to_name",
    "body_text": "body_text",
    "body_sender_text": "body_sender_text",
    "body_snippet": "body_snippet",
    "email_in_pdf": "email_in_pdf",
}

# Columns that should be SKIPPED (demoted to extended) because another
# column already maps to the same Supabase core column name.
# "Message ID" (snake: message_id) collides with "RFC 822 Message ID"
# (snake: rfc_822_message_id → message_id). The RFC 822 one is the real PK.
SKIP_TO_EXTENDED = {
    "message_id",  # from "Message ID" header — use rfc_822_message_id instead
}


def coerce_value(value, data_type: str):
    """Coerce a cell value to the declared type, preserving NULLs."""
    if value is None:
        return None
    if data_type == "INTEGER":
        if isinstance(value, (int, float)):
            return int(value)
        s = str(value).strip()
        if not s:
            return None
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None
    # TEXT — convert to string, but keep None as None
    if isinstance(value, str):
        return value if value.strip() else None
    return str(value)


def build_email_row(
    mapping: list[tuple[str, str, str]], values: list
) -> dict:
    """Build a Supabase row dict from an email mapping + raw values.

    Core columns become top-level keys. Everything else goes into
    the 'extended' JSONB column. None values are excluded from extended.
    """
    row = {}
    extended = {}

    for i, (original, snake, dtype) in enumerate(mapping):
        val = values[i] if i < len(values) else None
        val = coerce_value(val, dtype)

        # Map snake_case → Supabase column name
        supa_name = SNAKE_TO_SUPABASE.get(snake, snake)

        # Some snake names collide with a renamed column's target.
        # Force them to extended to avoid overwriting the real PK/core value.
        if snake in SKIP_TO_EXTENDED:
            if val is not None:
                extended[snake] = val
        elif supa_name in CORE_EMAIL_COLUMNS:
            row[supa_name] = val
        else:
            # Non-core → extended JSONB (skip nulls to keep blob clean)
            if val is not None:
                extended[snake] = val

    if extended:
        row["extended"] = extended

    return row


def build_attachment_row(
    mapping: list[tuple[str, str, str]], values: list
) -> dict:
    """Build a Supabase row dict from an attachment mapping + raw values.

    All attachment columns map directly (no extended JSONB needed).
    Adds extraction_status='pending' if not present.
    """
    row = {}

    for i, (original, snake, dtype) in enumerate(mapping):
        val = values[i] if i < len(values) else None
        val = coerce_value(val, dtype)

        # Map snake_case → Supabase column name
        supa_name = SNAKE_TO_SUPABASE.get(snake, snake)

        if supa_name in CORE_ATTACHMENT_COLUMNS:
            if val is not None:
                row[supa_name] = val

    # Default extraction_status
    if "extraction_status" not in row:
        row["extraction_status"] = "pending"

    return row


# ─── Excel reading (reused from sqlite loader) ───

def read_excel_headers(path: str, sheet_name: str | None = None) -> list[str]:
    """Read header row from an Excel file."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    return headers


def read_excel_data(path: str, sheet_name: str | None = None):
    """Yield data rows from an Excel file as lists of values."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(min_row=2)
    for row in rows:
        yield [cell.value for cell in row]
    wb.close()


# ─── Loading functions ───

BATCH_SIZE = 50  # Keep small to avoid Supabase free tier statement timeout


def load_emails(client, emails_path: str, sheet_name: str,
                email_mapping: list[tuple[str, str, str]]) -> int:
    """Load emails from Excel into Supabase emails_master. Returns row count."""
    batch = []
    count = 0

    for row_values in read_excel_data(emails_path, sheet_name):
        # Pad to match column count
        while len(row_values) < len(email_mapping):
            row_values.append(None)
        row_values = row_values[:len(email_mapping)]

        row = build_email_row(email_mapping, row_values)
        batch.append(row)
        count += 1

        if len(batch) >= BATCH_SIZE:
            client.table("emails_master").insert(batch).execute()
            print(f"    {count:,} emails inserted...", flush=True)
            batch = []

    if batch:
        client.table("emails_master").insert(batch).execute()

    return count


def load_attachments(client, attachments_path: str,
                     attach_mapping: list[tuple[str, str, str]],
                     valid_message_ids: set[str]) -> tuple[int, int]:
    """Load attachments from Excel into Supabase.

    Returns (total_loaded, orphan_count).
    Orphan attachments get extraction_status='pending_orphan'.
    """
    batch = []
    count = 0
    orphan_count = 0

    # Find the message_id column index
    manifest_snake = [snake for _, snake, _ in attach_mapping]
    msg_id_idx = manifest_snake.index("message_id")

    for row_values in read_excel_data(attachments_path):
        while len(row_values) < len(attach_mapping):
            row_values.append(None)
        row_values = row_values[:len(attach_mapping)]

        row = build_attachment_row(attach_mapping, row_values)

        # Check orphan status — PostgreSQL FK constraint rejects invalid
        # message_ids, so orphans get message_id=NULL + status flag.
        msg_id = row.get("message_id")
        if msg_id and msg_id not in valid_message_ids:
            row["extraction_status"] = "pending_orphan"
            row["message_id"] = None  # NULL FK is valid in PostgreSQL
            orphan_count += 1

        batch.append(row)
        count += 1

        if len(batch) >= BATCH_SIZE:
            client.table("attachments").insert(batch).execute()
            print(f"    {count:,} attachments inserted...", flush=True)
            batch = []

    if batch:
        client.table("attachments").insert(batch).execute()

    return count, orphan_count


def save_column_map(client, email_mapping, attach_mapping):
    """Write column mappings to _column_map table."""
    # Clear existing
    client.table("_column_map").delete().neq("domain", "__impossible__").execute()

    rows = []
    for original, snake, _ in email_mapping:
        supa_name = SNAKE_TO_SUPABASE.get(snake, snake)
        rows.append({
            "original_name": original,
            "mapped_name": supa_name,
            "domain": "emails_master",
        })
    for original, snake, _ in attach_mapping:
        supa_name = SNAKE_TO_SUPABASE.get(snake, snake)
        rows.append({
            "original_name": original,
            "mapped_name": supa_name,
            "domain": "attachments",
        })

    # Batch insert
    for i in range(0, len(rows), BATCH_SIZE):
        client.table("_column_map").insert(rows[i:i + BATCH_SIZE]).execute()


def truncate_email_domain(client):
    """Delete all rows from email domain tables (preserves entity tables)."""
    # Order matters for FK constraints
    client.table("extraction_log").delete().neq("log_id", -1).execute()
    client.table("attachments").delete().neq("attachment_id", -1).execute()
    client.table("emails_master").delete().neq("message_id", "__impossible__").execute()


def get_valid_message_ids(client) -> set[str]:
    """Fetch all message_ids from emails_master for orphan detection."""
    all_ids = set()
    offset = 0
    page_size = 1000

    while True:
        result = client.table("emails_master").select("message_id").range(
            offset, offset + page_size - 1
        ).execute()
        if not result.data:
            break
        for row in result.data:
            all_ids.add(row["message_id"])
        if len(result.data) < page_size:
            break
        offset += page_size

    return all_ids


def print_summary(client):
    """Print a summary of loaded data."""
    email_count = client.table("emails_master").select(
        "message_id", count="exact"
    ).limit(0).execute().count
    attach_count = client.table("attachments").select(
        "attachment_id", count="exact"
    ).limit(0).execute().count
    col_map_count = client.table("_column_map").select(
        "original_name", count="exact"
    ).limit(0).execute().count

    print("\n" + "=" * 60)
    print("LITIGATION CORPUS — SUPABASE LOAD SUMMARY")
    print("=" * 60)
    print(f"  emails_master:    {email_count:,} rows")
    print(f"  attachments:      {attach_count:,} rows")
    print(f"  _column_map:      {col_map_count:,} mappings")
    print(f"  FTS:              auto-populated (tsvector columns)")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(
        description="Load email + attachment data into Supabase"
    )
    parser.add_argument("--emails", required=True,
                        help="Path to emails Excel workbook (e.g. V11.xlsx)")
    parser.add_argument("--sheet", required=True,
                        help="Sheet name in emails workbook (e.g. Merge1)")
    parser.add_argument("--attachments", required=True,
                        help="Path to attachments manifest Excel workbook")
    args = parser.parse_args()

    # Validate inputs
    if not Path(args.emails).exists():
        print(f"ERROR: Emails file not found: {args.emails}")
        sys.exit(1)
    if not Path(args.attachments).exists():
        print(f"ERROR: Attachments file not found: {args.attachments}")
        sys.exit(1)

    print(f"Target: Supabase ({os.environ.get('SUPABASE_URL', 'NOT SET')})")
    print(f"Emails:   {args.emails} [sheet: {args.sheet}]")
    print(f"Attachments: {args.attachments}")
    print()

    # Step 1: Read headers and build column mappings
    print("Reading headers...")
    email_headers = read_excel_headers(args.emails, args.sheet)
    attach_headers = read_excel_headers(args.attachments)

    email_mapping = build_email_column_mapping(email_headers)
    attach_mapping = build_attachments_column_mapping(attach_headers)

    print(f"  Email columns: {len(email_mapping)} ({len(CORE_EMAIL_COLUMNS)} core + rest in extended JSONB)")
    print(f"  Attachment columns: {len(attach_mapping)}")

    # Step 2: Connect and truncate
    print("\nConnecting to Supabase...")
    client = get_supabase_client()

    print("Truncating email domain tables...")
    truncate_email_domain(client)

    # Step 3: Load emails
    print("\nLoading emails...")
    email_count = load_emails(client, args.emails, args.sheet, email_mapping)
    print(f"  Loaded {email_count:,} emails")

    # Step 4: Load attachments
    print("Loading attachments...")
    valid_ids = get_valid_message_ids(client)
    attach_count, orphan_count = load_attachments(
        client, args.attachments, attach_mapping, valid_ids
    )
    print(f"  Loaded {attach_count:,} attachments")
    if orphan_count:
        print(f"  WARNING: {orphan_count:,} orphan attachments (message_id not in emails_master)")

    # Step 5: Save column mappings
    print("Saving column mappings...")
    save_column_map(client, email_mapping, attach_mapping)

    # Step 6: Summary
    print_summary(client)
    print("\nFTS indexes are auto-maintained by Supabase (tsvector generated columns).")
    print("No manual FTS population needed.")


if __name__ == "__main__":
    main()
