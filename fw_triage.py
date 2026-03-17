"""
fw_triage.py
------------
Config-driven triage scoring for the filewalker evidence pipeline.

Reads Master_File_Inventory rows with ProcessingStatus = "classified",
scores each file using configurable weights, and writes back:
    TriageScore, TriageBand, ReasonFlagged, NextStep,
    ProcessingStatus → "triaged"

Also manages:
    Triage_Config   — scoring weights (SignalType × SignalValue → points)
    Triage_Bands    — band thresholds and NextStep routing
    Triage_History  — one-row-per-run audit log

Merged from: fw_triage_grp1 through fw_triage_grp4.

Writes to: C:\\Users\\arika\\OneDrive\\Litigation\\filewalker_master.xlsx
Python 3.11, Windows 10. Requires openpyxl.
"""

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from fw_walk_grp2 import _COL_INDEX


WORKBOOK_PATH = r"C:\Users\arika\OneDrive\Litigation\filewalker_master.xlsx"

# Shared header style
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ===========================================================================
# grp1 — Triage_Config + Triage_Bands sheets + scoring logic
# ===========================================================================

TRIAGE_CONFIG_SHEET = "Triage_Config"
TRIAGE_BANDS_SHEET  = "Triage_Bands"

# (SignalType, SignalValue, ScorePoints, Notes)
_DEFAULT_TRIAGE_CONFIG = [
    # --- DocType weights ---
    ("DocType", "LTC_Claim",            30, "Core LTC claim doc"),
    ("DocType", "Denial_Letter",        30, "Denial / adverse action"),
    ("DocType", "EOB",                  28, "Explanation of Benefits"),
    ("DocType", "Appeal_Letter",        25, "Claim appeal"),
    ("DocType", "Trust_Amendment",      25, "Amendment to trust"),
    ("DocType", "Power_of_Attorney",    25, "POA document"),
    ("DocType", "Wire_Transfer",        22, "Wire / ACH transfer record"),
    ("DocType", "Care_Assessment",      22, "Level-of-care assessment"),
    ("DocType", "Care_Plan",            22, "Care plan"),
    ("DocType", "Court_Filing",         22, "Court document / filing"),
    ("DocType", "Account_Statement",    20, "Bank / financial account statement"),
    ("DocType", "Bank_Statement",       70, "Bank statement"),
    ("DocType", "Financial_Statement",  70, "Financial statement (general)"),
    ("DocType", "Investment_Statement", 20, "Brokerage / retirement statement"),
    ("DocType", "Invoice",              65, "Invoice / bill"),
    ("DocType", "Provider_Invoice",     20, "Medical provider invoice"),
    ("DocType", "Tax_Document",         65, "Tax return / 1099 / W-2"),
    ("DocType", "Utility_Bill",         60, "Utility bill"),
    ("DocType", "Trust_Document",       20, "Trust agreement / deed"),
    ("DocType", "Legal_Correspondence", 18, "Legal letters"),
    ("DocType", "Medical_Record",       18, "Medical record"),
    ("DocType", "Insurance_Policy",     15, "Insurance policy document"),
    ("DocType", "Premium_Notice",       15, "Premium / payment notice"),
    ("DocType", "Doctor_Note",          16, "Physician note"),
    ("DocType", "Lab_Result",           14, "Lab / diagnostic result"),
    ("DocType", "Check_Copy",           15, "Check image / copy"),
    ("DocType", "Promissory_Note",      25, "Promissory note / loan doc"),
    ("DocType", "Will",                 20, "Last will and testament"),
    ("DocType", "Contract",             15, "General contract / agreement"),
    ("DocType", "Receipt",              12, "Receipt / proof of payment"),
    ("DocType", "Correspondence",       10, "General correspondence"),
    ("DocType", "Unknown",               0, "No signals matched"),
    # --- DocTypeConfidence weights ---
    ("DocTypeConfidence", "high",       15, "High confidence classification"),
    ("DocTypeConfidence", "medium",     10, "Medium confidence classification"),
    ("DocTypeConfidence", "low",         5, "Low confidence classification"),
    # --- General signal weights ---
    ("Signal", "MoneyDetected",         15, "Dollar amounts / currency found"),
    ("Signal", "DateDetected",          10, "Dates found in text"),
    ("Signal", "KeywordHits",           10, "Litigation keywords matched"),
    ("Signal", "NeedsOCR",              5,  "Scanned doc flagged for OCR"),
    ("Signal", "EntityHits",           20, "Known case entity detected in text"),
]

# (Band, MinScore, NextStep)
# Valid NextStep values: "Priority manual review", "Manual review", "Batch review",
# "Archive", "OCR then review", "Review: no extractor" (financial DocTypes with no extraction path)
_DEFAULT_TRIAGE_BANDS = [
    ("High",   60, "Priority manual review"),
    ("Medium", 35, "Manual review"),
    ("Low",    10, "Batch review"),
    ("Skip",    0, "Archive"),
]


def ensure_triage_config_sheet(wb: Workbook) -> Worksheet:
    """Create and style Triage_Config sheet if absent; return it. Idempotent.

    Columns: SignalType | SignalValue | ScorePoints | Notes
    Seeds default scoring weights on first creation.
    """
    if TRIAGE_CONFIG_SHEET in wb.sheetnames:
        return wb[TRIAGE_CONFIG_SHEET]

    ws = wb.create_sheet(title=TRIAGE_CONFIG_SHEET)

    headers    = ["SignalType", "SignalValue", "ScorePoints", "Notes"]
    col_widths = [18, 30, 12, 45]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    for row_idx, (sig_type, sig_val, points, notes) in enumerate(_DEFAULT_TRIAGE_CONFIG, start=2):
        ws.cell(row=row_idx, column=1, value=sig_type)
        ws.cell(row=row_idx, column=2, value=sig_val)
        ws.cell(row=row_idx, column=3, value=points)
        ws.cell(row=row_idx, column=4, value=notes)

    return ws


def ensure_triage_bands_sheet(wb: Workbook) -> Worksheet:
    """Create and style Triage_Bands sheet if absent; return it. Idempotent.

    Columns: Band | MinScore | NextStep
    Seeds default band thresholds on first creation.
    """
    if TRIAGE_BANDS_SHEET in wb.sheetnames:
        return wb[TRIAGE_BANDS_SHEET]

    ws = wb.create_sheet(title=TRIAGE_BANDS_SHEET)

    headers    = ["Band", "MinScore", "NextStep"]
    col_widths = [12, 10, 35]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    for row_idx, (band, min_score, next_step) in enumerate(_DEFAULT_TRIAGE_BANDS, start=2):
        ws.cell(row=row_idx, column=1, value=band)
        ws.cell(row=row_idx, column=2, value=min_score)
        ws.cell(row=row_idx, column=3, value=next_step)

    return ws


def load_triage_config(ws: Worksheet) -> dict:
    """Load Triage_Config into a nested dict.

    Returns:
        {
          "DocType":           {"LTC_Claim": 30, "Invoice": 20, ...},
          "DocTypeConfidence": {"high": 15, "medium": 10, "low": 5},
          "Signal":            {"MoneyDetected": 15, "DateDetected": 10, ...},
        }
    """
    config: dict[str, dict[str, int]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        sig_type = str(row[0]).strip()
        sig_val  = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        points   = row[2] if len(row) > 2 else 0

        if not sig_type or not sig_val:
            continue
        try:
            points = int(points) if points is not None else 0
        except (ValueError, TypeError):
            points = 0

        config.setdefault(sig_type, {})[sig_val] = points

    return config


def load_triage_bands(ws: Worksheet) -> list[tuple[str, int, str]]:
    """Load Triage_Bands as a list sorted by MinScore descending (highest first).

    Returns list of (band_name, min_score, next_step).
    """
    bands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        band      = str(row[0]).strip()
        min_score = row[1] if len(row) > 1 else 0
        next_step = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        if not band:
            continue
        try:
            min_score = int(min_score) if min_score is not None else 0
        except (ValueError, TypeError):
            min_score = 0

        bands.append((band, min_score, next_step))

    bands.sort(key=lambda x: x[1], reverse=True)
    return bands


def score_record(record: dict, config: dict) -> int:
    """Compute 0–100 triage score for a classified record.

    Additive: sums all matching signal points from the loaded config.
    Clamped to [0, 100].

    record keys used: DocType, DocTypeConfidence, MoneyDetected,
                      DateDetected, KeywordHits, NeedsOCR
    """
    total = 0

    doc_type_scores = config.get("DocType", {})
    conf_scores     = config.get("DocTypeConfidence", {})
    signal_scores   = config.get("Signal", {})

    doc_type = (record.get("DocType") or "").strip()
    if doc_type and doc_type in doc_type_scores:
        total += doc_type_scores[doc_type]

    confidence = (record.get("DocTypeConfidence") or "").strip().lower()
    if confidence and confidence in conf_scores:
        total += conf_scores[confidence]

    if (record.get("MoneyDetected") or "").strip().upper() == "Y":
        total += signal_scores.get("MoneyDetected", 0)

    if (record.get("DateDetected") or "").strip().upper() == "Y":
        total += signal_scores.get("DateDetected", 0)

    if (record.get("KeywordHits") or "").strip():
        total += signal_scores.get("KeywordHits", 0)

    if (record.get("NeedsOCR") or "").strip().upper() == "Y":
        total += signal_scores.get("NeedsOCR", 0)

    if (record.get("EntityHits") or "").strip():
        total += signal_scores.get("EntityHits", 0)

    return max(0, min(100, total))


def get_triage_band(score: int,
                    bands: list[tuple[str, int, str]]) -> tuple[str, str]:
    """Return (band_name, next_step) for a given score.

    Iterates bands sorted highest-first; returns first where score >= min_score.
    Falls back to ("Skip", "Archive") if bands list is empty.
    """
    for band_name, min_score, next_step in bands:
        if score >= min_score:
            return (band_name, next_step)
    return ("Skip", "Archive")


def get_reason_flagged(record: dict, config: dict) -> str:
    """Return semicolon-joined list of signals that scored > 0.

    e.g. "LTC_Claim;MoneyDetected;DateDetected"
    Returns "" if no signals contributed.
    """
    reasons = []
    doc_type_scores = config.get("DocType", {})
    signal_scores   = config.get("Signal", {})

    doc_type = (record.get("DocType") or "").strip()
    if doc_type and doc_type_scores.get(doc_type, 0) > 0:
        reasons.append(doc_type)

    if (record.get("MoneyDetected") or "").strip().upper() == "Y" \
            and signal_scores.get("MoneyDetected", 0) > 0:
        reasons.append("MoneyDetected")

    if (record.get("DateDetected") or "").strip().upper() == "Y" \
            and signal_scores.get("DateDetected", 0) > 0:
        reasons.append("DateDetected")

    if (record.get("KeywordHits") or "").strip() \
            and signal_scores.get("KeywordHits", 0) > 0:
        reasons.append("KeywordHits")

    if (record.get("NeedsOCR") or "").strip().upper() == "Y" \
            and signal_scores.get("NeedsOCR", 0) > 0:
        reasons.append("NeedsOCR")

    return ";".join(reasons)


def get_next_step(band: str, record: dict,
                  bands: list[tuple[str, int, str]]) -> str:
    """Return NextStep string for a band + record.

    Special case: Medium + NeedsOCR="Y" → "OCR then review".
    Otherwise looks up NextStep from the bands list.
    """
    if band == "Medium" and (record.get("NeedsOCR") or "").strip().upper() == "Y":
        return "OCR then review"

    for band_name, _, next_step in bands:
        if band_name == band:
            return next_step

    return "Archive"


# ===========================================================================
# grp2 — Master_File_Inventory Excel I/O
# ===========================================================================

_PROTECTED_COLS = {"ManualReviewStatus", "KeepForCase", "PossibleExhibit"}

_TRIAGE_FIELDS = [
    "FileID", "FilePath", "FileFamily",
    "DocType", "DocSubtype", "DocTypeConfidence",
    "MoneyDetected", "DateDetected", "KeywordHits",
    "LikelyTextBearing", "NeedsOCR",
    "ProcessingStatus",
]


def get_triageable_rows(ws) -> list[tuple[int, dict]]:
    """Return (row_num, record_dict) for rows where ProcessingStatus = 'classified'.

    Stops at first empty FileID (end of data).
    """
    result = []
    for row_num in range(2, ws.max_row + 1):
        file_id = ws.cell(row=row_num, column=_COL_INDEX["FileID"]).value
        if file_id is None or str(file_id).strip() == "":
            break

        status = ws.cell(row=row_num, column=_COL_INDEX["ProcessingStatus"]).value
        if str(status or "").strip() != "classified":
            continue

        record = {field: ws.cell(row=row_num, column=_COL_INDEX[field]).value
                  for field in _TRIAGE_FIELDS}
        result.append((row_num, record))

    return result


def write_triage_results(ws, row_num: int, results: dict) -> None:
    """Write triage result columns to a specific row.

    Writes any key in results that exists in _COL_INDEX (unless protected).
    """
    for key, value in results.items():
        if key in _PROTECTED_COLS:
            continue
        if key not in _COL_INDEX:
            continue
        ws.cell(row=row_num, column=_COL_INDEX[key], value=value)


def mark_row_triaged(ws, row_num: int) -> None:
    """Set ProcessingStatus = 'triaged' for the given row."""
    ws.cell(row=row_num, column=_COL_INDEX["ProcessingStatus"], value="triaged")


# ===========================================================================
# grp3 — Triage_History sheet
# ===========================================================================

TRIAGE_HISTORY_SHEET = "Triage_History"

_TH_COLUMNS = [
    ("RunID",          25),
    ("StartedAt",      22),
    ("CompletedAt",    22),
    ("RowsProcessed",  14),
    ("RowsTriaged",    12),
    ("RowsSkipped",    12),
    ("Errors",         10),
    ("Notes",          40),
]

_TH_COL_MAP = {name: idx + 1 for idx, (name, _) in enumerate(_TH_COLUMNS)}


def ensure_triage_history_sheet(wb: Workbook) -> Worksheet:
    """Create and style Triage_History sheet if absent; return it. Idempotent."""
    if TRIAGE_HISTORY_SHEET in wb.sheetnames:
        return wb[TRIAGE_HISTORY_SHEET]

    ws = wb.create_sheet(title=TRIAGE_HISTORY_SHEET)

    for col_idx, (name, width) in enumerate(_TH_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    return ws


def log_triage_run(wb: Workbook, run_id: str, started_at: str,
                   stats: dict, notes: str = "") -> int:
    """Append one triage-run summary row to Triage_History.

    stats keys (default 0): rows_processed, rows_triaged, rows_skipped, errors.
    Returns the row number written.
    """
    ws = ensure_triage_history_sheet(wb)

    row_num = ws.max_row + 1
    if row_num == 2 and ws.cell(row=2, column=1).value is None:
        row_num = 2

    ws.cell(row=row_num, column=_TH_COL_MAP["RunID"],         value=run_id)
    ws.cell(row=row_num, column=_TH_COL_MAP["StartedAt"],     value=started_at)
    ws.cell(row=row_num, column=_TH_COL_MAP["CompletedAt"],   value=datetime.now().isoformat())
    ws.cell(row=row_num, column=_TH_COL_MAP["RowsProcessed"], value=stats.get("rows_processed", 0))
    ws.cell(row=row_num, column=_TH_COL_MAP["RowsTriaged"],   value=stats.get("rows_triaged",   0))
    ws.cell(row=row_num, column=_TH_COL_MAP["RowsSkipped"],   value=stats.get("rows_skipped",   0))
    ws.cell(row=row_num, column=_TH_COL_MAP["Errors"],        value=stats.get("errors",         0))
    ws.cell(row=row_num, column=_TH_COL_MAP["Notes"],         value=notes)

    return row_num


# ===========================================================================
# grp4 — main() orchestration
# ===========================================================================

def main() -> None:
    """Run a complete fw_triage scoring pass.

    Steps:
      1.  Open or create the master workbook.
      2.  Ensure Triage_Config, Triage_Bands, Triage_History sheets exist.
      3.  Load triage config and band thresholds.
      4.  Generate run ID: TRIAGE_YYYYMMDD_HHMMSS.
      5.  Get triageable rows from Master_File_Inventory.
      6.  For each row: score → band → reason → next_step → write → mark triaged.
      7.  Log Triage_History row.
      8.  Save workbook.
      9.  Print final summary.
    """
    try:
        # ------------------------------------------------------------------
        # 1. Open or create workbook
        # ------------------------------------------------------------------
        wb_path = Path(WORKBOOK_PATH)
        if wb_path.exists():
            print(f"[fw_triage] Loading existing workbook: {wb_path}")
            wb = load_workbook(wb_path)
        else:
            print(f"[fw_triage] Creating new workbook: {wb_path}")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ------------------------------------------------------------------
        # 2. Ensure config and history sheets
        # ------------------------------------------------------------------
        ensure_triage_config_sheet(wb)
        ensure_triage_bands_sheet(wb)
        ensure_triage_history_sheet(wb)

        # ------------------------------------------------------------------
        # 3. Load config
        # ------------------------------------------------------------------
        config = load_triage_config(wb[TRIAGE_CONFIG_SHEET])
        bands  = load_triage_bands(wb[TRIAGE_BANDS_SHEET])
        print(f"[fw_triage] Loaded {sum(len(v) for v in config.values())} config entries, "
              f"{len(bands)} band(s).")

        # ------------------------------------------------------------------
        # 4. Generate run ID
        # ------------------------------------------------------------------
        run_id: str = datetime.now().strftime("TRIAGE_%Y%m%d_%H%M%S")
        print(f"[fw_triage] Run ID: {run_id}")

        # ------------------------------------------------------------------
        # 5. Get triageable rows
        # ------------------------------------------------------------------
        if "Master_File_Inventory" not in wb.sheetnames:
            print("[fw_triage] ERROR: Master_File_Inventory not found. "
                  "Run fw_walk + fw_classify first.")
            sys.exit(1)

        ws_inv = wb["Master_File_Inventory"]
        rows   = get_triageable_rows(ws_inv)
        print(f"[fw_triage] Rows to triage: {len(rows):,}")

        if not rows:
            print("[fw_triage] Nothing to triage.")
            sys.exit(0)

        # ------------------------------------------------------------------
        # 6. Score and write
        # ------------------------------------------------------------------
        started_at_str: str = datetime.now().isoformat()
        rows_processed = 0
        rows_triaged   = 0
        rows_skipped   = 0
        errors         = 0

        for row_num, record in rows:
            try:
                score             = score_record(record, config)
                band, _           = get_triage_band(score, bands)
                reason            = get_reason_flagged(record, config)
                step              = get_next_step(band, record, bands)

                write_triage_results(ws_inv, row_num, {
                    "TriageScore":   score,
                    "TriageBand":    band,
                    "ReasonFlagged": reason,
                    "NextStep":      step,
                })
                mark_row_triaged(ws_inv, row_num)
                rows_triaged += 1

            except Exception as exc:
                print(f"[fw_triage] ERROR row {row_num} ({record.get('FilePath', '')}): {exc}")
                errors += 1
            finally:
                rows_processed += 1

            if rows_processed % 100 == 0:
                print(f"[fw_triage]   {rows_processed:,}/{len(rows):,} processed...")

        # ------------------------------------------------------------------
        # 7. Log Triage_History
        # ------------------------------------------------------------------
        log_triage_run(
            wb=wb,
            run_id=run_id,
            started_at=started_at_str,
            stats={
                "rows_processed": rows_processed,
                "rows_triaged":   rows_triaged,
                "rows_skipped":   rows_skipped,
                "errors":         errors,
            },
        )

        # ------------------------------------------------------------------
        # 8. Save
        # ------------------------------------------------------------------
        wb.save(WORKBOOK_PATH)

        # ------------------------------------------------------------------
        # 9. Summary
        # ------------------------------------------------------------------
        print(f"\n[fw_triage] Run complete: {run_id}")
        print(f"  Rows processed: {rows_processed:,}")
        print(f"  Triaged:        {rows_triaged:,}")
        print(f"  Skipped:        {rows_skipped:,}")
        print(f"  Errors:         {errors:,}")
        print(f"Saved: {WORKBOOK_PATH}")

    except KeyboardInterrupt:
        print("\n[fw_triage] Interrupted.")
        sys.exit(0)


if __name__ == "__main__":
    main()
