"""
tests/test_fw_triage_grp3.py
-----------------------------
Unit tests for fw_triage_grp3.py:
    ensure_triage_history_sheet, log_triage_run
"""

import pytest
import openpyxl

import fw_triage_grp3 as grp3


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fresh_wb():
    wb = openpyxl.Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    return wb


_STATS = {
    "rows_processed": 20,
    "rows_triaged":   18,
    "rows_skipped":    1,
    "errors":          1,
}


# ---------------------------------------------------------------------------
# TestEnsureTriageHistorySheet
# ---------------------------------------------------------------------------

class TestEnsureTriageHistorySheet:

    def test_creates_sheet_when_absent(self):
        wb = _fresh_wb()
        grp3.ensure_triage_history_sheet(wb)
        assert grp3.TRIAGE_HISTORY_SHEET in wb.sheetnames

    def test_returns_worksheet(self):
        wb = _fresh_wb()
        ws = grp3.ensure_triage_history_sheet(wb)
        assert ws.title == grp3.TRIAGE_HISTORY_SHEET

    def test_idempotent(self):
        wb = _fresh_wb()
        ws1 = grp3.ensure_triage_history_sheet(wb)
        ws1.cell(row=2, column=1, value="MARKER")
        ws2 = grp3.ensure_triage_history_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "MARKER"

    def test_run_id_header(self):
        wb = _fresh_wb()
        ws = grp3.ensure_triage_history_sheet(wb)
        assert ws.cell(row=1, column=1).value == "RunID"

    def test_all_headers_present(self):
        wb = _fresh_wb()
        ws = grp3.ensure_triage_history_sheet(wb)
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(grp3._COLUMNS) + 1)]
        expected = [name for name, _ in grp3._COLUMNS]
        assert headers == expected

    def test_freeze_panes(self):
        wb = _fresh_wb()
        ws = grp3.ensure_triage_history_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_column_count(self):
        wb = _fresh_wb()
        ws = grp3.ensure_triage_history_sheet(wb)
        non_null = [ws.cell(row=1, column=c).value
                    for c in range(1, 20) if ws.cell(row=1, column=c).value]
        assert len(non_null) == len(grp3._COLUMNS)


# ---------------------------------------------------------------------------
# TestLogTriageRun
# ---------------------------------------------------------------------------

class TestLogTriageRun:

    def test_writes_run_id(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "TRIAGE_20260315_120000", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        assert ws.cell(row=2, column=grp3._COL_MAP["RunID"]).value == "TRIAGE_20260315_120000"

    def test_writes_started_at(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        assert ws.cell(row=2, column=grp3._COL_MAP["StartedAt"]).value == "2026-03-15T12:00:00"

    def test_writes_rows_processed(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        assert ws.cell(row=2, column=grp3._COL_MAP["RowsProcessed"]).value == 20

    def test_writes_rows_triaged(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        assert ws.cell(row=2, column=grp3._COL_MAP["RowsTriaged"]).value == 18

    def test_writes_errors(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        assert ws.cell(row=2, column=grp3._COL_MAP["Errors"]).value == 1

    def test_returns_row_number(self):
        wb = _fresh_wb()
        row = grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        assert row == 2

    def test_second_run_appends(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        row = grp3.log_triage_run(wb, "RUN2", "2026-03-15T13:00:00", _STATS)
        assert row == 3
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        assert ws.cell(row=3, column=grp3._COL_MAP["RunID"]).value == "RUN2"

    def test_notes_written(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS, notes="test run")
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        assert ws.cell(row=2, column=grp3._COL_MAP["Notes"]).value == "test run"

    def test_stats_default_to_zero(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", {})
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        assert ws.cell(row=2, column=grp3._COL_MAP["RowsProcessed"]).value == 0
        assert ws.cell(row=2, column=grp3._COL_MAP["RowsTriaged"]).value == 0
        assert ws.cell(row=2, column=grp3._COL_MAP["Errors"]).value == 0

    def test_completed_at_populated(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp3.TRIAGE_HISTORY_SHEET]
        completed = ws.cell(row=2, column=grp3._COL_MAP["CompletedAt"]).value
        assert completed is not None and len(completed) > 0

    def test_creates_sheet_if_missing(self):
        wb = _fresh_wb()
        grp3.log_triage_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        assert grp3.TRIAGE_HISTORY_SHEET in wb.sheetnames
