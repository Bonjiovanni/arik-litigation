"""
Tests for email_pipeline/read_xlsx.py
"""

import json
import sys
import subprocess
import pytest
from pathlib import Path
import openpyxl
import read_xlsx as rx


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx(tmp_path, sheets: dict, filename="test.xlsx") -> Path:
    """Create an xlsx file from {sheet_name: [[row], [row], ...]}"""
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / filename
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# TestReadXlsx
# ---------------------------------------------------------------------------

class TestReadXlsx:

    def test_valid_file_with_sheet_name_returns_correct_data(self, tmp_path):
        path = _make_xlsx(tmp_path, {"Sheet1": [["Name", "Age"], ["Alice", 30], ["Bob", 25]]})
        result = rx.read_xlsx(str(path), "Sheet1")
        assert "Sheet1" in result
        assert len(result["Sheet1"]) == 2
        assert result["Sheet1"][0]["Name"] == "Alice"
        assert result["Sheet1"][1]["Name"] == "Bob"

    def test_valid_file_no_sheet_name_returns_all_sheets(self, tmp_path):
        path = _make_xlsx(tmp_path, {
            "Sheet1": [["A"], ["1"]],
            "Sheet2": [["B"], ["2"]],
        })
        result = rx.read_xlsx(str(path))
        assert "Sheet1" in result
        assert "Sheet2" in result

    def test_missing_sheet_returns_error_string_not_exception(self, tmp_path):
        path = _make_xlsx(tmp_path, {"Sheet1": [["A"], ["1"]]})
        result = rx.read_xlsx(str(path), "NonExistent")
        assert "NonExistent" in result
        assert "ERROR" in result["NonExistent"]

    def test_blank_rows_are_skipped(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name"])
        ws.append(["Alice"])
        ws.append([None])   # blank row
        ws.append(["Bob"])
        path = tmp_path / "blanks.xlsx"
        wb.save(str(path))
        result = rx.read_xlsx(str(path), "Sheet1")
        assert len(result["Sheet1"]) == 2  # Alice and Bob; blank skipped

    def test_nonexistent_file_exits_with_code_1(self, tmp_path):
        script = str(Path(__file__).parent.parent / "email_pipeline" / "read_xlsx.py")
        fake_path = str(tmp_path / "nonexistent.xlsx")
        proc = subprocess.run(
            [sys.executable, script, fake_path],
            capture_output=True, text=True
        )
        assert proc.returncode == 1
        output = json.loads(proc.stdout)
        assert "error" in output

    def test_empty_sheet_returns_empty_list(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        path = tmp_path / "empty.xlsx"
        wb.save(str(path))
        result = rx.read_xlsx(str(path), "Empty")
        assert result["Empty"] == []

    def test_none_values_become_empty_string(self, tmp_path):
        path = _make_xlsx(tmp_path, {"Sheet1": [["Name", "Note"], ["Alice", None]]})
        result = rx.read_xlsx(str(path), "Sheet1")
        assert result["Sheet1"][0]["Note"] == ""
