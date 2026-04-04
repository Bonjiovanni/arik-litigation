"""
tests/test_fw_triage_grp4.py
-----------------------------
Namespace and smoke tests for fw_triage_grp4.py (the main() orchestrator).
"""

import pytest
import openpyxl

import fw_triage_grp4 as grp4


# ---------------------------------------------------------------------------
# TestNamespace
# ---------------------------------------------------------------------------

class TestNamespace:

    def test_module_importable(self):
        assert grp4 is not None

    def test_main_callable(self):
        assert callable(grp4.main)

    def test_workbook_path_constant(self):
        assert hasattr(grp4, "WORKBOOK_PATH")
        assert isinstance(grp4.WORKBOOK_PATH, str)
        assert grp4.WORKBOOK_PATH.endswith(".xlsx")

    def test_grp1_accessible(self):
        assert grp4.grp1 is not None
        assert callable(grp4.grp1.score_record)

    def test_grp2_accessible(self):
        assert grp4.grp2 is not None
        assert callable(grp4.grp2.get_triageable_rows)

    def test_grp3_accessible(self):
        assert grp4.grp3 is not None
        assert callable(grp4.grp3.log_triage_run)


# ---------------------------------------------------------------------------
# TestMainBranches
# ---------------------------------------------------------------------------

class TestMainBranches:

    def test_exits_when_no_mfi_sheet(self, tmp_path, monkeypatch):
        """main() should sys.exit(1) when Master_File_Inventory is missing."""
        wb_path = tmp_path / "test_master.xlsx"
        wb = openpyxl.Workbook()
        # Keep default Sheet so openpyxl can save — just no MFI sheet
        wb.save(str(wb_path))

        monkeypatch.setattr(grp4, "WORKBOOK_PATH", str(wb_path))

        with pytest.raises(SystemExit) as exc_info:
            grp4.main()
        assert exc_info.value.code == 1

    def test_exits_when_no_triageable_rows(self, tmp_path, monkeypatch):
        """main() should sys.exit(0) when MFI has no classified rows."""
        from fw_walk_grp2 import _COL_INDEX

        wb_path = tmp_path / "test_master.xlsx"
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        ws = wb.create_sheet("Master_File_Inventory")
        for name, col_idx in _COL_INDEX.items():
            ws.cell(row=1, column=col_idx, value=name)

        wb.save(str(wb_path))
        monkeypatch.setattr(grp4, "WORKBOOK_PATH", str(wb_path))

        with pytest.raises(SystemExit) as exc_info:
            grp4.main()
        assert exc_info.value.code == 0
