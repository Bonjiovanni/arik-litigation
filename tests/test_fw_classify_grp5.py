"""
tests/test_fw_classify_grp5.py
-------------------------------
Namespace and smoke tests for fw_classify_grp5.py (the main() orchestrator).

main() integrates all grp1-4 modules and reads from a hardcoded workbook path,
so direct invocation is not unit-tested here. Tests verify that the module
imports cleanly, constants are correct, and all group imports resolve.
"""

import pytest
import sys
from pathlib import Path

import fw_classify_grp5 as grp5


# ---------------------------------------------------------------------------
# TestNamespace — module-level sanity
# ---------------------------------------------------------------------------

class TestNamespace:

    def test_module_importable(self):
        assert grp5 is not None

    def test_main_is_callable(self):
        assert callable(grp5.main)

    def test_workbook_path_constant_exists(self):
        assert hasattr(grp5, "WORKBOOK_PATH")
        assert isinstance(grp5.WORKBOOK_PATH, str)
        assert grp5.WORKBOOK_PATH.endswith(".xlsx")

    def test_all_group_imports_resolved(self):
        """All four group imports must be accessible via grp5 namespace."""
        assert grp5.grp1 is not None
        assert grp5.grp2 is not None
        assert grp5.grp3 is not None
        assert grp5.grp4 is not None

    def test_grp1_functions_accessible(self):
        assert callable(grp5.grp1.get_text_sample)
        assert callable(grp5.grp1.detect_money)
        assert callable(grp5.grp1.detect_dates)
        assert callable(grp5.grp1.match_keywords)

    def test_grp2_functions_accessible(self):
        assert callable(grp5.grp2.get_classifiable_rows)
        assert callable(grp5.grp2.write_classify_signals)
        assert callable(grp5.grp2.mark_row_classified)

    def test_grp3_functions_accessible(self):
        assert callable(grp5.grp3.ensure_keywords_config_sheet)
        assert callable(grp5.grp3.load_keywords)
        assert callable(grp5.grp3.infer_likely_text_bearing)
        assert callable(grp5.grp3.infer_needs_ocr)
        assert callable(grp5.grp3.classify_doc_type)

    def test_grp4_functions_accessible(self):
        assert callable(grp5.grp4.ensure_classify_history_sheet)
        assert callable(grp5.grp4.log_classify_run)


# ---------------------------------------------------------------------------
# TestMainNoMFISheet — error branch when workbook has no MFI sheet
# ---------------------------------------------------------------------------

class TestMainNoMFISheet:

    def test_exits_when_no_mfi_sheet(self, tmp_path, monkeypatch):
        """main() should sys.exit(1) when Master_File_Inventory is missing."""
        import openpyxl

        wb_path = tmp_path / "test_master.xlsx"
        wb = openpyxl.Workbook()
        # Keep the default sheet so openpyxl can save — just don't add MFI
        wb.save(str(wb_path))

        monkeypatch.setattr(grp5, "WORKBOOK_PATH", str(wb_path))

        with pytest.raises(SystemExit) as exc_info:
            grp5.main()
        assert exc_info.value.code == 1

    def test_exits_when_no_rows_to_classify(self, tmp_path, monkeypatch):
        """main() should sys.exit(0) when MFI exists but has no classifiable rows."""
        import openpyxl
        from fw_walk_grp2 import _COL_INDEX

        wb_path = tmp_path / "test_master.xlsx"
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Create a minimal MFI sheet with headers only
        ws = wb.create_sheet("Master_File_Inventory")
        for name, col_idx in _COL_INDEX.items():
            ws.cell(row=1, column=col_idx, value=name)

        wb.save(str(wb_path))

        monkeypatch.setattr(grp5, "WORKBOOK_PATH", str(wb_path))

        with pytest.raises(SystemExit) as exc_info:
            grp5.main()
        assert exc_info.value.code == 0
