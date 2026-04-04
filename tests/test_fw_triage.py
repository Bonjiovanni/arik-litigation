"""
tests/test_fw_triage.py
------------------------
Module-level integration tests for fw_triage.py (the merged single-file module).

Covers:
  1. TestNamespace         — all public functions importable, constants present
  2. TestScoringPipeline   — config/bands/score/band/reason/next_step end-to-end
  3. TestSheetStructure    — all 3 sheets created, freeze panes, idempotency
  4. TestIOPipeline        — write_triage_results + mark_row_triaged + round-trip
  5. TestLoggingPipeline   — log_triage_run appends correctly, stats defaults
  6. TestRoundTrip         — save/reload workbook preserves all 3 triage sheets
"""

import pytest
import openpyxl
from pathlib import Path

import fw_triage as triage
from fw_walk_grp2 import _COL_INDEX


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fresh_wb():
    wb = openpyxl.Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    return wb


def _record(**kwargs):
    base = {
        "DocType": "", "DocTypeConfidence": "",
        "MoneyDetected": "N", "DateDetected": "N",
        "KeywordHits": "", "NeedsOCR": "N", "EntityHits": "",
    }
    base.update(kwargs)
    return base


def _make_mfi_ws(wb=None):
    if wb is None:
        wb = openpyxl.Workbook()
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]
    ws = wb.create_sheet("Master_File_Inventory")
    for name, col_idx in _COL_INDEX.items():
        ws.cell(row=1, column=col_idx, value=name)
    return wb, ws


def _write_mfi_row(ws, row_num, **kwargs):
    for col_name, value in kwargs.items():
        ws.cell(row=row_num, column=_COL_INDEX[col_name], value=value)


@pytest.fixture
def config_and_bands():
    wb = _fresh_wb()
    ws_cfg = triage.ensure_triage_config_sheet(wb)
    ws_bnd = triage.ensure_triage_bands_sheet(wb)
    config = triage.load_triage_config(ws_cfg)
    bands  = triage.load_triage_bands(ws_bnd)
    return config, bands


# ---------------------------------------------------------------------------
# 1. TestNamespace
# ---------------------------------------------------------------------------

class TestNamespace:

    def test_scoring_functions_present(self):
        for fn in ("score_record", "get_triage_band", "get_reason_flagged", "get_next_step"):
            assert callable(getattr(triage, fn))

    def test_sheet_functions_present(self):
        for fn in ("ensure_triage_config_sheet", "ensure_triage_bands_sheet",
                   "ensure_triage_history_sheet"):
            assert callable(getattr(triage, fn))

    def test_io_functions_present(self):
        for fn in ("get_triageable_rows", "write_triage_results", "mark_row_triaged"):
            assert callable(getattr(triage, fn))

    def test_logging_functions_present(self):
        assert callable(triage.log_triage_run)

    def test_load_functions_present(self):
        assert callable(triage.load_triage_config)
        assert callable(triage.load_triage_bands)

    def test_main_callable(self):
        assert callable(triage.main)

    def test_constants_present(self):
        assert triage.TRIAGE_CONFIG_SHEET == "Triage_Config"
        assert triage.TRIAGE_BANDS_SHEET  == "Triage_Bands"
        assert triage.TRIAGE_HISTORY_SHEET == "Triage_History"
        assert triage.WORKBOOK_PATH.endswith(".xlsx")


# ---------------------------------------------------------------------------
# 2. TestScoringPipeline
# ---------------------------------------------------------------------------

class TestScoringPipeline:

    def test_ltc_claim_scores_high(self, config_and_bands):
        config, bands = config_and_bands
        # LTC_Claim(30) + medium(10) + money(15) + dates(10) = 65 → High
        rec = _record(DocType="LTC_Claim", DocTypeConfidence="medium",
                      MoneyDetected="Y", DateDetected="Y")
        score = triage.score_record(rec, config)
        assert score == 65
        band, _ = triage.get_triage_band(score, bands)
        assert band == "High"

    def test_trust_doc_alone_scores_low(self, config_and_bands):
        config, bands = config_and_bands
        # Trust_Document(20) + low(5) = 25 → Low
        rec = _record(DocType="Trust_Document", DocTypeConfidence="low")
        score = triage.score_record(rec, config)
        assert score == 25
        band, _ = triage.get_triage_band(score, bands)
        assert band == "Low"

    def test_entity_hits_contributes(self, config_and_bands):
        config, bands = config_and_bands
        rec = _record(EntityHits="John Smith")
        assert triage.score_record(rec, config) == 20

    def test_reason_flagged_includes_doctype(self, config_and_bands):
        config, _ = config_and_bands
        rec = _record(DocType="LTC_Claim", MoneyDetected="Y")
        reason = triage.get_reason_flagged(rec, config)
        assert "LTC_Claim" in reason
        assert "MoneyDetected" in reason

    def test_next_step_medium_ocr_override(self, config_and_bands):
        _, bands = config_and_bands
        rec = _record(NeedsOCR="Y")
        step = triage.get_next_step("Medium", rec, bands)
        assert step == "OCR then review"

    def test_next_step_high_no_override(self, config_and_bands):
        _, bands = config_and_bands
        rec = _record(NeedsOCR="Y")
        step = triage.get_next_step("High", rec, bands)
        assert step == "Priority manual review"

    def test_score_clamped_at_100(self, config_and_bands):
        config, _ = config_and_bands
        rec = _record(DocType="Bank_Statement", DocTypeConfidence="high",
                      MoneyDetected="Y", DateDetected="Y",
                      KeywordHits="trust", EntityHits="Smith")
        # Bank_Statement(70) + high(15) + money(15) ... would exceed 100
        score = triage.score_record(rec, config)
        assert score == 100

    def test_empty_record_scores_zero(self, config_and_bands):
        config, _ = config_and_bands
        assert triage.score_record({}, config) == 0

    def test_skip_band_for_zero_score(self, config_and_bands):
        _, bands = config_and_bands
        band, step = triage.get_triage_band(0, bands)
        assert band == "Skip"
        assert step == "Archive"


# ---------------------------------------------------------------------------
# 3. TestSheetStructure
# ---------------------------------------------------------------------------

class TestSheetStructure:

    def test_triage_config_sheet_created(self):
        wb = _fresh_wb()
        triage.ensure_triage_config_sheet(wb)
        assert "Triage_Config" in wb.sheetnames

    def test_triage_bands_sheet_created(self):
        wb = _fresh_wb()
        triage.ensure_triage_bands_sheet(wb)
        assert "Triage_Bands" in wb.sheetnames

    def test_triage_history_sheet_created(self):
        wb = _fresh_wb()
        triage.ensure_triage_history_sheet(wb)
        assert "Triage_History" in wb.sheetnames

    def test_config_sheet_idempotent(self):
        wb = _fresh_wb()
        ws1 = triage.ensure_triage_config_sheet(wb)
        ws1.cell(row=2, column=1, value="MARKER")
        ws2 = triage.ensure_triage_config_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "MARKER"

    def test_bands_sheet_idempotent(self):
        wb = _fresh_wb()
        ws1 = triage.ensure_triage_bands_sheet(wb)
        ws1.cell(row=2, column=1, value="MARKER")
        ws2 = triage.ensure_triage_bands_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "MARKER"

    def test_history_sheet_idempotent(self):
        wb = _fresh_wb()
        ws1 = triage.ensure_triage_history_sheet(wb)
        ws1.cell(row=2, column=1, value="MARKER")
        ws2 = triage.ensure_triage_history_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "MARKER"

    def test_config_freeze_panes(self):
        wb = _fresh_wb()
        ws = triage.ensure_triage_config_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_bands_freeze_panes(self):
        wb = _fresh_wb()
        ws = triage.ensure_triage_bands_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_history_freeze_panes(self):
        wb = _fresh_wb()
        ws = triage.ensure_triage_history_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_all_three_sheets_coexist(self):
        wb = _fresh_wb()
        triage.ensure_triage_config_sheet(wb)
        triage.ensure_triage_bands_sheet(wb)
        triage.ensure_triage_history_sheet(wb)
        assert len(wb.sheetnames) == 3


# ---------------------------------------------------------------------------
# 4. TestIOPipeline
# ---------------------------------------------------------------------------

class TestIOPipeline:

    def test_get_triageable_rows_returns_classified(self):
        wb, ws = _make_mfi_ws()
        _write_mfi_row(ws, 2, FileID="F001", ProcessingStatus="classified",
                       FilePath="C:/a.pdf", FileFamily="pdf",
                       DocType="LTC_Claim", DocTypeConfidence="medium",
                       MoneyDetected="Y", DateDetected="N",
                       KeywordHits="trust", LikelyTextBearing="Y",
                       NeedsOCR="N", DocSubtype="")
        rows = triage.get_triageable_rows(ws)
        assert len(rows) == 1
        assert rows[0][1]["DocType"] == "LTC_Claim"

    def test_get_triageable_rows_skips_file_listed(self):
        wb, ws = _make_mfi_ws()
        _write_mfi_row(ws, 2, FileID="F001", ProcessingStatus="file_listed",
                       FilePath="C:/a.pdf", FileFamily="pdf",
                       DocType="", DocTypeConfidence="",
                       MoneyDetected="N", DateDetected="N",
                       KeywordHits="", LikelyTextBearing="", NeedsOCR="N", DocSubtype="")
        assert triage.get_triageable_rows(ws) == []

    def test_write_triage_results_writes_score(self):
        wb, ws = _make_mfi_ws()
        triage.write_triage_results(ws, 2, {"TriageScore": 65, "TriageBand": "High"})
        assert ws.cell(row=2, column=_COL_INDEX["TriageScore"]).value == 65
        assert ws.cell(row=2, column=_COL_INDEX["TriageBand"]).value == "High"

    def test_write_triage_results_protects_manual_review(self):
        wb, ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["ManualReviewStatus"], value="needs_review")
        triage.write_triage_results(ws, 2, {"ManualReviewStatus": "overwrite"})
        assert ws.cell(row=2, column=_COL_INDEX["ManualReviewStatus"]).value == "needs_review"

    def test_mark_row_triaged_sets_status(self):
        wb, ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"], value="classified")
        triage.mark_row_triaged(ws, 2)
        assert ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"]).value == "triaged"

    def test_full_io_sequence(self):
        """classify → get_triageable_rows → write_triage_results → mark_triaged → no longer returned."""
        wb, ws = _make_mfi_ws()
        _write_mfi_row(ws, 2, FileID="F001", ProcessingStatus="classified",
                       FilePath="C:/a.pdf", FileFamily="pdf",
                       DocType="Invoice", DocTypeConfidence="low",
                       MoneyDetected="Y", DateDetected="N",
                       KeywordHits="", LikelyTextBearing="Y",
                       NeedsOCR="N", DocSubtype="")

        rows = triage.get_triageable_rows(ws)
        assert len(rows) == 1

        triage.write_triage_results(ws, rows[0][0], {"TriageScore": 70, "TriageBand": "High"})
        triage.mark_row_triaged(ws, rows[0][0])

        # Should not be returned again
        rows2 = triage.get_triageable_rows(ws)
        assert rows2 == []


# ---------------------------------------------------------------------------
# 5. TestLoggingPipeline
# ---------------------------------------------------------------------------

class TestLoggingPipeline:

    def test_log_run_writes_run_id(self):
        wb = _fresh_wb()
        triage.log_triage_run(wb, "TRIAGE_20260316_090000", "2026-03-16T09:00:00",
                              {"rows_processed": 5, "rows_triaged": 5})
        ws = wb["Triage_History"]
        assert ws.cell(row=2, column=1).value == "TRIAGE_20260316_090000"

    def test_second_run_appends(self):
        wb = _fresh_wb()
        triage.log_triage_run(wb, "RUN1", "2026-03-16T09:00:00", {"rows_processed": 3})
        row = triage.log_triage_run(wb, "RUN2", "2026-03-16T10:00:00", {"rows_processed": 2})
        assert row == 3
        ws = wb["Triage_History"]
        assert ws.cell(row=3, column=1).value == "RUN2"

    def test_stats_default_to_zero(self):
        wb = _fresh_wb()
        triage.log_triage_run(wb, "RUN1", "2026-03-16T09:00:00", {})
        ws = wb["Triage_History"]
        # RowsProcessed is column 4 (RunID=1, StartedAt=2, CompletedAt=3, RowsProcessed=4)
        assert ws.cell(row=2, column=4).value == 0

    def test_creates_history_sheet_if_absent(self):
        wb = _fresh_wb()
        triage.log_triage_run(wb, "RUN1", "2026-03-16T09:00:00", {})
        assert "Triage_History" in wb.sheetnames


# ---------------------------------------------------------------------------
# 6. TestRoundTrip
# ---------------------------------------------------------------------------

class TestRoundTrip:

    def test_save_reload_preserves_all_sheets(self, tmp_path):
        wb_path = str(tmp_path / "triage_rt.xlsx")
        wb = _fresh_wb()
        triage.ensure_triage_config_sheet(wb)
        triage.ensure_triage_bands_sheet(wb)
        triage.ensure_triage_history_sheet(wb)
        triage.log_triage_run(wb, "RUN1", "2026-03-16T09:00:00",
                              {"rows_processed": 10, "rows_triaged": 8})
        wb.save(wb_path)

        wb2 = openpyxl.load_workbook(wb_path)
        assert "Triage_Config"  in wb2.sheetnames
        assert "Triage_Bands"   in wb2.sheetnames
        assert "Triage_History" in wb2.sheetnames

    def test_save_reload_preserves_run_log(self, tmp_path):
        wb_path = str(tmp_path / "triage_rt2.xlsx")
        wb = _fresh_wb()
        triage.log_triage_run(wb, "RUN_PERSIST", "2026-03-16T09:00:00",
                              {"rows_processed": 7})
        wb.save(wb_path)

        wb2 = openpyxl.load_workbook(wb_path)
        ws = wb2["Triage_History"]
        assert ws.cell(row=2, column=1).value == "RUN_PERSIST"

    def test_config_data_survives_reload(self, tmp_path):
        wb_path = str(tmp_path / "triage_rt3.xlsx")
        wb = _fresh_wb()
        triage.ensure_triage_config_sheet(wb)
        wb.save(wb_path)

        wb2 = openpyxl.load_workbook(wb_path)
        ws = wb2["Triage_Config"]
        config = triage.load_triage_config(ws)
        assert config["DocType"]["LTC_Claim"] == 30
        assert config["Signal"]["EntityHits"] == 20
