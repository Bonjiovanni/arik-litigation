"""
tests/test_fw_classify_grp2.py
-------------------------------
Unit tests for fw_classify_grp2.py:
    get_classifiable_rows, write_classify_signals, mark_row_classified

All tests use a real openpyxl Workbook with a fake Master_File_Inventory
worksheet. Column positions are driven by _COL_INDEX imported from fw_walk_grp2
so these tests stay in sync with the production schema.
"""

import pytest
import openpyxl

from fw_walk_grp2 import _COL_INDEX
import fw_classify_grp2 as grp2


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_mfi_ws():
    """Create an openpyxl worksheet with MFI column headers in row 1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master_File_Inventory"
    for name, col_idx in _COL_INDEX.items():
        ws.cell(row=1, column=col_idx, value=name)
    return ws


def _write_row(ws, row_num: int, **kwargs):
    """Write values to a data row using column names from _COL_INDEX."""
    for col_name, value in kwargs.items():
        ws.cell(row=row_num, column=_COL_INDEX[col_name], value=value)


# ---------------------------------------------------------------------------
# TestGetClassifiableRows
# ---------------------------------------------------------------------------

class TestGetClassifiableRows:

    def test_returns_file_listed_rows(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", FilePath="C:/a/b.pdf", FileName="b.pdf",
                   FileFamily="pdf", LikelyTextBearing="", NeedsOCR="",
                   ProcessingStatus="file_listed", SkipReason=None)
        rows = grp2.get_classifiable_rows(ws)
        assert len(rows) == 1
        assert rows[0][0] == 2
        assert rows[0][1]["FileID"] == "F001"

    def test_returns_null_status_rows(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F002", FilePath="C:/x.txt", FileName="x.txt",
                   FileFamily="text_file", LikelyTextBearing="", NeedsOCR="",
                   ProcessingStatus=None, SkipReason=None)
        rows = grp2.get_classifiable_rows(ws)
        assert len(rows) == 1

    def test_skips_classified_rows(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F003", FilePath="C:/c.pdf", FileName="c.pdf",
                   FileFamily="pdf", LikelyTextBearing="Y", NeedsOCR="N",
                   ProcessingStatus="classified", SkipReason=None)
        rows = grp2.get_classifiable_rows(ws)
        assert rows == []

    def test_skips_rows_with_skip_reason(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F004", FilePath="C:/d.tmp", FileName="d.tmp",
                   FileFamily="other", LikelyTextBearing="", NeedsOCR="",
                   ProcessingStatus="file_listed", SkipReason="temp_file")
        rows = grp2.get_classifiable_rows(ws)
        assert rows == []

    def test_stops_at_empty_file_id(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", ProcessingStatus="file_listed", SkipReason=None,
                   FilePath="C:/a.pdf", FileName="a.pdf", FileFamily="pdf",
                   LikelyTextBearing="", NeedsOCR="")
        # row 3 left empty — loop should stop
        rows = grp2.get_classifiable_rows(ws)
        assert len(rows) == 1

    def test_returns_multiple_classifiable_rows(self):
        ws = _make_mfi_ws()
        for i, fid in enumerate(["F001", "F002", "F003"], start=2):
            _write_row(ws, i, FileID=fid, FilePath=f"C:/{fid}.pdf", FileName=f"{fid}.pdf",
                       FileFamily="pdf", LikelyTextBearing="", NeedsOCR="",
                       ProcessingStatus="file_listed", SkipReason=None)
        rows = grp2.get_classifiable_rows(ws)
        assert len(rows) == 3

    def test_record_dict_has_expected_keys(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", FilePath="C:/a.pdf", FileName="a.pdf",
                   FileFamily="pdf", LikelyTextBearing="", NeedsOCR="",
                   ProcessingStatus="file_listed", SkipReason=None)
        rows = grp2.get_classifiable_rows(ws)
        record = rows[0][1]
        for key in ("FileID", "FilePath", "FileName", "FileFamily",
                    "LikelyTextBearing", "NeedsOCR", "ProcessingStatus"):
            assert key in record

    def test_mixed_rows_only_returns_eligible(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", ProcessingStatus="file_listed", SkipReason=None,
                   FilePath="C:/a.pdf", FileName="a.pdf", FileFamily="pdf",
                   LikelyTextBearing="", NeedsOCR="")
        _write_row(ws, 3, FileID="F002", ProcessingStatus="classified", SkipReason=None,
                   FilePath="C:/b.pdf", FileName="b.pdf", FileFamily="pdf",
                   LikelyTextBearing="Y", NeedsOCR="N")
        _write_row(ws, 4, FileID="F003", ProcessingStatus="file_listed", SkipReason="lock_file",
                   FilePath="C:/c.pdf", FileName="c.pdf", FileFamily="pdf",
                   LikelyTextBearing="", NeedsOCR="")
        _write_row(ws, 5, FileID="F004", ProcessingStatus=None, SkipReason=None,
                   FilePath="C:/d.txt", FileName="d.txt", FileFamily="text_file",
                   LikelyTextBearing="", NeedsOCR="")
        rows = grp2.get_classifiable_rows(ws)
        ids = [r[1]["FileID"] for r in rows]
        assert ids == ["F001", "F004"]


# ---------------------------------------------------------------------------
# TestWriteClassifySignals
# ---------------------------------------------------------------------------

class TestWriteClassifySignals:

    def test_writes_text_sample(self):
        ws = _make_mfi_ws()
        grp2.write_classify_signals(ws, 2, {"TextSample": "hello world"})
        assert ws.cell(row=2, column=_COL_INDEX["TextSample"]).value == "hello world"

    def test_writes_money_detected(self):
        ws = _make_mfi_ws()
        grp2.write_classify_signals(ws, 2, {"MoneyDetected": "Y"})
        assert ws.cell(row=2, column=_COL_INDEX["MoneyDetected"]).value == "Y"

    def test_writes_doc_type(self):
        ws = _make_mfi_ws()
        grp2.write_classify_signals(ws, 2, {"DocType": "Invoice"})
        assert ws.cell(row=2, column=_COL_INDEX["DocType"]).value == "Invoice"

    def test_does_not_write_manual_review_status(self):
        ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["ManualReviewStatus"], value="needs_review")
        grp2.write_classify_signals(ws, 2, {"ManualReviewStatus": "overwrite_attempt"})
        assert ws.cell(row=2, column=_COL_INDEX["ManualReviewStatus"]).value == "needs_review"

    def test_does_not_write_keep_for_case(self):
        ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["KeepForCase"], value="Y")
        grp2.write_classify_signals(ws, 2, {"KeepForCase": "N"})
        assert ws.cell(row=2, column=_COL_INDEX["KeepForCase"]).value == "Y"

    def test_does_not_write_possible_exhibit(self):
        ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["PossibleExhibit"], value="Y")
        grp2.write_classify_signals(ws, 2, {"PossibleExhibit": "N"})
        assert ws.cell(row=2, column=_COL_INDEX["PossibleExhibit"]).value == "Y"

    def test_unknown_key_silently_ignored(self):
        ws = _make_mfi_ws()
        # Should not raise even for non-existent column names
        grp2.write_classify_signals(ws, 2, {"NonExistentColumn": "value"})

    def test_writes_multiple_signals(self):
        ws = _make_mfi_ws()
        grp2.write_classify_signals(ws, 2, {
            "TextSample": "trust document",
            "MoneyDetected": "N",
            "DateDetected": "Y",
            "KeywordHits": "trust",
            "LikelyTextBearing": "Y",
            "NeedsOCR": "N",
            "DocType": "Trust_Document",
            "DocSubtype": "",
            "DocTypeConfidence": "medium",
        })
        assert ws.cell(row=2, column=_COL_INDEX["TextSample"]).value == "trust document"
        assert ws.cell(row=2, column=_COL_INDEX["DocType"]).value == "Trust_Document"
        assert ws.cell(row=2, column=_COL_INDEX["DocTypeConfidence"]).value == "medium"


# ---------------------------------------------------------------------------
# TestMarkRowClassified
# ---------------------------------------------------------------------------

class TestMarkRowClassified:

    def test_sets_processing_status(self):
        ws = _make_mfi_ws()
        grp2.mark_row_classified(ws, 2)
        assert ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"]).value == "classified"

    def test_overwrites_file_listed(self):
        ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"], value="file_listed")
        grp2.mark_row_classified(ws, 2)
        assert ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"]).value == "classified"

    def test_marks_specific_row(self):
        ws = _make_mfi_ws()
        ws.cell(row=5, column=_COL_INDEX["ProcessingStatus"], value="file_listed")
        grp2.mark_row_classified(ws, 5)
        assert ws.cell(row=5, column=_COL_INDEX["ProcessingStatus"]).value == "classified"
        # Row 2 unaffected
        assert ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"]).value is None
