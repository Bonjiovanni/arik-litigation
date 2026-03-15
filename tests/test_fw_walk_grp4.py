"""
tests/test_fw_walk_grp4.py

Unit tests for fw_walk_grp4.py — Walk_Coverage + Walk_History sheet management.
Run with: pytest tests/test_fw_walk_grp4.py
"""

import sys
import os
import pytest
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import fw_walk_grp4 as grp4


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fresh_wb():
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _headers(ws):
    return [c.value for c in ws[1] if c.value is not None]


def _cell(ws, row, col_name):
    h = _headers(ws)
    return ws.cell(row=row, column=h.index(col_name) + 1).value


# ---------------------------------------------------------------------------
# TestEnsureWalkCoverageSheet
# ---------------------------------------------------------------------------

class TestEnsureWalkCoverageSheet:

    EXPECTED_HEADERS = {"CoverageID", "RunID", "DirPath", "ProcessingLevel",
                        "FilesInserted", "FilesUpdated", "FilesSkipped",
                        "Errors", "WalkedAt"}

    def test_creates_sheet(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_coverage_sheet(wb)
        assert ws is not None
        assert "Walk_Coverage" in wb.sheetnames

    def test_idempotent(self):
        wb = _fresh_wb()
        ws1 = grp4.ensure_walk_coverage_sheet(wb)
        ws2 = grp4.ensure_walk_coverage_sheet(wb)
        assert ws1.title == ws2.title
        assert wb.sheetnames.count("Walk_Coverage") == 1

    def test_expected_headers(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_coverage_sheet(wb)
        headers = set(_headers(ws))
        assert self.EXPECTED_HEADERS.issubset(headers)

    def test_freeze_panes_A2(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_coverage_sheet(wb)
        assert str(ws.freeze_panes) == "A2"

    def test_existing_sheet_not_modified(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_coverage_sheet(wb)
        ws.cell(row=2, column=1, value="sentinel_value")
        ws2 = grp4.ensure_walk_coverage_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "sentinel_value"


# ---------------------------------------------------------------------------
# TestGetNextCoverageId
# ---------------------------------------------------------------------------

class TestGetNextCoverageId:

    def _ws_with_ids(self, id_values):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_coverage_sheet(wb)
        for i, val in enumerate(id_values, start=2):
            ws.cell(row=i, column=1, value=val)
        return ws

    def test_empty_sheet_returns_WC00001(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_coverage_sheet(wb)
        assert grp4.get_next_coverage_id(ws) == "WC00001"

    def test_increments_past_existing(self):
        ws = self._ws_with_ids(["WC00005", "WC00003"])
        assert grp4.get_next_coverage_id(ws) == "WC00006"

    def test_malformed_ids_ignored(self):
        ws = self._ws_with_ids([None, "garbage", "WC00002"])
        assert grp4.get_next_coverage_id(ws) == "WC00003"

    def test_zero_padded_5_digits(self):
        ws = self._ws_with_ids(["WC00099"])
        result = grp4.get_next_coverage_id(ws)
        assert result == "WC00100"
        assert result.startswith("WC")


# ---------------------------------------------------------------------------
# TestUpdateWalkCoverage
# ---------------------------------------------------------------------------

class TestUpdateWalkCoverage:

    def _setup(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_coverage_sheet(wb)
        return wb, ws

    def test_returns_row_number(self):
        _, ws = self._setup()
        row = grp4.update_walk_coverage(ws, "RUN1", "C:/Docs", "file_listed", 10, 2, 3, 0)
        assert row == 2

    def test_second_call_returns_next_row(self):
        _, ws = self._setup()
        grp4.update_walk_coverage(ws, "RUN1", "C:/Docs", "file_listed", 10, 2, 3, 0)
        row2 = grp4.update_walk_coverage(ws, "RUN1", "C:/Other", "file_listed", 5, 0, 0, 0)
        assert row2 == 3

    def test_coverage_id_is_set(self):
        _, ws = self._setup()
        row = grp4.update_walk_coverage(ws, "RUN1", "C:/Docs", "file_listed", 10, 2, 3, 0)
        cid = _cell(ws, row, "CoverageID")
        assert cid is not None and str(cid).startswith("WC")

    def test_run_id_stored(self):
        _, ws = self._setup()
        row = grp4.update_walk_coverage(ws, "RUN_XYZ", "C:/Docs", "file_listed", 0, 0, 0, 0)
        assert _cell(ws, row, "RunID") == "RUN_XYZ"

    def test_dir_path_stored(self):
        _, ws = self._setup()
        row = grp4.update_walk_coverage(ws, "RUN1", "C:/My/Dir", "file_listed", 0, 0, 0, 0)
        stored = _cell(ws, row, "DirPath")
        assert stored is not None
        assert "My" in str(stored) or "my" in str(stored)

    def test_files_inserted_stored(self):
        _, ws = self._setup()
        row = grp4.update_walk_coverage(ws, "RUN1", "C:/Docs", "file_listed", 42, 0, 0, 0)
        assert _cell(ws, row, "FilesInserted") == 42

    def test_errors_stored(self):
        _, ws = self._setup()
        row = grp4.update_walk_coverage(ws, "RUN1", "C:/Docs", "file_listed", 0, 0, 0, 7)
        assert _cell(ws, row, "Errors") == 7

    def test_walked_at_is_iso_datetime_string(self):
        _, ws = self._setup()
        row = grp4.update_walk_coverage(ws, "RUN1", "C:/Docs", "file_listed", 0, 0, 0, 0)
        walked_at = _cell(ws, row, "WalkedAt")
        assert walked_at is not None
        assert "T" in str(walked_at) or "-" in str(walked_at)


# ---------------------------------------------------------------------------
# TestEnsureWalkHistorySheet
# ---------------------------------------------------------------------------

class TestEnsureWalkHistorySheet:

    EXPECTED_HEADERS = {"RunID", "StartedAt", "CompletedAt", "ScanDirs",
                        "ProcessingLevel", "TotalInserted", "TotalUpdated",
                        "TotalSkippedFiles", "TotalSkippedDirs", "TotalErrors",
                        "OverlapAction", "Notes"}

    def test_creates_sheet(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_history_sheet(wb)
        assert ws is not None
        assert "Walk_History" in wb.sheetnames

    def test_idempotent(self):
        wb = _fresh_wb()
        ws1 = grp4.ensure_walk_history_sheet(wb)
        ws2 = grp4.ensure_walk_history_sheet(wb)
        assert ws1.title == ws2.title
        assert wb.sheetnames.count("Walk_History") == 1

    def test_expected_headers(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_history_sheet(wb)
        headers = set(_headers(ws))
        assert self.EXPECTED_HEADERS.issubset(headers)

    def test_freeze_panes_A2(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_history_sheet(wb)
        assert str(ws.freeze_panes) == "A2"

    def test_existing_sheet_not_modified(self):
        wb = _fresh_wb()
        ws = grp4.ensure_walk_history_sheet(wb)
        ws.cell(row=2, column=1, value="sentinel")
        ws2 = grp4.ensure_walk_history_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "sentinel"


# ---------------------------------------------------------------------------
# TestLogWalkRun
# ---------------------------------------------------------------------------

class TestLogWalkRun:

    def _stats(self, inserted=10, updated=2, skipped_files=3, skipped_dirs=1, errors=0):
        return {
            "inserted":      inserted,
            "updated":       updated,
            "skipped_files": skipped_files,
            "skipped_dirs":  skipped_dirs,
            "errors":        errors,
        }

    def _log(self, wb, **kwargs):
        defaults = dict(
            run_id="RUN_TEST",
            started_at="2026-01-01T10:00:00",
            scan_dirs=["C:/Docs"],
            processing_level="file_listed",
            stats=self._stats(),
            overlap_action="skip",
            notes="",
        )
        defaults.update(kwargs)
        return grp4.log_walk_run(wb=wb, **defaults)

    def test_creates_history_sheet_if_absent(self):
        wb = _fresh_wb()
        self._log(wb)
        assert "Walk_History" in wb.sheetnames

    def test_returns_row_number(self):
        wb = _fresh_wb()
        row = self._log(wb)
        assert row == 2

    def test_second_call_returns_row_3(self):
        wb = _fresh_wb()
        self._log(wb)
        row2 = self._log(wb, run_id="RUN2")
        assert row2 == 3

    def test_run_id_stored(self):
        wb = _fresh_wb()
        row = self._log(wb, run_id="MY_RUN_001")
        ws  = wb["Walk_History"]
        assert _cell(ws, row, "RunID") == "MY_RUN_001"

    def test_started_at_stored(self):
        wb = _fresh_wb()
        row = self._log(wb, started_at="2026-03-15T09:00:00")
        ws  = wb["Walk_History"]
        assert _cell(ws, row, "StartedAt") == "2026-03-15T09:00:00"

    def test_completed_at_is_set(self):
        wb = _fresh_wb()
        row = self._log(wb)
        ws  = wb["Walk_History"]
        completed = _cell(ws, row, "CompletedAt")
        assert completed is not None and completed != ""

    def test_scan_dirs_joined_with_pipe(self):
        wb  = _fresh_wb()
        row = self._log(wb, scan_dirs=["C:/A", "C:/B"])
        ws  = wb["Walk_History"]
        val = str(_cell(ws, row, "ScanDirs"))
        assert "|" in val

    def test_stats_inserted_stored(self):
        wb  = _fresh_wb()
        row = self._log(wb, stats=self._stats(inserted=99))
        ws  = wb["Walk_History"]
        assert _cell(ws, row, "TotalInserted") == 99

    def test_stats_errors_stored(self):
        wb  = _fresh_wb()
        row = self._log(wb, stats=self._stats(errors=5))
        ws  = wb["Walk_History"]
        assert _cell(ws, row, "TotalErrors") == 5

    def test_overlap_action_stored(self):
        wb  = _fresh_wb()
        row = self._log(wb, overlap_action="rewalk")
        ws  = wb["Walk_History"]
        assert _cell(ws, row, "OverlapAction") == "rewalk"

    def test_notes_stored(self):
        wb  = _fresh_wb()
        row = self._log(wb, notes="test run for CI")
        ws  = wb["Walk_History"]
        assert _cell(ws, row, "Notes") == "test run for CI"

    def test_missing_stats_keys_default_to_zero(self):
        wb  = _fresh_wb()
        row = grp4.log_walk_run(
            wb=wb,
            run_id="R1",
            started_at="2026-01-01T00:00:00",
            scan_dirs=["C:/X"],
            processing_level="file_listed",
            stats={},   # empty — all keys missing
            overlap_action="skip",
        )
        ws = wb["Walk_History"]
        assert _cell(ws, row, "TotalInserted") == 0
        assert _cell(ws, row, "TotalErrors")   == 0
