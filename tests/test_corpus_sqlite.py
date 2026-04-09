"""
Tests for corpus_sqlite_schema.py and corpus_sqlite_loader.py

Covers: schema creation, column mapping, data loading, FK enforcement,
FTS5 search, domain isolation, and rebuild safety.
"""

import os
import sqlite3
import tempfile

import pytest

# Add repo root to path
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from corpus_sqlite_schema import (
    to_snake_case,
    build_email_column_mapping,
    build_attachments_column_mapping,
    connect,
    create_shared_tables,
    create_email_tables,
    drop_email_tables,
    get_column_type,
    EMAILS_INTEGER_COLUMNS,
    ATTACHMENTS_INTEGER_COLUMNS,
)


# ─── Source file paths ───

EMAILS_PATH = r"C:\Users\arika\OneDrive\Litigation\Pipeline\label_LegalEmailExtracts - Ariks Version V11.xlsx"
ATTACHMENTS_PATH = r"C:\Users\arika\OneDrive\Litigation\Aid4Mail Exports\Aid4 to json with python attachments\Attachment_Manifest_GML_EML_MSG (1).xlsx"


# ─── Fixtures ───

@pytest.fixture
def tmp_db(tmp_path):
    """Create a temporary database for testing."""
    db_path = str(tmp_path / "test_corpus.db")
    conn = connect(db_path)
    yield conn, db_path
    conn.close()


@pytest.fixture
def db_with_shared_tables(tmp_db):
    """Database with shared tables created."""
    conn, db_path = tmp_db
    create_shared_tables(conn)
    return conn, db_path


@pytest.fixture
def sample_email_columns():
    """Minimal email column mapping for testing."""
    return [
        ("RFC 822 Message ID", "rfc_822_message_id", "TEXT"),
        ("From", "from_addr", "TEXT"),
        ("To", "to_addr", "TEXT"),
        ("Subject", "subject", "TEXT"),
        ("Date & Time Sent", "date_time_sent", "TEXT"),
        ("body_clean.1", "body_clean_1", "TEXT"),
        ("Attachments Count", "attachments_count", "INTEGER"),
        ("Email.Size", "email_size", "INTEGER"),
    ]


@pytest.fixture
def db_with_email_tables(db_with_shared_tables, sample_email_columns):
    """Database with shared + email tables created."""
    conn, db_path = db_with_shared_tables
    create_email_tables(conn, sample_email_columns)
    return conn, db_path


# ═══════════════════════════════════════════════
# Step 1: Source file verification
# ═══════════════════════════════════════════════

class TestSourceFiles:
    def test_emails_workbook_exists(self):
        assert os.path.exists(EMAILS_PATH), f"Emails workbook not found: {EMAILS_PATH}"

    def test_attachments_workbook_exists(self):
        assert os.path.exists(ATTACHMENTS_PATH), f"Attachments workbook not found: {ATTACHMENTS_PATH}"

    def test_emails_merge1_sheet_exists(self):
        import openpyxl
        wb = openpyxl.load_workbook(EMAILS_PATH, read_only=True)
        try:
            assert "Merge1" in wb.sheetnames, f"Merge1 sheet not found. Sheets: {wb.sheetnames}"
        finally:
            wb.close()

    def test_emails_merge1_column_count(self):
        import openpyxl
        wb = openpyxl.load_workbook(EMAILS_PATH, read_only=True)
        ws = wb["Merge1"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        assert len(headers) == 103, f"Expected 103 columns, got {len(headers)}"

    def test_emails_merge1_row_count(self):
        import openpyxl
        wb = openpyxl.load_workbook(EMAILS_PATH, read_only=True)
        ws = wb["Merge1"]
        row_count = sum(1 for _ in ws.iter_rows(min_row=2))
        wb.close()
        assert row_count == 1550, f"Expected 1550 data rows, got {row_count}"

    def test_attachments_row_count(self):
        import openpyxl
        wb = openpyxl.load_workbook(ATTACHMENTS_PATH, read_only=True)
        ws = wb.active
        row_count = sum(1 for _ in ws.iter_rows(min_row=2))
        wb.close()
        assert row_count == 2154, f"Expected 2154 data rows, got {row_count}"

    def test_attachments_column_count(self):
        import openpyxl
        wb = openpyxl.load_workbook(ATTACHMENTS_PATH, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        assert len(headers) == 17, f"Expected 17 columns, got {len(headers)}"


# ═══════════════════════════════════════════════
# Step 2: Column name mapping
# ═══════════════════════════════════════════════

class TestSnakeCase:
    def test_spaces(self):
        assert to_snake_case("Date & Time Sent") == "date_time_sent"

    def test_dots(self):
        assert to_snake_case("body_clean.1") == "body_clean_1"

    def test_parens(self):
        assert to_snake_case("From (name)") == "from_name"

    def test_rfc_message_id(self):
        assert to_snake_case("RFC 822 Message ID") == "rfc_822_message_id"

    def test_equals_sign(self):
        assert to_snake_case("original=saved") == "original_saved"

    def test_question_mark(self):
        assert to_snake_case("Is Real Reply?") == "is_real_reply"

    def test_already_snake(self):
        assert to_snake_case("message_id") == "message_id"

    def test_dots_prefix(self):
        assert to_snake_case("Header.X-Priority") == "header_x_priority"

    def test_empty_after_strip(self):
        # Edge case: all special chars
        result = to_snake_case("???")
        assert result == ""

    def test_no_leading_trailing_underscores(self):
        result = to_snake_case("  Subject  ")
        assert not result.startswith("_")
        assert not result.endswith("_")


class TestEmailColumnMapping:
    def test_all_snake_case_valid(self):
        """All mapped names should be valid SQLite identifiers."""
        import openpyxl
        wb = openpyxl.load_workbook(EMAILS_PATH, read_only=True)
        ws = wb["Merge1"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        mapping = build_email_column_mapping(headers)
        for original, snake, dtype in mapping:
            assert snake, f"Empty snake_case for: {original}"
            assert " " not in snake, f"Space in snake_case '{snake}' for: {original}"
            assert "." not in snake, f"Dot in snake_case '{snake}' for: {original}"
            assert snake == snake.lower(), f"Uppercase in snake_case '{snake}' for: {original}"

    def test_no_collisions(self):
        """No two headers should map to the same snake_case name."""
        import openpyxl
        wb = openpyxl.load_workbook(EMAILS_PATH, read_only=True)
        ws = wb["Merge1"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        mapping = build_email_column_mapping(headers)
        snake_names = [snake for _, snake, _ in mapping]
        assert len(snake_names) == len(set(snake_names)), \
            f"Duplicate snake names found: {[n for n in snake_names if snake_names.count(n) > 1]}"

    def test_has_103_columns(self):
        import openpyxl
        wb = openpyxl.load_workbook(EMAILS_PATH, read_only=True)
        ws = wb["Merge1"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        mapping = build_email_column_mapping(headers)
        assert len(mapping) == 103

    def test_rfc_message_id_is_pk(self):
        import openpyxl
        wb = openpyxl.load_workbook(EMAILS_PATH, read_only=True)
        ws = wb["Merge1"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        mapping = build_email_column_mapping(headers)
        snake_names = [snake for _, snake, _ in mapping]
        assert "rfc_822_message_id" in snake_names, \
            "rfc_822_message_id not found in mapping"

    def test_integer_columns_typed_correctly(self):
        import openpyxl
        wb = openpyxl.load_workbook(EMAILS_PATH, read_only=True)
        ws = wb["Merge1"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        mapping = build_email_column_mapping(headers)
        int_cols = {snake: dtype for _, snake, dtype in mapping if dtype == "INTEGER"}
        # At minimum these should be INTEGER
        for expected in ["attachments_count", "email_size"]:
            assert expected in int_cols, f"{expected} should be INTEGER, not found"

    def test_reserved_word_from_renamed(self):
        mapping = build_email_column_mapping(["From", "To", "Subject"])
        snake_names = [snake for _, snake, _ in mapping]
        assert "from_addr" in snake_names
        assert "to_addr" in snake_names
        assert "from" not in snake_names
        assert "to" not in snake_names


class TestAttachmentsColumnMapping:
    def test_all_17_columns_mapped(self):
        import openpyxl
        wb = openpyxl.load_workbook(ATTACHMENTS_PATH, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        mapping = build_attachments_column_mapping(headers)
        assert len(mapping) == 17

    def test_reserved_words_renamed(self):
        mapping = build_attachments_column_mapping(["from", "to", "subject"])
        snake_names = [snake for _, snake, _ in mapping]
        assert "from_addr" in snake_names
        assert "to_addr" in snake_names


# ═══════════════════════════════════════════════
# Schema creation
# ═══════════════════════════════════════════════

class TestSharedTables:
    def test_entity_master_exists(self, db_with_shared_tables):
        conn, _ = db_with_shared_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "entity_master" in tables

    def test_entity_candidates_exists(self, db_with_shared_tables):
        conn, _ = db_with_shared_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "entity_candidates" in tables

    def test_column_map_exists(self, db_with_shared_tables):
        conn, _ = db_with_shared_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "_column_map" in tables

    def test_schema_notes_exists(self, db_with_shared_tables):
        conn, _ = db_with_shared_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "_schema_notes" in tables

    def test_entity_master_columns(self, db_with_shared_tables):
        conn, _ = db_with_shared_tables
        cols = [r[1] for r in conn.execute("PRAGMA table_info(entity_master)").fetchall()]
        expected = {"entity_id", "canonical_name", "entity_type", "known_aliases",
                    "status", "first_seen_source", "notes"}
        assert set(cols) == expected

    def test_idempotent(self, db_with_shared_tables):
        """Running create_shared_tables twice should not error."""
        conn, _ = db_with_shared_tables
        create_shared_tables(conn)  # second call
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "entity_master" in tables


class TestEmailTables:
    def test_emails_master_exists(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "emails_master" in tables

    def test_attachments_exists(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "attachments" in tables

    def test_extraction_log_exists(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "extraction_log" in tables

    def test_attachments_has_extracted_text(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        cols = [r[1] for r in conn.execute("PRAGMA table_info(attachments)").fetchall()]
        assert "extracted_text" in cols

    def test_attachments_has_extraction_status(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        cols = [r[1] for r in conn.execute("PRAGMA table_info(attachments)").fetchall()]
        assert "extraction_status" in cols

    def test_extraction_log_columns(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        cols = [r[1] for r in conn.execute("PRAGMA table_info(extraction_log)").fetchall()]
        expected = {"log_id", "attachment_id", "status", "method",
                    "file_type", "timestamp", "error_message"}
        assert set(cols) == expected

    def test_fts5_emails_exists(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "_fts_emails" in tables

    def test_fts5_attachments_exists(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        assert "_fts_attachments" in tables


# ═══════════════════════════════════════════════
# FK enforcement
# ═══════════════════════════════════════════════

class TestForeignKeys:
    def test_fk_enabled(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        result = conn.execute("PRAGMA foreign_keys").fetchone()
        assert result[0] == 1, "Foreign keys not enabled"

    def test_orphan_attachment_rejected(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        with pytest.raises(sqlite3.IntegrityError):
            conn.execute("""
                INSERT INTO attachments (message_id, original_filename, extraction_status)
                VALUES ('FAKE_NONEXISTENT_ID', 'test.pdf', 'pending')
            """)

    def test_valid_attachment_accepted(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        # Insert a valid email first
        conn.execute("""
            INSERT INTO emails_master (rfc_822_message_id, subject)
            VALUES ('test_msg_001', 'Test Subject')
        """)
        # Now insert attachment referencing it — should succeed
        conn.execute("""
            INSERT INTO attachments (message_id, original_filename, extraction_status)
            VALUES ('test_msg_001', 'document.pdf', 'pending')
        """)
        count = conn.execute("SELECT COUNT(*) FROM attachments").fetchone()[0]
        assert count == 1


# ═══════════════════════════════════════════════
# Domain isolation
# ═══════════════════════════════════════════════

class TestDomainIsolation:
    def test_email_rebuild_preserves_entity_master(self, db_with_email_tables, sample_email_columns):
        conn, _ = db_with_email_tables
        # Insert an entity
        conn.execute("""
            INSERT INTO entity_master (canonical_name, entity_type)
            VALUES ('David Peterson', 'person')
        """)
        conn.commit()

        # Rebuild email tables
        create_email_tables(conn, sample_email_columns, rebuild=True)

        # Entity should still be there
        count = conn.execute("SELECT COUNT(*) FROM entity_master").fetchone()[0]
        assert count == 1, "Entity was lost during email rebuild!"

    def test_email_rebuild_preserves_entity_candidates(self, db_with_email_tables, sample_email_columns):
        conn, _ = db_with_email_tables
        conn.execute("""
            INSERT INTO entity_candidates (candidate_text, likely_entity_type)
            VALUES ('Gravel & Shea', 'law_firm')
        """)
        conn.commit()

        create_email_tables(conn, sample_email_columns, rebuild=True)

        count = conn.execute("SELECT COUNT(*) FROM entity_candidates").fetchone()[0]
        assert count == 1, "Entity candidate was lost during email rebuild!"

    def test_email_rebuild_clears_email_data(self, db_with_email_tables, sample_email_columns):
        conn, _ = db_with_email_tables
        # Insert some email data
        conn.execute("""
            INSERT INTO emails_master (rfc_822_message_id, subject)
            VALUES ('msg_001', 'Test')
        """)
        conn.commit()

        # Rebuild
        create_email_tables(conn, sample_email_columns, rebuild=True)

        count = conn.execute("SELECT COUNT(*) FROM emails_master").fetchone()[0]
        assert count == 0, "Email data survived rebuild!"


# ═══════════════════════════════════════════════
# FTS5 search (basic — with sample data)
# ═══════════════════════════════════════════════

class TestFTS5:
    def test_email_fts_search(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        # Insert a test email
        conn.execute("""
            INSERT INTO emails_master (rfc_822_message_id, subject, body_clean_1)
            VALUES ('msg_fts_test', 'HIPAA Authorization Form', 'Please sign the HIPAA form')
        """)
        # Populate FTS (standalone table, not content-synced)
        conn.execute("""
            INSERT INTO _fts_emails (subject, body_clean)
            VALUES ('HIPAA Authorization Form', 'Please sign the HIPAA form')
        """)
        conn.commit()

        results = conn.execute(
            "SELECT * FROM _fts_emails WHERE _fts_emails MATCH 'HIPAA'"
        ).fetchall()
        assert len(results) >= 1

    def test_attachment_fts_search(self, db_with_email_tables):
        conn, _ = db_with_email_tables
        # Need a parent email first
        conn.execute("""
            INSERT INTO emails_master (rfc_822_message_id, subject)
            VALUES ('msg_attach_fts', 'Test')
        """)
        conn.execute("""
            INSERT INTO attachments (message_id, original_filename, extraction_status)
            VALUES ('msg_attach_fts', 'W-9 Tax Form.pdf', 'pending')
        """)
        # Populate FTS (standalone table)
        conn.execute("""
            INSERT INTO _fts_attachments (original_filename, extracted_text)
            VALUES ('W-9 Tax Form.pdf', NULL)
        """)
        conn.commit()

        # Note: W-9 must be quoted — FTS5 treats '-' as NOT operator
        results = conn.execute(
            'SELECT * FROM _fts_attachments WHERE _fts_attachments MATCH \'"W-9"\'',
        ).fetchall()
        assert len(results) >= 1


# ═══════════════════════════════════════════════
# Data type mapping
# ═══════════════════════════════════════════════

class TestDataTypes:
    def test_attachments_count_is_integer(self):
        assert get_column_type("attachments_count", "emails_master") == "INTEGER"

    def test_email_size_is_integer(self):
        assert get_column_type("email_size", "emails_master") == "INTEGER"

    def test_size_bytes_is_integer(self):
        assert get_column_type("size_bytes", "attachments") == "INTEGER"

    def test_subject_is_text(self):
        assert get_column_type("subject", "emails_master") == "TEXT"

    def test_from_addr_is_text(self):
        assert get_column_type("from_addr", "emails_master") == "TEXT"
