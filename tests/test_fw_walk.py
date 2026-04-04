"""
tests/test_fw_walk.py

Module-level tests for the merged fw_walk.py entry point.

Section 1 (TestPublicAPI, TestCrossGroupConsistency):
    Smoke tests written at merge time — namespace and basic cross-group checks.

Sections 2-8 (added after merge):
    Integration tests covering the full pipeline, data contracts,
    insert/update behaviour, walk_files against a real temp directory,
    Walk_Coverage + Walk_History logging, sheet structure, and save/reload
    round-trip.

Interactive functions not tested here: prompt_overlap_decision(), main()
Run with: pytest tests/test_fw_walk.py
"""

import os
import sys
import tempfile
from datetime import datetime

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import fw_walk


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _fresh_wb():
    wb = openpyxl.Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    return wb

def _make_record(file_path="C:/test/document.pdf", filename="document.pdf",
                 file_family="pdf", **kwargs):
    base = {
        "file_path":         file_path,
        "parent_folder":     "C:/test",
        "filename":          filename,
        "scan_root_path":    "C:/test",
        "source_store":      "Local",
        "size_bytes":        1024,
        "created_time":      "2026-01-01T00:00:00",
        "modified_time":     "2026-01-01T00:00:00",
        "sha256":            "",
        "file_family":       file_family,
        "processing_status": "file_listed",
    }
    base.update(kwargs)
    return base


# ---------------------------------------------------------------------------
# Import-level smoke tests
# ---------------------------------------------------------------------------

class TestPublicAPI:

    def test_group1_functions_accessible(self):
        assert callable(fw_walk.load_file_family_config)
        assert callable(fw_walk.classify_file_family)
        assert callable(fw_walk.should_skip_file)
        assert callable(fw_walk.get_file_metadata)
        assert callable(fw_walk.compute_sha256)
        assert callable(fw_walk.peek_archive_contents)

    def test_group2_functions_accessible(self):
        assert callable(fw_walk.ensure_master_file_inventory_sheet)
        assert callable(fw_walk.get_next_file_id)
        assert callable(fw_walk.find_existing_file_row)
        assert callable(fw_walk.insert_file_record)
        assert callable(fw_walk.update_file_record)
        assert callable(fw_walk.write_or_update_file_record)
        assert callable(fw_walk.ensure_file_family_config_sheet)

    def test_group3_functions_accessible(self):
        assert isinstance(fw_walk.PROCESSING_LEVEL_ORDER, list)
        assert callable(fw_walk.get_covered_paths)
        assert callable(fw_walk.check_overlap)
        assert callable(fw_walk.prompt_overlap_decision)
        assert callable(fw_walk.walk_files)

    def test_group4_functions_accessible(self):
        assert callable(fw_walk.ensure_walk_coverage_sheet)
        assert callable(fw_walk.get_next_coverage_id)
        assert callable(fw_walk.update_walk_coverage)
        assert callable(fw_walk.ensure_walk_history_sheet)
        assert callable(fw_walk.log_walk_run)

    def test_group5_main_accessible(self):
        assert callable(fw_walk.main)

    def test_workbook_path_defined(self):
        assert hasattr(fw_walk, "WORKBOOK_PATH")
        assert "xlsx" in fw_walk.WORKBOOK_PATH


# ---------------------------------------------------------------------------
# Cross-group consistency tests (merged module)
# ---------------------------------------------------------------------------

class TestCrossGroupConsistency:

    def test_processing_level_order_has_file_listed(self):
        assert "file_listed" in fw_walk.PROCESSING_LEVEL_ORDER

    def test_classify_and_sheet_use_same_family_names(self):
        # Families in _FAMILY_MAP should all appear in FileFamily_Config defaults
        wb = openpyxl.Workbook()
        fc_ws = fw_walk.ensure_file_family_config_sheet(wb)
        config_families = {
            row[0] for row in fc_ws.iter_rows(min_row=2, values_only=True) if row and row[0]
        }
        # A few spot checks
        assert "pdf" in config_families
        assert "archive" in config_families
        assert "text_file" in config_families

    def test_master_inventory_has_archive_contents_column(self):
        wb = openpyxl.Workbook()
        ws = fw_walk.ensure_master_file_inventory_sheet(wb)
        headers = {c.value for c in ws[1] if c.value}
        assert "ArchiveContents" in headers

    def test_walk_coverage_and_history_created_together(self):
        wb = openpyxl.Workbook()
        fw_walk.ensure_walk_coverage_sheet(wb)
        fw_walk.ensure_walk_history_sheet(wb)
        assert "Walk_Coverage" in wb.sheetnames
        assert "Walk_History" in wb.sheetnames

    def test_write_or_update_creates_inventory_sheet_if_absent(self):
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        record = {"file_path": "C:/Test/file.pdf", "filename": "file.pdf", "size_bytes": 100}
        fw_walk.write_or_update_file_record(wb, record, "RUN1")
        assert "Master_File_Inventory" in wb.sheetnames

    def test_full_init_sequence_creates_all_sheets(self):
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        fw_walk.ensure_master_file_inventory_sheet(wb)
        fw_walk.ensure_file_family_config_sheet(wb)
        fw_walk.ensure_walk_coverage_sheet(wb)
        fw_walk.ensure_walk_history_sheet(wb)
        expected = {
            "Master_File_Inventory", "FileFamily_Config",
            "Walk_Coverage", "Walk_History",
        }
        assert expected.issubset(set(wb.sheetnames))


# ===========================================================================
# 3. PROCESSING_LEVEL_ORDER — ordering contract used across groups
# ===========================================================================

class TestProcessingLevelOrder:
    def test_file_listed_is_first(self):
        assert fw_walk.PROCESSING_LEVEL_ORDER[0] == "file_listed"

    def test_triaged_is_last(self):
        assert fw_walk.PROCESSING_LEVEL_ORDER[-1] == "triaged"

    def test_hashed_after_file_listed(self):
        levels = fw_walk.PROCESSING_LEVEL_ORDER
        assert levels.index("hashed") > levels.index("file_listed")

    def test_classified_after_hashed(self):
        levels = fw_walk.PROCESSING_LEVEL_ORDER
        assert levels.index("classified") > levels.index("hashed")


# ===========================================================================
# 4. FileFamily_Config ↔ load_file_family_config round-trip
# ===========================================================================

class TestFileFamilyConfigRoundTrip:
    """ensure_file_family_config_sheet writes defaults that load_file_family_config reads back."""

    def test_all_10_default_families_loaded(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_file_family_config_sheet(wb)
        family_map, _, _ = fw_walk.load_file_family_config(ws)
        assert len(family_map) == 10

    def test_pdf_extension_in_loaded_map(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_file_family_config_sheet(wb)
        family_map, _, _ = fw_walk.load_file_family_config(ws)
        assert ".pdf" in family_map.get("pdf", set())

    def test_archive_in_skip_families(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_file_family_config_sheet(wb)
        _, skip_families, _ = fw_walk.load_file_family_config(ws)
        assert "archive" in skip_families

    def test_email_file_in_skip_families(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_file_family_config_sheet(wb)
        _, skip_families, _ = fw_walk.load_file_family_config(ws)
        assert "email_file" in skip_families

    def test_pdf_not_in_skip_families(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_file_family_config_sheet(wb)
        _, skip_families, _ = fw_walk.load_file_family_config(ws)
        assert "pdf" not in skip_families

    def test_classify_uses_loaded_map_correctly(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_file_family_config_sheet(wb)
        family_map, _, _ = fw_walk.load_file_family_config(ws)
        assert fw_walk.classify_file_family("report.pdf", ".pdf", family_map) == "pdf"
        assert fw_walk.classify_file_family("data.xlsx", ".xlsx", family_map) == "spreadsheet"


# ===========================================================================
# 4b. likely_flags — third return value from load_file_family_config
# ===========================================================================

class TestLikelyFlags:
    """load_file_family_config returns correct likely_flags per family."""

    def _flags(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_file_family_config_sheet(wb)
        _, _, likely_flags = fw_walk.load_file_family_config(ws)
        return likely_flags

    def test_returns_dict(self):
        assert isinstance(self._flags(), dict)

    def test_all_10_families_have_flags(self):
        flags = self._flags()
        assert len(flags) == 10

    def test_pdf_likely_text_bearing(self):
        assert self._flags()["pdf"]["likely_text_bearing"] == "Y"

    def test_pdf_likely_document(self):
        assert self._flags()["pdf"]["likely_document"] == "Y"

    def test_pdf_likely_image_is_no(self):
        assert self._flags()["pdf"]["likely_image"] == "N"

    def test_pdf_likely_spreadsheet_is_no(self):
        assert self._flags()["pdf"]["likely_spreadsheet"] == "N"

    def test_image_likely_image(self):
        assert self._flags()["image"]["likely_image"] == "Y"

    def test_image_likely_text_bearing_is_no(self):
        assert self._flags()["image"]["likely_text_bearing"] == "N"

    def test_image_likely_document_is_no(self):
        assert self._flags()["image"]["likely_document"] == "N"

    def test_spreadsheet_likely_spreadsheet(self):
        assert self._flags()["spreadsheet"]["likely_spreadsheet"] == "Y"

    def test_spreadsheet_likely_text_bearing(self):
        assert self._flags()["spreadsheet"]["likely_text_bearing"] == "Y"

    def test_spreadsheet_likely_image_is_no(self):
        assert self._flags()["spreadsheet"]["likely_image"] == "N"

    def test_flags_keys_present(self):
        pdf_flags = self._flags()["pdf"]
        for key in ("likely_text_bearing", "likely_image", "likely_spreadsheet", "likely_document"):
            assert key in pdf_flags


# ===========================================================================
# 5. write_or_update_file_record — insert then update pipeline
# ===========================================================================

class TestInsertUpdatePipeline:
    def test_first_call_inserts(self):
        wb = _fresh_wb()
        action, _ = fw_walk.write_or_update_file_record(wb, _make_record(), "RUN001")
        assert action == "inserted"

    def test_second_call_same_path_updates(self):
        wb = _fresh_wb()
        fw_walk.write_or_update_file_record(wb, _make_record(), "RUN001")
        action, _ = fw_walk.write_or_update_file_record(wb, _make_record(), "RUN002")
        assert action == "updated"

    def test_two_different_paths_produce_two_rows(self):
        wb = _fresh_wb()
        fw_walk.write_or_update_file_record(wb, _make_record("C:/a.pdf", "a.pdf"), "R1")
        fw_walk.write_or_update_file_record(wb, _make_record("C:/b.pdf", "b.pdf"), "R1")
        ws = wb[fw_walk.SHEET_NAME]
        assert ws.max_row == 3  # header + 2 data rows

    def test_file_id_assigned_on_insert(self):
        wb = _fresh_wb()
        fw_walk.write_or_update_file_record(wb, _make_record(), "R1")
        ws = wb[fw_walk.SHEET_NAME]
        assert str(ws.cell(row=2, column=1).value).startswith("F")

    def test_run_id_last_seen_updated_on_second_walk(self):
        wb = _fresh_wb()
        fw_walk.write_or_update_file_record(wb, _make_record(), "RUN001")
        fw_walk.write_or_update_file_record(wb, _make_record(), "RUN002")
        ws = wb[fw_walk.SHEET_NAME]
        col = fw_walk._COL_INDEX["RunID_LastSeen"]
        assert ws.cell(row=2, column=col).value == "RUN002"

    def test_manual_review_status_not_overwritten(self):
        wb = _fresh_wb()
        fw_walk.write_or_update_file_record(wb, _make_record(), "R1")
        ws = wb[fw_walk.SHEET_NAME]
        mrc = fw_walk._COL_INDEX["ManualReviewStatus"]
        ws.cell(row=2, column=mrc, value="reviewed")
        fw_walk.write_or_update_file_record(wb, _make_record(), "R2")
        assert ws.cell(row=2, column=mrc).value == "reviewed"


# ===========================================================================
# 6. walk_files — end-to-end against a real temp directory
# ===========================================================================

class TestWalkFilesPipeline:
    @pytest.fixture
    def scan_dir(self, tmp_path):
        (tmp_path / "report.pdf").write_bytes(b"%PDF content")
        (tmp_path / "data.xlsx").write_bytes(b"PK fake xlsx")
        (tmp_path / "photo.jpg").write_bytes(b"\xff\xd8\xff fake jpeg")
        (tmp_path / "~$lock.docx").write_bytes(b"lock")   # office temp → skipped
        sub = tmp_path / "sub"
        sub.mkdir()
        (sub / "notes.txt").write_text("hello")
        return tmp_path

    def test_returns_stats_dict(self, scan_dir):
        wb = _fresh_wb()
        stats = fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        assert isinstance(stats, dict)
        for key in ("inserted", "updated", "skipped_files", "skipped_dirs", "errors"):
            assert key in stats

    def test_pdf_and_xlsx_counted_as_inserted(self, scan_dir):
        wb = _fresh_wb()
        stats = fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        assert stats["inserted"] >= 2

    def test_office_temp_file_is_skipped(self, scan_dir):
        wb = _fresh_wb()
        stats = fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        assert stats["skipped_files"] >= 1

    def test_mfi_sheet_created(self, scan_dir):
        wb = _fresh_wb()
        fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        assert fw_walk.SHEET_NAME in wb.sheetnames

    def test_mfi_has_data_rows(self, scan_dir):
        wb = _fresh_wb()
        fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        assert wb[fw_walk.SHEET_NAME].max_row >= 3  # header + ≥2 files

    def test_second_walk_updates_not_inserts(self, scan_dir):
        wb = _fresh_wb()
        fw_walk.walk_files([str(scan_dir)], "RUN1", "file_listed", wb)
        row_count = wb[fw_walk.SHEET_NAME].max_row
        stats2 = fw_walk.walk_files([str(scan_dir)], "RUN2", "file_listed", wb)
        assert stats2["updated"] > 0
        assert wb[fw_walk.SHEET_NAME].max_row == row_count  # no new rows

    def test_file_paths_use_forward_slashes(self, scan_dir):
        wb = _fresh_wb()
        fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        ws = wb[fw_walk.SHEET_NAME]
        col = fw_walk._COL_INDEX["FilePath"]
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                assert "\\" not in str(val)

    def test_jpg_gets_likely_image_y(self, scan_dir):
        wb = _fresh_wb()
        fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        ws = wb[fw_walk.SHEET_NAME]
        fp_col = fw_walk._COL_INDEX["FilePath"]
        li_col = fw_walk._COL_INDEX["LikelyImage"]
        for row in range(2, ws.max_row + 1):
            fp = str(ws.cell(row=row, column=fp_col).value or "")
            if fp.endswith(".jpg") or fp.endswith(".jpeg"):
                assert ws.cell(row=row, column=li_col).value == "Y"
                return
        pytest.skip("no .jpg file found in scan_dir fixture")

    def test_xlsx_gets_likely_spreadsheet_y(self, scan_dir):
        wb = _fresh_wb()
        fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        ws = wb[fw_walk.SHEET_NAME]
        fp_col = fw_walk._COL_INDEX["FilePath"]
        ls_col = fw_walk._COL_INDEX["LikelySpreadsheet"]
        for row in range(2, ws.max_row + 1):
            fp = str(ws.cell(row=row, column=fp_col).value or "")
            if fp.endswith(".xlsx"):
                assert ws.cell(row=row, column=ls_col).value == "Y"
                return
        pytest.skip("no .xlsx file found in scan_dir fixture")

    def test_pdf_gets_likely_document_y(self, scan_dir):
        wb = _fresh_wb()
        fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        ws = wb[fw_walk.SHEET_NAME]
        fp_col = fw_walk._COL_INDEX["FilePath"]
        ld_col = fw_walk._COL_INDEX["LikelyDocument"]
        for row in range(2, ws.max_row + 1):
            fp = str(ws.cell(row=row, column=fp_col).value or "")
            if fp.endswith(".pdf"):
                assert ws.cell(row=row, column=ld_col).value == "Y"
                return
        pytest.skip("no .pdf file found in scan_dir fixture")

    def test_pdf_gets_likely_text_bearing_y(self, scan_dir):
        wb = _fresh_wb()
        fw_walk.walk_files([str(scan_dir)], "TEST", "file_listed", wb)
        ws = wb[fw_walk.SHEET_NAME]
        fp_col = fw_walk._COL_INDEX["FilePath"]
        ltb_col = fw_walk._COL_INDEX["LikelyTextBearing"]
        for row in range(2, ws.max_row + 1):
            fp = str(ws.cell(row=row, column=fp_col).value or "")
            if fp.endswith(".pdf"):
                assert ws.cell(row=row, column=ltb_col).value == "Y"
                return
        pytest.skip("no .pdf file found in scan_dir fixture")


# ===========================================================================
# 7. Walk_Coverage + Walk_History logging pipeline
# ===========================================================================

class TestLoggingPipeline:
    def test_update_walk_coverage_appends_row(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_walk_coverage_sheet(wb)
        fw_walk.update_walk_coverage(ws, "R1", "C:/test", "file_listed", 5, 0, 2, 0)
        assert ws.max_row == 2  # header + 1

    def test_coverage_id_starts_at_WC00001(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_walk_coverage_sheet(wb)
        fw_walk.update_walk_coverage(ws, "R1", "C:/test", "file_listed", 0, 0, 0, 0)
        assert ws.cell(row=2, column=1).value == "WC00001"

    def test_coverage_id_sequential(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_walk_coverage_sheet(wb)
        fw_walk.update_walk_coverage(ws, "R1", "C:/a", "file_listed", 1, 0, 0, 0)
        fw_walk.update_walk_coverage(ws, "R1", "C:/b", "file_listed", 1, 0, 0, 0)
        assert ws.cell(row=2, column=1).value == "WC00001"
        assert ws.cell(row=3, column=1).value == "WC00002"

    def test_log_walk_run_writes_run_id(self):
        wb = _fresh_wb()
        fw_walk.log_walk_run(
            wb=wb, run_id="MY_RUN",
            started_at=datetime.now().isoformat(),
            scan_dirs=["C:/test"],
            processing_level="file_listed",
            stats={"inserted": 0, "updated": 0, "skipped_files": 0,
                   "skipped_dirs": 0, "errors": 0},
            overlap_action="proceed",
        )
        assert wb["Walk_History"].cell(row=2, column=1).value == "MY_RUN"

    def test_log_walk_run_scan_dirs_pipe_separated(self):
        wb = _fresh_wb()
        fw_walk.log_walk_run(
            wb=wb, run_id="R1",
            started_at=datetime.now().isoformat(),
            scan_dirs=["C:/dirA", "C:/dirB"],
            processing_level="file_listed",
            stats={"inserted": 0, "updated": 0, "skipped_files": 0,
                   "skipped_dirs": 0, "errors": 0},
            overlap_action="proceed",
        )
        col = fw_walk._WH_COL_INDEX["ScanDirs"]
        val = str(wb["Walk_History"].cell(row=2, column=col).value)
        assert "|" in val
        assert "C:/dirA" in val

    def test_second_log_appends_not_overwrites(self):
        wb = _fresh_wb()
        stats = {"inserted": 0, "updated": 0, "skipped_files": 0,
                 "skipped_dirs": 0, "errors": 0}
        fw_walk.log_walk_run(wb, "R1", datetime.now().isoformat(),
                             ["C:/t"], "file_listed", stats, "proceed")
        fw_walk.log_walk_run(wb, "R2", datetime.now().isoformat(),
                             ["C:/t"], "file_listed", stats, "proceed")
        assert wb["Walk_History"].max_row == 3  # header + 2 runs


# ===========================================================================
# 8. Sheet structure — freeze panes, headers, idempotency
# ===========================================================================

class TestSheetStructure:
    def test_mfi_freeze_panes(self):
        wb = _fresh_wb()
        fw_walk.ensure_master_file_inventory_sheet(wb)
        assert wb[fw_walk.SHEET_NAME].freeze_panes == "A2"

    def test_mfi_first_header_is_file_id(self):
        wb = _fresh_wb()
        ws = fw_walk.ensure_master_file_inventory_sheet(wb)
        assert ws.cell(row=1, column=1).value == "FileID"

    def test_mfi_idempotent(self):
        wb = _fresh_wb()
        fw_walk.ensure_master_file_inventory_sheet(wb)
        fw_walk.ensure_master_file_inventory_sheet(wb)
        assert wb.sheetnames.count(fw_walk.SHEET_NAME) == 1

    def test_file_family_config_freeze_panes(self):
        wb = _fresh_wb()
        fw_walk.ensure_file_family_config_sheet(wb)
        assert wb["FileFamily_Config"].freeze_panes == "A2"

    def test_walk_coverage_freeze_panes(self):
        wb = _fresh_wb()
        fw_walk.ensure_walk_coverage_sheet(wb)
        assert wb["Walk_Coverage"].freeze_panes == "A2"

    def test_walk_history_freeze_panes(self):
        wb = _fresh_wb()
        fw_walk.ensure_walk_history_sheet(wb)
        assert wb["Walk_History"].freeze_panes == "A2"

    def test_walk_coverage_idempotent(self):
        wb = _fresh_wb()
        fw_walk.ensure_walk_coverage_sheet(wb)
        fw_walk.ensure_walk_coverage_sheet(wb)
        assert wb.sheetnames.count("Walk_Coverage") == 1

    def test_walk_history_idempotent(self):
        wb = _fresh_wb()
        fw_walk.ensure_walk_history_sheet(wb)
        fw_walk.ensure_walk_history_sheet(wb)
        assert wb.sheetnames.count("Walk_History") == 1


# ===========================================================================
# 9. Round-trip: save → reload → data intact
# ===========================================================================

class TestRoundTrip:
    def test_all_four_sheets_survive_save_reload(self, tmp_path):
        with tempfile.TemporaryDirectory() as td:
            open(os.path.join(td, "doc.pdf"), "wb").write(b"%PDF")
            wb = _fresh_wb()
            stats = fw_walk.walk_files([td], "RT_RUN", "file_listed", wb)
            ws_cov = fw_walk.ensure_walk_coverage_sheet(wb)
            fw_walk.update_walk_coverage(
                ws_cov, "RT_RUN", td, "file_listed",
                stats["inserted"], stats["updated"], stats["skipped_files"], stats["errors"])
            fw_walk.log_walk_run(
                wb=wb, run_id="RT_RUN",
                started_at=datetime.now().isoformat(),
                scan_dirs=[td],
                processing_level="file_listed",
                stats=stats, overlap_action="proceed")
            xlsx_path = str(tmp_path / "rt.xlsx")
            wb.save(xlsx_path)

        wb2 = openpyxl.load_workbook(xlsx_path)
        assert fw_walk.SHEET_NAME in wb2.sheetnames
        assert "FileFamily_Config" in wb2.sheetnames
        assert "Walk_Coverage" in wb2.sheetnames
        assert "Walk_History" in wb2.sheetnames

    def test_mfi_data_survives_reload(self, tmp_path):
        with tempfile.TemporaryDirectory() as td:
            open(os.path.join(td, "doc.pdf"), "wb").write(b"%PDF")
            wb = _fresh_wb()
            fw_walk.walk_files([td], "RT2", "file_listed", wb)
            xlsx_path = str(tmp_path / "rt2.xlsx")
            wb.save(xlsx_path)

        wb2 = openpyxl.load_workbook(xlsx_path)
        assert wb2[fw_walk.SHEET_NAME].max_row >= 2
