"""
tests/test_fw_dirmap_grp4.py
Unit tests for fw_dirmap_grp4: open_or_create_workbook,
ensure_dir_inventory_sheet, write_dir_inventory_rows.
"""

import os
import sys

import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from fw_dirmap_grp4 import (
    open_or_create_workbook,
    ensure_dir_inventory_sheet,
    write_dir_inventory_rows,
    _HEADERS,
)


def _sample_record(n=1, **overrides):
    base = {
        "dir_id":           f"D{n:03d}",
        "scan_root":        "C:/TestRoot",
        "relative_dir":     "." if n == 1 else f"sub{n}",
        "dir_name":         "TestRoot" if n == 1 else f"sub{n}",
        "full_path":        "C:/TestRoot" if n == 1 else f"C:/TestRoot/sub{n}",
        "depth":            0 if n == 1 else 1,
        "file_count":       5,
        "subdir_count":     2,
        "processing_status": "not_scanned",
        "notes":            "",
    }
    base.update(overrides)
    return base


# ---------------------------------------------------------------------------
# open_or_create_workbook
# ---------------------------------------------------------------------------

class TestOpenOrCreateWorkbook:
    def test_creates_new_workbook_when_missing(self, tmp_path):
        path = str(tmp_path / "new.xlsx")
        wb = open_or_create_workbook(path)
        assert wb is not None

    def test_new_workbook_has_no_default_sheet(self, tmp_path):
        path = str(tmp_path / "new.xlsx")
        wb = open_or_create_workbook(path)
        # Default "Sheet" removed
        assert "Sheet" not in wb.sheetnames

    def test_opens_existing_workbook(self, tmp_path):
        path = str(tmp_path / "existing.xlsx")
        existing = openpyxl.Workbook()
        existing.create_sheet("MySheet")
        existing.save(path)
        wb = open_or_create_workbook(path)
        assert "MySheet" in wb.sheetnames


# ---------------------------------------------------------------------------
# ensure_dir_inventory_sheet
# ---------------------------------------------------------------------------

class TestEnsureDirInventorySheet:
    def test_creates_sheet_when_absent(self):
        wb = openpyxl.Workbook()
        ws = ensure_dir_inventory_sheet(wb)
        assert "Dir_Inventory" in wb.sheetnames

    def test_returns_existing_sheet_unchanged(self):
        wb = openpyxl.Workbook()
        ws1 = ensure_dir_inventory_sheet(wb)
        ws1["A1"] = "SENTINEL"
        ws2 = ensure_dir_inventory_sheet(wb)
        assert ws2["A1"].value == "SENTINEL"

    def test_header_row_has_correct_labels(self):
        wb = openpyxl.Workbook()
        ws = ensure_dir_inventory_sheet(wb)
        headers = [ws.cell(row=1, column=i+1).value for i in range(len(_HEADERS))]
        assert headers == _HEADERS

    def test_freeze_panes_set(self):
        wb = openpyxl.Workbook()
        ws = ensure_dir_inventory_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_auto_filter_set(self):
        wb = openpyxl.Workbook()
        ws = ensure_dir_inventory_sheet(wb)
        assert ws.auto_filter.ref is not None


# ---------------------------------------------------------------------------
# write_dir_inventory_rows
# ---------------------------------------------------------------------------

class TestWriteDirInventoryRows:
    def test_rows_appended(self):
        wb = openpyxl.Workbook()
        records = [_sample_record(1), _sample_record(2)]
        write_dir_inventory_rows(wb, records)
        ws = wb["Dir_Inventory"]
        # Row 1 = header, rows 2 and 3 = data
        assert ws.max_row == 3

    def test_dir_id_written_correctly(self):
        wb = openpyxl.Workbook()
        write_dir_inventory_rows(wb, [_sample_record(1)])
        ws = wb["Dir_Inventory"]
        assert ws["A2"].value == "D001"

    def test_depth_indents_dir_name(self):
        wb = openpyxl.Workbook()
        record = _sample_record(2, depth=2, dir_name="deep")
        write_dir_inventory_rows(wb, [record])
        ws = wb["Dir_Inventory"]
        dir_name_col = _HEADERS.index("DirName") + 1
        cell_val = ws.cell(row=2, column=dir_name_col).value
        assert cell_val.startswith("    ")  # 2 levels * 2 spaces

    def test_depth_zero_no_indent(self):
        wb = openpyxl.Workbook()
        record = _sample_record(1, depth=0, dir_name="Root")
        write_dir_inventory_rows(wb, [record])
        ws = wb["Dir_Inventory"]
        dir_name_col = _HEADERS.index("DirName") + 1
        cell_val = ws.cell(row=2, column=dir_name_col).value
        assert cell_val == "Root"

    def test_outline_level_set(self):
        wb = openpyxl.Workbook()
        record = _sample_record(2, depth=3)
        write_dir_inventory_rows(wb, [record])
        ws = wb["Dir_Inventory"]
        assert ws.row_dimensions[2].outline_level == 3

    def test_status_cell_has_fill(self):
        wb = openpyxl.Workbook()
        write_dir_inventory_rows(wb, [_sample_record(1, processing_status="classified")])
        ws = wb["Dir_Inventory"]
        status_col = _HEADERS.index("ProcessingStatus") + 1
        cell = ws.cell(row=2, column=status_col)
        assert cell.fill.fill_type == "solid"

    def test_default_status_is_not_scanned(self):
        wb = openpyxl.Workbook()
        record = _sample_record(1)
        del record["processing_status"]
        write_dir_inventory_rows(wb, [record])
        ws = wb["Dir_Inventory"]
        status_col = _HEADERS.index("ProcessingStatus") + 1
        assert ws.cell(row=2, column=status_col).value == "not_scanned"

    def test_multiple_records_sequential(self):
        wb = openpyxl.Workbook()
        records = [_sample_record(i) for i in range(1, 6)]
        write_dir_inventory_rows(wb, records)
        ws = wb["Dir_Inventory"]
        assert ws.max_row == 6  # 1 header + 5 data

    def test_idempotent_second_call_appends(self):
        """Calling write twice should append, not overwrite."""
        wb = openpyxl.Workbook()
        write_dir_inventory_rows(wb, [_sample_record(1)])
        write_dir_inventory_rows(wb, [_sample_record(2)])
        ws = wb["Dir_Inventory"]
        assert ws.max_row == 3  # header + 2 data rows
