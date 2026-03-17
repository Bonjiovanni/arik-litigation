"""
tests/test_fw_triage_grp2.py
-----------------------------
Unit tests for fw_triage_grp2.py:
    get_triageable_rows, write_triage_results, mark_row_triaged
"""

import pytest
import openpyxl

from fw_walk_grp2 import _COL_INDEX
import fw_triage_grp2 as grp2


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_mfi_ws():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master_File_Inventory"
    for name, col_idx in _COL_INDEX.items():
        ws.cell(row=1, column=col_idx, value=name)
    return ws


def _write_row(ws, row_num: int, **kwargs):
    for col_name, value in kwargs.items():
        ws.cell(row=row_num, column=_COL_INDEX[col_name], value=value)


# ---------------------------------------------------------------------------
# TestGetTriageableRows
# ---------------------------------------------------------------------------

class TestGetTriageableRows:

    def test_returns_classified_rows(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", FilePath="C:/a.pdf", FileFamily="pdf",
                   DocType="LTC_Claim", DocTypeConfidence="medium",
                   MoneyDetected="Y", DateDetected="N", KeywordHits="trust",
                   LikelyTextBearing="Y", NeedsOCR="N",
                   ProcessingStatus="classified", DocSubtype="")
        rows = grp2.get_triageable_rows(ws)
        assert len(rows) == 1
        assert rows[0][0] == 2
        assert rows[0][1]["FileID"] == "F001"

    def test_skips_file_listed_rows(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", ProcessingStatus="file_listed",
                   FilePath="C:/a.pdf", FileFamily="pdf", DocType="",
                   DocTypeConfidence="", MoneyDetected="N", DateDetected="N",
                   KeywordHits="", LikelyTextBearing="", NeedsOCR="N", DocSubtype="")
        rows = grp2.get_triageable_rows(ws)
        assert rows == []

    def test_skips_triaged_rows(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", ProcessingStatus="triaged",
                   FilePath="C:/a.pdf", FileFamily="pdf", DocType="Invoice",
                   DocTypeConfidence="medium", MoneyDetected="Y", DateDetected="Y",
                   KeywordHits="", LikelyTextBearing="Y", NeedsOCR="N", DocSubtype="")
        rows = grp2.get_triageable_rows(ws)
        assert rows == []

    def test_stops_at_empty_file_id(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", ProcessingStatus="classified",
                   FilePath="C:/a.pdf", FileFamily="pdf", DocType="Invoice",
                   DocTypeConfidence="low", MoneyDetected="N", DateDetected="N",
                   KeywordHits="", LikelyTextBearing="Y", NeedsOCR="N", DocSubtype="")
        # Row 3 empty — stop
        rows = grp2.get_triageable_rows(ws)
        assert len(rows) == 1

    def test_record_has_triage_fields(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", ProcessingStatus="classified",
                   FilePath="C:/a.pdf", FileFamily="pdf", DocType="LTC_Claim",
                   DocTypeConfidence="high", MoneyDetected="Y", DateDetected="Y",
                   KeywordHits="trust", LikelyTextBearing="Y", NeedsOCR="N", DocSubtype="")
        rows = grp2.get_triageable_rows(ws)
        record = rows[0][1]
        for field in ("FileID", "FilePath", "FileFamily", "DocType", "DocTypeConfidence",
                      "MoneyDetected", "DateDetected", "KeywordHits", "NeedsOCR"):
            assert field in record

    def test_multiple_classified_rows_returned(self):
        ws = _make_mfi_ws()
        for i, fid in enumerate(["F001", "F002", "F003"], start=2):
            _write_row(ws, i, FileID=fid, ProcessingStatus="classified",
                       FilePath=f"C:/{fid}.pdf", FileFamily="pdf", DocType="Invoice",
                       DocTypeConfidence="low", MoneyDetected="N", DateDetected="N",
                       KeywordHits="", LikelyTextBearing="Y", NeedsOCR="N", DocSubtype="")
        rows = grp2.get_triageable_rows(ws)
        assert len(rows) == 3

    def test_mixed_rows_only_classified_returned(self):
        ws = _make_mfi_ws()
        _write_row(ws, 2, FileID="F001", ProcessingStatus="classified",
                   FilePath="C:/a.pdf", FileFamily="pdf", DocType="Invoice",
                   DocTypeConfidence="low", MoneyDetected="N", DateDetected="N",
                   KeywordHits="", LikelyTextBearing="Y", NeedsOCR="N", DocSubtype="")
        _write_row(ws, 3, FileID="F002", ProcessingStatus="file_listed",
                   FilePath="C:/b.pdf", FileFamily="pdf", DocType="",
                   DocTypeConfidence="", MoneyDetected="N", DateDetected="N",
                   KeywordHits="", LikelyTextBearing="", NeedsOCR="N", DocSubtype="")
        _write_row(ws, 4, FileID="F003", ProcessingStatus="classified",
                   FilePath="C:/c.pdf", FileFamily="pdf", DocType="Trust_Document",
                   DocTypeConfidence="medium", MoneyDetected="Y", DateDetected="N",
                   KeywordHits="trust", LikelyTextBearing="Y", NeedsOCR="N", DocSubtype="")
        rows = grp2.get_triageable_rows(ws)
        ids = [r[1]["FileID"] for r in rows]
        assert ids == ["F001", "F003"]


# ---------------------------------------------------------------------------
# TestWriteTriageResults
# ---------------------------------------------------------------------------

class TestWriteTriageResults:

    def test_writes_triage_score(self):
        ws = _make_mfi_ws()
        grp2.write_triage_results(ws, 2, {"TriageScore": 65})
        assert ws.cell(row=2, column=_COL_INDEX["TriageScore"]).value == 65

    def test_writes_triage_band(self):
        ws = _make_mfi_ws()
        grp2.write_triage_results(ws, 2, {"TriageBand": "High"})
        assert ws.cell(row=2, column=_COL_INDEX["TriageBand"]).value == "High"

    def test_writes_reason_flagged(self):
        ws = _make_mfi_ws()
        grp2.write_triage_results(ws, 2, {"ReasonFlagged": "LTC_Claim;MoneyDetected"})
        assert ws.cell(row=2, column=_COL_INDEX["ReasonFlagged"]).value == "LTC_Claim;MoneyDetected"

    def test_writes_next_step(self):
        ws = _make_mfi_ws()
        grp2.write_triage_results(ws, 2, {"NextStep": "Priority manual review"})
        assert ws.cell(row=2, column=_COL_INDEX["NextStep"]).value == "Priority manual review"

    def test_does_not_overwrite_manual_review_status(self):
        ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["ManualReviewStatus"], value="needs_review")
        grp2.write_triage_results(ws, 2, {"ManualReviewStatus": "should_not_write"})
        assert ws.cell(row=2, column=_COL_INDEX["ManualReviewStatus"]).value == "needs_review"

    def test_does_not_overwrite_keep_for_case(self):
        ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["KeepForCase"], value="Y")
        grp2.write_triage_results(ws, 2, {"KeepForCase": "N"})
        assert ws.cell(row=2, column=_COL_INDEX["KeepForCase"]).value == "Y"

    def test_unknown_key_ignored(self):
        ws = _make_mfi_ws()
        grp2.write_triage_results(ws, 2, {"NotARealColumn": "value"})

    def test_writes_all_four_results(self):
        ws = _make_mfi_ws()
        grp2.write_triage_results(ws, 2, {
            "TriageScore": 42,
            "TriageBand": "Medium",
            "ReasonFlagged": "Trust_Document",
            "NextStep": "Manual review",
        })
        assert ws.cell(row=2, column=_COL_INDEX["TriageScore"]).value == 42
        assert ws.cell(row=2, column=_COL_INDEX["TriageBand"]).value == "Medium"
        assert ws.cell(row=2, column=_COL_INDEX["ReasonFlagged"]).value == "Trust_Document"
        assert ws.cell(row=2, column=_COL_INDEX["NextStep"]).value == "Manual review"


# ---------------------------------------------------------------------------
# TestMarkRowTriaged
# ---------------------------------------------------------------------------

class TestMarkRowTriaged:

    def test_sets_triaged_status(self):
        ws = _make_mfi_ws()
        grp2.mark_row_triaged(ws, 2)
        assert ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"]).value == "triaged"

    def test_overwrites_classified(self):
        ws = _make_mfi_ws()
        ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"], value="classified")
        grp2.mark_row_triaged(ws, 2)
        assert ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"]).value == "triaged"

    def test_marks_specific_row(self):
        ws = _make_mfi_ws()
        grp2.mark_row_triaged(ws, 7)
        assert ws.cell(row=7, column=_COL_INDEX["ProcessingStatus"]).value == "triaged"
        assert ws.cell(row=2, column=_COL_INDEX["ProcessingStatus"]).value is None
