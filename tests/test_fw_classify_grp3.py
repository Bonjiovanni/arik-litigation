"""
tests/test_fw_classify_grp3.py
-------------------------------
Unit tests for fw_classify_grp3.py:
    ensure_keywords_config_sheet, load_keywords,
    infer_likely_text_bearing, infer_needs_ocr, classify_doc_type
"""

import pytest
import openpyxl

import fw_classify_grp3 as grp3


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fresh_wb():
    wb = openpyxl.Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    return wb


# ---------------------------------------------------------------------------
# TestEnsureKeywordsConfigSheet
# ---------------------------------------------------------------------------

class TestEnsureKeywordsConfigSheet:

    def test_creates_sheet_when_absent(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        assert grp3.KEYWORDS_CONFIG_SHEET in wb.sheetnames

    def test_returns_worksheet(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        assert ws is not None
        assert ws.title == grp3.KEYWORDS_CONFIG_SHEET

    def test_idempotent_returns_existing(self):
        wb = _fresh_wb()
        ws1 = grp3.ensure_keywords_config_sheet(wb)
        ws1.cell(row=2, column=1, value="MARKER")
        ws2 = grp3.ensure_keywords_config_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "MARKER"

    def test_seeds_default_keywords(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        keywords_in_sheet = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        assert len(keywords_in_sheet) == len(grp3._DEFAULT_KEYWORDS)

    def test_default_includes_trust(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        keywords = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "trust" in keywords

    def test_default_includes_ltc(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        keywords = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "LTC" in keywords

    def test_headers_in_row_1(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        assert ws.cell(row=1, column=1).value == "Keyword"
        assert ws.cell(row=1, column=2).value == "Active"

    def test_freeze_panes_set(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_active_column_is_Y(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        # All default keywords should have Active = "Y"
        for row in range(2, ws.max_row + 1):
            kw = ws.cell(row=row, column=1).value
            if kw:
                assert ws.cell(row=row, column=2).value == "Y", f"Expected Y for {kw}"


# ---------------------------------------------------------------------------
# TestLoadKeywords
# ---------------------------------------------------------------------------

class TestLoadKeywords:

    def test_returns_active_keywords(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        keywords = grp3.load_keywords(ws)
        assert "trust" in keywords
        assert len(keywords) > 0

    def test_skips_inactive_rows(self):
        wb = _fresh_wb()
        ws = wb.create_sheet("Keywords_Config")
        ws.cell(row=1, column=1, value="Keyword")
        ws.cell(row=1, column=2, value="Active")
        ws.cell(row=2, column=1, value="active_kw")
        ws.cell(row=2, column=2, value="Y")
        ws.cell(row=3, column=1, value="inactive_kw")
        ws.cell(row=3, column=2, value="N")
        keywords = grp3.load_keywords(ws)
        assert "active_kw" in keywords
        assert "inactive_kw" not in keywords

    def test_skips_blank_keyword_rows(self):
        wb = _fresh_wb()
        ws = wb.create_sheet("Keywords_Config")
        ws.cell(row=1, column=1, value="Keyword")
        ws.cell(row=1, column=2, value="Active")
        ws.cell(row=2, column=1, value=None)
        ws.cell(row=2, column=2, value="Y")
        keywords = grp3.load_keywords(ws)
        assert keywords == []

    def test_case_insensitive_active_check(self):
        wb = _fresh_wb()
        ws = wb.create_sheet("Keywords_Config")
        ws.cell(row=1, column=1, value="Keyword")
        ws.cell(row=1, column=2, value="Active")
        ws.cell(row=2, column=1, value="mykw")
        ws.cell(row=2, column=2, value="y")  # lowercase y
        keywords = grp3.load_keywords(ws)
        assert "mykw" in keywords

    def test_all_defaults_loaded(self):
        wb = _fresh_wb()
        ws = grp3.ensure_keywords_config_sheet(wb)
        keywords = grp3.load_keywords(ws)
        assert len(keywords) == len(grp3._DEFAULT_KEYWORDS)


# ---------------------------------------------------------------------------
# TestInferLikelyTextBearing
# ---------------------------------------------------------------------------

class TestInferLikelyTextBearing:

    def test_text_file_always_Y(self):
        assert grp3.infer_likely_text_bearing("text_file", "", False) == "Y"

    def test_word_doc_always_Y(self):
        assert grp3.infer_likely_text_bearing("word_doc", "some text", True) == "Y"

    def test_spreadsheet_always_Y(self):
        assert grp3.infer_likely_text_bearing("spreadsheet", "", False) == "Y"

    def test_pdf_with_text_is_Y(self):
        assert grp3.infer_likely_text_bearing("pdf", "some text", True) == "Y"

    def test_pdf_no_text_is_N(self):
        assert grp3.infer_likely_text_bearing("pdf", "", False) == "N"

    def test_pdf_empty_sample_even_if_succeeded_is_N(self):
        assert grp3.infer_likely_text_bearing("pdf", "", True) == "N"

    def test_image_is_N(self):
        assert grp3.infer_likely_text_bearing("image", "", False) == "N"

    def test_audio_is_N(self):
        assert grp3.infer_likely_text_bearing("audio", "", False) == "N"

    def test_video_is_N(self):
        assert grp3.infer_likely_text_bearing("video", "", False) == "N"

    def test_email_file_is_Y(self):
        assert grp3.infer_likely_text_bearing("email_file", "", False) == "Y"

    def test_unknown_family_is_empty(self):
        assert grp3.infer_likely_text_bearing("archive", "", False) == ""

    def test_case_insensitive_family(self):
        assert grp3.infer_likely_text_bearing("PDF", "text", True) == "Y"


# ---------------------------------------------------------------------------
# TestInferNeedsOCR
# ---------------------------------------------------------------------------

class TestInferNeedsOCR:

    def test_pdf_with_text_is_N(self):
        assert grp3.infer_needs_ocr("pdf", "some text", True) == "N"

    def test_pdf_no_text_is_Y(self):
        assert grp3.infer_needs_ocr("pdf", "", False) == "Y"

    def test_image_always_Y(self):
        assert grp3.infer_needs_ocr("image", "", False) == "Y"

    def test_text_file_is_N(self):
        assert grp3.infer_needs_ocr("text_file", "text", True) == "N"

    def test_word_doc_is_N(self):
        assert grp3.infer_needs_ocr("word_doc", "text", True) == "N"

    def test_spreadsheet_is_N(self):
        assert grp3.infer_needs_ocr("spreadsheet", "", False) == "N"

    def test_email_file_is_N(self):
        assert grp3.infer_needs_ocr("email_file", "text", True) == "N"

    def test_audio_is_N(self):
        assert grp3.infer_needs_ocr("audio", "", False) == "N"

    def test_archive_is_N(self):
        assert grp3.infer_needs_ocr("archive", "", False) == "N"

    def test_unknown_family_is_empty(self):
        assert grp3.infer_needs_ocr("other_thing", "", False) == ""


# ---------------------------------------------------------------------------
# TestClassifyDocType
# ---------------------------------------------------------------------------

class TestClassifyDocType:

    def test_invoice_rule(self):
        result = grp3.classify_doc_type("This invoice is due", "", "pdf")
        assert result["DocType"] == "Invoice"
        assert result["DocTypeConfidence"] == "medium"

    def test_ltc_claim_with_explicit_ltc(self):
        result = grp3.classify_doc_type("LTC claim filed", "claim;LTC", "pdf")
        assert result["DocType"] == "LTC_Claim"

    def test_ltc_claim_long_term_care(self):
        result = grp3.classify_doc_type("long-term care claim submitted", "", "pdf")
        assert result["DocType"] == "LTC_Claim"
        assert result["DocTypeConfidence"] == "medium"

    def test_ltc_claim_benefit_plus_claim(self):
        result = grp3.classify_doc_type("benefit and claim information", "", "pdf")
        assert result["DocType"] == "LTC_Claim"
        assert result["DocTypeConfidence"] == "low"

    def test_trust_document_with_corroborating_entity(self):
        result = grp3.classify_doc_type("trust beneficiary distribution", "trust;beneficiary", "pdf")
        assert result["DocType"] == "Trust_Document"
        assert result["DocTypeConfidence"] == "medium"

    def test_trust_document_trustee(self):
        result = grp3.classify_doc_type("The trustee of the trust", "", "pdf")
        assert result["DocType"] == "Trust_Document"
        assert result["DocTypeConfidence"] == "medium"

    def test_trust_document_alone_low_confidence(self):
        result = grp3.classify_doc_type("trust only text here", "", "pdf")
        assert result["DocType"] == "Trust_Document"
        assert result["DocTypeConfidence"] == "low"

    def test_amendment_rule(self):
        result = grp3.classify_doc_type("First amendment to trust", "", "pdf")
        assert result["DocType"] == "Trust_Document"
        assert result["DocSubtype"] == "Amendment"

    def test_restatement_rule(self):
        result = grp3.classify_doc_type("Restatement of trust", "", "pdf")
        assert result["DocType"] == "Trust_Document"
        assert result["DocSubtype"] == "Amendment"

    def test_insurance_policy_premium(self):
        result = grp3.classify_doc_type("monthly premium due", "", "pdf")
        assert result["DocType"] == "Insurance_Policy"

    def test_no_match_returns_empty(self):
        result = grp3.classify_doc_type("random unrelated text", "", "pdf")
        assert result["DocType"] == ""
        assert result["DocSubtype"] == ""
        assert result["DocTypeConfidence"] == ""

    def test_invoice_takes_priority_over_trust(self):
        # invoice rule fires first
        result = grp3.classify_doc_type("trust invoice billing", "", "pdf")
        assert result["DocType"] == "Invoice"

    def test_empty_inputs(self):
        result = grp3.classify_doc_type("", "", "pdf")
        assert result["DocType"] == ""

    def test_none_inputs(self):
        result = grp3.classify_doc_type(None, None, "pdf")
        assert result["DocType"] == ""

    def test_returns_all_required_keys(self):
        result = grp3.classify_doc_type("some text", "", "pdf")
        assert "DocType" in result
        assert "DocSubtype" in result
        assert "DocTypeConfidence" in result
