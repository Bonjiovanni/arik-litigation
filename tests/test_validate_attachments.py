"""
Tests for email_pipeline/validate_attachments.py

Pure-function tests only — no real Excel index or Aid4Mail exports required.
"""

import hashlib
import json
import os
import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

import validate_attachments as va


# ---------------------------------------------------------------------------
# TestNormalizeName
# ---------------------------------------------------------------------------
class TestNormalizeName:
    def test_lowercase_on_windows(self):
        if os.name == "nt":
            assert va.normalize_name("File.PDF") == "file.pdf"
        else:
            assert va.normalize_name("File.PDF") == "File.PDF"

    def test_already_lowercase_unchanged(self):
        assert va.normalize_name("already_lower.pdf") == "already_lower.pdf"

    def test_empty_string(self):
        assert va.normalize_name("") == ""

    def test_mixed_case(self):
        name = "Invoice_123.PDF"
        result = va.normalize_name(name)
        if os.name == "nt":
            assert result == name.lower()
        else:
            assert result == name


# ---------------------------------------------------------------------------
# TestHashFile
# ---------------------------------------------------------------------------
class TestHashFile:
    def test_known_content(self, tmp_path):
        content = b"hello world"
        f = tmp_path / "test.bin"
        f.write_bytes(content)
        expected = hashlib.sha256(content).hexdigest()
        assert va.hash_file(f) == expected

    def test_empty_file(self, tmp_path):
        f = tmp_path / "empty.bin"
        f.write_bytes(b"")
        expected = hashlib.sha256(b"").hexdigest()
        assert va.hash_file(f) == expected

    def test_large_content_is_hex(self, tmp_path):
        content = b"x" * 200_000
        f = tmp_path / "large.bin"
        f.write_bytes(content)
        result = va.hash_file(f)
        assert len(result) == 64
        assert all(c in "0123456789abcdef" for c in result)


# ---------------------------------------------------------------------------
# TestLoadJson
# ---------------------------------------------------------------------------
class TestLoadJson:
    def _write(self, tmp_path, data, name="test.json"):
        f = tmp_path / name
        f.write_text(json.dumps(data), encoding="utf-8")
        return str(f)

    def test_list_input_returned_as_is(self, tmp_path):
        data = [{"id": 1}, {"id": 2}]
        result = va.load_json(self._write(tmp_path, data))
        assert result == data

    def test_dict_with_emails_key(self, tmp_path):
        data = {"emails": [{"id": 1}], "meta": "ignored"}
        result = va.load_json(self._write(tmp_path, data))
        assert result == [{"id": 1}]

    def test_dict_with_messages_key(self, tmp_path):
        data = {"messages": [{"id": 2}]}
        result = va.load_json(self._write(tmp_path, data))
        assert result == [{"id": 2}]

    def test_dict_with_records_key(self, tmp_path):
        data = {"records": [{"id": 3}]}
        result = va.load_json(self._write(tmp_path, data))
        assert result == [{"id": 3}]

    def test_dict_with_data_key(self, tmp_path):
        data = {"data": [{"id": 4}]}
        result = va.load_json(self._write(tmp_path, data))
        assert result == [{"id": 4}]

    def test_dict_unknown_keys_returns_first_list_value(self, tmp_path):
        data = {"unknown_key": [{"id": 5}], "other": "string"}
        result = va.load_json(self._write(tmp_path, data))
        assert result == [{"id": 5}]

    def test_dict_no_list_values_returns_dict_in_list(self, tmp_path):
        data = {"key1": "val1", "key2": "val2"}
        result = va.load_json(self._write(tmp_path, data))
        assert result == [data]

    def test_empty_list(self, tmp_path):
        result = va.load_json(self._write(tmp_path, []))
        assert result == []


# ---------------------------------------------------------------------------
# TestScanFolder
# ---------------------------------------------------------------------------
class TestScanFolder:
    def test_existing_folder_returns_files(self, tmp_path):
        (tmp_path / "file1.pdf").write_bytes(b"x")
        (tmp_path / "file2.txt").write_bytes(b"y")
        result = va.scan_folder(str(tmp_path))
        assert len(result) == 2

    def test_basename_keys(self, tmp_path):
        (tmp_path / "Invoice_001.pdf").write_bytes(b"x")
        result = va.scan_folder(str(tmp_path))
        expected_key = va.normalize_name("Invoice_001.pdf")
        assert expected_key in result

    def test_nonexistent_folder_returns_empty(self, capsys):
        result = va.scan_folder("/nonexistent/path/xyzabc99999")
        assert len(result) == 0
        captured = capsys.readouterr()
        assert "WARNING" in captured.out

    def test_recursive_finds_nested_files(self, tmp_path):
        subdir = tmp_path / "subdir"
        subdir.mkdir()
        (subdir / "nested.pdf").write_bytes(b"x")
        result = va.scan_folder(str(tmp_path))
        expected_key = va.normalize_name("nested.pdf")
        assert expected_key in result

    def test_directories_not_included(self, tmp_path):
        subdir = tmp_path / "asubdir"
        subdir.mkdir()
        result = va.scan_folder(str(tmp_path))
        # The subdir itself should not appear; only files
        assert va.normalize_name("asubdir") not in result


# ---------------------------------------------------------------------------
# TestWriteDfToSheet
# ---------------------------------------------------------------------------
class TestWriteDfToSheet:
    def _fresh_wb(self):
        wb = Workbook()
        wb.remove(wb.active)
        return wb

    def test_empty_df_writes_no_issues(self):
        wb = self._fresh_wb()
        va.write_df_to_sheet(wb, "Empty", pd.DataFrame())
        ws = wb["Empty"]
        assert ws.cell(1, 1).value == "No issues found."

    def test_nonempty_df_writes_header_row(self):
        wb = self._fresh_wb()
        df = pd.DataFrame({"Name": ["file.pdf"], "Size": [1024]})
        va.write_df_to_sheet(wb, "Data", df)
        ws = wb["Data"]
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(1, 2).value == "Size"

    def test_nonempty_df_writes_data_rows(self):
        wb = self._fresh_wb()
        df = pd.DataFrame({"Name": ["file.pdf"], "Size": [1024]})
        va.write_df_to_sheet(wb, "Data", df)
        ws = wb["Data"]
        assert ws.cell(2, 1).value == "file.pdf"
        assert ws.cell(2, 2).value == 1024

    def test_header_row_has_fill(self):
        wb = self._fresh_wb()
        df = pd.DataFrame({"Col": ["val"]})
        va.write_df_to_sheet(wb, "Sheet", df)
        ws = wb["Sheet"]
        fill = ws.cell(1, 1).fill
        assert fill.fgColor.rgb != "00000000"

    def test_column_widths_set(self):
        wb = self._fresh_wb()
        df = pd.DataFrame({"LongColumnName": ["some value"], "B": ["x"]})
        va.write_df_to_sheet(wb, "Sheet", df)
        ws = wb["Sheet"]
        # Column A should have width > 0
        assert ws.column_dimensions["A"].width > 0
