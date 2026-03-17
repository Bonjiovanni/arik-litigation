"""
tests/test_fw_triage_grp1.py
-----------------------------
Unit tests for fw_triage_grp1.py:
    ensure_triage_config_sheet, ensure_triage_bands_sheet,
    load_triage_config, load_triage_bands,
    score_record, get_triage_band, get_reason_flagged, get_next_step
"""

import pytest
import openpyxl

import fw_triage_grp1 as grp1


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fresh_wb():
    wb = openpyxl.Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    return wb


def _default_config():
    """Build the default config dict directly without a worksheet."""
    wb = _fresh_wb()
    ws = grp1.ensure_triage_config_sheet(wb)
    return grp1.load_triage_config(ws)


def _default_bands():
    """Build the default bands list directly without a worksheet."""
    wb = _fresh_wb()
    ws = grp1.ensure_triage_bands_sheet(wb)
    return grp1.load_triage_bands(ws)


def _record(**kwargs):
    base = {
        "DocType": "", "DocTypeConfidence": "",
        "MoneyDetected": "N", "DateDetected": "N",
        "KeywordHits": "", "NeedsOCR": "N",
    }
    base.update(kwargs)
    return base


# ---------------------------------------------------------------------------
# TestEnsureTriageConfigSheet
# ---------------------------------------------------------------------------

class TestEnsureTriageConfigSheet:

    def test_creates_sheet(self):
        wb = _fresh_wb()
        grp1.ensure_triage_config_sheet(wb)
        assert grp1.TRIAGE_CONFIG_SHEET in wb.sheetnames

    def test_returns_worksheet(self):
        wb = _fresh_wb()
        ws = grp1.ensure_triage_config_sheet(wb)
        assert ws.title == grp1.TRIAGE_CONFIG_SHEET

    def test_idempotent(self):
        wb = _fresh_wb()
        ws1 = grp1.ensure_triage_config_sheet(wb)
        ws1.cell(row=2, column=1, value="MARKER")
        ws2 = grp1.ensure_triage_config_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "MARKER"

    def test_headers_row_1(self):
        wb = _fresh_wb()
        ws = grp1.ensure_triage_config_sheet(wb)
        assert ws.cell(row=1, column=1).value == "SignalType"
        assert ws.cell(row=1, column=2).value == "SignalValue"
        assert ws.cell(row=1, column=3).value == "ScorePoints"

    def test_seeds_default_rows(self):
        wb = _fresh_wb()
        ws = grp1.ensure_triage_config_sheet(wb)
        data_rows = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)
                     if ws.cell(row=r, column=1).value]
        assert len(data_rows) == len(grp1._DEFAULT_TRIAGE_CONFIG)

    def test_freeze_panes(self):
        wb = _fresh_wb()
        ws = grp1.ensure_triage_config_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_ltc_claim_present(self):
        wb = _fresh_wb()
        ws = grp1.ensure_triage_config_sheet(wb)
        sig_vals = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert "LTC_Claim" in sig_vals


# ---------------------------------------------------------------------------
# TestEnsureTriageBandsSheet
# ---------------------------------------------------------------------------

class TestEnsureTriageBandsSheet:

    def test_creates_sheet(self):
        wb = _fresh_wb()
        grp1.ensure_triage_bands_sheet(wb)
        assert grp1.TRIAGE_BANDS_SHEET in wb.sheetnames

    def test_idempotent(self):
        wb = _fresh_wb()
        ws1 = grp1.ensure_triage_bands_sheet(wb)
        ws1.cell(row=2, column=1, value="MARKER")
        ws2 = grp1.ensure_triage_bands_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "MARKER"

    def test_headers_row_1(self):
        wb = _fresh_wb()
        ws = grp1.ensure_triage_bands_sheet(wb)
        assert ws.cell(row=1, column=1).value == "Band"
        assert ws.cell(row=1, column=2).value == "MinScore"
        assert ws.cell(row=1, column=3).value == "NextStep"

    def test_seeds_four_default_bands(self):
        wb = _fresh_wb()
        ws = grp1.ensure_triage_bands_sheet(wb)
        data_rows = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)
                     if ws.cell(row=r, column=1).value]
        assert len(data_rows) == len(grp1._DEFAULT_TRIAGE_BANDS)

    def test_freeze_panes(self):
        wb = _fresh_wb()
        ws = grp1.ensure_triage_bands_sheet(wb)
        assert ws.freeze_panes == "A2"


# ---------------------------------------------------------------------------
# TestLoadTriageConfig
# ---------------------------------------------------------------------------

class TestLoadTriageConfig:

    def test_returns_nested_dict(self):
        config = _default_config()
        assert isinstance(config, dict)
        assert "DocType" in config
        assert "DocTypeConfidence" in config
        assert "Signal" in config

    def test_doctype_ltc_claim_score(self):
        config = _default_config()
        assert config["DocType"]["LTC_Claim"] == 30

    def test_doctype_invoice_score(self):
        config = _default_config()
        assert config["DocType"]["Invoice"] == 65

    def test_confidence_high_score(self):
        config = _default_config()
        assert config["DocTypeConfidence"]["high"] == 15

    def test_confidence_medium_score(self):
        config = _default_config()
        assert config["DocTypeConfidence"]["medium"] == 10

    def test_confidence_low_score(self):
        config = _default_config()
        assert config["DocTypeConfidence"]["low"] == 5

    def test_signal_money_detected(self):
        config = _default_config()
        assert config["Signal"]["MoneyDetected"] == 15

    def test_signal_date_detected(self):
        config = _default_config()
        assert config["Signal"]["DateDetected"] == 10

    def test_signal_keyword_hits(self):
        config = _default_config()
        assert config["Signal"]["KeywordHits"] == 10

    def test_signal_needs_ocr(self):
        config = _default_config()
        assert config["Signal"]["NeedsOCR"] == 5

    def test_signal_entity_hits(self):
        config = _default_config()
        assert config["Signal"]["EntityHits"] == 20

    def test_unknown_doctype_zero(self):
        config = _default_config()
        assert config["DocType"].get("Unknown", 0) == 0

    def test_skips_blank_rows(self):
        wb = _fresh_wb()
        ws = wb.create_sheet("Triage_Config")
        ws.cell(row=1, column=1, value="SignalType")
        ws.cell(row=2, column=1, value=None)  # blank row
        ws.cell(row=3, column=1, value="DocType")
        ws.cell(row=3, column=2, value="Invoice")
        ws.cell(row=3, column=3, value=20)
        config = grp1.load_triage_config(ws)
        assert config.get("DocType", {}).get("Invoice") == 20


# ---------------------------------------------------------------------------
# TestLoadTriageBands
# ---------------------------------------------------------------------------

class TestLoadTriageBands:

    def test_returns_list_of_tuples(self):
        bands = _default_bands()
        assert isinstance(bands, list)
        assert all(isinstance(b, tuple) and len(b) == 3 for b in bands)

    def test_sorted_highest_first(self):
        bands = _default_bands()
        scores = [b[1] for b in bands]
        assert scores == sorted(scores, reverse=True)

    def test_high_band_first(self):
        bands = _default_bands()
        assert bands[0][0] == "High"

    def test_skip_band_last(self):
        bands = _default_bands()
        assert bands[-1][0] == "Skip"

    def test_high_threshold_is_60(self):
        bands = _default_bands()
        high = next(b for b in bands if b[0] == "High")
        assert high[1] == 60

    def test_medium_threshold_is_35(self):
        bands = _default_bands()
        medium = next(b for b in bands if b[0] == "Medium")
        assert medium[1] == 35

    def test_four_bands_loaded(self):
        bands = _default_bands()
        assert len(bands) == 4


# ---------------------------------------------------------------------------
# TestScoreRecord
# ---------------------------------------------------------------------------

class TestScoreRecord:

    def test_ltc_claim_medium_full_signals(self):
        # LTC_Claim(30) + medium(10) + money(15) + dates(10) = 65
        config = _default_config()
        rec = _record(DocType="LTC_Claim", DocTypeConfidence="medium",
                      MoneyDetected="Y", DateDetected="Y")
        assert grp1.score_record(rec, config) == 65

    def test_invoice_low_confidence(self):
        # Invoice(65) + low(5) = 70
        config = _default_config()
        rec = _record(DocType="Invoice", DocTypeConfidence="low")
        assert grp1.score_record(rec, config) == 70

    def test_unknown_doctype_zero(self):
        config = _default_config()
        rec = _record(DocType="Unknown", DocTypeConfidence="")
        assert grp1.score_record(rec, config) == 0

    def test_keyword_hits_adds_points(self):
        # just keyword signal: 10
        config = _default_config()
        rec = _record(KeywordHits="trust;beneficiary")
        assert grp1.score_record(rec, config) == 10

    def test_needs_ocr_adds_points(self):
        config = _default_config()
        rec = _record(NeedsOCR="Y")
        assert grp1.score_record(rec, config) == 5

    def test_entity_hits_adds_points(self):
        # EntityHits signal: 20
        config = _default_config()
        rec = _record(EntityHits="John Smith;Jane Doe")
        assert grp1.score_record(rec, config) == 20

    def test_all_signals_sum(self):
        # LTC_Claim(30) + high(15) + money(15) + dates(10) + keywords(10) + ocr(5) = 85
        config = _default_config()
        rec = _record(DocType="LTC_Claim", DocTypeConfidence="high",
                      MoneyDetected="Y", DateDetected="Y",
                      KeywordHits="trust", NeedsOCR="Y")
        assert grp1.score_record(rec, config) == 85

    def test_score_clamps_at_100(self):
        # Inject a huge custom config to force overflow
        config = {
            "DocType": {"BigDoc": 80},
            "DocTypeConfidence": {"high": 80},
            "Signal": {"MoneyDetected": 80},
        }
        rec = _record(DocType="BigDoc", DocTypeConfidence="high", MoneyDetected="Y")
        assert grp1.score_record(rec, config) == 100

    def test_empty_record_scores_zero(self):
        config = _default_config()
        assert grp1.score_record({}, config) == 0

    def test_case_insensitive_confidence(self):
        # The code does .lower() on confidence
        config = _default_config()
        rec = _record(DocType="Invoice", DocTypeConfidence="High")  # capital H
        score = grp1.score_record(rec, config)
        assert score == 65 + 15  # Invoice + high


# ---------------------------------------------------------------------------
# TestGetTriageBand
# ---------------------------------------------------------------------------

class TestGetTriageBand:

    def test_score_60_is_high(self):
        bands = _default_bands()
        band, _ = grp1.get_triage_band(60, bands)
        assert band == "High"

    def test_score_above_60_is_high(self):
        bands = _default_bands()
        band, _ = grp1.get_triage_band(85, bands)
        assert band == "High"

    def test_score_59_is_medium(self):
        bands = _default_bands()
        band, _ = grp1.get_triage_band(59, bands)
        assert band == "Medium"

    def test_score_35_is_medium(self):
        bands = _default_bands()
        band, _ = grp1.get_triage_band(35, bands)
        assert band == "Medium"

    def test_score_34_is_low(self):
        bands = _default_bands()
        band, _ = grp1.get_triage_band(34, bands)
        assert band == "Low"

    def test_score_10_is_low(self):
        bands = _default_bands()
        band, _ = grp1.get_triage_band(10, bands)
        assert band == "Low"

    def test_score_9_is_skip(self):
        bands = _default_bands()
        band, _ = grp1.get_triage_band(9, bands)
        assert band == "Skip"

    def test_score_0_is_skip(self):
        bands = _default_bands()
        band, _ = grp1.get_triage_band(0, bands)
        assert band == "Skip"

    def test_returns_next_step(self):
        bands = _default_bands()
        _, next_step = grp1.get_triage_band(60, bands)
        assert next_step == "Priority manual review"

    def test_empty_bands_fallback(self):
        band, step = grp1.get_triage_band(50, [])
        assert band == "Skip"
        assert step == "Archive"


# ---------------------------------------------------------------------------
# TestGetReasonFlagged
# ---------------------------------------------------------------------------

class TestGetReasonFlagged:

    def test_doc_type_in_reasons(self):
        config = _default_config()
        rec = _record(DocType="LTC_Claim")
        reason = grp1.get_reason_flagged(rec, config)
        assert "LTC_Claim" in reason

    def test_money_in_reasons(self):
        config = _default_config()
        rec = _record(MoneyDetected="Y")
        reason = grp1.get_reason_flagged(rec, config)
        assert "MoneyDetected" in reason

    def test_dates_in_reasons(self):
        config = _default_config()
        rec = _record(DateDetected="Y")
        reason = grp1.get_reason_flagged(rec, config)
        assert "DateDetected" in reason

    def test_keywords_in_reasons(self):
        config = _default_config()
        rec = _record(KeywordHits="trust")
        reason = grp1.get_reason_flagged(rec, config)
        assert "KeywordHits" in reason

    def test_ocr_in_reasons(self):
        config = _default_config()
        rec = _record(NeedsOCR="Y")
        reason = grp1.get_reason_flagged(rec, config)
        assert "NeedsOCR" in reason

    def test_empty_record_returns_empty(self):
        config = _default_config()
        reason = grp1.get_reason_flagged(_record(), config)
        assert reason == ""

    def test_zero_score_doctype_excluded(self):
        # "Unknown" maps to 0 in config — should not appear in reasons
        config = _default_config()
        rec = _record(DocType="Unknown")
        reason = grp1.get_reason_flagged(rec, config)
        assert "Unknown" not in reason

    def test_multiple_reasons_semicolon_joined(self):
        config = _default_config()
        rec = _record(DocType="LTC_Claim", MoneyDetected="Y", DateDetected="Y")
        reason = grp1.get_reason_flagged(rec, config)
        parts = reason.split(";")
        assert len(parts) >= 3


# ---------------------------------------------------------------------------
# TestGetNextStep
# ---------------------------------------------------------------------------

class TestGetNextStep:

    def test_high_band_standard_next_step(self):
        bands = _default_bands()
        step = grp1.get_next_step("High", _record(), bands)
        assert step == "Priority manual review"

    def test_low_band_standard_next_step(self):
        bands = _default_bands()
        step = grp1.get_next_step("Low", _record(), bands)
        assert step == "Batch review"

    def test_skip_band_archive(self):
        bands = _default_bands()
        step = grp1.get_next_step("Skip", _record(), bands)
        assert step == "Archive"

    def test_medium_with_needs_ocr_override(self):
        bands = _default_bands()
        rec = _record(NeedsOCR="Y")
        step = grp1.get_next_step("Medium", rec, bands)
        assert step == "OCR then review"

    def test_medium_without_needs_ocr_is_normal(self):
        bands = _default_bands()
        rec = _record(NeedsOCR="N")
        step = grp1.get_next_step("Medium", rec, bands)
        assert step == "Manual review"

    def test_high_with_needs_ocr_no_override(self):
        # OCR override only fires for Medium band
        bands = _default_bands()
        rec = _record(NeedsOCR="Y")
        step = grp1.get_next_step("High", rec, bands)
        assert step == "Priority manual review"

    def test_unknown_band_returns_archive(self):
        bands = _default_bands()
        step = grp1.get_next_step("Nonexistent", _record(), bands)
        assert step == "Archive"
