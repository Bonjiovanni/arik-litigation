"""
tests/test_fw_dirmap.py
=======================
Module-level integration tests for fw_dirmap.py (the merged single-file entry point).

These tests verify that all groups work correctly *together* when imported from
the integrated module.  They complement the per-group unit tests but do NOT
duplicate them — focus is on cross-group contracts and end-to-end pipeline
correctness.

Interactive functions skipped: pick_scan_dirs(), main()
"""

import json
import os
import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import fw_dirmap


# ===========================================================================
# Helpers shared across test classes
# ===========================================================================

def _fake_validate(raw):
    return raw.replace("\\", "/").rstrip("/")

def _fake_detect_store(path):
    return "Local"

def _fake_count_contents(path):
    return (2, 1)

def _fake_walk_three(root, recursive, max_depth):
    """Yields a fixed 3-node tree: root + sub1 + sub1/child."""
    norm = root.replace("\\", "/").rstrip("/")
    yield (norm, 0)
    yield (norm + "/sub1", 1)
    yield (norm + "/sub1/child", 2)

def _fake_walk_single(root, recursive, max_depth):
    norm = root.replace("\\", "/").rstrip("/")
    yield (norm, 0)

def _build_records(roots=None, walk_fn=_fake_walk_three, run_id="MOD_TEST_001"):
    if roots is None:
        roots = ["C:/Root"]
    return fw_dirmap.build_dir_records(
        roots=roots,
        recursive=True,
        max_depth=None,
        run_id=run_id,
        _validate=_fake_validate,
        _detect_store=_fake_detect_store,
        _count_contents=_fake_count_contents,
        _walk=walk_fn,
    )

def _make_stats(elapsed_sec=3.0, dir_count=3, root_count=1):
    start = datetime(2026, 3, 15, 12, 0, 0)
    return {
        "start_time": start,
        "end_time":   start + timedelta(seconds=elapsed_sec),
        "dir_count":  dir_count,
        "root_count": root_count,
    }


# ===========================================================================
# 1. Namespace / import smoke tests
# ===========================================================================

class TestNamespace:
    """All public functions must be importable from fw_dirmap."""

    def test_validate_and_normalize_path(self):
        assert callable(fw_dirmap.validate_and_normalize_path)

    def test_generate_run_id(self):
        assert callable(fw_dirmap.generate_run_id)

    def test_detect_source_store(self):
        assert callable(fw_dirmap.detect_source_store)

    def test_count_dir_contents(self):
        assert callable(fw_dirmap.count_dir_contents)

    def test_walk_directories(self):
        assert callable(fw_dirmap.walk_directories)

    def test_build_dir_records(self):
        assert callable(fw_dirmap.build_dir_records)

    def test_open_or_create_workbook(self):
        assert callable(fw_dirmap.open_or_create_workbook)

    def test_ensure_dir_inventory_sheet(self):
        assert callable(fw_dirmap.ensure_dir_inventory_sheet)

    def test_write_dir_inventory_rows(self):
        assert callable(fw_dirmap.write_dir_inventory_rows)

    def test_ensure_dir_processing_status_sheet(self):
        assert callable(fw_dirmap.ensure_dir_processing_status_sheet)

    def test_initialize_processing_status_rows(self):
        assert callable(fw_dirmap.initialize_processing_status_rows)

    def test_log_dirmap_run(self):
        assert callable(fw_dirmap.log_dirmap_run)

    def test_print_run_summary(self):
        assert callable(fw_dirmap.print_run_summary)


# ===========================================================================
# 2. End-to-end pipeline — all three sheets written to one workbook
# ===========================================================================

class TestFullPipeline:
    """Build records → write all sheets → log run → verify workbook state."""

    @pytest.fixture
    def pipeline_wb(self):
        records = _build_records()
        wb = openpyxl.Workbook()
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]

        fw_dirmap.write_dir_inventory_rows(wb, records)
        fw_dirmap.initialize_processing_status_rows(wb, records)
        fw_dirmap.log_dirmap_run(
            wb, "MOD_TEST_001",
            {"roots": ["C:/Root"], "recursive": True, "max_depth": None},
            _make_stats(dir_count=len(records), root_count=1),
        )
        return wb, records

    def test_all_three_sheets_exist(self, pipeline_wb):
        wb, _ = pipeline_wb
        assert "Dir_Inventory" in wb.sheetnames
        assert "Dir_Processing_Status" in wb.sheetnames
        assert "Walk_History" in wb.sheetnames

    def test_dir_inventory_row_count(self, pipeline_wb):
        wb, records = pipeline_wb
        ws = wb["Dir_Inventory"]
        # header row + one data row per record
        assert ws.max_row == 1 + len(records)

    def test_dps_row_count_is_records_times_nine(self, pipeline_wb):
        """Dir_Processing_Status has 9 file-family rows per directory."""
        wb, records = pipeline_wb
        ws = wb["Dir_Processing_Status"]
        expected_data_rows = len(records) * 9
        assert ws.max_row == 1 + expected_data_rows

    def test_walk_history_has_one_data_row(self, pipeline_wb):
        wb, _ = pipeline_wb
        ws = wb["Walk_History"]
        assert ws.max_row == 2

    def test_dir_ids_consistent_across_inventory_and_dps(self, pipeline_wb):
        """Every DirID in Dir_Inventory must also appear in Dir_Processing_Status."""
        wb, _ = pipeline_wb
        inv_ids = {
            wb["Dir_Inventory"].cell(row=r, column=1).value
            for r in range(2, wb["Dir_Inventory"].max_row + 1)
        }
        dps_ids = {
            wb["Dir_Processing_Status"].cell(row=r, column=1).value
            for r in range(2, wb["Dir_Processing_Status"].max_row + 1)
        }
        assert inv_ids <= dps_ids  # every inventory ID has DPS rows


# ===========================================================================
# 3. Cross-group data contracts
# ===========================================================================

class TestDataContracts:
    """build_dir_records output feeds correctly into the writer functions."""

    def test_all_required_record_keys_present(self):
        records = _build_records()
        required = {
            "dir_id", "scan_root", "relative_dir", "dir_name",
            "full_path", "depth", "file_count", "subdir_count",
            "source_store", "run_id",
        }
        for record in records:
            assert required <= record.keys(), \
                f"Missing keys in record: {required - record.keys()}"

    def test_dir_inventory_first_data_cell_is_dir_id(self):
        records = _build_records()
        wb = openpyxl.Workbook()
        fw_dirmap.write_dir_inventory_rows(wb, records)
        ws = wb["Dir_Inventory"]
        assert ws.cell(row=2, column=1).value == "D001"

    def test_dps_dir_id_column_matches_first_record(self):
        records = _build_records(walk_fn=_fake_walk_single)
        wb = openpyxl.Workbook()
        fw_dirmap.initialize_processing_status_rows(wb, records)
        ws = wb["Dir_Processing_Status"]
        # All 9 rows for the single directory must share the same DirID
        dps_ids = [ws.cell(row=r, column=1).value for r in range(2, 11)]
        assert all(v == "D001" for v in dps_ids)

    def test_dps_full_path_matches_inventory(self):
        records = _build_records(walk_fn=_fake_walk_single)
        wb = openpyxl.Workbook()
        fw_dirmap.write_dir_inventory_rows(wb, records)
        fw_dirmap.initialize_processing_status_rows(wb, records)

        inv_path = wb["Dir_Inventory"].cell(row=2, column=5).value  # FullPath col
        dps_path = wb["Dir_Processing_Status"].cell(row=2, column=2).value
        assert inv_path == dps_path

    def test_nine_distinct_file_families_per_directory(self):
        records = _build_records(walk_fn=_fake_walk_single)
        wb = openpyxl.Workbook()
        fw_dirmap.initialize_processing_status_rows(wb, records)
        ws = wb["Dir_Processing_Status"]
        families = [ws.cell(row=r, column=3).value for r in range(2, 11)]
        assert len(set(families)) == 9

    def test_log_run_dir_count_matches_records(self):
        records = _build_records()
        wb = openpyxl.Workbook()
        fw_dirmap.log_dirmap_run(
            wb, "R1",
            {"roots": ["C:/Root"], "recursive": True, "max_depth": None},
            _make_stats(dir_count=len(records)),
        )
        ws = wb["Walk_History"]
        # DirCount is column G (index 7)
        assert ws.cell(row=2, column=7).value == len(records)


# ===========================================================================
# 4. Workbook round-trip (save → reload → data intact)
# ===========================================================================

class TestRoundTrip:
    def test_save_and_reload_preserves_sheets(self, tmp_path):
        records = _build_records()
        wb = openpyxl.Workbook()
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]
        fw_dirmap.write_dir_inventory_rows(wb, records)
        fw_dirmap.initialize_processing_status_rows(wb, records)
        fw_dirmap.log_dirmap_run(
            wb, "ROUND_TRIP",
            {"roots": ["C:/Root"], "recursive": True, "max_depth": None},
            _make_stats(dir_count=3),
        )

        xlsx_path = str(tmp_path / "test_output.xlsx")
        wb.save(xlsx_path)

        wb2 = openpyxl.load_workbook(xlsx_path)
        assert "Dir_Inventory" in wb2.sheetnames
        assert "Dir_Processing_Status" in wb2.sheetnames
        assert "Walk_History" in wb2.sheetnames

    def test_save_and_reload_dir_inventory_row_count(self, tmp_path):
        records = _build_records()
        wb = openpyxl.Workbook()
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]
        fw_dirmap.write_dir_inventory_rows(wb, records)
        xlsx_path = str(tmp_path / "inv_only.xlsx")
        wb.save(xlsx_path)

        wb2 = openpyxl.load_workbook(xlsx_path)
        ws = wb2["Dir_Inventory"]
        assert ws.max_row == 1 + len(records)

    def test_open_or_create_workbook_creates_new_file(self, tmp_path):
        path = str(tmp_path / "fresh.xlsx")
        assert not os.path.exists(path)
        wb = fw_dirmap.open_or_create_workbook(path)
        assert isinstance(wb, openpyxl.Workbook)
        # New workbook has no sheets yet (default sheet was removed)
        assert len(wb.sheetnames) == 0

    def test_open_or_create_workbook_opens_existing(self, tmp_path):
        path = str(tmp_path / "existing.xlsx")
        wb_init = openpyxl.Workbook()
        wb_init.create_sheet("MySheet")
        wb_init.save(path)

        wb2 = fw_dirmap.open_or_create_workbook(path)
        assert "MySheet" in wb2.sheetnames


# ===========================================================================
# 5. Multi-root integration
# ===========================================================================

class TestMultiRoot:
    def test_dir_ids_are_continuous_across_two_roots(self):
        records = _build_records(roots=["C:/RootA", "C:/RootB"])
        ids = [r["dir_id"] for r in records]
        # 3 dirs per root → D001 … D006, no reset
        assert ids[0] == "D001"
        assert ids[3] == "D004"
        assert ids[-1] == "D006"

    def test_inventory_sheet_has_rows_from_all_roots(self):
        records = _build_records(roots=["C:/RootA", "C:/RootB"])
        wb = openpyxl.Workbook()
        fw_dirmap.write_dir_inventory_rows(wb, records)
        ws = wb["Dir_Inventory"]
        scan_roots = {ws.cell(row=r, column=2).value
                      for r in range(2, ws.max_row + 1)}
        assert "C:/RootA" in scan_roots
        assert "C:/RootB" in scan_roots

    def test_dps_has_rows_for_all_roots(self):
        records = _build_records(roots=["C:/RootA", "C:/RootB"])
        wb = openpyxl.Workbook()
        fw_dirmap.initialize_processing_status_rows(wb, records)
        ws = wb["Dir_Processing_Status"]
        # 2 roots × 3 dirs × 9 families = 54 data rows
        assert ws.max_row == 1 + 54


# ===========================================================================
# 6. Header and sheet structure
# ===========================================================================

class TestSheetStructure:
    def test_dir_inventory_header_row(self):
        wb = openpyxl.Workbook()
        fw_dirmap.ensure_dir_inventory_sheet(wb)
        ws = wb["Dir_Inventory"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 11)]
        assert headers[0] == "DirID"
        assert headers[-1] == "Notes"

    def test_dps_header_row(self):
        wb = openpyxl.Workbook()
        fw_dirmap.ensure_dir_processing_status_sheet(wb)
        ws = wb["Dir_Processing_Status"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers[0] == "DirID"
        assert "FileFamily" in headers

    def test_dir_inventory_freeze_panes(self):
        wb = openpyxl.Workbook()
        fw_dirmap.ensure_dir_inventory_sheet(wb)
        assert wb["Dir_Inventory"].freeze_panes == "A2"

    def test_dps_freeze_panes(self):
        wb = openpyxl.Workbook()
        fw_dirmap.ensure_dir_processing_status_sheet(wb)
        assert wb["Dir_Processing_Status"].freeze_panes == "A2"

    def test_walk_history_freeze_panes(self):
        wb = openpyxl.Workbook()
        fw_dirmap.log_dirmap_run(
            wb, "R1",
            {"roots": [], "recursive": True, "max_depth": None},
            _make_stats(),
        )
        assert wb["Walk_History"].freeze_panes == "A2"

    def test_idempotent_ensure_dir_inventory(self):
        """Calling ensure_dir_inventory_sheet twice must not add a second sheet."""
        wb = openpyxl.Workbook()
        fw_dirmap.ensure_dir_inventory_sheet(wb)
        fw_dirmap.ensure_dir_inventory_sheet(wb)
        assert wb.sheetnames.count("Dir_Inventory") == 1

    def test_idempotent_ensure_dps(self):
        wb = openpyxl.Workbook()
        fw_dirmap.ensure_dir_processing_status_sheet(wb)
        fw_dirmap.ensure_dir_processing_status_sheet(wb)
        assert wb.sheetnames.count("Dir_Processing_Status") == 1
