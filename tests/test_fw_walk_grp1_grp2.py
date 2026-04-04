"""
tests/test_fw_walk_grp1_grp2.py

Unit tests for fw_walk_grp1.py and fw_walk_grp2.py.
Run with: pytest tests/test_fw_walk_grp1_grp2.py
"""

import sys
import os
import hashlib
import tempfile
import zipfile

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import fw_walk_grp1 as grp1
import fw_walk_grp2 as grp2


# ---------------------------------------------------------------------------
# Mock worksheet helper
# ---------------------------------------------------------------------------

class _MockWs:
    """Minimal worksheet mock: rows[0] = header, rows[1:] = data."""

    def __init__(self, rows):
        self._rows = rows

    def iter_rows(self, min_row=1, values_only=False):
        for row in self._rows[min_row - 1:]:
            yield row


def _make_mock_ws(header, data_rows):
    return _MockWs([header] + data_rows)


# ---------------------------------------------------------------------------
# Temp file helpers
# ---------------------------------------------------------------------------

def _temp_bytes(content: bytes) -> str:
    f = tempfile.NamedTemporaryFile(delete=False)
    f.write(content)
    f.flush()
    f.close()
    return f.name


# ---------------------------------------------------------------------------
# TestLoadFileFamilyConfig
# ---------------------------------------------------------------------------
# family_map shape: {family_name: {".ext", ...}}
# skip_families:    set of family name strings

class TestLoadFileFamilyConfig:
    HEADER = ("Family", "Extensions", "ShouldSkip", "LikelyTextBearing",
              "LikelyImage", "LikelySpreadsheet", "LikelyDocument", "Notes")

    def test_empty_worksheet_returns_empty(self):
        ws = _make_mock_ws(self.HEADER, [])
        fm, sf, _ = grp1.load_file_family_config(ws)
        assert fm == {}
        assert sf == set()

    def test_single_family_appears_in_map(self):
        ws = _make_mock_ws(self.HEADER, [("pdf", ".pdf", "N", "Y", "N", "N", "Y", "")])
        fm, _, _ = grp1.load_file_family_config(ws)
        assert "pdf" in fm
        assert ".pdf" in fm["pdf"]

    def test_should_skip_Y_adds_to_skip_families(self):
        ws = _make_mock_ws(self.HEADER, [("archive", ".zip", "Y", "N", "N", "N", "N", "")])
        _, sf, _ = grp1.load_file_family_config(ws)
        assert "archive" in sf

    def test_should_skip_N_does_not_add(self):
        ws = _make_mock_ws(self.HEADER, [("pdf", ".pdf", "N", "Y", "N", "N", "Y", "")])
        _, sf, _ = grp1.load_file_family_config(ws)
        assert "pdf" not in sf

    def test_extension_without_leading_dot_gets_dot_added(self):
        ws = _make_mock_ws(self.HEADER, [("pdf", "pdf", "N", "Y", "N", "N", "Y", "")])
        fm, _, _ = grp1.load_file_family_config(ws)
        assert "pdf" in fm
        assert ".pdf" in fm["pdf"]

    def test_semicolon_delimited_extensions_are_split(self):
        ws = _make_mock_ws(self.HEADER, [("image", ".jpg;.png;.gif", "N", "N", "Y", "N", "N", "")])
        fm, _, _ = grp1.load_file_family_config(ws)
        assert "image" in fm
        assert ".jpg" in fm["image"]
        assert ".png" in fm["image"]
        assert ".gif" in fm["image"]

    def test_rows_with_none_family_are_skipped(self):
        ws = _make_mock_ws(self.HEADER, [
            (None, ".xyz", "N", "N", "N", "N", "N", ""),
            ("pdf", ".pdf", "N", "Y", "N", "N", "Y", ""),
        ])
        fm, _, _ = grp1.load_file_family_config(ws)
        # None-family row should not produce an "xyz" family or similar
        assert ".xyz" not in {ext for exts in fm.values() for ext in exts}
        assert "pdf" in fm

    def test_none_should_skip_defaults_to_not_skip(self):
        ws = _make_mock_ws(self.HEADER, [("weirdtype", ".weird", None, "N", "N", "N", "N", "")])
        _, sf, _ = grp1.load_file_family_config(ws)
        assert "weirdtype" not in sf

    def test_multiple_families(self):
        ws = _make_mock_ws(self.HEADER, [
            ("archive", ".zip", "Y", "N", "N", "N", "N", ""),
            ("video",   ".mp4", "Y", "N", "Y", "N", "N", ""),
            ("pdf",     ".pdf", "N", "Y", "N", "N", "Y", ""),
        ])
        fm, sf, _ = grp1.load_file_family_config(ws)
        assert "archive" in sf
        assert "video" in sf
        assert "pdf" not in sf
        assert ".zip" in fm["archive"]
        assert ".mp4" in fm["video"]
        assert ".pdf" in fm["pdf"]


# ---------------------------------------------------------------------------
# TestClassifyFileFamily
# ---------------------------------------------------------------------------

class TestClassifyFileFamily:

    def test_pdf(self):
        assert grp1.classify_file_family("report.pdf", ".pdf") == "pdf"

    def test_docx_word_doc(self):
        assert grp1.classify_file_family("doc.docx", ".docx") == "word_doc"

    def test_xlsx_spreadsheet(self):
        assert grp1.classify_file_family("data.xlsx", ".xlsx") == "spreadsheet"

    def test_mp3_audio(self):
        assert grp1.classify_file_family("song.mp3", ".mp3") == "audio"

    def test_txt_text_file(self):
        assert grp1.classify_file_family("readme.txt", ".txt") == "text_file"

    def test_json_text_file(self):
        assert grp1.classify_file_family("data.json", ".json") == "text_file"

    def test_zip_archive(self):
        assert grp1.classify_file_family("archive.zip", ".zip") == "archive"

    def test_unknown_extension_returns_other(self):
        assert grp1.classify_file_family("file.xyzzy", ".xyzzy") == "other"

    def test_case_insensitive_extension(self):
        assert grp1.classify_file_family("report.PDF", ".PDF") == "pdf"

    def test_custom_family_map(self):
        custom = {"custom_fam": {".pdf"}}
        assert grp1.classify_file_family("r.pdf", ".pdf", family_map=custom) == "custom_fam"

    def test_custom_map_unmatched_returns_other(self):
        custom = {"pdf_only": {".pdf"}}
        assert grp1.classify_file_family("f.docx", ".docx", family_map=custom) == "other"

    def test_none_family_map_uses_builtins(self):
        assert grp1.classify_file_family("r.pdf", ".pdf", family_map=None) == "pdf"

    def test_jpg_image(self):
        assert grp1.classify_file_family("photo.jpg", ".jpg") == "image"

    def test_mp4_video(self):
        assert grp1.classify_file_family("clip.mp4", ".mp4") == "video"

    def test_msg_email_file(self):
        assert grp1.classify_file_family("email.msg", ".msg") == "email_file"


# ---------------------------------------------------------------------------
# TestShouldSkipFile
# ---------------------------------------------------------------------------

class TestShouldSkipFile:

    def test_email_file_skipped(self):
        skip, reason = grp1.should_skip_file("email.msg", ".msg", "email_file")
        assert skip is True
        assert reason == "email_file"

    def test_archive_skipped(self):
        skip, reason = grp1.should_skip_file("data.zip", ".zip", "archive")
        assert skip is True
        assert reason == "archive"

    def test_office_temp_tilde_dollar(self):
        skip, reason = grp1.should_skip_file("~$document.docx", ".docx", "word_doc")
        assert skip is True
        assert reason == "office_temp"

    def test_hidden_file_dot_prefix(self):
        skip, reason = grp1.should_skip_file(".hidden", "", "other")
        assert skip is True
        assert reason == "hidden_file"

    def test_tmp_extension(self):
        skip, reason = grp1.should_skip_file("tempfile.tmp", ".tmp", "other")
        assert skip is True
        assert reason == "system_file"

    def test_normal_pdf_not_skipped(self):
        skip, reason = grp1.should_skip_file("report.pdf", ".pdf", "pdf")
        assert skip is False
        assert reason == ""

    def test_custom_skip_families_video(self):
        skip, _ = grp1.should_skip_file("clip.mp4", ".mp4", "video", skip_families={"video"})
        assert skip is True

    def test_custom_skip_families_archive_not_skipped(self):
        # Only video in custom skip; archive should NOT be skipped
        skip, _ = grp1.should_skip_file("data.zip", ".zip", "archive", skip_families={"video"})
        assert skip is False

    def test_empty_skip_families_no_family_skip(self):
        # email_file normally skipped, but skip_families=set() disables family-based skip
        skip, reason = grp1.should_skip_file("email.msg", ".msg", "email_file", skip_families=set())
        # Family-based skip should not fire; but other rules (tilde, dot, tmp) still work
        if skip:
            assert reason != "email_file"

    def test_normal_docx_not_skipped(self):
        skip, reason = grp1.should_skip_file("report.docx", ".docx", "word_doc")
        assert skip is False

    def test_tilde_dollar_overrides_empty_skip_families(self):
        skip, reason = grp1.should_skip_file("~$sheet.xlsx", ".xlsx", "spreadsheet",
                                              skip_families=set())
        assert skip is True
        assert reason == "office_temp"

    def test_dot_hidden_overrides_empty_skip_families(self):
        skip, reason = grp1.should_skip_file(".DS_Store", "", "other", skip_families=set())
        assert skip is True
        assert reason == "hidden_file"


# ---------------------------------------------------------------------------
# TestGetFileMetadata
# ---------------------------------------------------------------------------

class TestGetFileMetadata:

    EXPECTED_KEYS = {"filename", "extension", "parent_folder", "file_path",
                     "size_bytes", "created_time", "modified_time", "accessed_time"}

    def test_metadata_keys_present(self):
        path = _temp_bytes(b"hello world")
        try:
            meta = grp1.get_file_metadata(path)
            assert self.EXPECTED_KEYS.issubset(set(meta.keys()))
        finally:
            os.unlink(path)

    def test_extension_lowercased(self):
        f = tempfile.NamedTemporaryFile(suffix=".TXT", delete=False)
        f.write(b"data")
        f.close()
        try:
            meta = grp1.get_file_metadata(f.name)
            assert meta["extension"] == ".txt"
        finally:
            os.unlink(f.name)

    def test_path_uses_forward_slashes(self):
        path = _temp_bytes(b"data")
        try:
            meta = grp1.get_file_metadata(path)
            assert "\\" not in meta["file_path"]
        finally:
            os.unlink(path)

    def test_size_bytes_positive_int(self):
        path = _temp_bytes(b"hello world")
        try:
            meta = grp1.get_file_metadata(path)
            assert isinstance(meta["size_bytes"], int)
            assert meta["size_bytes"] > 0
        finally:
            os.unlink(path)

    def test_nonexistent_path_empty_numeric_fields(self):
        meta = grp1.get_file_metadata("/nonexistent/does_not_exist.pdf")
        assert meta["size_bytes"] == ""
        assert meta["created_time"] == ""
        assert meta["modified_time"] == ""

    def test_filename_is_basename(self):
        path = _temp_bytes(b"x")
        try:
            meta = grp1.get_file_metadata(path)
            assert meta["filename"] == os.path.basename(path)
        finally:
            os.unlink(path)


# ---------------------------------------------------------------------------
# TestComputeSha256
# ---------------------------------------------------------------------------

class TestComputeSha256:

    def test_returns_64_char_hex_complete_read(self):
        path = _temp_bytes(b"hello world data")
        try:
            digest, complete = grp1.compute_sha256(path, max_bytes=10 * 1024 * 1024)
            assert len(digest) == 64
            assert complete is True
        finally:
            os.unlink(path)

    def test_max_bytes_truncation_is_complete_false(self):
        path = _temp_bytes(b"x" * 2000)
        try:
            digest, complete = grp1.compute_sha256(path, max_bytes=1000)
            assert complete is False
            assert len(digest) == 64
        finally:
            os.unlink(path)

    def test_nonexistent_file_returns_empty_false(self):
        digest, complete = grp1.compute_sha256("/no/such/file.bin", max_bytes=1024)
        assert digest == ""
        assert complete is False

    def test_permission_error_returns_empty_false(self, monkeypatch):
        def _raise(*args, **kwargs):
            raise PermissionError("no access")
        monkeypatch.setattr("builtins.open", _raise)
        digest, complete = grp1.compute_sha256("/some/file.bin", max_bytes=1024)
        assert digest == ""
        assert complete is False

    def test_known_hash_of_hello(self):
        path = _temp_bytes(b"hello")
        try:
            expected = hashlib.sha256(b"hello").hexdigest()
            digest, complete = grp1.compute_sha256(path, max_bytes=10 * 1024 * 1024)
            assert digest == expected
            assert complete is True
        finally:
            os.unlink(path)

    def test_empty_file_known_hash(self):
        path = _temp_bytes(b"")
        try:
            expected = hashlib.sha256(b"").hexdigest()
            digest, complete = grp1.compute_sha256(path, max_bytes=1024)
            assert digest == expected
            assert complete is True
        finally:
            os.unlink(path)


# ---------------------------------------------------------------------------
# TestPeekArchiveContents
# ---------------------------------------------------------------------------

class TestPeekArchiveContents:

    def test_rar_returns_unsupported(self):
        f = tempfile.NamedTemporaryFile(suffix=".rar", delete=False)
        f.write(b"Rar!\x1a\x07\x00fake")
        f.close()
        try:
            result = grp1.peek_archive_contents(f.name)
            assert "unsupported" in result.lower() or "rar" in result.lower()
        finally:
            os.unlink(f.name)

    def test_nonexistent_zip_returns_error(self):
        result = grp1.peek_archive_contents("/no/such/archive.zip")
        lower = result.lower()
        assert any(w in lower for w in ("corrupt", "error", "not found", "os error"))

    def test_valid_zip_file_count(self):
        f = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        f.close()
        try:
            with zipfile.ZipFile(f.name, "w") as zf:
                zf.writestr("photo1.jpg", b"fake jpg 1")
                zf.writestr("photo2.jpg", b"fake jpg 2")
                zf.writestr("photo3.jpg", b"fake jpg 3")
                zf.writestr("document.pdf", b"fake pdf")
            result = grp1.peek_archive_contents(f.name)
            assert result.startswith("4 files:")
        finally:
            os.unlink(f.name)

    def test_valid_zip_extensions_in_result(self):
        f = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        f.close()
        try:
            with zipfile.ZipFile(f.name, "w") as zf:
                zf.writestr("photo.jpg", b"fake jpg")
                zf.writestr("report.pdf", b"fake pdf")
            result = grp1.peek_archive_contents(f.name)
            assert ".jpg" in result
            assert ".pdf" in result
        finally:
            os.unlink(f.name)

    def test_empty_zip(self):
        f = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        f.close()
        try:
            with zipfile.ZipFile(f.name, "w"):
                pass
            result = grp1.peek_archive_contents(f.name)
            assert "0 files" in result or "empty" in result.lower()
        finally:
            os.unlink(f.name)

    def test_corrupt_zip_returns_error(self):
        f = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        f.write(b"this is not a zip file at all")
        f.close()
        try:
            result = grp1.peek_archive_contents(f.name)
            lower = result.lower()
            assert any(w in lower for w in ("corrupt", "error", "bad", "os error"))
        finally:
            os.unlink(f.name)


# ---------------------------------------------------------------------------
# TestEnsureMasterFileInventorySheet
# ---------------------------------------------------------------------------

class TestEnsureMasterFileInventorySheet:

    SPOT_HEADERS = {"FileID", "FilePath", "ArchiveContents", "Notes"}

    def _headers(self, ws):
        return {c.value for c in ws[1] if c.value}

    def test_creates_sheet(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_master_file_inventory_sheet(wb)
        assert ws is not None

    def test_idempotent(self):
        wb = openpyxl.Workbook()
        ws1 = grp2.ensure_master_file_inventory_sheet(wb)
        ws2 = grp2.ensure_master_file_inventory_sheet(wb)
        assert ws1.title == ws2.title

    def test_header_count_matches_columns(self):
        from fw_walk_grp2 import COLUMNS
        wb = openpyxl.Workbook()
        ws = grp2.ensure_master_file_inventory_sheet(wb)
        row1 = [c.value for c in ws[1] if c.value is not None]
        assert len(row1) == len(COLUMNS)

    def test_spot_check_headers(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_master_file_inventory_sheet(wb)
        headers = self._headers(ws)
        assert self.SPOT_HEADERS.issubset(headers)

    def test_freeze_panes_A2(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_master_file_inventory_sheet(wb)
        assert str(ws.freeze_panes) == "A2"


# ---------------------------------------------------------------------------
# TestEnsureFileFamilyConfigSheet
# ---------------------------------------------------------------------------

class TestEnsureFileFamilyConfigSheet:

    EXPECTED_FAMILIES = {
        "pdf", "word_doc", "spreadsheet", "text_file", "presentation",
        "image", "email_file", "archive", "audio", "video",
    }
    EXPECTED_HEADERS = {
        "Family", "Extensions", "ShouldSkip", "LikelyTextBearing",
        "LikelyImage", "LikelySpreadsheet", "LikelyDocument", "Notes",
    }

    def _families(self, ws):
        return {row[0] for row in ws.iter_rows(min_row=2, values_only=True)
                if row and row[0]}

    def _headers(self, ws):
        return {c.value for c in ws[1] if c.value}

    def _get_row_for_family(self, ws, family_name):
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == family_name:
                return dict(zip(headers, row))
        return None

    def test_creates_sheet(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_file_family_config_sheet(wb)
        assert ws is not None

    def test_idempotent(self):
        wb = openpyxl.Workbook()
        ws1 = grp2.ensure_file_family_config_sheet(wb)
        ws2 = grp2.ensure_file_family_config_sheet(wb)
        assert ws1.title == ws2.title

    def test_all_10_default_families(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_file_family_config_sheet(wb)
        assert self.EXPECTED_FAMILIES.issubset(self._families(ws))

    def test_expected_headers(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_file_family_config_sheet(wb)
        assert self.EXPECTED_HEADERS.issubset(self._headers(ws))

    def test_archive_should_skip_Y(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_file_family_config_sheet(wb)
        row = self._get_row_for_family(ws, "archive")
        assert row is not None
        assert row["ShouldSkip"] == "Y"

    def test_pdf_should_skip_N(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_file_family_config_sheet(wb)
        row = self._get_row_for_family(ws, "pdf")
        assert row is not None
        assert row["ShouldSkip"] == "N"

    def test_email_file_should_skip_Y(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_file_family_config_sheet(wb)
        row = self._get_row_for_family(ws, "email_file")
        assert row is not None
        assert row["ShouldSkip"] == "Y"


# ---------------------------------------------------------------------------
# TestGetNextFileId
# ---------------------------------------------------------------------------

class TestGetNextFileId:

    def _ws_with_ids(self, id_values):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_master_file_inventory_sheet(wb)
        # FileID is column 1 (col A)
        for i, val in enumerate(id_values, start=2):
            ws.cell(row=i, column=1, value=val)
        return ws

    def test_empty_sheet_returns_F00001(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_master_file_inventory_sheet(wb)
        assert grp2.get_next_file_id(ws) == "F00001"

    def test_existing_ids_returns_next(self):
        ws = self._ws_with_ids(["F00005", "F00003"])
        assert grp2.get_next_file_id(ws) == "F00006"

    def test_malformed_values_ignored(self):
        ws = self._ws_with_ids([None, "garbage", "F00002"])
        assert grp2.get_next_file_id(ws) == "F00003"

    def test_increments_past_single_id(self):
        ws = self._ws_with_ids(["F00010"])
        assert grp2.get_next_file_id(ws) == "F00011"

    def test_zero_padded_to_5_digits(self):
        ws = self._ws_with_ids(["F00099"])
        result = grp2.get_next_file_id(ws)
        assert result == "F00100"
        assert result.startswith("F")


# ---------------------------------------------------------------------------
# TestFindExistingFileRow
# ---------------------------------------------------------------------------

class TestFindExistingFileRow:

    def _ws_with_paths(self, paths):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_master_file_inventory_sheet(wb)
        # FilePath = column 4 (D)
        fp_col = 4
        for i, p in enumerate(paths, start=2):
            ws.cell(row=i, column=fp_col, value=p)
        return ws

    def test_exact_match(self):
        ws = self._ws_with_paths(["C:/Foo/Bar.pdf", "C:/Baz/Qux.docx"])
        assert grp2.find_existing_file_row(ws, "C:/Foo/Bar.pdf") == 2

    def test_case_insensitive(self):
        ws = self._ws_with_paths(["C:/Foo/Bar.pdf"])
        assert grp2.find_existing_file_row(ws, "c:/foo/bar.pdf") == 2

    def test_backslash_normalized(self):
        ws = self._ws_with_paths(["C:/Foo/Bar.pdf"])
        assert grp2.find_existing_file_row(ws, "C:\\Foo\\Bar.pdf") == 2

    def test_not_found_returns_none(self):
        ws = self._ws_with_paths(["C:/Foo/Bar.pdf"])
        assert grp2.find_existing_file_row(ws, "C:/Other/File.pdf") is None

    def test_empty_sheet_returns_none(self):
        wb = openpyxl.Workbook()
        ws = grp2.ensure_master_file_inventory_sheet(wb)
        assert grp2.find_existing_file_row(ws, "C:/Any/File.pdf") is None

    def test_second_row_match(self):
        ws = self._ws_with_paths(["C:/A.pdf", "C:/B.pdf", "C:/C.pdf"])
        assert grp2.find_existing_file_row(ws, "C:/B.pdf") == 3


# ---------------------------------------------------------------------------
# TestInsertFileRecord
# ---------------------------------------------------------------------------

class TestInsertFileRecord:

    def _ws(self):
        wb = openpyxl.Workbook()
        return grp2.ensure_master_file_inventory_sheet(wb)

    def _cell(self, ws, row, col_name):
        headers = [c.value for c in ws[1]]
        return ws.cell(row=row, column=headers.index(col_name) + 1).value

    def _record(self, path="C:/Test/file.pdf", filename="file.pdf", size=1234):
        return {"file_id": "F00001", "file_path": path, "filename": filename, "size_bytes": size}

    def test_first_insert_returns_row_2(self):
        ws = self._ws()
        assert grp2.insert_file_record(ws, self._record(), "RUN1") == 2

    def test_second_insert_returns_row_3(self):
        ws = self._ws()
        r1 = self._record("C:/A.pdf", "A.pdf")
        r2 = self._record("C:/B.pdf", "B.pdf")
        r2["file_id"] = "F00002"
        grp2.insert_file_record(ws, r1, "RUN1")
        assert grp2.insert_file_record(ws, r2, "RUN1") == 3

    def test_manual_review_status_defaults_to_unreviewed(self):
        ws = self._ws()
        row = grp2.insert_file_record(ws, self._record(), "RUN1")
        assert self._cell(ws, row, "ManualReviewStatus") == "unreviewed"

    def test_processing_status_defaults_to_file_listed(self):
        ws = self._ws()
        row = grp2.insert_file_record(ws, self._record(), "RUN1")
        assert self._cell(ws, row, "ProcessingStatus") == "file_listed"

    def test_run_id_first_seen_and_last_seen_set(self):
        ws = self._ws()
        row = grp2.insert_file_record(ws, self._record(), "RUN_ABC")
        assert self._cell(ws, row, "RunID_FirstSeen") == "RUN_ABC"
        assert self._cell(ws, row, "RunID_LastSeen") == "RUN_ABC"


# ---------------------------------------------------------------------------
# TestUpdateFileRecord
# ---------------------------------------------------------------------------

class TestUpdateFileRecord:

    def _ws(self):
        wb = openpyxl.Workbook()
        return grp2.ensure_master_file_inventory_sheet(wb)

    def _cell(self, ws, row, col_name):
        headers = [c.value for c in ws[1]]
        return ws.cell(row=row, column=headers.index(col_name) + 1).value

    def _set_cell(self, ws, row, col_name, value):
        headers = [c.value for c in ws[1]]
        ws.cell(row=row, column=headers.index(col_name) + 1).value = value

    def test_run_id_last_seen_updated(self):
        ws = self._ws()
        record = {"file_id": "F00001", "file_path": "C:/f.pdf", "size_bytes": 100}
        row = grp2.insert_file_record(ws, record, "RUN1")
        grp2.update_file_record(ws, row, {"size_bytes": 200}, "RUN2")
        assert self._cell(ws, row, "RunID_LastSeen") == "RUN2"

    def test_size_bytes_updated(self):
        ws = self._ws()
        record = {"file_id": "F00001", "file_path": "C:/f.pdf", "size_bytes": 100}
        row = grp2.insert_file_record(ws, record, "RUN1")
        grp2.update_file_record(ws, row, {"size_bytes": 9999}, "RUN2")
        assert self._cell(ws, row, "SizeBytes") == 9999

    def test_manual_review_status_not_overwritten(self):
        ws = self._ws()
        record = {"file_id": "F00001", "file_path": "C:/f.pdf", "size_bytes": 100}
        row = grp2.insert_file_record(ws, record, "RUN1")
        self._set_cell(ws, row, "ManualReviewStatus", "reviewed")
        grp2.update_file_record(ws, row, {"size_bytes": 200}, "RUN2")
        assert self._cell(ws, row, "ManualReviewStatus") == "reviewed"

    def test_doc_type_not_touched(self):
        ws = self._ws()
        record = {"file_id": "F00001", "file_path": "C:/f.pdf", "size_bytes": 100}
        row = grp2.insert_file_record(ws, record, "RUN1")
        self._set_cell(ws, row, "DocType", "invoice")
        grp2.update_file_record(ws, row, {"size_bytes": 200}, "RUN2")
        assert self._cell(ws, row, "DocType") == "invoice"

    def test_keep_for_case_not_overwritten(self):
        ws = self._ws()
        record = {"file_id": "F00001", "file_path": "C:/f.pdf", "size_bytes": 100}
        row = grp2.insert_file_record(ws, record, "RUN1")
        self._set_cell(ws, row, "KeepForCase", "Y")
        grp2.update_file_record(ws, row, {"size_bytes": 200, "keep_for_case": "N"}, "RUN2")
        assert self._cell(ws, row, "KeepForCase") == "Y"


# ---------------------------------------------------------------------------
# TestWriteOrUpdateFileRecord
# ---------------------------------------------------------------------------

class TestWriteOrUpdateFileRecord:

    def _wb(self):
        wb = openpyxl.Workbook()
        grp2.ensure_master_file_inventory_sheet(wb)
        return wb

    def _record(self, path="C:/Docs/file.pdf"):
        return {"file_path": path, "filename": "file.pdf", "size_bytes": 512}

    def test_new_record_inserted(self):
        wb = self._wb()
        action, row = grp2.write_or_update_file_record(wb, self._record(), "RUN1")
        assert action == "inserted"
        assert row == 2

    def test_same_path_second_call_updated(self):
        wb = self._wb()
        grp2.write_or_update_file_record(wb, self._record(), "RUN1")
        action, row = grp2.write_or_update_file_record(wb, self._record(), "RUN2")
        assert action == "updated"
        assert row == 2

    def test_missing_file_path_raises_value_error(self):
        wb = self._wb()
        with pytest.raises(ValueError):
            grp2.write_or_update_file_record(wb, {"filename": "file.pdf"}, "RUN1")

    def test_two_different_paths_insert_two_rows(self):
        wb = self._wb()
        a1, r1 = grp2.write_or_update_file_record(wb, self._record("C:/A.pdf"), "RUN1")
        a2, r2 = grp2.write_or_update_file_record(wb, self._record("C:/B.pdf"), "RUN1")
        assert a1 == "inserted" and a2 == "inserted"
        assert r1 == 2 and r2 == 3

    def test_empty_file_path_string_raises_value_error(self):
        wb = self._wb()
        with pytest.raises(ValueError):
            grp2.write_or_update_file_record(wb, {"file_path": "", "filename": "x"}, "RUN1")
