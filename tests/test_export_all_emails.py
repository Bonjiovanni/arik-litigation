"""
Tests for email_pipeline/export_all_emails.py
"""

import json
import sys
import pytest
from pathlib import Path
import export_all_emails as exp


# ---------------------------------------------------------------------------
# TestSafeStr
# ---------------------------------------------------------------------------

class TestSafeStr:

    def test_none_returns_empty(self):
        assert exp.safe_str(None) == ""

    def test_truncates_at_32000(self):
        result = exp.safe_str("x" * 40000)
        assert len(result) == 32000

    def test_strips_bad_control_chars(self):
        # chr(1) is SOH — a bad control char
        result = exp.safe_str("hello\x01world")
        assert "\x01" not in result
        assert "helloworld" in result

    def test_keeps_newline(self):
        assert "\n" in exp.safe_str("a\nb")

    def test_keeps_carriage_return(self):
        assert "\r" in exp.safe_str("a\rb")

    def test_keeps_tab(self):
        assert "\t" in exp.safe_str("a\tb")

    def test_normal_string_unchanged(self):
        assert exp.safe_str("hello world") == "hello world"

    def test_number_converted_to_string(self):
        assert exp.safe_str(42) == "42"

    def test_unicode_preserved(self):
        assert exp.safe_str("café") == "café"


# ---------------------------------------------------------------------------
# TestColumnOrdering
# ---------------------------------------------------------------------------

class TestColumnOrdering:

    def _build_columns(self, records):
        all_keys = set()
        for r in records:
            all_keys.update(r.keys())
        priority_present = [c for c in exp.PRIORITY_COLS if c in all_keys]
        rest = sorted(all_keys - set(exp.PRIORITY_COLS))
        return priority_present + rest

    def test_priority_cols_appear_before_others(self):
        records = [{"Header.Date": "d", "ZZZ_other": "x", "Header.Subject": "s"}]
        columns = self._build_columns(records)
        assert columns.index("Header.Date") < columns.index("ZZZ_other")
        assert columns.index("Header.Subject") < columns.index("ZZZ_other")

    def test_non_priority_cols_sorted_alphabetically(self):
        records = [{"ZZZ": "z", "AAA": "a", "MMM": "m"}]
        columns = self._build_columns(records)
        rest = [c for c in columns if c not in exp.PRIORITY_COLS]
        assert rest == sorted(rest)

    def test_priority_cols_maintain_defined_order(self):
        records = [{"Header.Subject": "s", "Header.Date": "d", "Address.Sender": "x"}]
        columns = self._build_columns(records)
        priority_in_output = [c for c in columns if c in exp.PRIORITY_COLS]
        expected_order = [c for c in exp.PRIORITY_COLS if c in {"Header.Subject", "Header.Date", "Address.Sender"}]
        assert priority_in_output == expected_order

    def test_missing_priority_cols_not_included(self):
        records = [{"Header.Date": "d"}]
        columns = self._build_columns(records)
        # Header.Subject not in records — should not appear
        assert "Header.Subject" not in columns


# ---------------------------------------------------------------------------
# TestExcelIntegration
# ---------------------------------------------------------------------------

class TestExcelIntegration:

    def test_writes_xlsx_with_correct_row_count(self, tmp_path, monkeypatch):
        data = {"emails": [
            {"Header.Date": "2024-01-01", "Header.Subject": "Test 1", "body_clean": "Hello"},
            {"Header.Date": "2024-01-02", "Header.Subject": "Test 2", "body_clean": "World"},
        ]}
        input_path = tmp_path / "combined.json"
        input_path.write_text(json.dumps(data), encoding="utf-8")
        out_path = tmp_path / "output.xlsx"

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", lambda *a, **kw: out_path)
        monkeypatch.setattr(sys, "argv", ["export_all_emails.py", str(input_path)])

        exp.main()

        assert out_path.exists()
        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3  # 1 header + 2 data rows

    def test_writes_xlsx_with_correct_headers(self, tmp_path, monkeypatch):
        data = {"emails": [
            {"Header.Date": "2024-01-01", "Header.Subject": "Test", "body_clean": "Hi"},
        ]}
        input_path = tmp_path / "combined.json"
        input_path.write_text(json.dumps(data), encoding="utf-8")
        out_path = tmp_path / "output.xlsx"

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", lambda *a, **kw: out_path)
        monkeypatch.setattr(sys, "argv", ["export_all_emails.py", str(input_path)])

        exp.main()

        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Header.Date" in headers
        assert "Header.Subject" in headers
        assert "body_clean" in headers


# ---------------------------------------------------------------------------
# TestHeadlessMode (argv[2] = output path, no GUI)
# ---------------------------------------------------------------------------

class TestHeadlessMode:
    """Tests for fully non-interactive mode: argv[1]=input, argv[2]=output."""

    def _make_input_json(self, tmp_path, records=None):
        if records is None:
            records = [
                {"Header.Date": "2024-01-01", "Header.Subject": "Test 1", "body_clean": "Hello"},
                {"Header.Date": "2024-01-02", "Header.Subject": "Test 2", "body_clean": "World"},
            ]
        data = {"emails": records}
        input_path = tmp_path / "input.json"
        input_path.write_text(json.dumps(data), encoding="utf-8")
        return input_path

    def test_argv2_produces_xlsx_without_gui(self, tmp_path, monkeypatch):
        """When both argv[1] and argv[2] are provided, output is written
        without calling pick_output_file (no tkinter GUI)."""
        input_path = self._make_input_json(tmp_path)
        out_path = tmp_path / "headless_output.xlsx"

        picker_called = {"value": False}
        original_pick = exp.pick_output_file
        def spy_pick(*a, **kw):
            picker_called["value"] = True
            return original_pick(*a, **kw)

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", spy_pick)
        monkeypatch.setattr(sys, "argv", [
            "export_all_emails.py", str(input_path), str(out_path)
        ])

        exp.main()

        assert out_path.exists(), "Output xlsx should be created"
        assert not picker_called["value"], "pick_output_file should NOT be called in headless mode"

    def test_argv2_correct_row_count(self, tmp_path, monkeypatch):
        """Headless mode writes the correct number of rows."""
        input_path = self._make_input_json(tmp_path)
        out_path = tmp_path / "headless_rows.xlsx"

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", lambda *a, **kw: (_ for _ in ()).throw(AssertionError("GUI picker should not be called")))
        monkeypatch.setattr(sys, "argv", [
            "export_all_emails.py", str(input_path), str(out_path)
        ])

        exp.main()

        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3  # 1 header + 2 data

    def test_argv2_correct_headers(self, tmp_path, monkeypatch):
        """Headless mode writes the correct column headers."""
        input_path = self._make_input_json(tmp_path)
        out_path = tmp_path / "headless_headers.xlsx"

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", lambda *a, **kw: (_ for _ in ()).throw(AssertionError("GUI picker should not be called")))
        monkeypatch.setattr(sys, "argv", [
            "export_all_emails.py", str(input_path), str(out_path)
        ])

        exp.main()

        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Header.Date" in headers
        assert "Header.Subject" in headers
        assert "body_clean" in headers

    def test_argv2_creates_parent_directories(self, tmp_path, monkeypatch):
        """Headless mode creates parent directories if they don't exist."""
        input_path = self._make_input_json(tmp_path)
        out_path = tmp_path / "subdir" / "nested" / "output.xlsx"

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", lambda *a, **kw: (_ for _ in ()).throw(AssertionError("GUI picker should not be called")))
        monkeypatch.setattr(sys, "argv", [
            "export_all_emails.py", str(input_path), str(out_path)
        ])

        exp.main()

        assert out_path.exists()

    def test_argv2_overwrites_existing_file(self, tmp_path, monkeypatch):
        """Headless mode overwrites existing output without prompting."""
        input_path = self._make_input_json(tmp_path)
        out_path = tmp_path / "existing.xlsx"
        out_path.write_text("placeholder", encoding="utf-8")

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", lambda *a, **kw: (_ for _ in ()).throw(AssertionError("GUI picker should not be called")))
        monkeypatch.setattr(sys, "argv", [
            "export_all_emails.py", str(input_path), str(out_path)
        ])

        exp.main()

        assert out_path.exists()
        # Verify it's a real xlsx, not the placeholder text
        import openpyxl
        wb = openpyxl.load_workbook(str(out_path))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3

    def test_argv1_only_still_uses_picker(self, tmp_path, monkeypatch):
        """When only argv[1] is provided (no argv[2]), the output file
        picker is still called — existing behavior preserved."""
        input_path = self._make_input_json(tmp_path)
        out_path = tmp_path / "picker_output.xlsx"

        picker_called = {"value": False}
        def mock_pick(*a, **kw):
            picker_called["value"] = True
            return out_path

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", mock_pick)
        monkeypatch.setattr(sys, "argv", ["export_all_emails.py", str(input_path)])

        exp.main()

        assert picker_called["value"], "pick_output_file SHOULD be called when argv[2] is absent"
        assert out_path.exists()

    def test_argv2_returns_output_path(self, tmp_path, monkeypatch):
        """main() returns the output Path in headless mode (for chaining)."""
        input_path = self._make_input_json(tmp_path)
        out_path = tmp_path / "return_check.xlsx"

        monkeypatch.setattr(exp, "load_config", lambda: {})
        monkeypatch.setattr(exp, "save_config", lambda cfg: None)
        monkeypatch.setattr(exp, "pick_output_file", lambda *a, **kw: (_ for _ in ()).throw(AssertionError("GUI picker should not be called")))
        monkeypatch.setattr(sys, "argv", [
            "export_all_emails.py", str(input_path), str(out_path)
        ])

        result = exp.main()

        assert result == out_path
