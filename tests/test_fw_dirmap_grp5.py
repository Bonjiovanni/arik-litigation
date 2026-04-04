"""
tests/test_fw_dirmap_grp5.py
Unit tests for fw_dirmap_grp5: ensure_dir_processing_status_sheet,
initialize_processing_status_rows.
"""

import os
import sys

import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from fw_dirmap_grp5 import (
    ensure_dir_processing_status_sheet,
    initialize_processing_status_rows,
    SHEET_NAME,
    HEADERS,
    FILE_FAMILIES,
)


def _sample_records(n=2):
    return [
        {"dir_id": f"D{i:03d}", "full_path": f"C:/Root/dir{i}"}
        for i in range(1, n + 1)
    ]


# ---------------------------------------------------------------------------
# ensure_dir_processing_status_sheet
# ---------------------------------------------------------------------------

class TestEnsureDirProcessingStatusSheet:
    def test_creates_sheet_when_absent(self):
        wb = openpyxl.Workbook()
        ensure_dir_processing_status_sheet(wb)
        assert SHEET_NAME in wb.sheetnames

    def test_idempotent(self):
        wb = openpyxl.Workbook()
        ws1 = ensure_dir_processing_status_sheet(wb)
        ws1["A1"] = "SENTINEL"
        ws2 = ensure_dir_processing_status_sheet(wb)
        assert ws2["A1"].value == "SENTINEL"

    def test_header_labels_correct(self):
        wb = openpyxl.Workbook()
        ws = ensure_dir_processing_status_sheet(wb)
        actual = [ws.cell(row=1, column=i+1).value for i in range(len(HEADERS))]
        assert actual == HEADERS

    def test_freeze_panes(self):
        wb = openpyxl.Workbook()
        ws = ensure_dir_processing_status_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_auto_filter(self):
        wb = openpyxl.Workbook()
        ws = ensure_dir_processing_status_sheet(wb)
        assert ws.auto_filter.ref is not None


# ---------------------------------------------------------------------------
# initialize_processing_status_rows
# ---------------------------------------------------------------------------

class TestInitializeProcessingStatusRows:
    def test_appends_9_rows_per_directory(self):
        wb = openpyxl.Workbook()
        records = _sample_records(2)
        initialize_processing_status_rows(wb, records)
        ws = wb[SHEET_NAME]
        # 1 header + 2 dirs * 9 families = 19 rows
        assert ws.max_row == 1 + 2 * len(FILE_FAMILIES)

    def test_all_nine_families_present(self):
        wb = openpyxl.Workbook()
        records = _sample_records(1)
        initialize_processing_status_rows(wb, records)
        ws = wb[SHEET_NAME]
        family_col = HEADERS.index("FileFamily") + 1
        families = [ws.cell(row=r, column=family_col).value
                    for r in range(2, ws.max_row + 1)]
        assert set(families) == set(FILE_FAMILIES)

    def test_default_status_is_not_scanned(self):
        wb = openpyxl.Workbook()
        initialize_processing_status_rows(wb, _sample_records(1))
        ws = wb[SHEET_NAME]
        status_col = HEADERS.index("ProcessingStatus") + 1
        statuses = [ws.cell(row=r, column=status_col).value
                    for r in range(2, ws.max_row + 1)]
        assert all(s == "not_scanned" for s in statuses)

    def test_email_export_row_has_yellow_fill(self):
        wb = openpyxl.Workbook()
        initialize_processing_status_rows(wb, _sample_records(1))
        ws = wb[SHEET_NAME]
        family_col = HEADERS.index("FileFamily") + 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=family_col).value == "email_export":
                fill = ws.cell(row=row, column=1).fill
                assert fill.fgColor.rgb.upper().endswith("FFF2CC"), \
                    f"email_export row fill is {fill.fgColor.rgb}, expected FFF2CC"
                break

    def test_dir_id_written(self):
        wb = openpyxl.Workbook()
        initialize_processing_status_rows(wb, [{"dir_id": "D001", "full_path": "C:/x"}])
        ws = wb[SHEET_NAME]
        dir_id_col = HEADERS.index("DirID") + 1
        assert ws.cell(row=2, column=dir_id_col).value == "D001"

    def test_empty_records_writes_only_header(self):
        wb = openpyxl.Workbook()
        initialize_processing_status_rows(wb, [])
        ws = wb[SHEET_NAME]
        assert ws.max_row == 1
