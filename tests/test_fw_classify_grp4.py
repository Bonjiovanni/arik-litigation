"""
tests/test_fw_classify_grp4.py
-------------------------------
Unit tests for fw_classify_grp4.py:
    ensure_classify_history_sheet, log_classify_run
"""

import pytest
import openpyxl

import fw_classify_grp4 as grp4


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fresh_wb():
    wb = openpyxl.Workbook()
    if wb.sheetnames:
        del wb[wb.sheetnames[0]]
    return wb


_STATS = {
    "rows_processed": 10,
    "rows_updated": 8,
    "rows_skipped": 1,
    "errors": 1,
}


# ---------------------------------------------------------------------------
# TestEnsureClassifyHistorySheet
# ---------------------------------------------------------------------------

class TestEnsureClassifyHistorySheet:

    def test_creates_sheet_when_absent(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        assert grp4.CLASSIFY_HISTORY_SHEET in wb.sheetnames

    def test_returns_worksheet(self):
        wb = _fresh_wb()
        ws = grp4.ensure_classify_history_sheet(wb)
        assert ws.title == grp4.CLASSIFY_HISTORY_SHEET

    def test_idempotent_second_call(self):
        wb = _fresh_wb()
        ws1 = grp4.ensure_classify_history_sheet(wb)
        ws1.cell(row=2, column=1, value="MARKER")
        ws2 = grp4.ensure_classify_history_sheet(wb)
        assert ws2.cell(row=2, column=1).value == "MARKER"

    def test_header_row(self):
        wb = _fresh_wb()
        ws = grp4.ensure_classify_history_sheet(wb)
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(grp4._COLUMNS) + 1)]
        expected = [name for name, _ in grp4._COLUMNS]
        assert headers == expected

    def test_freeze_panes_set(self):
        wb = _fresh_wb()
        ws = grp4.ensure_classify_history_sheet(wb)
        assert ws.freeze_panes == "A2"

    def test_column_count_matches_schema(self):
        wb = _fresh_wb()
        ws = grp4.ensure_classify_history_sheet(wb)
        header_values = [ws.cell(row=1, column=c).value for c in range(1, 20)]
        non_null = [v for v in header_values if v is not None]
        assert len(non_null) == len(grp4._COLUMNS)

    def test_run_id_header_present(self):
        wb = _fresh_wb()
        ws = grp4.ensure_classify_history_sheet(wb)
        assert ws.cell(row=1, column=1).value == "RunID"


# ---------------------------------------------------------------------------
# TestLogClassifyRun
# ---------------------------------------------------------------------------

class TestLogClassifyRun:

    def test_writes_run_id(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        grp4.log_classify_run(wb, "CLASSIFY_20260315_120000", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp4.CLASSIFY_HISTORY_SHEET]
        assert ws.cell(row=2, column=1).value == "CLASSIFY_20260315_120000"

    def test_writes_started_at(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        grp4.log_classify_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp4.CLASSIFY_HISTORY_SHEET]
        assert ws.cell(row=2, column=2).value == "2026-03-15T12:00:00"

    def test_writes_stats(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        grp4.log_classify_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp4.CLASSIFY_HISTORY_SHEET]
        col_map = {name: idx + 1 for idx, (name, _) in enumerate(grp4._COLUMNS)}
        assert ws.cell(row=2, column=col_map["RowsProcessed"]).value == 10
        assert ws.cell(row=2, column=col_map["RowsUpdated"]).value == 8
        assert ws.cell(row=2, column=col_map["RowsSkipped"]).value == 1
        assert ws.cell(row=2, column=col_map["Errors"]).value == 1

    def test_returns_row_number(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        row_num = grp4.log_classify_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        assert row_num == 2

    def test_second_run_appends(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        grp4.log_classify_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        row_num = grp4.log_classify_run(wb, "RUN2", "2026-03-15T13:00:00", _STATS)
        assert row_num == 3
        ws = wb[grp4.CLASSIFY_HISTORY_SHEET]
        assert ws.cell(row=3, column=1).value == "RUN2"

    def test_notes_written(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        grp4.log_classify_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS, notes="test run")
        ws = wb[grp4.CLASSIFY_HISTORY_SHEET]
        col_map = {name: idx + 1 for idx, (name, _) in enumerate(grp4._COLUMNS)}
        assert ws.cell(row=2, column=col_map["Notes"]).value == "test run"

    def test_stats_defaults_to_zero(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        grp4.log_classify_run(wb, "RUN1", "2026-03-15T12:00:00", {})
        ws = wb[grp4.CLASSIFY_HISTORY_SHEET]
        col_map = {name: idx + 1 for idx, (name, _) in enumerate(grp4._COLUMNS)}
        assert ws.cell(row=2, column=col_map["RowsProcessed"]).value == 0
        assert ws.cell(row=2, column=col_map["Errors"]).value == 0

    def test_creates_sheet_if_missing(self):
        wb = _fresh_wb()
        # Do NOT call ensure_classify_history_sheet first
        grp4.log_classify_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        assert grp4.CLASSIFY_HISTORY_SHEET in wb.sheetnames

    def test_completed_at_is_populated(self):
        wb = _fresh_wb()
        grp4.ensure_classify_history_sheet(wb)
        grp4.log_classify_run(wb, "RUN1", "2026-03-15T12:00:00", _STATS)
        ws = wb[grp4.CLASSIFY_HISTORY_SHEET]
        col_map = {name: idx + 1 for idx, (name, _) in enumerate(grp4._COLUMNS)}
        completed_at = ws.cell(row=2, column=col_map["CompletedAt"]).value
        assert completed_at is not None
        assert len(completed_at) > 0
