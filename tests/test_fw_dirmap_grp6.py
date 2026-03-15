"""
tests/test_fw_dirmap_grp6.py
Unit tests for fw_dirmap_grp6: log_dirmap_run, print_run_summary.
main() is interactive and not unit-tested here.
"""

import os
import sys
import json
from datetime import datetime, timedelta

import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from fw_dirmap_grp6 import log_dirmap_run, print_run_summary


def _make_stats(elapsed_sec=5.0):
    start = datetime(2026, 3, 15, 14, 30, 0)
    end   = start + timedelta(seconds=elapsed_sec)
    return {
        "start_time": start,
        "end_time":   end,
        "dir_count":  42,
        "root_count": 2,
    }

def _make_config():
    return {
        "roots":     ["C:/Root1", "C:/Root2"],
        "recursive": True,
        "max_depth": None,
    }

def _sample_records():
    return [
        {"dir_id": "D001", "full_path": "C:/Root1", "depth": 0},
        {"dir_id": "D002", "full_path": "C:/Root1/sub1", "depth": 1},
        {"dir_id": "D003", "full_path": "C:/Root2", "depth": 0},
    ]


# ---------------------------------------------------------------------------
# log_dirmap_run
# ---------------------------------------------------------------------------

class TestLogDirmapRun:
    def test_creates_walk_history_sheet(self):
        wb = openpyxl.Workbook()
        log_dirmap_run(wb, "DIRMAP_20260315_143000", _make_config(), _make_stats())
        assert "Walk_History" in wb.sheetnames

    def test_appends_one_data_row(self):
        wb = openpyxl.Workbook()
        log_dirmap_run(wb, "DIRMAP_20260315_143000", _make_config(), _make_stats())
        ws = wb["Walk_History"]
        assert ws.max_row == 2  # header + 1 data row

    def test_run_type_is_dirmap(self):
        wb = openpyxl.Workbook()
        log_dirmap_run(wb, "RUN001", _make_config(), _make_stats())
        ws = wb["Walk_History"]
        # RunType is column B (index 2)
        assert ws.cell(row=2, column=2).value == "DIRMAP"

    def test_run_id_written(self):
        wb = openpyxl.Workbook()
        log_dirmap_run(wb, "MY_RUN_ID", _make_config(), _make_stats())
        ws = wb["Walk_History"]
        assert ws.cell(row=2, column=1).value == "MY_RUN_ID"

    def test_elapsed_seconds_correct(self):
        wb = openpyxl.Workbook()
        log_dirmap_run(wb, "R1", _make_config(), _make_stats(elapsed_sec=7.5))
        ws = wb["Walk_History"]
        # ElapsedSeconds is column E (index 5)
        assert ws.cell(row=2, column=5).value == pytest.approx(7.5, abs=0.01)

    def test_config_stored_as_json(self):
        wb = openpyxl.Workbook()
        config = _make_config()
        log_dirmap_run(wb, "R1", config, _make_stats())
        ws = wb["Walk_History"]
        # Config is column H (index 8)
        raw = ws.cell(row=2, column=8).value
        parsed = json.loads(raw)
        assert parsed["recursive"] is True

    def test_second_run_appends_not_overwrites(self):
        wb = openpyxl.Workbook()
        log_dirmap_run(wb, "R1", _make_config(), _make_stats())
        log_dirmap_run(wb, "R2", _make_config(), _make_stats())
        ws = wb["Walk_History"]
        assert ws.max_row == 3  # header + 2 runs

    def test_freeze_panes_on_new_sheet(self):
        wb = openpyxl.Workbook()
        log_dirmap_run(wb, "R1", _make_config(), _make_stats())
        ws = wb["Walk_History"]
        assert ws.freeze_panes == "A2"


# ---------------------------------------------------------------------------
# print_run_summary
# ---------------------------------------------------------------------------

class TestPrintRunSummary:
    def test_runs_without_error(self, capsys):
        records = _sample_records()
        start = datetime(2026, 3, 15, 14, 30, 0)
        end   = start + timedelta(seconds=3.5)
        print_run_summary(records, "DIRMAP_20260315_143000", start, end)
        out = capsys.readouterr().out
        assert "fw_dirmap" in out

    def test_shows_run_id(self, capsys):
        records = _sample_records()
        start = datetime(2026, 3, 15, 14, 30, 0)
        end   = start + timedelta(seconds=1)
        print_run_summary(records, "DIRMAP_TESTID", start, end)
        out = capsys.readouterr().out
        assert "DIRMAP_TESTID" in out

    def test_shows_dir_count(self, capsys):
        records = _sample_records()
        start = datetime(2026, 3, 15, 14, 30, 0)
        end   = start + timedelta(seconds=1)
        print_run_summary(records, "R1", start, end)
        out = capsys.readouterr().out
        assert "3" in out  # 3 total records

    def test_only_depth_zero_listed_as_top_level(self, capsys):
        records = _sample_records()
        start = datetime(2026, 3, 15, 14, 30, 0)
        end   = start + timedelta(seconds=1)
        print_run_summary(records, "R1", start, end)
        out = capsys.readouterr().out
        # sub1 (depth=1) should NOT appear in top-level list
        assert "sub1" not in out
