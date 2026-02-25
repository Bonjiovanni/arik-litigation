"""
build_workbook.py
=================
Generates workbook.xlsx with Android-safe Excel structured table formulas.

Pipeline:
  1. openpyxl  — builds sheets, tables, styles, seed data, placeholder formulas
  2. post_process() — opens the ZIP and injects:
       a. <calculatedColumnFormula> into each tblMasterIndex column in table XML
       b. Shared-formula t="shared" pattern in worksheet XML

This two-step pipeline is permanent by design.  openpyxl cannot produce
<calculatedColumnFormula> or t="shared" natively; the post-processor is not
a repair step, it IS the build step.  Always regenerate through this script —
never hand-edit the .xlsx.
"""

import copy
import io
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "workbook.xlsx"

# ---------------------------------------------------------------------------
# Seed data
# ---------------------------------------------------------------------------

ENTITIES = [
    # Entity_ID, Canonical_Name, Entity_Type, DOB,        DOD,        Notes
    ("E001", "Alice Johnson",  "Person",       "1980-03-15", "",           ""),
    ("E002", "Acme Corp",      "Organization", "",           "",           ""),
    ("E003", "Bob Smith",      "Person",       "1975-07-22", "2023-01-10", "Deceased"),
]

ALIASES = [
    # Entity_ID, Alias_Value
    ("E001", "AJ"),
    ("E003", "Robert Smith"),
]

CONTACTS = [
    # Entity_ID, Contact_Type,  Contact_Value
    ("E001", "Email",        "alice@example.com"),
    ("E001", "Phone Mobile", "555-0101"),
    ("E002", "Email",        "info@acme.com"),
    ("E002", "URL",          "https://acme.com"),
    ("E003", "Phone Main",   "555-0202"),
]

ENTITY_ROLES = [
    # Entity_ID, Matter,      Role_Context,  Notes
    ("E001", "Matter-001", "Lead Counsel", ""),
    ("E002", "Matter-001", "Client",       ""),
    ("E003", "Matter-002", "Witness",      ""),
]

# ---------------------------------------------------------------------------
# tblMasterIndex calculated column formulas
# Key: column header (must match exactly)
# Value: formula WITHOUT leading "=" (for calculatedColumnFormula XML element)
#        WITH leading "=" (for worksheet cell formula, openpyxl placeholder)
# ---------------------------------------------------------------------------

MASTER_FORMULAS = {
    "Entity_ID":      "INDEX(tblEntities[Entity_ID],ROW()-1)",
    "Canonical_Name": 'IFERROR(XLOOKUP([@Entity_ID],tblEntities[Entity_ID],tblEntities[Canonical_Name]),"")',
    "Entity_Type":    'IFERROR(XLOOKUP([@Entity_ID],tblEntities[Entity_ID],tblEntities[Entity_Type]),"")',
    "DOB":            'IFERROR(XLOOKUP([@Entity_ID],tblEntities[Entity_ID],tblEntities[DOB]),"")',
    "DOD":            'IFERROR(XLOOKUP([@Entity_ID],tblEntities[Entity_ID],tblEntities[DOD]),"")',
    "Aliases":        'IFERROR(TEXTJOIN("; ",TRUE,FILTER(tblAliases[Alias_Value],tblAliases[Entity_ID]=[@Entity_ID])),"")',
    "Emails":         'IFERROR(TEXTJOIN("; ",TRUE,FILTER(tblContacts[Contact_Value],(tblContacts[Entity_ID]=[@Entity_ID])*(tblContacts[Contact_Type]="Email"))),"")',
    "Phone_Main":     'IFERROR(TEXTJOIN("; ",TRUE,FILTER(tblContacts[Contact_Value],(tblContacts[Entity_ID]=[@Entity_ID])*(tblContacts[Contact_Type]="Phone Main"))),"")',
    "Phone_Mobile":   'IFERROR(TEXTJOIN("; ",TRUE,FILTER(tblContacts[Contact_Value],(tblContacts[Entity_ID]=[@Entity_ID])*(tblContacts[Contact_Type]="Phone Mobile"))),"")',
    "URLs":           'IFERROR(TEXTJOIN("; ",TRUE,FILTER(tblContacts[Contact_Value],(tblContacts[Entity_ID]=[@Entity_ID])*(tblContacts[Contact_Type]="URL"))),"")',
}

MASTER_COLUMNS = list(MASTER_FORMULAS.keys())

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT     = Font(name="Calibri", size=11)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center")

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row_num, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER_ALIGN

def add_table(ws, name, headers, data, start_row=1):
    """Write headers + data, register an Excel Table, return the Table."""
    # Headers
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, len(headers))

    # Data rows
    for r, row in enumerate(data, start_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = LEFT_ALIGN

    end_row = start_row + len(data)  # last data row (or header if no data)
    if len(data) == 0:
        # openpyxl requires at least 2 rows for a table; add a blank row
        end_row = start_row + 1

    end_col = get_column_letter(len(headers))
    ref = f"A{start_row}:{end_col}{end_row}"

    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,   showColumnStripes=False,
    )
    ws.add_table(tbl)
    return tbl

# ---------------------------------------------------------------------------
# Phase 1 — build with openpyxl
# ---------------------------------------------------------------------------

def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---- 01_Entities -------------------------------------------------------
    ws_ent = wb.create_sheet("01_Entities")
    ent_headers = ["Entity_ID", "Canonical_Name", "Entity_Type", "DOB", "DOD", "Notes"]
    add_table(ws_ent, "tblEntities", ent_headers, ENTITIES)
    ws_ent.column_dimensions["A"].width = 12
    ws_ent.column_dimensions["B"].width = 22
    ws_ent.column_dimensions["C"].width = 16
    ws_ent.column_dimensions["D"].width = 14
    ws_ent.column_dimensions["E"].width = 14
    ws_ent.column_dimensions["F"].width = 30

    # ---- 02_Aliases --------------------------------------------------------
    ws_ali = wb.create_sheet("02_Aliases")
    ali_headers = ["Entity_ID", "Alias_Value"]
    add_table(ws_ali, "tblAliases", ali_headers, ALIASES)
    ws_ali.column_dimensions["A"].width = 12
    ws_ali.column_dimensions["B"].width = 25

    # ---- 03_Contacts -------------------------------------------------------
    ws_con = wb.create_sheet("03_Contacts")
    con_headers = ["Entity_ID", "Contact_Type", "Contact_Value"]
    add_table(ws_con, "tblContacts", con_headers, CONTACTS)
    ws_con.column_dimensions["A"].width = 12
    ws_con.column_dimensions["B"].width = 16
    ws_con.column_dimensions["C"].width = 35

    # ---- 04_EntityRoles ----------------------------------------------------
    ws_rol = wb.create_sheet("04_EntityRoles")
    rol_headers = ["Entity_ID", "Matter", "Role_Context", "Notes"]
    add_table(ws_rol, "tblEntityRoles", rol_headers, ENTITY_ROLES)
    ws_rol.column_dimensions["A"].width = 12
    ws_rol.column_dimensions["B"].width = 16
    ws_rol.column_dimensions["C"].width = 20
    ws_rol.column_dimensions["D"].width = 30

    # ---- 90_MasterIndex ----------------------------------------------------
    ws_mst = wb.create_sheet("90_MasterIndex")
    n_entities = len(ENTITIES)

    # Write placeholder formulas — post-processor will convert to t="shared"
    ws_mst.append(MASTER_COLUMNS)                 # row 1: headers
    for row_idx in range(1, n_entities + 1):       # rows 2..N+1
        ws_row = []
        for col_name in MASTER_COLUMNS:
            formula = "=" + MASTER_FORMULAS[col_name]
            ws_row.append(formula)
        ws_mst.append(ws_row)

    style_header_row(ws_mst, 1, len(MASTER_COLUMNS))

    col_widths = {
        "Entity_ID": 12, "Canonical_Name": 22, "Entity_Type": 16,
        "DOB": 14, "DOD": 14, "Aliases": 30, "Emails": 35,
        "Phone_Main": 18, "Phone_Mobile": 18, "URLs": 35,
    }
    for i, col_name in enumerate(MASTER_COLUMNS, 1):
        ltr = get_column_letter(i)
        ws_mst.column_dimensions[ltr].width = col_widths.get(col_name, 18)
        cell = ws_mst.cell(row=1, column=i)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN

    end_row_mst = 1 + n_entities
    end_col_mst = get_column_letter(len(MASTER_COLUMNS))
    mst_ref = f"A1:{end_col_mst}{end_row_mst}"
    tbl_mst = Table(displayName="tblMasterIndex", ref=mst_ref)
    tbl_mst.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium6",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,   showColumnStripes=False,
    )
    ws_mst.add_table(tbl_mst)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ---------------------------------------------------------------------------
# Phase 2 — post-process the ZIP
# ---------------------------------------------------------------------------

# Namespace map for OOXML
NS = {
    "ss":  "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r":   "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "mc":  "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
}
SS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

def _reg_ns():
    """Register namespaces so ET serialises them without ns0: prefixes."""
    ET.register_namespace("",      SS)
    ET.register_namespace("r",     NS["r"])
    ET.register_namespace("mc",    NS["mc"])
    ET.register_namespace("x14ac", NS["x14ac"])

def _col_letter(col_name):
    idx = MASTER_COLUMNS.index(col_name) + 1
    return get_column_letter(idx)

def _cell_ref(col_letter, row_num):
    return f"{col_letter}{row_num}"

def fix_table_xml(xml_bytes):
    """
    Inject <calculatedColumnFormula> into each tblMasterIndex tableColumn.
    Returns modified XML bytes.
    """
    _reg_ns()
    root = ET.fromstring(xml_bytes)

    # Confirm this is tblMasterIndex
    disp = root.get("displayName", "")
    if disp != "tblMasterIndex":
        return xml_bytes  # leave other tables untouched

    table_columns_el = root.find(f"{{{SS}}}tableColumns")
    if table_columns_el is None:
        return xml_bytes

    for col_el in table_columns_el.findall(f"{{{SS}}}tableColumn"):
        col_name = col_el.get("name", "")
        if col_name not in MASTER_FORMULAS:
            continue

        formula_text = MASTER_FORMULAS[col_name]

        # Remove any existing calculatedColumnFormula child
        existing = col_el.find(f"{{{SS}}}calculatedColumnFormula")
        if existing is not None:
            col_el.remove(existing)

        # Create and insert <calculatedColumnFormula>
        ccf = ET.SubElement(col_el, f"{{{SS}}}calculatedColumnFormula")
        ccf.text = formula_text

    return ET.tostring(root, encoding="unicode", xml_declaration=False).encode("utf-8")

def col_addr_to_num(col_letter):
    """Convert column letter(s) to 1-based integer."""
    col_letter = col_letter.upper()
    num = 0
    for ch in col_letter:
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num

def parse_cell_ref(ref):
    """Return (col_letter, row_number) from e.g. 'AB12'."""
    m = re.match(r'^([A-Z]+)(\d+)$', ref.upper())
    if not m:
        raise ValueError(f"Cannot parse cell ref: {ref!r}")
    return m.group(1), int(m.group(2))

def fix_worksheet_xml(xml_bytes, n_data_rows):
    """
    Convert formula cells in tblMasterIndex columns from plain <f>formula</f>
    to the shared-formula pattern:
      - First data row:  <f t="shared" si="N" ref="ColR:ColLastR">formula</f>
      - Other data rows: <f t="shared" si="N"/>

    n_data_rows: number of data rows in tblMasterIndex (excluding header).
    """
    _reg_ns()
    root = ET.fromstring(xml_bytes)

    sheet_data = root.find(f"{{{SS}}}sheetData")
    if sheet_data is None:
        return xml_bytes

    # Build a map: col_letter -> (formula_string, si_index)
    col_formulas = {}
    for si_idx, col_name in enumerate(MASTER_COLUMNS):
        ltr = _col_letter(col_name)
        col_formulas[ltr] = (MASTER_FORMULAS[col_name], si_idx)

    first_data_row = 2         # header is row 1
    last_data_row  = 1 + n_data_rows

    for row_el in sheet_data.findall(f"{{{SS}}}row"):
        row_num_str = row_el.get("r", "")
        try:
            row_num = int(row_num_str)
        except ValueError:
            continue

        if row_num < first_data_row or row_num > last_data_row:
            continue

        is_first = (row_num == first_data_row)

        for c_el in row_el.findall(f"{{{SS}}}c"):
            cell_ref = c_el.get("r", "")
            try:
                col_ltr, _ = parse_cell_ref(cell_ref)
            except ValueError:
                continue

            if col_ltr not in col_formulas:
                continue

            formula_text, si_idx = col_formulas[col_ltr]

            # Remove all existing <f> and <v> children
            for child_tag in [f"{{{SS}}}f", f"{{{SS}}}v"]:
                for child in c_el.findall(child_tag):
                    c_el.remove(child)

            # Build new <f> element
            f_el = ET.Element(f"{{{SS}}}f")
            f_el.set("t", "shared")
            f_el.set("si", str(si_idx))
            if is_first:
                ref_range = f"{col_ltr}{first_data_row}:{col_ltr}{last_data_row}"
                f_el.set("ref", ref_range)
                f_el.text = formula_text
            # else: no text, no ref — just t="shared" si="N"

            # Add <v> placeholder (empty string value)
            v_el = ET.Element(f"{{{SS}}}v")

            c_el.insert(0, f_el)
            c_el.append(v_el)

    return ET.tostring(root, encoding="unicode", xml_declaration=False).encode("utf-8")

def post_process(xlsx_bytes, n_data_rows):
    """
    Open the .xlsx ZIP, fix table XML and worksheet XML, return fixed bytes.
    """
    in_buf  = io.BytesIO(xlsx_bytes)
    out_buf = io.BytesIO()

    with zipfile.ZipFile(in_buf, "r") as zin, \
         zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:

        # We need to know which sheet is 90_MasterIndex.
        # Read workbook.xml to find sheetId -> rId -> sheet file name.
        # Also read workbook.xml.rels to map rId -> file path.
        # Simpler: scan all sheets for the one whose name is "90_MasterIndex".

        # First pass: collect all names
        names = zin.namelist()

        # Parse workbook.xml to find sheet order
        wb_xml = zin.read("xl/workbook.xml")
        wb_root = ET.fromstring(wb_xml)
        sheets_el = wb_root.find(f"{{{SS}}}sheets")
        sheet_rids = {}  # sheetName -> rId
        if sheets_el is not None:
            for sh in sheets_el.findall(f"{{{SS}}}sheet"):
                sname = sh.get("name", "")
                rid   = sh.get(f"{{{NS['r']}}}id", "")
                sheet_rids[sname] = rid

        # Parse xl/_rels/workbook.xml.rels to map rId -> xl/worksheets/sheetN.xml
        rels_xml  = zin.read("xl/_rels/workbook.xml.rels")
        rels_root = ET.fromstring(rels_xml)
        rid_to_path = {}
        RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
        for rel in rels_root.findall(f"{{{RELS_NS}}}Relationship"):
            rid_to_path[rel.get("Id", "")] = "xl/" + rel.get("Target", "").lstrip("/")

        master_sheet_path = None
        master_rid = sheet_rids.get("90_MasterIndex")
        if master_rid:
            raw_target = rid_to_path.get(master_rid, "")
            # normalise: xl/xl/... can appear if Target already has xl/ prefix
            if raw_target.startswith("xl/xl/"):
                raw_target = raw_target[3:]
            master_sheet_path = raw_target

        # Find tblMasterIndex table file
        # Tables are in xl/tables/tableN.xml; we match by displayName inside.
        master_table_path = None
        for name in names:
            if name.startswith("xl/tables/") and name.endswith(".xml"):
                data = zin.read(name)
                if b"tblMasterIndex" in data:
                    master_table_path = name
                    break

        # Second pass: rewrite
        for item in names:
            data = zin.read(item)

            if item == master_table_path:
                data = fix_table_xml(data)
            elif master_sheet_path and item == master_sheet_path:
                data = fix_worksheet_xml(data, n_data_rows)

            zout.writestr(item, data)

    out_buf.seek(0)
    return out_buf.read()

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Phase 1: building workbook with openpyxl...")
    xlsx_bytes = build_workbook()

    print("Phase 2: post-processing OOXML (inject calculatedColumnFormula + shared formulas)...")
    n_data_rows = len(ENTITIES)
    fixed_bytes = post_process(xlsx_bytes, n_data_rows)

    with open(OUTPUT_PATH, "wb") as f:
        f.write(fixed_bytes)

    print(f"Done: {OUTPUT_PATH}  ({len(fixed_bytes):,} bytes)")

    # Verification: confirm <calculatedColumnFormula> is present
    with zipfile.ZipFile(OUTPUT_PATH) as z:
        names = z.namelist()
        for name in names:
            if name.startswith("xl/tables/") and name.endswith(".xml"):
                data = z.read(name)
                if b"tblMasterIndex" in data:
                    count = data.count(b"<calculatedColumnFormula")
                    print(f"  [{name}] calculatedColumnFormula elements: {count} "
                          f"(expected {len(MASTER_COLUMNS)})")
        for name in names:
            if name.startswith("xl/worksheets/"):
                data = z.read(name)
                if b"90_MasterIndex" in data or True:
                    shared_count = data.count(b't="shared"')
                    if shared_count:
                        print(f"  [{name}] t=\"shared\" formula cells: {shared_count}")

if __name__ == "__main__":
    main()
