"""Generate sample Excel comparing directory inventory format options."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_header(ws, num_cols):
    """Apply header styling: bold white text on dark blue background."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def auto_width(ws, num_cols, max_width=50):
    """Set column widths based on content."""
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, max_width)

# Simulated Windows directory tree
SAMPLE_DIRS = [
    # (full_path, depth_from_root, file_count, subdir_count, status)
    (r"C:\Users\arika\OneDrive", 0, 5, 4, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Documents", 1, 23, 3, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Documents\Tax2024", 2, 15, 0, "file_walked"),
    (r"C:\Users\arika\OneDrive\Documents\Contracts", 2, 8, 1, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Documents\Contracts\Signed", 3, 4, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Documents\Personal", 2, 34, 2, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Documents\Personal\Medical", 3, 12, 0, "triaged"),
    (r"C:\Users\arika\OneDrive\Documents\Personal\Financial", 3, 22, 0, "deep_processed"),
    (r"C:\Users\arika\OneDrive\Litigation", 1, 12, 3, "file_walked"),
    (r"C:\Users\arika\OneDrive\Litigation\ClientA", 2, 47, 3, "triaged"),
    (r"C:\Users\arika\OneDrive\Litigation\ClientA\Financials", 3, 23, 0, "deep_processed"),
    (r"C:\Users\arika\OneDrive\Litigation\ClientA\Medical", 3, 89, 2, "classified"),
    (r"C:\Users\arika\OneDrive\Litigation\ClientA\Medical\2024", 4, 34, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Litigation\ClientA\Medical\2023", 4, 55, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Litigation\ClientA\Correspondence", 3, 31, 0, "file_walked"),
    (r"C:\Users\arika\OneDrive\Litigation\PropertyDocs", 2, 15, 1, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Litigation\PropertyDocs\Photos", 3, 203, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Litigation\Insurance", 2, 28, 0, "classified"),
    (r"C:\Users\arika\OneDrive\Pictures", 1, 156, 2, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Pictures\Screenshots", 2, 89, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Pictures\Family", 2, 67, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Downloads", 1, 312, 5, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Downloads\Aid4Mail_Export", 2, 1547, 3, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Downloads\Aid4Mail_Export\Inbox", 3, 823, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Downloads\Aid4Mail_Export\Sent", 3, 512, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Downloads\Aid4Mail_Export\Drafts", 3, 212, 0, "not_scanned"),
    (r"C:\Users\arika\OneDrive\Downloads\Misc", 2, 45, 0, "not_scanned"),
]

STATUS_COLORS = {
    "not_scanned": "F2F2F2",      # light gray
    "file_walked": "DAEEF3",      # light blue
    "classified": "E2EFDA",       # light green
    "triaged": "FFF2CC",          # light yellow
    "deep_processed": "D5A6BD",   # light purple
    "reviewed": "B6D7A8",         # green
    "skipped": "F4CCCC",          # light red
}

def build_option_a(wb):
    """Option A: Full path in one cell + metadata columns."""
    ws = wb.create_sheet("Option_A_FullPath")

    headers = ["DirID", "DirPath", "DirName", "ParentPath", "Depth",
               "FileCount", "SubdirCount", "ProcessingStatus", "Notes"]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, (full_path, depth, files, subdirs, status) in enumerate(SAMPLE_DIRS, start=1):
        parts = full_path.replace("\\", "/").split("/")
        dir_name = parts[-1]
        parent = "\\".join(parts[:-1]) if len(parts) > 1 else ""

        row_num = i + 1  # +1 for header
        row = [f"D{i:03d}", full_path, dir_name, parent, depth,
               files, subdirs, status, ""]
        ws.append(row)

        # Color the status cell
        status_cell = ws.cell(row=row_num, column=8)
        if status in STATUS_COLORS:
            status_cell.fill = PatternFill(start_color=STATUS_COLORS[status],
                                           end_color=STATUS_COLORS[status],
                                           fill_type="solid")

        # Indent DirName by depth
        name_cell = ws.cell(row=row_num, column=3)
        name_cell.value = ("  " * depth) + dir_name
        name_cell.alignment = Alignment(horizontal="left")

        # Row grouping for expand/collapse (group children under parents)
        if depth > 0:
            ws.row_dimensions[row_num].outline_level = depth

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    auto_width(ws, len(headers))
    # Make DirPath column wider
    ws.column_dimensions["B"].width = 55


def build_option_c(wb):
    """Option C: Hybrid — ScanRoot + RelativeDir + FullPath, depth relative to scan root."""
    ws = wb.create_sheet("Option_C_Hybrid")

    # Simulate scan root = C:\Users\arika\OneDrive
    scan_root = r"C:\Users\arika\OneDrive"

    headers = ["DirID", "ScanRoot", "RelativeDir", "DirName", "FullPath",
               "Depth", "FileCount", "SubdirCount", "ProcessingStatus", "Notes"]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, (full_path, depth, files, subdirs, status) in enumerate(SAMPLE_DIRS, start=1):
        parts = full_path.replace("\\", "/").split("/")
        dir_name = parts[-1]

        # Relative dir from scan root
        if full_path == scan_root:
            rel_dir = "."
        elif full_path.startswith(scan_root + "\\"):
            rel_dir = full_path[len(scan_root) + 1:]
        else:
            rel_dir = full_path

        row_num = i + 1
        row = [f"D{i:03d}", scan_root, rel_dir, dir_name, full_path,
               depth, files, subdirs, status, ""]
        ws.append(row)

        # Color the status cell
        status_cell = ws.cell(row=row_num, column=9)
        if status in STATUS_COLORS:
            status_cell.fill = PatternFill(start_color=STATUS_COLORS[status],
                                           end_color=STATUS_COLORS[status],
                                           fill_type="solid")

        # Indent DirName by depth
        name_cell = ws.cell(row=row_num, column=4)
        name_cell.value = ("  " * depth) + dir_name
        name_cell.alignment = Alignment(horizontal="left")

        # Row grouping for expand/collapse
        if depth > 0:
            ws.row_dimensions[row_num].outline_level = depth

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    auto_width(ws, len(headers))
    ws.column_dimensions["E"].width = 55


def build_processing_status_sample(wb):
    """Dir_Processing_Status sheet — tracks processing per directory per file type."""
    ws = wb.create_sheet("Dir_Processing_Status")

    headers = ["DirID", "DirPath", "FileFamily", "ProcessingLevel",
               "FileCount", "LastRunID", "LastProcessedAt", "Notes"]
    ws.append(headers)
    style_header(ws, len(headers))

    sample_data = [
        ("D010", r"...\Litigation\ClientA", "pdf", "text_extracted", 12, "RUN_20260310", "2026-03-10", ""),
        ("D010", r"...\Litigation\ClientA", "image", "ocr_complete", 8, "RUN_20260312", "2026-03-12", ""),
        ("D010", r"...\Litigation\ClientA", "word_doc", "not_processed", 15, "", "", ""),
        ("D010", r"...\Litigation\ClientA", "spreadsheet", "classified", 5, "RUN_20260311", "2026-03-11", ""),
        ("D010", r"...\Litigation\ClientA", "email_file", "skipped", 7, "RUN_20260310", "2026-03-10", "Already in Aid4Mail"),
        ("D011", r"...\Litigation\ClientA\Financials", "pdf", "deep_processed", 18, "RUN_20260313", "2026-03-13", "Form parser applied"),
        ("D011", r"...\Litigation\ClientA\Financials", "spreadsheet", "triaged", 5, "RUN_20260312", "2026-03-12", ""),
        ("D011", r"...\Litigation\ClientA\Financials", "image", "not_processed", 3, "", "", "Screenshots of statements"),
        ("D017", r"...\Litigation\PropertyDocs\Photos", "image", "not_processed", 203, "", "", "Mostly property condition photos"),
    ]

    processing_colors = {
        "not_processed": "F2F2F2",
        "file_walked": "DAEEF3",
        "classified": "E2EFDA",
        "triaged": "FFF2CC",
        "text_extracted": "D9EAD3",
        "ocr_complete": "B6D7A8",
        "deep_processed": "D5A6BD",
        "skipped": "F4CCCC",
    }

    for row_data in sample_data:
        ws.append(row_data)
        row_num = ws.max_row
        level = row_data[3]
        cell = ws.cell(row=row_num, column=4)
        if level in processing_colors:
            cell.fill = PatternFill(start_color=processing_colors[level],
                                     end_color=processing_colors[level],
                                     fill_type="solid")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    auto_width(ws, len(headers))


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_option_a(wb)
    build_option_c(wb)
    build_processing_status_sample(wb)

    output_path = "/home/user/Repo-for-Claude-android/dir_format_sample.xlsx"
    wb.save(output_path)
    print(f"Sample Excel saved to: {output_path}")
    print("Sheets: Option_A_FullPath, Option_C_Hybrid, Dir_Processing_Status")
    print("\nOpen in Excel to test the expand/collapse row grouping!")


if __name__ == "__main__":
    main()
