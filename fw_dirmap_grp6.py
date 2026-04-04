"""
fw_dirmap_grp6.py
=================
Group 6 of the fw_dirmap directory mapper agent.

Contains the run-logging, summary-printing, and main orchestration functions:
  - log_dirmap_run()
  - print_run_summary()
  - main()

Part of the multi-file litigation file-walker system.
Python 3.11 | Windows 10 | openpyxl
"""

import json
import os
from datetime import datetime

from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# FUNCTION 1: log_dirmap_run
# ---------------------------------------------------------------------------

def log_dirmap_run(wb, run_id: str, config: dict, stats: dict) -> None:
    """Write a single summary row to the Walk_History sheet.

    Creates the sheet with a styled header if it does not yet exist,
    then appends one row for the current run.

    Parameters
    ----------
    wb : openpyxl.Workbook
        The open master workbook.
    run_id : str
        Unique identifier for this run (e.g. "DIRMAP_20260315_143022").
    config : dict
        Run configuration with keys:
            roots (list), recursive (bool), max_depth (int | None)
    stats : dict
        Run statistics with keys:
            start_time (datetime), end_time (datetime),
            dir_count (int), root_count (int)
    """
    SHEET_NAME   = "Walk_History"
    HEADER_FONT  = Font(bold=True, color="FFFFFF")
    HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

    COLUMNS = [
        "RunID",
        "RunType",
        "StartTime",
        "EndTime",
        "ElapsedSeconds",
        "RootCount",
        "DirCount",
        "Config",
    ]

    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        ws.freeze_panes = "A2"
    else:
        ws = wb[SHEET_NAME]

    start_time: datetime = stats["start_time"]
    end_time: datetime   = stats["end_time"]
    elapsed = round((end_time - start_time).total_seconds(), 2)

    row_data = [
        run_id,
        "DIRMAP",
        start_time.isoformat(),
        end_time.isoformat(),
        elapsed,
        stats["root_count"],
        stats["dir_count"],
        json.dumps(config, default=str),
    ]

    ws.append(row_data)


# ---------------------------------------------------------------------------
# FUNCTION 2: print_run_summary
# ---------------------------------------------------------------------------

def print_run_summary(records: list, run_id: str, start_time, end_time) -> None:
    """Print a clean terminal summary after a dirmap run.

    Lists run metadata and enumerates the top-level directories
    (depth == 0) discovered during the walk.

    Parameters
    ----------
    records : list of dict
        Directory records as produced by build_dir_records().
    run_id : str
        Unique identifier for this run.
    start_time : datetime
        Timestamp recorded just before build_dir_records() was called.
    end_time : datetime
        Timestamp recorded just after build_dir_records() returned.
    """
    elapsed    = (end_time - start_time).total_seconds()
    dir_count  = len(records)
    top_level  = [r for r in records if r.get("depth", -1) == 0]
    root_count = len(top_level)

    divider      = "=" * 60
    thin_divider = "-" * 60

    print(divider)
    print("fw_dirmap Run Summary")
    print(divider)
    print(f"Run ID      : {run_id}")
    print(f"Directories : {dir_count}")
    print(f"Roots       : {root_count}")
    print(f"Elapsed     : {elapsed:.2f} seconds")
    print(thin_divider)
    print("Top-level directories:")

    for rec in top_level:
        dir_id   = rec.get("dir_id", "?")
        abs_path = rec.get("full_path", rec.get("abs_path", "?"))
        print(f"  {dir_id:<6}  {abs_path}")

    print(divider)


# ---------------------------------------------------------------------------
# FUNCTION 3: main
# ---------------------------------------------------------------------------

def main() -> None:
    """Orchestrate a full fw_dirmap directory-mapping run.

    Workflow
    --------
    1.  Print banner.
    2.  Prompt user to pick root directories via pick_scan_dirs().
    3.  Ask whether to scan recursively (default: Y).
    4.  Ask for optional max depth limit.
    5.  Generate a unique run ID.
    6.  Resolve the master workbook path (same directory as this script).
    7.  Open or create the workbook.
    8.  Walk directories and build records (timed).
    9.  Write inventory and processing-status rows.
    10. Log the run to Walk_History.
    11. Save the workbook.
    12. Print a run summary to the terminal.
    """
    from fw_dirmap_grp1 import pick_scan_dirs, validate_and_normalize_path, generate_run_id  # noqa: F401
    from fw_dirmap_grp2 import detect_source_store, count_dir_contents, walk_directories      # noqa: F401
    from fw_dirmap_grp3 import build_dir_records
    from fw_dirmap_grp4 import open_or_create_workbook, ensure_dir_inventory_sheet, write_dir_inventory_rows  # noqa: F401
    from fw_dirmap_grp5 import ensure_dir_processing_status_sheet, initialize_processing_status_rows          # noqa: F401

    # 1. Banner
    print("=" * 60)
    print("fw_dirmap — Directory Mapper Agent")
    print("=" * 60)

    # 2. Pick root directories
    roots = pick_scan_dirs()
    if not roots:
        print("No directories selected. Exiting.")
        return

    # 3. Recursive scan?
    recursive_input = input("Scan recursively? [Y/n]: ").strip().lower()
    recursive = recursive_input not in ("n", "no")

    # 4. Max depth
    max_depth_input = input("Max depth (blank = unlimited): ").strip()
    if max_depth_input == "":
        max_depth = None
    else:
        try:
            max_depth = int(max_depth_input)
        except ValueError:
            print(f"Invalid depth '{max_depth_input}' — treating as unlimited.")
            max_depth = None

    # 5. Generate run ID
    run_id = generate_run_id()

    # 6. Resolve workbook path
    script_dir = os.path.dirname(os.path.abspath(__file__))
    wb_path = os.path.join(script_dir, "filewalker_master.xlsx")

    # 7. Open or create workbook
    wb = open_or_create_workbook(wb_path)

    # 8. Build directory records (timed)
    start_time = datetime.now()
    records = build_dir_records(roots, recursive, max_depth, run_id)
    end_time = datetime.now()

    # 9a. Guard: nothing found
    if not records:
        print("Warning: no directory records were produced. Nothing to save.")
        return

    # 9b. Write inventory and processing-status rows
    write_dir_inventory_rows(wb, records)
    initialize_processing_status_rows(wb, records)

    # 10. Log the run
    config = {
        "roots":     [str(r) for r in roots],
        "recursive": recursive,
        "max_depth": max_depth,
    }
    stats = {
        "start_time": start_time,
        "end_time":   end_time,
        "dir_count":  len(records),
        "root_count": len([r for r in records if r.get("depth", -1) == 0]),
    }
    log_dirmap_run(wb, run_id, config, stats)

    # 11. Save workbook
    wb.save(wb_path)
    print(f"Saved: {wb_path}")

    # 12. Print run summary
    print_run_summary(records, run_id, start_time, end_time)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    main()
