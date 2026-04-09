"""
Read file-type column from both Google Sheets, count occurrences,
write results to an Excel file with one tab per sheet.
"""
import sys
from collections import Counter
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import openpyxl

TOKEN = r"C:\Users\arika\Repo-for-Claude-android\token_sheets.json"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
OUTPUT = r"C:\Users\arika\OneDrive\Documents\type_export.xlsx"

SHEET_FULL = "1JLt5IMkIyKYey8dT3VEDav8g9YZuHgsg5A46xK8pmDk"
SHEET_OD   = "1LKyP637uGOvmvfXmmNaoN5sMZ23WcpkOfg4afIPrzxY"

def get_service():
    creds = Credentials.from_authorized_user_file(TOKEN, SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return build("sheets", "v4", credentials=creds)

def get_all_tabs(svc, sheet_id):
    meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
    return [s["properties"]["title"] for s in meta["sheets"]]

def read_extension_column(svc, sheet_id):
    """Read the extension/file-type column (D) from all tabs, return flat list."""
    tabs = get_all_tabs(svc, sheet_id)
    print(f"  Tabs found: {tabs}")
    all_vals = []
    for tab in tabs:
        # Read header row first to find the extension column
        hdr = svc.spreadsheets().values().get(
            spreadsheetId=sheet_id, range=f"'{tab}'!1:1"
        ).execute().get("values", [[]])
        if not hdr:
            continue
        header = [h.lower().strip() for h in hdr[0]]
        # Look for extension column
        ext_col = None
        for i, h in enumerate(header):
            if h in ("extension", "ext", "file_type", "filetype", "type"):
                ext_col = i
                break
        if ext_col is None:
            print(f"  Tab '{tab}': no extension column found in header: {hdr[0]}")
            continue
        
        col_letter = chr(65 + ext_col) if ext_col < 26 else "D"
        print(f"  Tab '{tab}': extension column = {col_letter} ('{hdr[0][ext_col]}')")
        
        # Read all values in that column
        result = svc.spreadsheets().values().get(
            spreadsheetId=sheet_id, range=f"'{tab}'!{col_letter}:{col_letter}"
        ).execute()
        values = result.get("values", [])
        # Skip header row
        for row in values[1:]:
            if row:
                all_vals.append(row[0].strip())
            else:
                all_vals.append("(blank)")
    return all_vals

def main():
    svc = get_service()
    
    print("Reading Full C: Drive sheet...")
    full_vals = read_extension_column(svc, SHEET_FULL)
    print(f"  Total values: {len(full_vals)}")
    
    print("Reading OneDrive sheet...")
    od_vals = read_extension_column(svc, SHEET_OD)
    print(f"  Total values: {len(od_vals)}")
    
    full_counts = Counter(full_vals).most_common()
    od_counts = Counter(od_vals).most_common()
    
    # Write to Excel
    wb = openpyxl.Workbook()
    
    # Tab 1: Full C Drive
    ws1 = wb.active
    ws1.title = "Full_C_Drive"
    ws1.append(["File Type", "Count"])
    for ext, count in full_counts:
        ws1.append([ext, count])
    
    # Tab 2: OneDrive
    ws2 = wb.create_sheet("OneDrive")
    ws2.append(["File Type", "Count"])
    for ext, count in od_counts:
        ws2.append([ext, count])
    
    # Bold headers
    from openpyxl.styles import Font
    bold = Font(bold=True)
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.font = bold
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
    
    wb.save(OUTPUT)
    print(f"\nDone! Saved to: {OUTPUT}")
    print(f"  Full C Drive: {len(full_counts)} unique types, {len(full_vals)} total files")
    print(f"  OneDrive: {len(od_counts)} unique types, {len(od_vals)} total files")

if __name__ == "__main__":
    main()
