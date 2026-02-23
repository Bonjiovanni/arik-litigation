"""
Google Drive File Metadata Extractor

Lists all filenames, sizes, and extensions from the 'email/all' folder
in Google Drive.

Setup:
1. Go to https://console.cloud.google.com/
2. Create a project and enable the Google Drive API
3. Create OAuth 2.0 credentials (Desktop app) and download as 'credentials.json'
4. Place credentials.json in the same directory as this script
5. Run: pip install -r requirements.txt
6. Run: python drive_file_metadata.py
"""

import os
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# If modifying scopes, delete token.json
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# Target folder path: email/all
FOLDER_PATH = ["email", "all"]


def authenticate():
    """Authenticate and return a Google Drive service instance."""
    creds = None

    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists("credentials.json"):
                raise FileNotFoundError(
                    "credentials.json not found. Download it from Google Cloud Console "
                    "(APIs & Services > Credentials > OAuth 2.0 Client IDs)."
                )
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
            creds = flow.run_local_server(port=8080, open_browser=False)

        with open("token.json", "w") as token:
            token.write(creds.to_json())

    return build("drive", "v3", credentials=creds)


def find_folder(service, folder_name, parent_id="root"):
    """Find a folder by name within a parent folder and return its ID."""
    query = (
        f"name = '{folder_name}' "
        f"and mimeType = 'application/vnd.google-apps.folder' "
        f"and '{parent_id}' in parents "
        f"and trashed = false"
    )
    results = service.files().list(q=query, fields="files(id, name)").execute()
    folders = results.get("files", [])

    if not folders:
        raise FileNotFoundError(
            f"Folder '{folder_name}' not found under parent ID '{parent_id}'."
        )
    return folders[0]["id"]


def resolve_folder_path(service, path_parts):
    """Resolve a folder path list (e.g. ['email', 'all']) to a folder ID."""
    parent_id = "root"
    for part in path_parts:
        parent_id = find_folder(service, part, parent_id)
    return parent_id


def list_files_in_folder(service, folder_id):
    """Return a list of all files (not folders) inside the given folder ID."""
    files = []
    page_token = None

    while True:
        query = (
            f"'{folder_id}' in parents "
            f"and mimeType != 'application/vnd.google-apps.folder' "
            f"and trashed = false"
        )
        response = service.files().list(
            q=query,
            fields="nextPageToken, files(id, name, size, mimeType)",
            pageSize=1000,
            pageToken=page_token,
        ).execute()

        files.extend(response.get("files", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break

    return files


def get_extension(filename):
    """Extract the file extension from a filename, or return '' if none."""
    suffix = Path(filename).suffix
    return suffix.lstrip(".").lower() if suffix else ""


def format_size(size_bytes):
    """Format a byte count into a human-readable string."""
    if size_bytes is None:
        return "N/A"
    size_bytes = int(size_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} PB"


def save_to_excel(results):
    """Save file metadata to a formatted Excel workbook and return the filename."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File Metadata"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["#", "Filename", "Extension", "Size (Bytes)", "Size (Human)"]
    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Alternating row colours
    fill_light = PatternFill(fill_type="solid", fgColor="DEEAF1")
    fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")

    for row_idx, r in enumerate(results, start=1):
        row_fill = fill_light if row_idx % 2 == 0 else fill_white
        values = [
            row_idx,
            r["filename"],
            r["extension"],
            r["size_bytes"],
            r["size_human"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx + 1, column=col, value=value)
            cell.fill = row_fill

    # Auto-fit column widths
    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 4, 80)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Add an auto-filter
    ws.auto_filter.ref = ws.dimensions

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"file_metadata_{timestamp}.xlsx"
    wb.save(output_file)
    return output_file


def main():
    print("Authenticating with Google Drive...")
    service = authenticate()

    folder_display = "/".join(FOLDER_PATH)
    print(f"Resolving folder path: {folder_display}")
    folder_id = resolve_folder_path(service, FOLDER_PATH)
    print(f"Found folder ID: {folder_id}")

    print("Fetching file list...\n")
    files = list_files_in_folder(service, folder_id)

    if not files:
        print(f"No files found in '{folder_display}'.")
        return

    # Build results
    results = []
    for f in files:
        name = f["name"]
        size_bytes = f.get("size")
        ext = get_extension(name)
        results.append({
            "filename": name,
            "extension": ext if ext else "(none)",
            "size_bytes": int(size_bytes) if size_bytes is not None else None,
            "size_human": format_size(size_bytes),
        })

    # Sort by filename
    results.sort(key=lambda x: x["filename"].lower())

    # Print table
    col_name = max(len(r["filename"]) for r in results)
    col_ext = max(len(r["extension"]) for r in results)
    col_size = max(len(r["size_human"]) for r in results)

    header = (
        f"{'Filename':<{col_name}}  {'Ext':<{col_ext}}  {'Size':>{col_size}}"
    )
    separator = "-" * len(header)

    print(header)
    print(separator)
    for r in results:
        print(
            f"{r['filename']:<{col_name}}  "
            f"{r['extension']:<{col_ext}}  "
            f"{r['size_human']:>{col_size}}"
        )

    print(separator)
    print(f"Total: {len(results)} file(s)")

    # Save to Excel
    output_file = save_to_excel(results)
    print(f"\nResults saved to {output_file}")


if __name__ == "__main__":
    main()
