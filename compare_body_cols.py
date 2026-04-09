"""
Compare 4 body text columns in Merge1 tab of V11 workbook.
Strip HTML/formatting, check if all 4 are identical per row.
Output summary + new Excel with message ID, 4 cols, and match flag.
"""
import re
import sys
from html.parser import HTMLParser
import openpyxl
from openpyxl.styles import Font, PatternFill

SRC = r"C:\Users\arika\OneDrive\Litigation\Pipeline\label_LegalEmailExtracts - Ariks Version V11.xlsx"
OUT = r"C:\Users\arika\OneDrive\Documents\body_column_comparison.xlsx"
SHEET = "Merge1"

# Columns we need
TARGET_COLS = ["body_clean.1", "Body.HTML", "Body.SenderText", "Body.Text"]
ID_COL = "rfc823.message.id"

class HTMLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []
    def handle_data(self, d):
        self.parts.append(d)
    def get_text(self):
        return "".join(self.parts)

def strip_html(s):
    if not s:
        return ""
    s = str(s)
    h = HTMLStripper()
    try:
        h.feed(s)
        text = h.get_text()
    except:
        text = re.sub(r'<[^>]+>', '', s)
    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def main():
    print(f"Loading {SRC} ...")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[SHEET]

    # Read header row
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip() if h else "" for h in row]
        break

    print(f"Found {len(headers)} columns")

    # Find column indices
    col_map = {}
    for target in TARGET_COLS + [ID_COL]:
        found = False
        for i, h in enumerate(headers):
            if h.lower() == target.lower():
                col_map[target] = i
                found = True
                break
        if not found:
            # Try partial match
            for i, h in enumerate(headers):
                if target.lower() in h.lower():
                    col_map[target] = i
                    found = True
                    print(f"  Partial match: '{target}' -> col {i} '{h}'")
                    break
        if not found:
            print(f"  WARNING: Column '{target}' not found!")

    print(f"Column mapping: { {k: headers[v] for k,v in col_map.items()} }")

    if ID_COL not in col_map:
        # Try alternate names
        for i, h in enumerate(headers):
            if "message" in h.lower() and "id" in h.lower():
                col_map[ID_COL] = i
                print(f"  Using '{h}' as message ID column")
                break

    # Read all data rows
    rows_data = []
    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        msg_id = str(row[col_map[ID_COL]]) if ID_COL in col_map and row[col_map[ID_COL]] else ""
        vals = {}
        for col_name in TARGET_COLS:
            if col_name in col_map:
                raw = row[col_map[col_name]]
                vals[col_name] = str(raw) if raw else ""
            else:
                vals[col_name] = ""
        rows_data.append((msg_id, vals))

    wb.close()
    print(f"Read {len(rows_data)} data rows")

    # Compare
    match_count = 0
    mismatch_count = 0
    empty_count = 0
    results = []

    for msg_id, vals in rows_data:
        stripped = {k: strip_html(v) for k, v in vals.items()}
        unique_texts = set(t for t in stripped.values() if t)

        if len(unique_texts) == 0:
            flag = "ALL_EMPTY"
            empty_count += 1
        elif len(unique_texts) == 1:
            flag = "MATCH"
            match_count += 1
        else:
            flag = "MISMATCH"
            mismatch_count += 1

        results.append((msg_id, vals, flag))

    total = len(results)
    print(f"\n=== SUMMARY ===")
    print(f"Total rows: {total}")
    print(f"All 4 match: {match_count} ({100*match_count/total:.1f}%)")
    print(f"Mismatch:    {mismatch_count} ({100*mismatch_count/total:.1f}%)")
    print(f"All empty:   {empty_count} ({100*empty_count/total:.1f}%)")

    # Write output Excel
    print(f"\nWriting {OUT} ...")
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Body Comparison"

    # Headers
    out_headers = [ID_COL] + TARGET_COLS + ["Match?"]
    out_ws.append(out_headers)

    bold = Font(bold=True)
    for cell in out_ws[1]:
        cell.font = bold

    green = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
    red = PatternFill(start_color="FADADD", end_color="FADADD", fill_type="solid")
    gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for msg_id, vals, flag in results:
        row_vals = [msg_id]
        for col_name in TARGET_COLS:
            row_vals.append(vals.get(col_name, ""))
        row_vals.append(flag)
        out_ws.append(row_vals)

        # Color the flag cell
        last_row = out_ws.max_row
        flag_cell = out_ws.cell(row=last_row, column=len(out_headers))
        if flag == "MATCH":
            flag_cell.fill = green
        elif flag == "MISMATCH":
            flag_cell.fill = red
        else:
            flag_cell.fill = gray

    # Auto-width for first and last columns
    out_ws.column_dimensions["A"].width = 45
    out_ws.column_dimensions["F"].width = 12

    out_wb.save(OUT)
    print(f"Done! Saved to {OUT}")

if __name__ == "__main__":
    main()
