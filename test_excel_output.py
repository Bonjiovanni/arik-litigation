"""Quick test: verifies save_to_excel() and helper logic with mock data."""
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# --- helpers copied inline to avoid importing broken auth deps ---

def get_extension(filename):
    suffix = Path(filename).suffix
    return suffix.lstrip(".").lower() if suffix else ""


def format_size(size_bytes):
    if size_bytes is None:
        return "N/A"
    size_bytes = int(size_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} PB"


def save_to_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File Metadata"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["#", "Filename", "Extension", "Size (Bytes)", "Size (Human)"]
    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    fill_light = PatternFill(fill_type="solid", fgColor="DEEAF1")
    fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")

    for row_idx, r in enumerate(results, start=1):
        row_fill = fill_light if row_idx % 2 == 0 else fill_white
        values = [row_idx, r["filename"], r["extension"], r["size_bytes"], r["size_human"]]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx + 1, column=col, value=value)
            cell.fill = row_fill

    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col if cell.value is not None), default=0
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 4, 80)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"file_metadata_{timestamp}.xlsx"
    wb.save(output_file)
    return output_file


# --- mock data ---

mock_files = [
    {"filename": "invoice_march.pdf",  "size": "204800"},
    {"filename": "budget_2025.xlsx",   "size": "51200"},
    {"filename": "notes",              "size": "1024"},
    {"filename": "photo_holiday.jpg",  "size": "3145728"},
    {"filename": "report_final.docx",  "size": "98304"},
    {"filename": "archive.tar.gz",     "size": "10485760"},
    {"filename": "readme.txt",         "size": "512"},
    {"filename": "data_export.csv",    "size": "2097152"},
]

results = []
for f in mock_files:
    name = f["filename"]
    ext = get_extension(name)
    results.append({
        "filename": name,
        "extension": ext if ext else "(none)",
        "size_bytes": int(f["size"]),
        "size_human": format_size(f["size"]),
    })

results.sort(key=lambda x: x["filename"].lower())

output = save_to_excel(results)
print(f"Excel file created : {output}")
print(f"Rows written       : {len(results)}")

# Verify by reading it back
wb2 = openpyxl.load_workbook(output)
ws2 = wb2.active
print(f"Sheet name         : {ws2.title}")
print(f"Total rows (incl. header): {ws2.max_row}")
print("\nHeader row:", [ws2.cell(1, c).value for c in range(1, 6)])
print("First data row:", [ws2.cell(2, c).value for c in range(1, 6)])
print("\nAll rows:")
for row in ws2.iter_rows(min_row=2, values_only=True):
    print(" ", row)
