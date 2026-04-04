"""
fw_dirmap_grp5.py
=================
Part of the multi-file litigation file-walker system.
Handles the `Dir_Processing_Status` sheet in `filewalker_master.xlsx`.

This sheet tracks, for each directory, how far processing has gone per file
family. Each directory occupies 9 consecutive rows (one per file family).

Functions
---------
ensure_dir_processing_status_sheet(wb)
    Idempotent setup of the sheet — creates it with headers, styling, column
    widths, freeze panes, and auto-filter if it does not already exist;
    otherwise returns the existing sheet unchanged.

initialize_processing_status_rows(wb, records)
    Appends 9 rows per DirRecord dict in `records` to the sheet, one row per
    file family, with default values and alternating / flag-colour row shading.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SHEET_NAME = "Dir_Processing_Status"

HEADERS = [
    "DirID",
    "FullPath",
    "FileFamily",
    "FileCount",
    "ProcessingStatus",
    "LastProcessed",
    "Notes",
]

FILE_FAMILIES = [
    "pdf",
    "image",
    "word_doc",
    "spreadsheet",
    "presentation",
    "email_export",
    "video",
    "audio",
    "other",
]

COLUMN_WIDTHS = {
    "A": 8,
    "B": 60,
    "C": 14,
    "D": 10,
    "E": 16,
    "F": 16,
    "G": 25,
}

# Header styling
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# Row shading fills
_FILL_EVEN  = PatternFill(fill_type="solid", fgColor="F9F9F9")
_FILL_ODD   = PatternFill(fill_type="solid", fgColor="FFFFFF")
_FILL_EMAIL = PatternFill(fill_type="solid", fgColor="FFF2CC")

_THIN_SIDE   = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)


# ---------------------------------------------------------------------------
# FUNCTION 1: ensure_dir_processing_status_sheet
# ---------------------------------------------------------------------------

def ensure_dir_processing_status_sheet(wb: Workbook) -> Worksheet:
    """Create and style the `Dir_Processing_Status` sheet, or return it as-is.

    Idempotent: if the sheet already exists the function returns it immediately
    without modifying it.  If it does not exist the function creates it with:

    * A bold white-on-dark-blue header row (row 1).
    * Freeze panes at A2 so the header stays visible while scrolling.
    * Auto-filter spanning all header columns.
    * Fixed column widths as specified in ``COLUMN_WIDTHS``.

    Parameters
    ----------
    wb : openpyxl.Workbook
        The open workbook to operate on.

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet
        The ``Dir_Processing_Status`` worksheet (existing or newly created).
    """
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]

    ws = wb.create_sheet(title=SHEET_NAME)

    # Write and style the header row
    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border    = _CELL_BORDER

    # Column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze panes below header
    ws.freeze_panes = "A2"

    # Auto-filter on header row
    last_col_letter = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    return ws


# ---------------------------------------------------------------------------
# FUNCTION 2: initialize_processing_status_rows
# ---------------------------------------------------------------------------

def initialize_processing_status_rows(wb: Workbook, records: list) -> None:
    """Append 9 default rows per directory record to `Dir_Processing_Status`.

    For every dict in *records* the function writes one row per file family
    (9 families total, in the order defined by ``FILE_FAMILIES``). Default
    values are:

    * **FileCount** -> ``0``  (to be updated by the file walker later)
    * **ProcessingStatus** -> ``"not_scanned"``
    * **LastProcessed** / **Notes** -> empty string

    Row shading
    -----------
    * ``email_export`` rows receive a light-yellow fill (#FFF2CC) on every
      cell to visually flag that these families are skipped (not processed).
    * All other rows alternate between #F9F9F9 (even ``ws.max_row``) and
      #FFFFFF (odd ``ws.max_row``), evaluated *after* the row is appended.

    Progress is printed to stdout every 100 directories.

    Parameters
    ----------
    wb : openpyxl.Workbook
        The open workbook containing (or about to contain) the sheet.
    records : list[dict]
        Each dict must have at least the keys ``"dir_id"`` and
        ``"full_path"``.
    """
    ws = ensure_dir_processing_status_sheet(wb)

    for dir_index, record in enumerate(records, start=1):
        dir_id    = record["dir_id"]
        full_path = record["full_path"]

        for family in FILE_FAMILIES:
            row_data = [
                dir_id,
                full_path,
                family,
                0,              # FileCount
                "not_scanned",  # ProcessingStatus
                "",             # LastProcessed
                "",             # Notes
            ]
            ws.append(row_data)

            current_row = ws.max_row
            if family == "email_export":
                row_fill = _FILL_EMAIL
            elif current_row % 2 == 0:
                row_fill = _FILL_EVEN
            else:
                row_fill = _FILL_ODD

            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=current_row, column=col_idx).fill = row_fill

        if dir_index % 100 == 0:
            print(f"  [fw_dirmap_grp5] Initialized processing-status rows "
                  f"for {dir_index} directories...")

    print(f"  [fw_dirmap_grp5] Done — {len(records)} director"
          f"{'y' if len(records) == 1 else 'ies'} written to '{SHEET_NAME}'.")
