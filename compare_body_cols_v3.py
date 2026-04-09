"""
Compare 4 body text columns in Merge1 tab of V11 workbook. v3
Strip HTML/formatting, compare all pairs, show char counts.
"""
import re
from html.parser import HTMLParser
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill

SRC = r"C:\Users\arika\OneDrive\Litigation\Pipeline\label_LegalEmailExtracts - Ariks Version V11.xlsx"
OUT = r"C:\Users\arika\OneDrive\Documents\body_column_comparison.xlsx"
SHEET = "Merge1"
TARGET_COLS = ["body_clean.1", "Body.HTML", "Body.SenderText", "Body.Text"]

class HTMLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []
    def handle_data(self, d):
        self.parts.append(d)
    def get_text(self):
        return "".join(self.parts)

def strip_html(s):
    if not s:
        return ""
    s = str(s)
    h = HTMLStripper()
    try:
        h.feed(s)
        text = h.get_text()
    except:
        text = re.sub(r'<[^>]+>', '', s)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def main():
    print(f"Loading {SRC} ...")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[SHEET]

    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip() if h else "" for h in row]
        break

    col_map = {}
    for target in TARGET_COLS:
        for i, h in enumerate(headers):
            if h.lower() == target.lower():
                col_map[target] = i
                break

    id_col = None
    for i, h in enumerate(headers):
        if "message" in h.lower() and "id" in h.lower():
            id_col = i
            print(f"Message ID column: col {i} = '{h}'")
            break

    print(f"Body columns: { {k: headers[v] for k,v in col_map.items()} }")

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        msg_id = str(row[id_col]) if id_col is not None and row[id_col] else ""
        raw_vals = {}
        stripped_vals = {}
        for col_name in TARGET_COLS:
            if col_name in col_map:
                raw = row[col_map[col_name]]
                raw_vals[col_name] = str(raw) if raw else ""
                stripped_vals[col_name] = strip_html(raw)
            else:
                raw_vals[col_name] = ""
                stripped_vals[col_name] = ""
        results.append((msg_id, raw_vals, stripped_vals))

    wb.close()
    print(f"Read {len(results)} rows\n")

    output_rows = []
    counts = {"ALL_MATCH": 0, "ALL_EMPTY": 0, "NONE_MATCH": 0, "PARTIAL": 0}
    pattern_counts = {}

    for msg_id, raw_vals, stripped_vals in results:
        texts = [stripped_vals[c] for c in TARGET_COLS]
        char_counts = [len(t) for t in texts]
        non_empty = [t for t in texts if t]

        if not non_empty:
            code = "----"
            flag = "ALL_EMPTY"
        else:
            freq = Counter(non_empty)
            majority_text = freq.most_common(1)[0][0]
            code_chars = []
            for t in texts:
                if not t:
                    code_chars.append("E")
                elif t == majority_text:
                    code_chars.append("Y")
                else:
                    code_chars.append("N")
            code = "".join(code_chars)
            y_count = code.count("Y")
            n_count = code.count("N")
            e_count = code.count("E")
            if n_count == 0 and e_count == 0:
                flag = "ALL_MATCH"
            elif y_count + e_count == len(code_chars):
                flag = "MATCH (some empty)"
            else:
                unique_non_empty = set(non_empty)
                if len(unique_non_empty) == len(non_empty):
                    flag = "NONE_MATCH"
                else:
                    flag = "PARTIAL"

        if flag == "ALL_MATCH":
            counts["ALL_MATCH"] += 1
        elif flag == "ALL_EMPTY":
            counts["ALL_EMPTY"] += 1
        elif flag == "NONE_MATCH":
            counts["NONE_MATCH"] += 1
        else:
            counts["PARTIAL"] += 1

        pattern_counts[code] = pattern_counts.get(code, 0) + 1
        output_rows.append((msg_id, raw_vals, code, flag, char_counts))

    total = len(output_rows)
    print("=== SUMMARY ===")
    print(f"Total rows:          {total}")
    print(f"ALL_MATCH (YYYY):    {counts['ALL_MATCH']} ({100*counts['ALL_MATCH']/total:.1f}%)")
    print(f"PARTIAL match:       {counts['PARTIAL']} ({100*counts['PARTIAL']/total:.1f}%)")
    print(f"NONE_MATCH:          {counts['NONE_MATCH']} ({100*counts['NONE_MATCH']/total:.1f}%)")
    print(f"ALL_EMPTY:           {counts['ALL_EMPTY']} ({100*counts['ALL_EMPTY']/total:.1f}%)")
    print(f"\nMatch pattern breakdown (code = {'/'.join(TARGET_COLS)}):")
    print(f"  Y=matches majority, N=different, E=empty")
    for code, cnt in sorted(pattern_counts.items(), key=lambda x: -x[1]):
        print(f"  {code}: {cnt} rows")

    # Write Excel
    print(f"\nWriting {OUT} ...")
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Body Comparison"

    out_headers = (
        ["Message ID"]
        + TARGET_COLS
        + ["Match Code", "Result"]
        + [f"Len:{c}" for c in TARGET_COLS]
    )
    out_ws.append(out_headers)

    bold = Font(bold=True)
    for cell in out_ws[1]:
        cell.font = bold

    green = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
    red = PatternFill(start_color="FADADD", end_color="FADADD", fill_type="solid")
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for msg_id, raw_vals, code, flag, char_counts in output_rows:
        row_data = [msg_id]
        for c in TARGET_COLS:
            row_data.append(raw_vals.get(c, ""))
        row_data.append(code)
        row_data.append(flag)
        for cc in char_counts:
            row_data.append(cc)
        out_ws.append(row_data)

        r = out_ws.max_row
        code_cell = out_ws.cell(row=r, column=6)
        flag_cell = out_ws.cell(row=r, column=7)
        if flag == "ALL_MATCH":
            code_cell.fill = green
            flag_cell.fill = green
        elif flag == "NONE_MATCH":
            code_cell.fill = red
            flag_cell.fill = red
        elif "PARTIAL" in flag or "some empty" in flag:
            code_cell.fill = yellow
            flag_cell.fill = yellow
        else:
            code_cell.fill = gray
            flag_cell.fill = gray

    out_ws.column_dimensions["A"].width = 45
    out_ws.column_dimensions["F"].width = 14
    out_ws.column_dimensions["G"].width = 20
    for col_letter in ["H", "I", "J", "K"]:
        out_ws.column_dimensions[col_letter].width = 18

    out_wb.save(OUT)
    print(f"Done! Saved to {OUT}")

if __name__ == "__main__":
    main()
