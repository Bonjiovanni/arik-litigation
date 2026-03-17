"""
fw_classify_grp3.py
-------------------
Keywords_Config sheet management and DocType inference logic for the fw_classify pipeline.
Litigation evidence management system — trust/LTC insurance documents.

Functions:
    ensure_keywords_config_sheet  — creates/returns Keywords_Config sheet with defaults
    load_keywords                 — returns list of active keywords from the sheet
    infer_likely_text_bearing     — "Y"/"N"/"" based on file family + extraction result
    infer_needs_ocr               — "Y"/"N"/"" based on file family + extraction result
    classify_doc_type             — lightweight keyword-driven DocType inference
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# Module constants
# ---------------------------------------------------------------------------

KEYWORDS_CONFIG_SHEET = "Keywords_Config"

# (Keyword, Active, Category, Notes)
_DEFAULT_KEYWORDS = [
    ("trust",          "Y", "entity",    ""),
    ("trustee",        "Y", "entity",    ""),
    ("beneficiary",    "Y", "entity",    ""),
    ("distribution",   "Y", "financial", ""),
    ("fiduciary",      "Y", "entity",    ""),
    ("amendment",      "Y", "legal",     ""),
    ("restatement",    "Y", "legal",     ""),
    ("invoice",        "Y", "financial", ""),
    ("premium",        "Y", "financial", ""),
    ("claim",          "Y", "ltc",       ""),
    ("long-term care", "Y", "ltc",       ""),
    ("LTC",            "Y", "ltc",       ""),
    ("benefit",        "Y", "ltc",       ""),
    ("policy",         "Y", "ltc",       ""),
    ("coverage",       "Y", "ltc",       ""),
    ("caregiver",      "Y", "ltc",       ""),
    ("facility",       "Y", "ltc",       ""),
    ("nursing",        "Y", "ltc",       ""),
    ("home health",    "Y", "ltc",       ""),
]

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# FUNCTION 1: ensure_keywords_config_sheet
# ---------------------------------------------------------------------------

def ensure_keywords_config_sheet(wb: Workbook) -> Worksheet:
    """Create and style the Keywords_Config sheet if absent; return it.

    Idempotent — returns the existing sheet unchanged if already present.
    Seeds default litigation keywords on first creation.
    """
    if KEYWORDS_CONFIG_SHEET in wb.sheetnames:
        return wb[KEYWORDS_CONFIG_SHEET]

    ws = wb.create_sheet(title=KEYWORDS_CONFIG_SHEET)

    headers = ["Keyword", "Active", "Category", "Notes"]
    col_widths = [25, 8, 14, 40]
    for col_idx, (header_text, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A2"

    for row_idx, (keyword, active, category, notes) in enumerate(_DEFAULT_KEYWORDS, start=2):
        ws.cell(row=row_idx, column=1, value=keyword)
        ws.cell(row=row_idx, column=2, value=active)
        ws.cell(row=row_idx, column=3, value=category)
        ws.cell(row=row_idx, column=4, value=notes)

    return ws


# ---------------------------------------------------------------------------
# FUNCTION 2: load_keywords
# ---------------------------------------------------------------------------

def load_keywords(ws: Worksheet) -> list[str]:
    """Return list of keyword strings where Active == 'Y' (case-insensitive).

    Skips header row (row 1) and rows with None/empty Keyword.
    Column A = Keyword, Column B = Active.
    """
    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        keyword = row[0] if len(row) > 0 else None
        active  = row[1] if len(row) > 1 else None
        if not keyword:
            continue
        if active is not None and str(active).strip().upper() == "Y":
            keywords.append(str(keyword))
    return keywords


# ---------------------------------------------------------------------------
# FUNCTION 3: infer_likely_text_bearing
# ---------------------------------------------------------------------------

def infer_likely_text_bearing(file_family: str, text_sample: str,
                               extraction_succeeded: bool) -> str:
    """Infer whether the file likely contains extractable text.

    Returns "Y", "N", or "" (unknown/not applicable).
    """
    fam = (file_family or "").lower()

    if fam in ("text_file", "word_doc", "spreadsheet", "presentation"):
        return "Y"
    if fam == "pdf":
        return "Y" if (extraction_succeeded and text_sample) else "N"
    if fam in ("image", "audio", "video"):
        return "N"
    if fam == "email_file":
        return "Y"
    return ""


# ---------------------------------------------------------------------------
# FUNCTION 4: infer_needs_ocr
# ---------------------------------------------------------------------------

def infer_needs_ocr(file_family: str, text_sample: str,
                    extraction_succeeded: bool) -> str:
    """Infer whether the file needs OCR processing.

    Returns "Y", "N", or "" (not applicable).
    """
    fam = (file_family or "").lower()

    if fam == "pdf":
        return "N" if (extraction_succeeded and text_sample) else "Y"
    if fam == "image":
        return "Y"
    if fam in ("text_file", "word_doc", "spreadsheet", "presentation", "email_file"):
        return "N"
    if fam in ("audio", "video", "archive"):
        return "N"
    return ""


# ---------------------------------------------------------------------------
# FUNCTION 5: classify_doc_type
# ---------------------------------------------------------------------------

def classify_doc_type(text_sample: str, keyword_hits: str,
                      file_family: str) -> dict:
    """Lightweight keyword-driven DocType inference.

    Returns dict with keys: DocType, DocSubtype, DocTypeConfidence.
    First matching rule wins.
    """
    ts  = (text_sample  or "").lower()
    kh  = (keyword_hits or "").lower()
    sig = ts + " " + kh

    # Rule 1: Invoice
    if "invoice" in sig:
        return {"DocType": "Invoice", "DocSubtype": "", "DocTypeConfidence": "medium"}

    # Rule 2: LTC Claim — claim + explicit LTC signal
    if "claim" in sig and ("long-term care" in sig or "ltc" in sig):
        return {"DocType": "LTC_Claim", "DocSubtype": "", "DocTypeConfidence": "medium"}

    # Rule 3: LTC Claim — benefit/coverage + claim
    if ("benefit" in sig or "coverage" in sig) and "claim" in sig:
        return {"DocType": "LTC_Claim", "DocSubtype": "", "DocTypeConfidence": "low"}

    # Rule 4: Trust Document — Amendment/Restatement
    if "amendment" in sig or "restatement" in sig:
        return {"DocType": "Trust_Document", "DocSubtype": "Amendment", "DocTypeConfidence": "medium"}

    # Rule 5: Trust Document — trust + corroborating entity
    if "trust" in sig and ("trustee" in sig or "beneficiary" in sig or "fiduciary" in sig):
        return {"DocType": "Trust_Document", "DocSubtype": "", "DocTypeConfidence": "medium"}

    # Rule 6: Trust Document — trust alone
    if "trust" in sig:
        return {"DocType": "Trust_Document", "DocSubtype": "", "DocTypeConfidence": "low"}

    # Rule 7: Insurance Policy
    if "premium" in sig or ("policy" in sig and "long-term care" in sig):
        return {"DocType": "Insurance_Policy", "DocSubtype": "", "DocTypeConfidence": "low"}

    # Rule 8: No match
    return {"DocType": "", "DocSubtype": "", "DocTypeConfidence": ""}


def load_entity_aliases(xlsm_path: str) -> dict:
    """Load alias→Entity_ID lookup from 02_Aliases sheet of EntityIndex xlsm.

    Returns {alias_value_lower: entity_id} for all rows.
    Returns {} if file not found or openpyxl fails.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsm_path, read_only=True, keep_vba=True)
        ws = wb["02_Aliases"]
        alias_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[1] is None or row[2] is None:
                continue
            entity_id  = str(row[1]).strip()
            alias_val  = str(row[2]).strip()
            if entity_id and alias_val:
                alias_map[alias_val.lower()] = entity_id
        wb.close()
        return alias_map
    except Exception:
        return {}


def detect_entity_hits(text_sample: str, alias_map: dict) -> str:
    """Return semicolon-joined Entity_IDs found in text_sample.

    Case-insensitive substring match against alias_map keys.
    Returns "" if no hits or text_sample is empty.
    """
    if not text_sample or not alias_map:
        return ""
    text_lower = text_sample.lower()
    seen = set()
    for alias, entity_id in alias_map.items():
        if alias in text_lower and entity_id not in seen:
            seen.add(entity_id)
    return ";".join(sorted(seen))
