"""
fw_triage_grp3.py
-----------------
Triage_History sheet management for fw_triage.

Functions:
    ensure_triage_history_sheet  — creates/returns the Triage_History sheet
    log_triage_run               — appends one run summary row
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


TRIAGE_HISTORY_SHEET = "Triage_History"

_COLUMNS = [
    ("RunID",          25),
    ("StartedAt",      22),
    ("CompletedAt",    22),
    ("RowsProcessed",  14),
    ("RowsTriaged",    12),
    ("RowsSkipped",    12),
    ("Errors",         10),
    ("Notes",          40),
]

_COL_MAP = {name: idx + 1 for idx, (name, _) in enumerate(_COLUMNS)}

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


def ensure_triage_history_sheet(wb: Workbook) -> Worksheet:
    """Create and style the Triage_History sheet if absent; return it. Idempotent."""
    if TRIAGE_HISTORY_SHEET in wb.sheetnames:
        return wb[TRIAGE_HISTORY_SHEET]

    ws = wb.create_sheet(title=TRIAGE_HISTORY_SHEET)

    for col_idx, (name, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    return ws


def log_triage_run(wb: Workbook, run_id: str, started_at: str,
                   stats: dict, notes: str = "") -> int:
    """Append one triage-run summary row to Triage_History.

    stats keys (default 0): rows_processed, rows_triaged, rows_skipped, errors.
    Returns the row number written.
    """
    ws = ensure_triage_history_sheet(wb)

    row_num = ws.max_row + 1
    if row_num == 2 and ws.cell(row=2, column=1).value is None:
        row_num = 2

    ws.cell(row=row_num, column=_COL_MAP["RunID"],         value=run_id)
    ws.cell(row=row_num, column=_COL_MAP["StartedAt"],     value=started_at)
    ws.cell(row=row_num, column=_COL_MAP["CompletedAt"],   value=datetime.now().isoformat())
    ws.cell(row=row_num, column=_COL_MAP["RowsProcessed"], value=stats.get("rows_processed", 0))
    ws.cell(row=row_num, column=_COL_MAP["RowsTriaged"],   value=stats.get("rows_triaged",   0))
    ws.cell(row=row_num, column=_COL_MAP["RowsSkipped"],   value=stats.get("rows_skipped",   0))
    ws.cell(row=row_num, column=_COL_MAP["Errors"],        value=stats.get("errors",         0))
    ws.cell(row=row_num, column=_COL_MAP["Notes"],         value=notes)

    return row_num
