"""
fw_triage_grp4.py
-----------------
main() orchestration for fw_triage.

Coordinates:
    grp1 — Triage_Config + Triage_Bands sheets + scoring logic
    grp2 — Master_File_Inventory Excel I/O
    grp3 — Triage_History logging

Reads: Master_File_Inventory rows with ProcessingStatus = "classified"
Writes: TriageScore, TriageBand, ReasonFlagged, NextStep,
        ProcessingStatus → "triaged"

Workbook: C:\\Users\\arika\\OneDrive\\Litigation\\filewalker_master.xlsx
Python 3.11, Windows 10.
"""

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook

import fw_triage_grp1 as grp1
import fw_triage_grp2 as grp2
import fw_triage_grp3 as grp3


WORKBOOK_PATH = r"C:\Users\arika\OneDrive\Litigation\filewalker_master.xlsx"


def main() -> None:
    """Run a complete fw_triage scoring pass.

    Steps:
      1.  Open or create the master workbook.
      2.  Ensure Triage_Config, Triage_Bands, Triage_History sheets exist.
      3.  Load triage config and band thresholds.
      4.  Generate run ID: TRIAGE_YYYYMMDD_HHMMSS.
      5.  Get triageable rows from Master_File_Inventory.
      6.  For each row: score → band → reason → next_step → write → mark triaged.
      7.  Log Triage_History row.
      8.  Save workbook.
      9.  Print final summary.
    """
    try:
        # ------------------------------------------------------------------
        # 1. Open or create workbook
        # ------------------------------------------------------------------
        wb_path = Path(WORKBOOK_PATH)
        if wb_path.exists():
            print(f"[fw_triage] Loading existing workbook: {wb_path}")
            wb = load_workbook(wb_path)
        else:
            print(f"[fw_triage] Creating new workbook: {wb_path}")
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ------------------------------------------------------------------
        # 2. Ensure config and history sheets exist
        # ------------------------------------------------------------------
        grp1.ensure_triage_config_sheet(wb)
        grp1.ensure_triage_bands_sheet(wb)
        grp3.ensure_triage_history_sheet(wb)

        # ------------------------------------------------------------------
        # 3. Load config
        # ------------------------------------------------------------------
        ws_config = wb[grp1.TRIAGE_CONFIG_SHEET]
        ws_bands  = wb[grp1.TRIAGE_BANDS_SHEET]
        config = grp1.load_triage_config(ws_config)
        bands  = grp1.load_triage_bands(ws_bands)
        print(f"[fw_triage] Loaded {sum(len(v) for v in config.values())} config entries, "
              f"{len(bands)} band(s).")

        # ------------------------------------------------------------------
        # 4. Generate run ID
        # ------------------------------------------------------------------
        run_id: str = datetime.now().strftime("TRIAGE_%Y%m%d_%H%M%S")
        print(f"[fw_triage] Run ID: {run_id}")

        # ------------------------------------------------------------------
        # 5. Get triageable rows
        # ------------------------------------------------------------------
        if "Master_File_Inventory" not in wb.sheetnames:
            print("[fw_triage] ERROR: Master_File_Inventory not found. Run fw_walk + fw_classify first.")
            sys.exit(1)

        ws_inv = wb["Master_File_Inventory"]
        rows   = grp2.get_triageable_rows(ws_inv)
        print(f"[fw_triage] Rows to triage: {len(rows):,}")

        if not rows:
            print("[fw_triage] Nothing to triage.")
            sys.exit(0)

        # ------------------------------------------------------------------
        # 6. Score and write
        # ------------------------------------------------------------------
        started_at_str: str = datetime.now().isoformat()
        rows_processed = 0
        rows_triaged   = 0
        rows_skipped   = 0
        errors         = 0

        for row_num, record in rows:
            try:
                score  = grp1.score_record(record, config)
                band, default_next_step = grp1.get_triage_band(score, bands)
                reason = grp1.get_reason_flagged(record, config)
                step   = grp1.get_next_step(band, record, bands)

                results = {
                    "TriageScore":    score,
                    "TriageBand":     band,
                    "ReasonFlagged":  reason,
                    "NextStep":       step,
                }
                grp2.write_triage_results(ws_inv, row_num, results)
                grp2.mark_row_triaged(ws_inv, row_num)
                rows_triaged += 1

            except Exception as exc:
                filepath = record.get("FilePath", "")
                print(f"[fw_triage] ERROR row {row_num} ({filepath}): {exc}")
                errors += 1
            finally:
                rows_processed += 1

            if rows_processed % 100 == 0:
                print(f"[fw_triage]   {rows_processed:,}/{len(rows):,} processed...")

        # ------------------------------------------------------------------
        # 7. Log Triage_History
        # ------------------------------------------------------------------
        grp3.log_triage_run(
            wb=wb,
            run_id=run_id,
            started_at=started_at_str,
            stats={
                "rows_processed": rows_processed,
                "rows_triaged":   rows_triaged,
                "rows_skipped":   rows_skipped,
                "errors":         errors,
            },
        )

        # ------------------------------------------------------------------
        # 8. Save
        # ------------------------------------------------------------------
        wb.save(WORKBOOK_PATH)

        # ------------------------------------------------------------------
        # 9. Summary
        # ------------------------------------------------------------------
        print(f"\n[fw_triage] Run complete: {run_id}")
        print(f"  Rows processed: {rows_processed:,}")
        print(f"  Triaged:        {rows_triaged:,}")
        print(f"  Skipped:        {rows_skipped:,}")
        print(f"  Errors:         {errors:,}")
        print(f"Saved: {WORKBOOK_PATH}")

    except KeyboardInterrupt:
        print("\n[fw_triage] Interrupted.")
        sys.exit(0)


if __name__ == "__main__":
    main()
