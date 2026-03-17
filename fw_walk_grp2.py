"""
fw_walk_grp2.py
---------------
Manages the Master_File_Inventory sheet and FileFamily_Config sheet
in filewalker_master.xlsx.

One row per unique file. Handles creation, lookup, insert, and update logic.

Column layout: A (FileID) through BI (Notes), 61 columns total.
Header styling: bold white on dark blue (#2F5496), centered, frozen row 1,
auto-filter enabled.

NEVER overwrite: ManualReviewStatus, KeepForCase, PossibleExhibit
Classification/triage columns are written only by fw_classify and fw_triage.

Also provides:
    ensure_file_family_config_sheet — creates/returns the FileFamily_Config sheet,
        pre-populating it with built-in defaults on first creation.
"""

import re

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Column definitions (name, width) — order = column order in sheet
# ---------------------------------------------------------------------------

COLUMNS = [
    # --- A. Location / identity ---
    ("FileID",               8),
    ("RunID_FirstSeen",      20),
    ("RunID_LastSeen",       20),
    ("FilePath",             70),
    ("ParentFolder",         60),
    ("FileName",             35),
    ("ScanRootPath",         50),
    ("SourceStore",          14),
    # --- B. File metadata ---
    ("SizeBytes",            12),
    ("CreatedTime",          20),
    ("ModifiedTime",         20),
    ("SHA256",               66),
    ("IsDuplicateExact",     10),
    ("DuplicateGroupID",     14),
    # --- C. Physical/source classification ---
    ("FileFamily",           14),
    ("SourceType",           20),
    ("LikelyTextBearing",    10),
    ("LikelyImage",          10),
    ("LikelySpreadsheet",    10),
    ("LikelyDocument",       10),
    ("LikelyScreenshot",     10),
    ("NeedsOCR",             10),
    ("HasFormFields",        10),
    ("ExtractionMethod",     20),
    ("IsContainerType",      10),
    ("SkipReason",           14),
    ("ArchiveContents",      45),   # plain-text summary for zip/rar/7z contents
    # --- D. Existing structured joins ---
    ("MsgID",                20),
    ("AttachmentID",         20),
    ("AttachmentKind",       20),
    ("RelatedEmailSubject",  20),
    ("RelatedEmailFrom",     20),
    ("RelatedEmailDate",     20),
    ("RelatedEmailPDFPath",  20),
    ("RelatedTextThreadID",  20),
    # --- E. Cheap first-pass extraction signals ---
    ("TextSample",           20),
    ("OCRSnippet",           20),
    ("KeywordHits",          20),
    ("EntityHits",           40),
    ("MoneyDetected",        20),
    ("DateDetected",         20),
    ("SignatureLike",        20),
    ("HandwritingLike",      20),
    ("LogoLetterheadLike",   20),
    ("TableLayoutLike",      20),
    # --- F. Classification / semantics ---
    ("DocType",              25),
    ("DocSubtype",           25),
    ("DocTypeConfidence",    20),
    ("CategoryTags",         30),
    ("ReviewGroup",          20),
    ("PrimaryEntity",        25),
    ("MatchedEntities",      25),
    ("NewEntityCandidates",  25),
    ("EntityMatchSource",    25),
    ("EntityConfidence",     20),
    # --- G. Triage / routing ---
    ("TriageScore",          18),
    ("TriageBand",           18),
    ("ReasonFlagged",        18),
    ("NextStep",             18),
    ("ManualReviewStatus",   18),
    ("KeepForCase",          12),
    ("PossibleExhibit",      12),
    ("ProcessingStatus",     16),
    ("Notes",                30),
]

SHEET_NAME = "Master_File_Inventory"

# Map header name -> 1-based column index
_COL_INDEX: dict[str, int] = {name: idx + 1 for idx, (name, _) in enumerate(COLUMNS)}

# Columns always overwritten on update
_ALWAYS_UPDATE_COLS = {
    "RunID_LastSeen", "SizeBytes", "ModifiedTime", "SHA256",
    "IsDuplicateExact", "DuplicateGroupID", "ProcessingStatus",
}

# Columns never overwritten if already populated
_NEVER_OVERWRITE_COLS = {"ManualReviewStatus", "KeepForCase", "PossibleExhibit"}

# Columns owned by fw_classify / fw_triage — untouched on update
_CLASSIFICATION_COLS = {
    "TextSample", "OCRSnippet", "KeywordHits", "MoneyDetected", "DateDetected",
    "SignatureLike", "HandwritingLike", "LogoLetterheadLike", "TableLayoutLike",
    "DocType", "DocSubtype", "DocTypeConfidence", "CategoryTags", "ReviewGroup",
    "PrimaryEntity", "MatchedEntities", "NewEntityCandidates", "EntityMatchSource",
    "EntityConfidence", "TriageScore", "TriageBand", "ReasonFlagged", "NextStep",
}

_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _normalize_path(path: str) -> str:
    return path.replace("\\", "/").lower()


def _col_name_to_key(col_name: str) -> str:
    """Convert CamelCase column name to snake_case record key.
    e.g. 'ManualReviewStatus' -> 'manual_review_status'
    """
    return re.sub(r"(?<=[a-z])(?=[A-Z])", "_", col_name).lower()


# ---------------------------------------------------------------------------
# FUNCTION 1: ensure_master_file_inventory_sheet
# ---------------------------------------------------------------------------

def ensure_master_file_inventory_sheet(wb: Workbook) -> Worksheet:
    """Create and style the Master_File_Inventory sheet if absent; return it.

    Idempotent — returns the existing sheet unchanged if it already exists.

    Parameters
    ----------
    wb : Workbook

    Returns
    -------
    Worksheet
    """
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]

    ws = wb.create_sheet(title=SHEET_NAME)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    return ws


# ---------------------------------------------------------------------------
# FUNCTION 2: get_next_file_id
# ---------------------------------------------------------------------------

def get_next_file_id(ws: Worksheet) -> str:
    """Return the next available FileID (e.g. "F00248") by scanning column A.

    Returns "F00001" if the sheet has no existing data rows or no parseable IDs.

    Parameters
    ----------
    ws : Worksheet

    Returns
    -------
    str
    """
    max_num = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        val = str(row[0] or "").strip().upper()
        if val.startswith("F") and val[1:].isdigit():
            num = int(val[1:])
            if num > max_num:
                max_num = num
    return f"F{max_num + 1:05d}"


# ---------------------------------------------------------------------------
# FUNCTION 3: find_existing_file_row
# ---------------------------------------------------------------------------

def find_existing_file_row(ws: Worksheet, file_path: str) -> "int | None":
    """Search column D (FilePath) for an exact normalized match.

    Comparison is case-insensitive and forward-slash normalized.

    Parameters
    ----------
    ws : Worksheet
    file_path : str

    Returns
    -------
    int or None
        1-based row number, or None if not found.
    """
    col_d  = _COL_INDEX["FilePath"]
    target = _normalize_path(file_path)

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, min_col=col_d, max_col=col_d, values_only=True),
        start=2,
    ):
        val = row[0]
        if val is not None and _normalize_path(str(val)) == target:
            return row_idx
    return None


# ---------------------------------------------------------------------------
# FUNCTION 4: insert_file_record
# ---------------------------------------------------------------------------

def insert_file_record(ws: Worksheet, record: dict, run_id: str) -> int:
    """Append a new file record row. Returns the 1-based row number written.

    All columns not in record default to "".
    ManualReviewStatus defaults to "unreviewed".
    ProcessingStatus defaults to "file_listed".
    RunID_FirstSeen and RunID_LastSeen are both set to run_id.

    Parameters
    ----------
    ws : Worksheet
    record : dict
    run_id : str

    Returns
    -------
    int
    """
    next_row = max(ws.max_row + 1, 2)

    # Build full row dict defaulting everything to ""
    row_data: dict[str, object] = {col_name: "" for col_name, _ in COLUMNS}

    # Populate from record
    row_data["FileID"]           = record.get("file_id", "")
    row_data["RunID_FirstSeen"]  = run_id
    row_data["RunID_LastSeen"]   = run_id
    row_data["FilePath"]         = record.get("file_path", "").replace("\\", "/")
    row_data["ParentFolder"]     = record.get("parent_folder", "").replace("\\", "/")
    row_data["FileName"]         = record.get("filename", "")
    row_data["ScanRootPath"]     = record.get("scan_root_path", "").replace("\\", "/")
    row_data["SourceStore"]      = record.get("source_store", "")
    row_data["SizeBytes"]        = record.get("size_bytes", "")
    row_data["CreatedTime"]      = record.get("created_time", "")
    row_data["ModifiedTime"]     = record.get("modified_time", "")
    row_data["SHA256"]           = record.get("sha256", "")
    row_data["IsDuplicateExact"] = record.get("is_duplicate_exact", "")
    row_data["DuplicateGroupID"] = record.get("duplicate_group_id", "")
    row_data["FileFamily"]       = record.get("file_family", "")
    row_data["SourceType"]       = record.get("source_type", "standalone_file")
    row_data["LikelyTextBearing"]= record.get("likely_text_bearing", "")
    row_data["LikelyImage"]      = record.get("likely_image", "")
    row_data["LikelySpreadsheet"]= record.get("likely_spreadsheet", "")
    row_data["LikelyDocument"]   = record.get("likely_document", "")
    row_data["LikelyScreenshot"] = record.get("likely_screenshot", "")
    row_data["NeedsOCR"]         = record.get("needs_ocr", "")
    row_data["IsContainerType"]  = record.get("is_container_type", "")
    row_data["SkipReason"]       = record.get("skip_reason", "")
    row_data["ArchiveContents"]  = record.get("archive_contents", "")
    row_data["ManualReviewStatus"] = record.get("manual_review_status", "unreviewed")
    row_data["ProcessingStatus"] = record.get("processing_status", "file_listed")

    for col_name, col_idx in _COL_INDEX.items():
        ws.cell(row=next_row, column=col_idx, value=row_data[col_name])

    return next_row


# ---------------------------------------------------------------------------
# FUNCTION 5: update_file_record
# ---------------------------------------------------------------------------

def update_file_record(ws: Worksheet, row: int, record: dict, run_id: str) -> None:
    """Update an existing row with fresh scan data.

    Rules:
    - Always update: RunID_LastSeen, SizeBytes, ModifiedTime, SHA256,
      IsDuplicateExact, DuplicateGroupID, ProcessingStatus.
    - Never overwrite if already populated: ManualReviewStatus,
      KeepForCase, PossibleExhibit.
    - Leave untouched: all classification/triage columns
      (written only by fw_classify / fw_triage).

    Parameters
    ----------
    ws : Worksheet
    row : int  — 1-based row number
    record : dict
    run_id : str
    """
    def _set(col_name: str, value) -> None:
        ws.cell(row=row, column=_COL_INDEX[col_name], value=value)

    def _current(col_name: str):
        return ws.cell(row=row, column=_COL_INDEX[col_name]).value

    # Always update
    _set("RunID_LastSeen",   run_id)
    _set("SizeBytes",        record.get("size_bytes", ""))
    _set("ModifiedTime",     record.get("modified_time", ""))
    _set("SHA256",           record.get("sha256", ""))
    _set("IsDuplicateExact", record.get("is_duplicate_exact", ""))
    _set("DuplicateGroupID", record.get("duplicate_group_id", ""))
    _set("ProcessingStatus", record.get("processing_status", "file_listed"))

    # Update ArchiveContents if a fresh peek was done
    if record.get("archive_contents"):
        _set("ArchiveContents", record["archive_contents"])

    # Never overwrite non-empty review fields
    for col_name in _NEVER_OVERWRITE_COLS:
        current_val = _current(col_name)
        if current_val is None or str(current_val).strip() == "":
            key     = _col_name_to_key(col_name)
            new_val = record.get(key, "")
            if new_val:
                _set(col_name, new_val)

    # Classification/triage columns: not touched here


# ---------------------------------------------------------------------------
# FUNCTION 6: write_or_update_file_record
# ---------------------------------------------------------------------------

def write_or_update_file_record(
    wb: Workbook, record: dict, run_id: str
) -> "tuple[str, int]":
    """Insert a new file record or update an existing one.

    Ensures the sheet exists, checks for an existing row by FilePath,
    and routes to insert or update accordingly.

    Parameters
    ----------
    wb : Workbook
    record : dict  — must contain "file_path"
    run_id : str

    Returns
    -------
    tuple[str, int]
        ("inserted", row_number) or ("updated", row_number)

    Raises
    ------
    ValueError
        If record does not contain a non-empty "file_path".
    """
    file_path = record.get("file_path", "")
    if not file_path:
        raise ValueError("record must contain a non-empty 'file_path' key.")

    ws = ensure_master_file_inventory_sheet(wb)

    existing_row = find_existing_file_row(ws, file_path)
    if existing_row is not None:
        update_file_record(ws, existing_row, record, run_id)
        return ("updated", existing_row)

    record = dict(record)  # don't mutate caller's dict
    record["file_id"] = get_next_file_id(ws)
    inserted_row = insert_file_record(ws, record, run_id)
    return ("inserted", inserted_row)


# ---------------------------------------------------------------------------
# FUNCTION 7: ensure_file_family_config_sheet
# ---------------------------------------------------------------------------

# Default family config rows: (Family, Extensions, ShouldSkip,
#   LikelyTextBearing, LikelyImage, LikelySpreadsheet, LikelyDocument, Notes)
_FAMILY_CONFIG_DEFAULTS = [
    ("pdf",          ".pdf",
     "N", "Y", "N", "N", "Y", ""),
    ("word_doc",     ".doc;.docx;.odt;.rtf;.dot;.dotx",
     "N", "Y", "N", "N", "Y", ""),
    ("spreadsheet",  ".xls;.xlsx;.xlsm;.xlsb;.csv;.ods;.numbers",
     "N", "Y", "N", "Y", "N", ""),
    ("text_file",    ".txt;.md;.log;.json;.xml;.html;.htm;.yaml;.yml;.ini;.toml;.cfg",
     "N", "Y", "N", "N", "N", ""),
    ("presentation", ".ppt;.pptx;.odp;.key",
     "N", "Y", "N", "N", "Y", ""),
    ("image",        ".jpg;.jpeg;.png;.gif;.bmp;.tiff;.tif;.heic;.heif;.webp;.svg;.raw;.cr2;.nef;.arw;.dng;.ico",
     "N", "N", "Y", "N", "N", ""),
    ("email_file",   ".eml;.msg;.pst;.ost;.mbox",
     "Y", "N", "N", "N", "N", "Processed by separate email pipeline"),
    ("archive",      ".zip;.rar;.7z;.tar;.gz;.bz2;.xz;.cab",
     "Y", "N", "N", "N", "N", "Content description captured in ArchiveContents column"),
    ("audio",        ".mp3;.wav;.m4a;.flac;.aac;.ogg;.wma;.aiff",
     "N", "N", "N", "N", "N", ""),
    ("video",        ".mp4;.mov;.avi;.mkv;.wmv;.flv;.m4v;.mpg;.mpeg;.webm",
     "N", "N", "N", "N", "N", ""),
]

_FC_SHEET_NAME = "FileFamily_Config"

_FC_COLUMNS = [
    ("Family",           18),
    ("Extensions",       80),   # semicolon-delimited
    ("ShouldSkip",       11),   # Y / N
    ("LikelyTextBearing", 16),
    ("LikelyImage",      13),
    ("LikelySpreadsheet", 16),
    ("LikelyDocument",   15),
    ("Notes",            50),
]


def ensure_file_family_config_sheet(wb: Workbook) -> Worksheet:
    """Create and populate the FileFamily_Config sheet if absent; return it.

    Idempotent — returns the existing sheet unchanged if it already exists.

    The sheet uses the same header style as Master_File_Inventory (bold white
    on dark blue, centered, frozen row 1, auto-filter). On first creation,
    all built-in default families are written as data rows.

    The sheet is intended to be human-editable: operators can add new rows,
    add semicolon-separated extensions, or flip ShouldSkip/signal columns
    without touching code.  fw_walk_grp1.load_file_family_config() reads it
    at runtime.

    Parameters
    ----------
    wb : Workbook

    Returns
    -------
    Worksheet
    """
    if _FC_SHEET_NAME in wb.sheetnames:
        return wb[_FC_SHEET_NAME]

    ws = wb.create_sheet(title=_FC_SHEET_NAME)

    # --- Header row ---
    for col_idx, (col_name, col_width) in enumerate(_FC_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_FC_COLUMNS))}1"

    # --- Default data rows ---
    for row_idx, row_data in enumerate(_FAMILY_CONFIG_DEFAULTS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    return ws
